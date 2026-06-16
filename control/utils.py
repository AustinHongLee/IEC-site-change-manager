# -*- coding: utf-8 -*-
"""
utils.py — 工具函數模組

包含：
- 檔案操作工具
- 指紋計算
- 完整性檢查
- PDF 合併
- 錯誤記錄
"""

import os
import re
import hashlib
import time
from typing import List, Optional
from dataclasses import dataclass

from openpyxl import load_workbook
from zipfile import BadZipFile

from config import RUNTIME


# ========= 欄位名稱同義字匹配 =========
_COL_SYNONYMS = {"焊": "銲", "銲": "焊"}


def _normalize_col(s: str) -> str:
    """去除空白/換行並轉小寫"""
    return s.replace(" ", "").replace("\n", "").lower()


def _synonym_variants(s: str) -> list[str]:
    """產生同義字替換的所有變體（如 焊→銲）"""
    variants = [s]
    for orig, repl in _COL_SYNONYMS.items():
        if orig in s:
            variants.append(s.replace(orig, repl))
    return variants


def resolve_col(name: str, actual_keys) -> str:
    """
    將設定欄位名對應到 Excel 實際欄位名。
    依序嘗試：完全一致 → 正規化後一致 → 同義字替換 → 子字串包含。

    Args:
        name: 設定中的欄位名（如 "焊口編號"）
        actual_keys: Excel 實際欄位名集合（set / list / dict.keys()）

    Returns:
        匹配到的實際欄位名；若找不到則原樣回傳 name
    """
    if name in actual_keys:
        return name
    norm = _normalize_col(name)
    for k in actual_keys:
        if _normalize_col(k) == norm:
            return k
    for variant in _synonym_variants(norm):
        for k in actual_keys:
            if _normalize_col(k) == variant:
                return k
    for k in actual_keys:
        kn = _normalize_col(k)
        if norm in kn or kn in norm:
            return k
    return name


def resolve_col_map(col_map: dict, name: str) -> int | None:
    """
    在 col_map ({header: index}) 中以 fuzzy 方式找欄位索引。

    Returns:
        index (int) 或 None
    """
    resolved = resolve_col(name, col_map.keys())
    return col_map.get(resolved)


# ========= 檔案操作工具 =========
def safe_load_workbook(path: str, max_retries: int = 20, delay: float = 0.3, **kwargs):
    """安全載入 Excel（自動重試）"""
    for attempt in range(max_retries):
        try:
            return load_workbook(path, **kwargs)
        except (BadZipFile, PermissionError, OSError):
            time.sleep(delay)
    return load_workbook(path, **kwargs)


def atomic_save_wb(wb, path: str):
    """原子性儲存 Excel（避免損壞）"""
    tmp = path + ".tmp"
    wb.save(tmp)
    os.replace(tmp, path)


def safe_remove(path: str, retries: int = 3, delay: float = 0.3) -> bool:
    """安全刪除檔案"""
    for _ in range(retries):
        try:
            if os.path.exists(path):
                os.remove(path)
            return True
        except PermissionError:
            time.sleep(delay)
    return False


def wait_for_stable_file(path: str, tries: int = 30, interval: float = 0.2) -> bool:
    """等待檔案穩定（不再變動）"""
    last = -1
    for _ in range(tries):
        if os.path.exists(path):
            size = os.path.getsize(path)
            if size > 0 and size == last:
                return True
            last = size
        time.sleep(interval)
    return os.path.exists(path)


# ========= 完整性檢查 =========
def _file_ok_min_kb(path: str, kb: float) -> bool:
    """檢查檔案存在且大於指定 KB"""
    try:
        return os.path.exists(path) and (os.path.getsize(path) >= kb * 1024)
    except Exception:
        return False


def _pdf_readable_has_pages(path: str) -> bool:
    """檢查 PDF 可讀且有頁面"""
    try:
        from pypdf import PdfReader
        r = PdfReader(path)
        return len(r.pages) >= 1
    except Exception:
        return False


def check_integrity(
    xlsm_path: Optional[str] = None,
    pdf_path: Optional[str] = None,
    level: int = 1,
    with_attachments: Optional[List[str]] = None
) -> bool:
    """
    分級完整性檢查
    
    Levels:
        0: 檢查檔案存在且大小足夠
        1: +PDF 可讀取且有頁面
        2: +附件 PDF 也要可讀
    """
    ok = True
    
    if xlsm_path:
        ok = ok and _file_ok_min_kb(xlsm_path, RUNTIME.min_xlsm_kb)
    
    if pdf_path:
        ok = ok and _file_ok_min_kb(pdf_path, RUNTIME.min_pdf_kb)
        if level >= 1:
            ok = ok and _pdf_readable_has_pages(pdf_path)
        if level >= 2 and with_attachments:
            for ap in with_attachments:
                ok = ok and _pdf_readable_has_pages(ap)
                if not ok:
                    break
    
    return ok


def check_output_integrity(
    output_file: str,
    pdf_file: str,
    min_size: int = 512,
    verbose: bool = True
) -> bool:
    """舊版完整性檢查（向後相容）"""
    ok = True
    
    if os.path.exists(output_file) and os.path.getsize(output_file) >= 1024:
        if verbose:
            print(f"✅ XLSM OK: {output_file}")
    else:
        if verbose:
            print(f"🚨 XLSM 不存在或過小：{output_file}")
        ok = False
    
    if os.path.exists(pdf_file) and os.path.getsize(pdf_file) >= max(512, min_size):
        if verbose:
            print(f"✅ PDF OK: {pdf_file}")
    else:
        if verbose:
            print(f"🚨 PDF 不存在或過小：{pdf_file}")
        ok = False
    
    if not ok and verbose:
        print("💡 Suggestion: Manually fix or delete record to force redo.")
    
    return ok


# ========= 錯誤記錄 =========
def write_error_marker(folder_path: str, err: str):
    """寫入錯誤標記檔案"""
    try:
        p = os.path.join(folder_path, "_ERROR.txt")
        with open(p, "a", encoding="utf-8") as f:
            f.write(f"[{time.strftime('%Y-%m-%d %H:%M:%S')}] {err}\n")
    except Exception:
        pass


def clear_error_marker(folder_path: str) -> bool:
    """清除錯誤標記檔案"""
    try:
        p = os.path.join(folder_path, "_ERROR.txt")
        if os.path.exists(p):
            os.remove(p)
        return True
    except Exception:
        return False


# ========= PDF 處理 =========
def find_attachment_pdf(folder_path: str, series_no: str) -> Optional[str]:
    """尋找附件 PDF"""
    try:
        pdfs = [
            os.path.join(folder_path, f)
            for f in os.listdir(folder_path)
            if f.lower().endswith(".pdf") and os.path.isfile(os.path.join(folder_path, f))
        ]
    except Exception:
        return None
    
    if not pdfs:
        return None
    
    series_variants = {series_no, series_no.lstrip('0')}
    candidates = [p for p in pdfs if any(v in os.path.basename(p) for v in series_variants)]
    return sorted(candidates or pdfs)[0]


def copy_prefab_pdf(folder_path: str, series_no: str, prefab_dir: str = "") -> Optional[str]:
    """從預製圖目錄複製符合 series_no 的 PDF 到 folder_path。

    PDF 命名慣例: ``{series_no}.DW-xxxx-xx-xxxx-xx-x.pdf``
    匹配規則: 檔名以 ``{series_no}.`` 開頭（不分大小寫）。

    Returns:
        複製後的檔案完整路徑，若無匹配或目錄未設定則回傳 None。
    """
    if not prefab_dir:
        from settings_manager import get_prefab_drawing_dir
        prefab_dir = get_prefab_drawing_dir()

    if not prefab_dir or not os.path.isdir(prefab_dir):
        return None

    if not series_no:
        return None

    prefix = f"{series_no}.".lower()
    try:
        matches = [
            f for f in os.listdir(prefab_dir)
            if f.lower().startswith(prefix) and f.lower().endswith(".pdf")
        ]
    except OSError:
        return None

    if not matches:
        return None

    # 取第一個符合的（通常只有一個）
    src_name = sorted(matches)[0]
    src_path = os.path.join(prefab_dir, src_name)
    dst_path = os.path.join(folder_path, src_name)

    # 若目標已存在且大小相同，不重複複製
    if os.path.exists(dst_path):
        try:
            if os.path.getsize(src_path) == os.path.getsize(dst_path):
                return dst_path
        except OSError:
            pass

    import shutil
    try:
        shutil.copy2(src_path, dst_path)
        return dst_path
    except OSError:
        return None


def merge_into_second_page(exported_pdf_path: str, attach_pdf_path: str):
    """將附件 PDF 合併到第二頁"""
    try:
        from pypdf import PdfReader, PdfWriter
    except ImportError:
        try:
            from PyPDF2 import PdfReader, PdfWriter
        except ImportError:
            raise RuntimeError("請先安裝 pypdf 或 PyPDF2：pip install pypdf")
    
    reader_export = PdfReader(exported_pdf_path)
    reader_attach = PdfReader(attach_pdf_path)
    
    writer = PdfWriter()
    if len(reader_export.pages) == 0:
        raise RuntimeError(f"匯出 PDF 無頁面：{exported_pdf_path}")
    
    writer.add_page(reader_export.pages[0])
    for p in reader_attach.pages:
        writer.add_page(p)
    for i in range(1, len(reader_export.pages)):
        writer.add_page(reader_export.pages[i])
    
    tmp_out = exported_pdf_path + ".tmp"
    with open(tmp_out, "wb") as f:
        writer.write(f)
    os.replace(tmp_out, exported_pdf_path)


# ========= 指紋計算 =========
def _file_content_hash(path: str) -> str:
    """計算檔案內容的 MD5（只讀前 64KB 加速）"""
    h = hashlib.md5()
    with open(path, "rb") as f:
        h.update(f.read(65536))
    return h.hexdigest()


# ========= 指紋計算 =========
def _fingerprint_groupweld(folder_path: str, parts: List[str]):
    """計算 GroupWeld.txt 的指紋"""
    p = os.path.join(folder_path, "GroupWeld.txt")
    if os.path.exists(p):
        try:
            with open(p, "rb") as f:
                data = f.read()
            parts.append(f"GroupWeld:{len(data)}")
            parts.append("GW_MD5:" + hashlib.md5(data).hexdigest())
        except Exception:
            parts.append("GroupWeld:ERR")
    else:
        parts.append("GroupWeld:NONE")


def compute_fingerprint(
    date_str: str,
    folder: str,
    series_no: str,
    suffixes_raw: List[str],
    note_text: str,
    materials_text: str,
    folder_path: str,
    is_group: bool = False,
    use_dual_images: bool = False
) -> str:
    """
    計算內容指紋（用於變更偵測）

    使用檔案大小 + 內容雜湊（前 64KB），不依賴 mtime，
    避免 Google Drive 同步或圖片預處理導致指紋不穩定。
    """
    parts = [date_str, folder, "||".join(suffixes_raw), note_text, materials_text]
    
    image_names = [
        "before_1.jpg", "before_2.jpg", "after_1.jpg", "after_2.jpg"
    ] if use_dual_images else [
        "before.jpg", "after.jpg"
    ]
    
    for imgname in image_names:
        p = os.path.join(folder_path, imgname)
        if os.path.exists(p):
            try:
                sz = os.path.getsize(p)
                ch = _file_content_hash(p)
                parts.append(f"{imgname}:{sz}:{ch}")
            except Exception:
                parts.append(f"{imgname}:ERR")
        else:
            parts.append(f"{imgname}:NONE")
    
    ap = find_attachment_pdf(folder_path, series_no)
    if ap and os.path.exists(ap):
        try:
            sz = os.path.getsize(ap)
            ch = _file_content_hash(ap)
            parts.append(f"attachpdf:{os.path.basename(ap)}:{sz}:{ch}")
        except Exception:
            parts.append(f"attachpdf:{os.path.basename(ap)}:ERR")
    else:
        parts.append("attachpdf:NONE")
    
    if is_group:
        _fingerprint_groupweld(folder_path, parts)
    
    blob = "||".join(parts).encode("utf-8")
    return hashlib.md5(blob).hexdigest()


# ========= 報告編號解析 =========
def parse_seq_from_report_id(report_id: str) -> Optional[int]:
    """從報告編號解析序號"""
    m = re.fullmatch(r'(\d{8})-(\d{2})', str(report_id))
    return int(m.group(2)) if m else None


# ========= 日期資料夾掃描 =========
def scan_date_folders(attachments_root: str) -> List[str]:
    """掃描日期資料夾（YYYYMMDD 格式）"""
    if not os.path.exists(attachments_root):
        return []
    
    date_dirs = [
        d for d in os.listdir(attachments_root)
        if os.path.isdir(os.path.join(attachments_root, d)) and re.fullmatch(r'\d{8}', d)
    ]
    date_dirs.sort()
    return date_dirs


def scan_subfolders(date_folder_path: str) -> List[str]:
    """掃描日期資料夾內的子資料夾"""
    if not os.path.exists(date_folder_path):
        return []
    
    return sorted(
        d for d in os.listdir(date_folder_path)
        if os.path.isdir(os.path.join(date_folder_path, d))
    )


# ========= 統計彙整 =========
@dataclass
class ProcessingSummary:
    """處理結果彙整"""
    total: int = 0
    success: int = 0
    skipped: int = 0
    failed: int = 0
    failed_list: List[str] = None
    
    def __post_init__(self):
        if self.failed_list is None:
            self.failed_list = []
    
    def add_success(self):
        self.total += 1
        self.success += 1
    
    def add_skipped(self):
        self.total += 1
        self.skipped += 1
    
    def add_failed(self, path: str):
        self.total += 1
        self.failed += 1
        self.failed_list.append(path)
    
    def print_summary(self):
        """印出執行摘要"""
        print("\n" + "═" * 50)
        print("📊 執行摘要")
        print("─" * 50)
        print(f"✅ 成功產出: {self.success}")
        print(f"⏭️ 略過（未變更）: {self.skipped}")
        print(f"❌ 失敗: {self.failed}")
        if self.failed_list:
            print("   失敗清單:")
            for p in self.failed_list[:10]:  # 最多顯示 10 筆
                print(f"   - {p}")
            if len(self.failed_list) > 10:
                print(f"   ... 還有 {len(self.failed_list) - 10} 筆")
        print("═" * 50)
