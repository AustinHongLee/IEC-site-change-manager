# -*- coding: utf-8 -*-
"""
xlsx_template_renderer.py - Render CanonicalReport data into an xlsx template.

This renderer implements the first concrete target for the template mapping
layer. It uses openpyxl only and never imports Excel COM.
"""

from __future__ import annotations

import os
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.cell import coordinate_to_tuple

from template_dry_run import dry_run_template_for_report
from template_mapping import resolve_field_path


EXCEL_MAX_ROW = 1048576
EXCEL_MAX_COL = 16384


def render_xlsx_template_for_report(
    report: dict[str, Any],
    template: dict[str, Any],
    output_path: str,
    *,
    template_dir: str | os.PathLike[str] | None = None,
) -> dict[str, Any]:
    dry_run = dry_run_template_for_report(report, template)
    if not dry_run["ok"]:
        return {
            "ok": False,
            "path": "",
            "dry_run": dry_run,
            "summary": {"text": 0, "image": 0, "table": 0, "rows": 0},
            "issues": list(dry_run.get("issues", [])),
        }
    layout_validation = validate_xlsx_template_layout(report, template)
    if not layout_validation["ok"]:
        return {
            "ok": False,
            "path": "",
            "dry_run": dry_run,
            "layout_validation": layout_validation,
            "summary": {"text": 0, "image": 0, "table": 0, "rows": 0},
            "issues": list(dry_run.get("issues", [])) + list(layout_validation.get("issues", [])),
        }

    wb = _load_template_workbook(template, template_dir=template_dir)
    ws = _resolve_sheet(wb, template)
    result = {
        "ok": True,
        "path": str(output_path),
        "dry_run": dry_run,
        "layout_validation": layout_validation,
        "summary": {"text": 0, "image": 0, "table": 0, "rows": 0},
        "issues": list(dry_run.get("issues", [])),
    }

    for idx, mapping in enumerate(template.get("fields", []) or [], start=1):
        mapping_type = str(mapping.get("type", "")).strip()
        if mapping_type == "text":
            _render_text(ws, report, mapping, idx, result)
        elif mapping_type == "image":
            _render_image(ws, report, mapping, idx, result)
        elif mapping_type == "table":
            _render_table(ws, report, mapping, idx, result)

    if not result["ok"]:
        result["path"] = ""
        wb.close()
        return result

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    _atomic_save_workbook(wb, output_path)
    wb.close()
    post_validation = validate_rendered_xlsx_workbook(output_path, report, template)
    result["post_validation"] = post_validation
    result["issues"].extend(post_validation["issues"])
    if not post_validation["ok"]:
        result["ok"] = False
    return result


def validate_xlsx_template_layout(report: dict[str, Any], template: dict[str, Any]) -> dict[str, Any]:
    result = {
        "ok": True,
        "regions": [],
        "issues": [],
    }
    for idx, mapping in enumerate(template.get("fields", []) or [], start=1):
        region = _mapping_region(mapping, idx)
        if region.get("error"):
            _add_layout_issue(
                result,
                "error",
                region["error"],
                idx,
                mapping,
                region["message"],
            )
            continue
        if not region:
            continue
        result["regions"].append(region)
        if not _region_in_bounds(region):
            _add_layout_issue(
                result,
                "error",
                "layout_out_of_bounds",
                idx,
                mapping,
                f"{region['label']} 超出 Excel 工作表邊界",
            )

    regions = result["regions"]
    for left_idx, left in enumerate(regions):
        for right in regions[left_idx + 1:]:
            if _regions_overlap(left, right):
                _add_layout_issue(
                    result,
                    "error",
                    "layout_overlap",
                    left["field_index"],
                    left["source"],
                    f"{left['label']} 與 {right['label']} 重疊",
                )
    result["ok"] = not any(issue.get("severity") == "error" for issue in result["issues"])
    return result


def validate_rendered_xlsx_workbook(
    workbook_path: str,
    report: dict[str, Any],
    template: dict[str, Any],
) -> dict[str, Any]:
    result = {
        "ok": True,
        "checked": {"text": 0, "image": 0, "table": 0, "rows": 0},
        "issues": [],
    }
    try:
        wb = load_workbook(workbook_path, data_only=False)
    except Exception as exc:
        return {
            "ok": False,
            "checked": {"text": 0, "image": 0, "table": 0, "rows": 0},
            "issues": [{
                "severity": "error",
                "code": "post_validation_load_failed",
                "field_index": 0,
                "source": "",
                "message": f"輸出後無法重新讀取 workbook：{exc}",
            }],
        }

    try:
        ws = _resolve_sheet_for_validation(wb, template)
        if ws is None:
            _add_post_issue(result, "error", "post_validation_sheet_missing", 0, "", "輸出後找不到目標工作表")
            return result
        for idx, mapping in enumerate(template.get("fields", []) or [], start=1):
            mapping_type = str(mapping.get("type", "")).strip()
            if mapping_type == "text":
                _validate_text(ws, report, mapping, idx, result)
            elif mapping_type == "image":
                _validate_image(ws, report, mapping, idx, result)
            elif mapping_type == "table":
                _validate_table(ws, report, mapping, idx, result)
    finally:
        wb.close()
    result["ok"] = not any(issue.get("severity") == "error" for issue in result["issues"])
    return result


def _load_template_workbook(template: dict[str, Any], *, template_dir: str | os.PathLike[str] | None):
    workbook_path = str(template.get("workbook", "") or "").strip()
    if not workbook_path:
        return Workbook()
    path = Path(workbook_path)
    if not path.is_absolute() and template_dir:
        path = Path(template_dir) / path
    return load_workbook(path)


def _resolve_sheet(wb, template: dict[str, Any]):
    sheet_name = str(template.get("sheet", "") or "").strip()
    if not sheet_name:
        return wb.active
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    return wb.create_sheet(sheet_name)


def _resolve_sheet_for_validation(wb, template: dict[str, Any]):
    sheet_name = str(template.get("sheet", "") or "").strip()
    if not sheet_name:
        return wb.active
    return wb[sheet_name] if sheet_name in wb.sheetnames else None


def _mapping_region(mapping: dict[str, Any], idx: int) -> dict[str, Any]:
    mapping_type = str(mapping.get("type", "")).strip()
    source = str(mapping.get("source", "") or "")
    if mapping_type == "text":
        cell = str(mapping.get("cell", "") or "").strip()
        if not cell:
            return {"error": "layout_missing_cell", "message": "text mapping 缺少 cell"}
        return _region_from_cell(cell, 1, 1, idx, mapping_type, source)
    if mapping_type == "image":
        anchor = str(mapping.get("anchor") or mapping.get("cell") or "").strip()
        if not anchor:
            return {"error": "layout_missing_anchor", "message": "image mapping 缺少 anchor"}
        cols, rows = _image_size_cells(mapping)
        if cols <= 0 or rows <= 0:
            return {"error": "layout_invalid_size_cells", "message": "image size_cells 必須是正整數 [cols, rows]"}
        return _region_from_cell(anchor, rows, cols, idx, mapping_type, source)
    if mapping_type == "table":
        start_cell = str(mapping.get("start_cell", "") or "").strip()
        if not start_cell:
            return {"error": "layout_missing_start_cell", "message": "table mapping 缺少 start_cell"}
        max_rows = _positive_int(mapping.get("max_rows") or mapping.get("rows_per_page"))
        columns = mapping.get("columns", []) or []
        row_count = max_rows + (1 if bool(mapping.get("write_header", False)) else 0)
        col_count = max(1, len(columns))
        return _region_from_cell(start_cell, row_count, col_count, idx, mapping_type, source)
    return {}


def _region_from_cell(
    cell: str,
    row_span: int,
    col_span: int,
    idx: int,
    mapping_type: str,
    source: str,
) -> dict[str, Any]:
    try:
        row, col = coordinate_to_tuple(cell)
    except ValueError:
        return {"error": "layout_invalid_cell", "message": f"無效儲存格座標：{cell}"}
    row_span = max(1, int(row_span))
    col_span = max(1, int(col_span))
    row2 = row + row_span - 1
    col2 = col + col_span - 1
    label = f"fields[{idx}] {mapping_type} {cell}:{_cell_label(row2, col2)}"
    return {
        "field_index": idx,
        "type": mapping_type,
        "source": source,
        "row1": row,
        "col1": col,
        "row2": row2,
        "col2": col2,
        "label": label,
    }


def _image_size_cells(mapping: dict[str, Any]) -> tuple[int, int]:
    value = mapping.get("size_cells")
    if value is None:
        return 1, 1
    if not isinstance(value, (list, tuple)) or len(value) != 2:
        return 0, 0
    cols = _positive_int(value[0])
    rows = _positive_int(value[1])
    return cols, rows


def _region_in_bounds(region: dict[str, Any]) -> bool:
    return (
        1 <= region["row1"] <= region["row2"] <= EXCEL_MAX_ROW
        and 1 <= region["col1"] <= region["col2"] <= EXCEL_MAX_COL
    )


def _regions_overlap(left: dict[str, Any], right: dict[str, Any]) -> bool:
    return not (
        left["row2"] < right["row1"]
        or right["row2"] < left["row1"]
        or left["col2"] < right["col1"]
        or right["col2"] < left["col1"]
    )


def _cell_label(row: int, col: int) -> str:
    from openpyxl.utils.cell import get_column_letter
    return f"{get_column_letter(col)}{row}"


def _render_text(ws, report: dict[str, Any], mapping: dict[str, Any], idx: int, result: dict[str, Any]) -> None:
    cell = str(mapping.get("cell", "") or "").strip()
    if not cell:
        _add_issue(result, "error", "missing_text_cell", idx, mapping, "text mapping 缺少 cell")
        result["ok"] = False
        return
    value = resolve_field_path(report, str(mapping.get("source", "")).strip())
    ws[cell] = _cell_value(value)
    result["summary"]["text"] += 1


def _render_image(ws, report: dict[str, Any], mapping: dict[str, Any], idx: int, result: dict[str, Any]) -> None:
    anchor = str(mapping.get("anchor") or mapping.get("cell") or "").strip()
    if not anchor:
        _add_issue(result, "error", "missing_image_anchor", idx, mapping, "image mapping 缺少 anchor")
        result["ok"] = False
        return

    value = resolve_field_path(report, str(mapping.get("source", "")).strip())
    path = _image_path(value)
    if not path:
        ws[anchor] = "缺圖片"
        _add_issue(result, "warning", "missing_image_value", idx, mapping, "圖片欄位沒有路徑")
        result["summary"]["image"] += 1
        return
    if not os.path.exists(path):
        ws[anchor] = "找不到圖片"
        _add_issue(result, "warning", "missing_image_file", idx, mapping, f"圖片檔不存在：{path}")
        result["summary"]["image"] += 1
        return

    try:
        img = XLImage(path)
    except (ImportError, OSError, ValueError):
        ws[anchor] = "無法讀取圖片"
        _add_issue(result, "warning", "unreadable_image_file", idx, mapping, f"無法讀取圖片：{path}")
        result["summary"]["image"] += 1
        return

    max_width = _positive_int(mapping.get("max_width_px")) or 260
    max_height = _positive_int(mapping.get("max_height_px")) or 180
    img.width, img.height = _fit_image_size(img.width, img.height, max_width, max_height)
    ws.add_image(img, anchor)
    result["summary"]["image"] += 1


def _render_table(ws, report: dict[str, Any], mapping: dict[str, Any], idx: int, result: dict[str, Any]) -> None:
    start_cell = str(mapping.get("start_cell", "") or "").strip()
    if not start_cell:
        _add_issue(result, "error", "missing_table_start_cell", idx, mapping, "table mapping 缺少 start_cell")
        result["ok"] = False
        return
    rows = resolve_field_path(report, str(mapping.get("source", "")).strip(), default=[])
    rows = rows if isinstance(rows, list) else []
    row_limit = _positive_int(mapping.get("max_rows") or mapping.get("rows_per_page"))
    if row_limit and len(rows) > row_limit:
        _add_issue(
            result,
            "error",
            "table_overflow",
            idx,
            mapping,
            f"表格資料 {len(rows)} 列超過預留 {row_limit} 列，已停止寫過界",
        )
        result["ok"] = False
        rows = rows[:row_limit]
    columns = [_column_source(column) for column in mapping.get("columns", []) or []]
    start_row, start_col = coordinate_to_tuple(start_cell)
    write_header = bool(mapping.get("write_header", False))

    row_offset = 0
    if write_header:
        for col_offset, column in enumerate(mapping.get("columns", []) or []):
            ws.cell(
                row=start_row,
                column=start_col + col_offset,
                value=_column_header(column),
            )
        row_offset = 1

    for row_idx, row_data in enumerate(rows, start=0):
        for col_offset, column_source in enumerate(columns):
            value = resolve_field_path(row_data, column_source)
            ws.cell(
                row=start_row + row_offset + row_idx,
                column=start_col + col_offset,
                value=_cell_value(value),
            )
    result["summary"]["table"] += 1
    result["summary"]["rows"] += len(rows)


def _validate_text(ws, report: dict[str, Any], mapping: dict[str, Any], idx: int, result: dict[str, Any]) -> None:
    cell = str(mapping.get("cell", "") or "").strip()
    if not cell:
        return
    expected = _cell_value(resolve_field_path(report, str(mapping.get("source", "")).strip()))
    actual = ws[cell].value
    result["checked"]["text"] += 1
    if not _values_match(expected, actual):
        _add_post_issue(
            result,
            "error",
            "post_validation_text_mismatch",
            idx,
            mapping,
            f"{cell} 預期「{expected}」，實際「{actual}」",
        )


def _validate_image(ws, report: dict[str, Any], mapping: dict[str, Any], idx: int, result: dict[str, Any]) -> None:
    anchor = str(mapping.get("anchor") or mapping.get("cell") or "").strip()
    if not anchor:
        return
    value = resolve_field_path(report, str(mapping.get("source", "")).strip())
    placeholder = _expected_image_placeholder(value)
    result["checked"]["image"] += 1
    if placeholder:
        actual = ws[anchor].value
        if actual != placeholder:
            _add_post_issue(
                result,
                "error",
                "post_validation_image_placeholder_mismatch",
                idx,
                mapping,
                f"{anchor} 預期「{placeholder}」，實際「{actual}」",
            )
        return
    if not _worksheet_has_image_at(ws, anchor):
        _add_post_issue(
            result,
            "error",
            "post_validation_image_missing",
            idx,
            mapping,
            f"輸出後沒有在 {anchor} 找到圖片",
        )


def _validate_table(ws, report: dict[str, Any], mapping: dict[str, Any], idx: int, result: dict[str, Any]) -> None:
    start_cell = str(mapping.get("start_cell", "") or "").strip()
    if not start_cell:
        return
    rows = resolve_field_path(report, str(mapping.get("source", "")).strip(), default=[])
    rows = rows if isinstance(rows, list) else []
    columns = [_column_source(column) for column in mapping.get("columns", []) or []]
    start_row, start_col = coordinate_to_tuple(start_cell)
    row_offset = 1 if bool(mapping.get("write_header", False)) else 0
    result["checked"]["table"] += 1
    result["checked"]["rows"] += len(rows)
    for row_idx, row_data in enumerate(rows, start=0):
        for col_offset, column_source in enumerate(columns):
            expected = _cell_value(resolve_field_path(row_data, column_source))
            cell = ws.cell(row=start_row + row_offset + row_idx, column=start_col + col_offset)
            if not _values_match(expected, cell.value):
                _add_post_issue(
                    result,
                    "error",
                    "post_validation_table_mismatch",
                    idx,
                    mapping,
                    f"{cell.coordinate} 預期「{expected}」，實際「{cell.value}」",
                )


def _image_path(value: Any) -> str:
    if isinstance(value, dict):
        return str(value.get("path", "") or "").strip()
    return str(value or "").strip()


def _cell_value(value: Any) -> Any:
    if isinstance(value, (str, int, float)) or value is None:
        return value
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, list):
        return "、".join(str(item) for item in value)
    if isinstance(value, dict):
        return str(value)
    return value


def _values_match(expected: Any, actual: Any) -> bool:
    if expected is None:
        return actual in (None, "")
    if isinstance(expected, (int, float)) and isinstance(actual, (int, float)):
        return float(expected) == float(actual)
    return str(expected) == str(actual)


def _expected_image_placeholder(value: Any) -> str:
    path = _image_path(value)
    if not path:
        return "缺圖片"
    if not os.path.exists(path):
        return "找不到圖片"
    try:
        XLImage(path)
    except (ImportError, OSError, ValueError):
        return "無法讀取圖片"
    return ""


def _worksheet_has_image_at(ws, anchor: str) -> bool:
    expected_row, expected_col = coordinate_to_tuple(anchor)
    for image in getattr(ws, "_images", []) or []:
        image_anchor = image.anchor
        if isinstance(image_anchor, str):
            row, col = coordinate_to_tuple(image_anchor)
            if row == expected_row and col == expected_col:
                return True
            continue
        marker = getattr(image_anchor, "_from", None)
        if marker and marker.row + 1 == expected_row and marker.col + 1 == expected_col:
            return True
    return False


def _column_source(column: Any) -> str:
    if isinstance(column, str):
        return column.strip()
    if isinstance(column, dict):
        return str(column.get("source", "") or "").strip()
    return ""


def _column_header(column: Any) -> str:
    if isinstance(column, str):
        return column.strip()
    if isinstance(column, dict):
        return str(column.get("header") or column.get("source") or "").strip()
    return ""


def _positive_int(value: Any) -> int:
    try:
        number = int(value)
    except (TypeError, ValueError):
        return 0
    return number if number > 0 else 0


def _fit_image_size(width: Any, height: Any, max_width: int, max_height: int) -> tuple[int, int]:
    try:
        width_f = float(width)
        height_f = float(height)
    except (TypeError, ValueError):
        return max_width, max_height
    if width_f <= 0 or height_f <= 0:
        return max_width, max_height
    scale = min(max_width / width_f, max_height / height_f, 1.0)
    return max(1, int(width_f * scale)), max(1, int(height_f * scale))


def _add_issue(
    result: dict[str, Any],
    severity: str,
    code: str,
    idx: int,
    mapping: dict[str, Any],
    message: str,
) -> None:
    result["issues"].append({
        "severity": severity,
        "code": code,
        "field_index": idx,
        "source": str(mapping.get("source", "") or ""),
        "message": message,
    })


def _add_post_issue(
    result: dict[str, Any],
    severity: str,
    code: str,
    idx: int,
    mapping: dict[str, Any] | str,
    message: str,
) -> None:
    source = mapping if isinstance(mapping, str) else str(mapping.get("source", "") or "")
    result["issues"].append({
        "severity": severity,
        "code": code,
        "field_index": idx,
        "source": source,
        "message": message,
    })


def _add_layout_issue(
    result: dict[str, Any],
    severity: str,
    code: str,
    idx: int,
    mapping: dict[str, Any] | str,
    message: str,
) -> None:
    source = mapping if isinstance(mapping, str) else str(mapping.get("source", "") or "")
    result["issues"].append({
        "severity": severity,
        "code": code,
        "field_index": idx,
        "source": source,
        "message": message,
    })


def _atomic_save_workbook(wb, path: str) -> None:
    tmp = path + ".tmp"
    wb.save(tmp)
    os.replace(tmp, path)
