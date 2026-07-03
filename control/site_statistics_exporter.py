# -*- coding: utf-8 -*-
"""
site_statistics_exporter.py - 現場修改統計單匯出

輸出重點是「現場資料統計」，不是請款或報價。
工作簿分成兩區：
- 開發者看的：資料總覽、原始清單、照片索引、問題清單
- 要繳出去的報告：總覽、修改單清單、焊口統計、用料統計、照片表
"""

from __future__ import annotations

import os
from collections import defaultdict
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from canonical_report import collect_canonical_report_set
from config import BASE_DIR
from utils import atomic_save_wb


SITE_STATISTICS_FILENAME_PREFIX = "現場修改統計單"

WORKBOOK_SECTIONS = [
    {
        "title": "開發者看的",
        "tab_color": "64748B",
        "sheets": [
            ("開發_資料總覽", "內部檢查用彙總，包含專案路徑、完整度與問題數。"),
            ("開發_修改單原始清單", "保留資料夾路徑與缺漏欄位，方便追查來源。"),
            ("開發_照片索引", "照片檔案索引與本機路徑，方便追查圖片來源。"),
            ("開發_問題清單", "資料缺漏、解析警示與需回頭修正的項目。"),
        ],
    },
    {
        "title": "要繳出去的報告",
        "tab_color": "1860AB",
        "sheets": [
            ("報告_總覽", "交付用統計封面，不顯示本機路徑。"),
            ("報告_修改單清單", "交付用修改單明細。"),
            ("報告_焊口統計", "交付用焊口彙總。"),
            ("報告_用料統計", "交付用材料彙總。"),
            ("報告_照片表", "交付用 before / after 照片表。"),
        ],
    },
]

WORKBOOK_SHEET_ORDER = [
    "目錄",
    *(sheet for section in WORKBOOK_SECTIONS for sheet, _description in section["sheets"]),
]


def export_site_statistics_workbook(
    output_path: str | None = None,
    *,
    report_set: dict[str, Any] | None = None,
    project_root: str | None = None,
    attachments_root: str | None = None,
    store: dict[str, Any] | None = None,
) -> str:
    if report_set is None:
        report_set = collect_canonical_report_set(
            project_root=project_root,
            attachments_root=attachments_root,
            store=store,
        )
    if output_path is None:
        output_path = os.path.join(
            BASE_DIR,
            "records",
            f"{SITE_STATISTICS_FILENAME_PREFIX}_{datetime.now().strftime('%Y%m%d')}.xlsx",
        )
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

    wb = Workbook()
    _write_index_sheet(wb.active, report_set)

    _write_overview_sheet(
        wb.create_sheet("開發_資料總覽"),
        report_set,
        sheet_title="開發_資料總覽",
        table_title="開發者檢查 - 資料總覽",
    )
    _write_report_list_sheet(
        wb.create_sheet("開發_修改單原始清單"),
        report_set,
        title="開發者檢查 - 修改單原始清單",
    )
    _write_photo_index_sheet(
        wb.create_sheet("開發_照片索引"),
        report_set,
        title="開發者檢查 - 照片索引",
    )
    _write_issue_sheet(
        wb.create_sheet("開發_問題清單"),
        report_set,
        title="開發者檢查 - 問題清單",
    )

    _write_deliverable_overview_sheet(wb.create_sheet("報告_總覽"), report_set)
    _write_deliverable_report_list_sheet(wb.create_sheet("報告_修改單清單"), report_set)
    _write_weld_summary_sheet(wb.create_sheet("報告_焊口統計"), report_set, title="報告 - 焊口統計")
    _write_material_summary_sheet(wb.create_sheet("報告_用料統計"), report_set, title="報告 - 用料統計")
    _write_photo_sheet(
        wb.create_sheet("報告_照片表"),
        report_set,
        title="報告 - 照片表",
        include_path_columns=False,
    )
    _apply_workbook_tab_colors(wb)

    atomic_save_wb(wb, output_path)
    wb.close()
    return output_path


def build_overview_rows(report_set: dict[str, Any]) -> list[list[Any]]:
    agg = report_set.get("aggregates", {})
    completeness = agg.get("completeness_counts", {}) or {}
    statuses = agg.get("status_counts", {}) or {}
    issues = report_set.get("issues", []) or []
    return [
        ["統計時間", report_set.get("project", {}).get("collected_at", ""), ""],
        ["專案資料夾", report_set.get("project", {}).get("root", ""), ""],
        ["修改單總數", agg.get("report_count", 0), ""],
        ["已產出", statuses.get("produced", 0), ""],
        ["未產出", statuses.get("unproduced", 0), ""],
        ["需重產", statuses.get("needs_rebuild", 0), ""],
        ["完整", completeness.get("complete", 0), ""],
        ["不完整", completeness.get("incomplete", 0), ""],
        ["空資料", completeness.get("empty", 0), ""],
        ["焊口總數", agg.get("weld_count", 0), ""],
        ["用料明細筆數", agg.get("material_row_count", 0), ""],
        ["before 照片", agg.get("before_photo_count", 0), ""],
        ["after 照片", agg.get("after_photo_count", 0), ""],
        ["問題筆數", len(issues), ""],
    ]


def build_deliverable_overview_rows(report_set: dict[str, Any]) -> list[list[Any]]:
    agg = report_set.get("aggregates", {}) or {}
    statuses = agg.get("status_counts", {}) or {}
    before_count = agg.get("before_photo_count", 0)
    after_count = agg.get("after_photo_count", 0)
    return [
        ["統計時間", report_set.get("project", {}).get("collected_at", ""), ""],
        ["修改單總數", agg.get("report_count", 0), ""],
        ["已產出", statuses.get("produced", 0), ""],
        ["未產出", statuses.get("unproduced", 0), ""],
        ["需重產", statuses.get("needs_rebuild", 0), ""],
        ["焊口總數", agg.get("weld_count", 0), ""],
        ["用料明細筆數", agg.get("material_row_count", 0), ""],
        ["照片張數", (before_count or 0) + (after_count or 0), f"before {before_count} / after {after_count}"],
    ]


def flatten_report_rows(report_set: dict[str, Any]) -> list[list[Any]]:
    rows = []
    for report in report_set.get("reports", []) or []:
        info = report.get("report", {})
        completeness = report.get("completeness", {})
        rows.append([
            info.get("report_id", ""),
            info.get("date", ""),
            info.get("series", ""),
            info.get("folder", ""),
            _status_label(info.get("status", "")),
            info.get("change_type", ""),
            report.get("welds", {}).get("summary", ""),
            report.get("welds", {}).get("count", 0),
            report.get("materials", {}).get("count", 0),
            len(report.get("photos", {}).get("before", []) or []),
            len(report.get("photos", {}).get("after", []) or []),
            "有" if report.get("attachment_pdf", {}).get("exists") else "無",
            completeness.get("level", ""),
            "、".join(completeness.get("missing", []) or []),
            info.get("description", ""),
            info.get("folder_path", ""),
        ])
    return rows


def flatten_deliverable_report_rows(report_set: dict[str, Any]) -> list[list[Any]]:
    rows = []
    for report in report_set.get("reports", []) or []:
        info = report.get("report", {})
        rows.append([
            info.get("report_id", ""),
            info.get("date", ""),
            info.get("series", ""),
            info.get("folder", ""),
            _status_label(info.get("status", "")),
            info.get("change_type", ""),
            report.get("welds", {}).get("summary", ""),
            report.get("welds", {}).get("count", 0),
            report.get("materials", {}).get("count", 0),
            len(report.get("photos", {}).get("before", []) or []),
            len(report.get("photos", {}).get("after", []) or []),
            "有" if report.get("attachment_pdf", {}).get("exists") else "無",
            info.get("description", ""),
        ])
    return rows


def flatten_weld_summary_rows(report_set: dict[str, Any]) -> list[list[Any]]:
    rows = []
    by_size = defaultdict(lambda: {"count": 0, "reports": set(), "total_size": 0.0})
    for report in report_set.get("reports", []) or []:
        report_label = report.get("report", {}).get("report_id") or report.get("report", {}).get("folder", "")
        for weld in report.get("welds", {}).get("rows", []) or []:
            size = str(weld.get("size", "")).strip() or "未填"
            by_size[size]["count"] += 1
            by_size[size]["reports"].add(report_label)
            try:
                by_size[size]["total_size"] += float(weld.get("size") or 0)
            except (TypeError, ValueError):
                pass
    for size, item in sorted(by_size.items(), key=lambda kv: str(kv[0])):
        rows.append([
            size,
            item["count"],
            item["total_size"] if item["total_size"] else "",
            len(item["reports"]),
            "、".join(sorted(item["reports"])),
        ])
    return rows


def flatten_photo_index_rows(report_set: dict[str, Any]) -> list[list[Any]]:
    rows = []
    for report in report_set.get("reports", []) or []:
        info = report.get("report", {})
        report_label = info.get("report_id") or info.get("folder", "")
        for kind in ("before", "after"):
            for idx, photo in enumerate(report.get("photos", {}).get(kind, []) or [], start=1):
                rows.append([
                    report_label,
                    info.get("date", ""),
                    info.get("series", ""),
                    info.get("folder", ""),
                    kind,
                    idx,
                    photo.get("name", ""),
                    photo.get("w", ""),
                    photo.get("h", ""),
                    photo.get("path", ""),
                ])
    return rows


def flatten_material_summary_rows(report_set: dict[str, Any]) -> list[list[Any]]:
    rows = []
    grouped = defaultdict(lambda: {"qty": 0.0, "reports": set()})
    for report in report_set.get("reports", []) or []:
        report_label = report.get("report", {}).get("report_id") or report.get("report", {}).get("folder", "")
        for material in report.get("materials", {}).get("rows", []) or []:
            key = (
                material.get("component", ""),
                material.get("size", ""),
                material.get("sch", ""),
                material.get("material", ""),
                material.get("unit", ""),
                material.get("category", ""),
            )
            grouped[key]["qty"] += _to_float(material.get("qty"))
            grouped[key]["reports"].add(report_label)
    for key, item in sorted(grouped.items(), key=lambda kv: tuple(str(part) for part in kv[0])):
        rows.append([
            key[0],
            key[1],
            key[2],
            key[3],
            key[5],
            item["qty"] if item["qty"] else "",
            key[4],
            len(item["reports"]),
            "、".join(sorted(item["reports"])),
        ])
    return rows


def flatten_issue_rows(report_set: dict[str, Any]) -> list[list[Any]]:
    return [
        [
            issue.get("report", ""),
            issue.get("code", ""),
            issue.get("message", ""),
        ]
        for issue in report_set.get("issues", []) or []
    ]


def _write_index_sheet(ws, report_set: dict[str, Any]) -> None:
    ws.title = "目錄"
    navy = "0C2F5E"
    blue = "1860AB"
    section_fill = PatternFill(start_color="EAF2F8", end_color="EAF2F8", fill_type="solid")
    header_fill = PatternFill(start_color=navy, end_color=navy, fill_type="solid")
    border = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )
    agg = report_set.get("aggregates", {}) or {}

    ws.cell(row=1, column=1, value="現場修改統計單 - 工作簿目錄")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.cell(row=1, column=1).font = Font(name="Microsoft JhengHei UI", bold=True, size=15, color=navy)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.cell(row=2, column=1, value="本檔分成「開發者看的」與「要繳出去的報告」兩區；交付區不顯示本機路徑。")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    ws.cell(row=2, column=1).font = Font(name="Microsoft JhengHei UI", size=10, color="475569")

    metrics = [
        ("修改單", agg.get("report_count", 0)),
        ("焊口", agg.get("weld_count", 0)),
        ("材料筆數", agg.get("material_row_count", 0)),
        ("照片", agg.get("photo_count", 0)),
    ]
    for col, (label, value) in enumerate(metrics, start=1):
        cell = ws.cell(row=4, column=col, value=f"{label}: {value}")
        cell.font = Font(name="Microsoft JhengHei UI", bold=True, size=10, color=blue)
        cell.fill = section_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row = 6
    for section in WORKBOOK_SECTIONS:
        ws.cell(row=row, column=1, value=section["title"])
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.cell(row=row, column=1).font = Font(name="Microsoft JhengHei UI", bold=True, size=12, color=navy)
        ws.cell(row=row, column=1).fill = section_fill
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center")
        row += 1

        for col, header in enumerate(["Sheet", "用途", "區域", "備註"], start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(name="Microsoft JhengHei UI", bold=True, size=10, color="FFFFFF")
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

        for sheet_name, description in section["sheets"]:
            sheet_cell = ws.cell(row=row, column=1, value=sheet_name)
            sheet_cell.hyperlink = f"#'{sheet_name}'!A1"
            sheet_cell.style = "Hyperlink"
            sheet_cell.font = Font(name="Microsoft JhengHei UI", size=10, color=blue, underline="single")
            ws.cell(row=row, column=2, value=description)
            ws.cell(row=row, column=3, value=section["title"])
            ws.cell(row=row, column=4, value="")
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            row += 1
        row += 1

    for col, width in enumerate([26, 62, 18, 28], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A6"


def _write_overview_sheet(
    ws,
    report_set: dict[str, Any],
    *,
    sheet_title: str = "總覽",
    table_title: str = "現場修改統計單 - 總覽",
) -> None:
    ws.title = sheet_title
    rows = build_overview_rows(report_set)
    _write_table(
        ws,
        table_title,
        ["項目", "數值", "備註"],
        rows,
        [22, 42, 24],
        number_cols=[2],
    )


def _write_deliverable_overview_sheet(ws, report_set: dict[str, Any]) -> None:
    _write_table(
        ws,
        "報告 - 總覽",
        ["項目", "數值", "備註"],
        build_deliverable_overview_rows(report_set),
        [22, 28, 34],
        number_cols=[2],
    )


def _write_report_list_sheet(ws, report_set: dict[str, Any], *, title: str = "修改單清單") -> None:
    headers = [
        "報告編號", "日期", "Series", "資料夾", "狀態", "變更類型", "焊口摘要",
        "焊口數", "用料筆數", "before", "after", "附件PDF", "完整度", "缺漏",
        "說明", "資料夾路徑",
    ]
    _write_table(
        ws,
        title,
        headers,
        flatten_report_rows(report_set),
        [14, 12, 10, 18, 10, 12, 30, 8, 10, 8, 8, 10, 10, 22, 38, 45],
        number_cols=[8, 9, 10, 11],
    )
    _apply_path_hyperlinks(ws, 16, start_row=3)


def _write_deliverable_report_list_sheet(ws, report_set: dict[str, Any]) -> None:
    headers = [
        "報告編號", "日期", "Series", "資料夾", "狀態", "變更類型", "焊口摘要",
        "焊口數", "用料筆數", "before", "after", "附件PDF", "說明",
    ]
    _write_table(
        ws,
        "報告 - 修改單清單",
        headers,
        flatten_deliverable_report_rows(report_set),
        [14, 12, 10, 18, 10, 12, 30, 8, 10, 8, 8, 10, 38],
        number_cols=[8, 9, 10, 11],
    )


def _write_weld_summary_sheet(ws, report_set: dict[str, Any], *, title: str = "焊口統計") -> None:
    _write_table(
        ws,
        title,
        ["尺寸", "焊口數", "尺寸合計", "來源報告數", "來源報告"],
        flatten_weld_summary_rows(report_set),
        [12, 10, 12, 12, 50],
        number_cols=[2, 3, 4],
    )


def _write_photo_index_sheet(ws, report_set: dict[str, Any], *, title: str = "照片索引") -> None:
    _write_table(
        ws,
        title,
        ["報告", "日期", "Series", "資料夾", "類型", "序號", "檔名", "寬", "高", "路徑"],
        flatten_photo_index_rows(report_set),
        [16, 12, 10, 18, 10, 8, 24, 8, 8, 55],
        number_cols=[6, 8, 9],
    )
    _apply_path_hyperlinks(ws, 10, start_row=3)


def _write_photo_sheet(
    ws,
    report_set: dict[str, Any],
    *,
    title: str = "照片表",
    include_path_columns: bool = True,
) -> None:
    if include_path_columns:
        headers = ["報告", "日期", "Series", "資料夾", "組", "before", "after", "說明", "before路徑", "after路徑"]
        widths = [16, 12, 10, 18, 8, 34, 34, 36, 46, 46]
    else:
        headers = ["報告", "日期", "Series", "資料夾", "組", "before", "after", "說明"]
        widths = [16, 12, 10, 18, 8, 34, 34, 36]
    navy = "1F3864"
    stripe_fill = PatternFill(start_color="EAF2F8", end_color="EAF2F8", fill_type="solid")
    header_fill = PatternFill(start_color=navy, end_color=navy, fill_type="solid")
    border = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )

    ws.cell(row=1, column=1, value=title)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1).font = Font(name="Microsoft JhengHei UI", bold=True, size=13, color=navy)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(name="Microsoft JhengHei UI", bold=True, size=10, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    row_idx = 3
    for report in report_set.get("reports", []) or []:
        info = report.get("report", {})
        photos = report.get("photos", {}) or {}
        before_photos = photos.get("before", []) or []
        after_photos = photos.get("after", []) or []
        pair_count = max(len(before_photos), len(after_photos), 1)
        report_label = info.get("report_id") or info.get("folder", "")

        for pair_idx in range(pair_count):
            before = before_photos[pair_idx] if pair_idx < len(before_photos) else {}
            after = after_photos[pair_idx] if pair_idx < len(after_photos) else {}
            values = [
                report_label,
                info.get("date", ""),
                info.get("series", ""),
                info.get("folder", ""),
                pair_idx + 1,
                "",
                "",
                info.get("description", ""),
            ]
            if include_path_columns:
                values.extend([before.get("path", ""), after.get("path", "")])
            for col, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.font = Font(name="Microsoft JhengHei UI", size=9)
                cell.alignment = Alignment(
                    horizontal="center" if col in (5, 6, 7) else "left",
                    vertical="center",
                    wrap_text=True,
                )
                cell.border = border
                if (row_idx - 3) % 2 == 0:
                    cell.fill = stripe_fill

            _write_photo_cell(ws, row_idx, 6, before, missing_label="缺 before")
            _write_photo_cell(ws, row_idx, 7, after, missing_label="缺 after")
            row_idx += 1

    if row_idx == 3:
        ws.cell(row=3, column=1, value="無照片資料")
        ws.cell(row=3, column=1).font = Font(name="Microsoft JhengHei UI", size=9)
        ws.cell(row=3, column=1).border = border
        row_idx = 4

    data_end = max(3, row_idx - 1)
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in range(3, data_end + 1):
        ws.row_dimensions[row].height = 122
    if include_path_columns:
        _apply_path_hyperlinks(ws, 9, start_row=3)
        _apply_path_hyperlinks(ws, 10, start_row=3)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{data_end}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def _write_material_summary_sheet(ws, report_set: dict[str, Any], *, title: str = "用料統計") -> None:
    _write_table(
        ws,
        title,
        ["零件類型", "尺寸", "SCH", "材質", "類別", "數量", "單位", "來源報告數", "來源報告"],
        flatten_material_summary_rows(report_set),
        [22, 10, 10, 22, 10, 10, 8, 12, 50],
        number_cols=[6, 8],
    )


def _write_issue_sheet(ws, report_set: dict[str, Any], *, title: str = "問題清單") -> None:
    _write_table(
        ws,
        title,
        ["報告/資料夾", "問題代碼", "說明"],
        flatten_issue_rows(report_set),
        [20, 18, 46],
    )


def _apply_workbook_tab_colors(wb) -> None:
    if "目錄" in wb.sheetnames:
        wb["目錄"].sheet_properties.tabColor = "0C2F5E"
    for section in WORKBOOK_SECTIONS:
        for sheet_name, _description in section["sheets"]:
            if sheet_name in wb.sheetnames:
                wb[sheet_name].sheet_properties.tabColor = section["tab_color"]


def _write_table(
    ws,
    title: str,
    headers: list[str],
    rows: list[list[Any]],
    widths: list[int],
    *,
    number_cols: list[int] | None = None,
) -> None:
    number_cols = set(number_cols or [])
    navy = "1F3864"
    stripe_fill = PatternFill(start_color="EAF2F8", end_color="EAF2F8", fill_type="solid")
    header_fill = PatternFill(start_color=navy, end_color=navy, fill_type="solid")
    border = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )
    ws.cell(row=1, column=1, value=title)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1).font = Font(name="Microsoft JhengHei UI", bold=True, size=13, color=navy)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(name="Microsoft JhengHei UI", bold=True, size=10, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_idx, row in enumerate(rows, start=3):
        for col, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = Font(name="Microsoft JhengHei UI", size=9)
            cell.alignment = Alignment(
                horizontal="right" if col in number_cols else "left",
                vertical="center",
                wrap_text=True,
            )
            cell.border = border
            if (row_idx - 3) % 2 == 0:
                cell.fill = stripe_fill

    data_end = max(3, len(rows) + 2)
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(headers))}{data_end}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def _apply_path_hyperlinks(ws, col_idx: int, *, start_row: int) -> None:
    for row in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        path = str(cell.value or "").strip()
        if not path:
            continue
        cell.hyperlink = path
        cell.style = "Hyperlink"


def _write_photo_cell(ws, row_idx: int, col_idx: int, photo: dict[str, Any], *, missing_label: str) -> None:
    path = str(photo.get("path", "") or "").strip()
    cell = ws.cell(row=row_idx, column=col_idx)
    if not path:
        cell.value = missing_label
        return
    if not os.path.exists(path):
        cell.value = "找不到圖片"
        return
    try:
        img = XLImage(path)
    except (ImportError, OSError, ValueError):
        cell.value = "無法讀取圖片"
        return
    img.width, img.height = _fit_image_size(img.width, img.height, 220, 150)
    ws.add_image(img, f"{get_column_letter(col_idx)}{row_idx}")


def _fit_image_size(width: Any, height: Any, max_width: int, max_height: int) -> tuple[int, int]:
    try:
        width_f = float(width)
        height_f = float(height)
    except (TypeError, ValueError):
        return max_width, max_height
    if width_f <= 0 or height_f <= 0:
        return max_width, max_height
    scale = min(max_width / width_f, max_height / height_f, 1.0)
    return max(1, int(width_f * scale)), max(1, int(height_f * scale))


def _status_label(status: str) -> str:
    return {
        "produced": "已產出",
        "unproduced": "未產出",
        "needs_rebuild": "需重產",
        "missing_folder": "缺資料夾",
    }.get(str(status or ""), str(status or ""))


def _to_float(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0
