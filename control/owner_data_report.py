# -*- coding: utf-8 -*-
"""Owner-facing data report package.

The owner package is a deliverable folder, not just one workbook:
- owner_data_report/index.xlsx
- owner_data_report/<report>/before/*
- owner_data_report/<report>/after/*
- owner_data_report/<report>/pdf/*

The workbook uses relative links so the whole folder can be moved or sent out.
"""

from __future__ import annotations

import os
import re
import shutil
import subprocess
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from utils import atomic_save_wb


OWNER_DATA_REPORT_DIRNAME = "owner_data_report"
OWNER_DATA_INDEX_FILENAME = "owner_data_index.xlsx"
THUMB_WIDTH = 260
THUMB_HEIGHT = 180
PREVIEW_ROW_HEIGHT = 145
_AUTO_WELD_LOOKUP = object()
_AUTO_DRAWING_LOOKUP = object()
_DN_TO_INCH = {
    15: "0.5",
    20: "0.75",
    25: "1",
    32: "1.25",
    40: "1.5",
    50: "2",
    65: "2.5",
    80: "3",
    100: "4",
    125: "5",
    150: "6",
    200: "8",
    250: "10",
    300: "12",
    350: "14",
    400: "16",
    450: "18",
    500: "20",
}


def build_owner_data_report_package(
    output_root: str | os.PathLike[str],
    report_set: dict[str, Any],
    *,
    dirname: str = OWNER_DATA_REPORT_DIRNAME,
    weld_lookup: Any = _AUTO_WELD_LOOKUP,
    drawing_lookup: Any = _AUTO_DRAWING_LOOKUP,
) -> dict[str, Any]:
    root = Path(output_root).resolve()
    package_root = root / dirname
    _reset_child_dir(root, package_root)
    package_root.mkdir(parents=True, exist_ok=True)
    if weld_lookup is _AUTO_WELD_LOOKUP:
        weld_lookup = _make_weld_lookup()
    if drawing_lookup is _AUTO_DRAWING_LOOKUP:
        drawing_lookup = _make_drawing_lookup()

    used_names: set[str] = set()
    entries = []
    for idx, report in enumerate(report_set.get("reports", []) or [], start=1):
        entry = _copy_report_assets(
            package_root,
            report,
            idx=idx,
            used_names=used_names,
            weld_lookup=weld_lookup,
            drawing_lookup=drawing_lookup,
        )
        entries.append(entry)

    index_path = package_root / OWNER_DATA_INDEX_FILENAME
    export_owner_data_index_workbook(index_path, report_set, entries, weld_lookup=weld_lookup)
    return {
        "ok": True,
        "package_root": str(package_root),
        "index_xlsx": str(index_path),
        "report_count": len(entries),
        "reports": entries,
    }


def _make_weld_lookup():
    try:
        from weld_lookup import WeldLookup

        return WeldLookup()
    except Exception:
        return None


def _make_drawing_lookup():
    try:
        from record_manager import load_drawing_map

        return load_drawing_map()
    except Exception:
        return {}


def export_owner_data_index_workbook(
    output_path: str | os.PathLike[str],
    report_set: dict[str, Any],
    entries: list[dict[str, Any]],
    *,
    weld_lookup: Any = None,
) -> str:
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    with tempfile.TemporaryDirectory(prefix="iec_owner_xlsx_") as tmp:
        media_dir = Path(tmp)
        index_sheet = wb.active
        index_sheet.title = "資料索引"
        _write_index_sheet(index_sheet, report_set, entries, media_dir=media_dir, base_dir=output.parent)
        _write_weld_sheet(wb.create_sheet("焊口統計"), report_set, weld_lookup=weld_lookup)
        _write_material_sheet(wb.create_sheet("用料統計"), report_set)
        _apply_tab_colors(wb)
        atomic_save_wb(wb, str(output))
    wb.close()
    return str(output)


def _owner_project_name(report_set: dict[str, Any]) -> str:
    project = report_set.get("project", {}) or {}
    for key in ("name", "project_name", "title", "display_name", "工程名稱", "專案名稱"):
        value = str(project.get(key, "") or "").strip()
        if value:
            return value
    return "HP6精濾區配管工事"


def _reset_child_dir(parent: Path, child: Path) -> None:
    parent = parent.resolve()
    child = child.resolve()
    if child.exists():
        if child.parent != parent:
            raise RuntimeError(f"拒絕清除非輸出根目錄底下的資料夾：{child}")
        shutil.rmtree(child)


def _copy_report_assets(
    package_root: Path,
    report: dict[str, Any],
    *,
    idx: int,
    used_names: set[str],
    weld_lookup: Any,
    drawing_lookup: Any,
) -> dict[str, Any]:
    info = report.get("report", {}) or {}
    drawing_info = _lookup_drawing_info(drawing_lookup, info.get("series", ""))
    label = _report_label(info)
    report_dir_name = _unique_name(_safe_filename(label, fallback=f"report_{idx:03d}"), used_names)
    report_dir = package_root / report_dir_name
    before_dir = report_dir / "before"
    after_dir = report_dir / "after"
    pdf_dir = report_dir / "pdf"
    before_dir.mkdir(parents=True, exist_ok=True)
    after_dir.mkdir(parents=True, exist_ok=True)
    pdf_dir.mkdir(parents=True, exist_ok=True)

    before_files = _copy_photo_files(report.get("photos", {}).get("before", []) or [], before_dir, "before")
    after_files = _copy_photo_files(report.get("photos", {}).get("after", []) or [], after_dir, "after")
    pdf_files = _copy_pdf_file(report.get("attachment_pdf", {}) or {}, pdf_dir)

    return {
        "report_id": label,
        "date": info.get("date", ""),
        "date_raw": info.get("date_raw", ""),
        "series": info.get("series", ""),
        "dwg_no": _first_text(info.get("dwg_no"), drawing_info.get("dwg_no")),
        "line_number": _first_text(info.get("line_number"), info.get("line_no"), drawing_info.get("line_number")),
        "folder": info.get("folder", ""),
        "folder_label": _display_folder_label(info),
        "folder_path": info.get("folder_path", ""),
        "report_dir": str(report_dir),
        "report_rel": report_dir_name,
        "before_dir": str(before_dir),
        "after_dir": str(after_dir),
        "pdf_dir": str(pdf_dir),
        "before_files": before_files,
        "after_files": after_files,
        "pdf_files": pdf_files,
        "description": info.get("description", ""),
        "change_type": info.get("change_type", ""),
        "status": info.get("status", ""),
        "weld_count": report.get("welds", {}).get("count", 0),
        "weld_summary": _owner_weld_summary(report.get("welds", {}) or {}, series=info.get("series", ""), weld_lookup=weld_lookup),
        "material_count": report.get("materials", {}).get("count", 0),
        "material_summary": report.get("materials", {}).get("summary", ""),
    }


def _series_lookup_keys(series: Any) -> list[str]:
    raw = str(series or "").strip()
    digits = re.sub(r"\D", "", raw)
    keys: list[str] = []
    for value in (raw, digits, digits.lstrip("0") if digits else "", digits.zfill(3) if digits else "", digits.zfill(4) if digits else ""):
        text = str(value or "").strip()
        if text and text not in keys:
            keys.append(text)
    return keys


def _lookup_drawing_info(drawing_lookup: Any, series: Any) -> dict[str, str]:
    if not drawing_lookup:
        return {}
    if callable(drawing_lookup):
        try:
            drawing_lookup = drawing_lookup()
        except Exception:
            return {}
    if not isinstance(drawing_lookup, dict):
        return {}
    for key in _series_lookup_keys(series):
        value = drawing_lookup.get(key)
        if isinstance(value, (list, tuple)):
            line_number = value[0] if len(value) > 0 else ""
            dwg_no = value[1] if len(value) > 1 else ""
            return {"line_number": str(line_number or ""), "dwg_no": str(dwg_no or "")}
        if isinstance(value, dict):
            return {
                "line_number": _first_text(value.get("line_number"), value.get("line_no"), value.get("LINE NUMBER"), value.get("Line No.")),
                "dwg_no": _first_text(value.get("dwg_no"), value.get("DWG NO"), value.get("圖號")),
            }
    return {}


def _owner_weld_summary(welds: dict[str, Any], *, series: Any = "", weld_lookup: Any = None) -> str:
    rows = [row for row in (welds.get("rows", []) or []) if isinstance(row, dict)]
    if not rows:
        return _summary_to_cell_lines(welds.get("summary", ""))
    lines = []
    for row in sorted(rows, key=_owner_weld_sort_key):
        line = _owner_weld_line(row, series=series, weld_lookup=weld_lookup)
        if line:
            lines.append(line)
    if welds.get("count"):
        lines.append(f"（共{welds.get('count')}口）")
    return "\n".join(lines)


def _owner_weld_line(row: dict[str, Any], *, series: Any, weld_lookup: Any) -> str:
    info = _lookup_weld_info(weld_lookup, series, row)
    code = _weld_code(row)
    size_raw = _first_text(info.get("size"), row.get("size"))
    material = _first_text(info.get("material"), row.get("material"))
    thickness = _first_text(info.get("sch"), row.get("thickness"), row.get("sch"))
    size_display = _format_weld_size(size_raw)
    db_display = _format_db_text(
        _first_text(
            info.get("db"),
            row.get("db"),
            row.get("db_count"),
            row.get("DB數"),
            row.get("DB"),
            row.get("DI"),
        ),
        size_raw,
    )
    budget_display = _format_budget_text(
        _first_text(
            info.get("budget_no"),
            row.get("budget_no"),
            row.get("預算編號"),
        ),
    )
    specs = [item for item in (size_display, material, thickness, db_display, budget_display) if item]
    if code and specs:
        return f"{code}（{' / '.join(specs)}）"
    if code:
        return code
    return " / ".join(specs)


def _lookup_weld_info(weld_lookup: Any, series: Any, row: dict[str, Any]) -> dict[str, Any]:
    if weld_lookup is None:
        return {}
    lookup_info = getattr(weld_lookup, "lookup_info", None)
    if not callable(lookup_info):
        return {}
    base = str(row.get("weld_no") or "").strip()
    if not base:
        base = _split_weld_code(_weld_code(row)).get("weld_no", "")
    if not base:
        return {}
    try:
        info = lookup_info(series, base)
    except Exception:
        return {}
    return info if isinstance(info, dict) else {}


def _weld_code(row: dict[str, Any]) -> str:
    code = str(row.get("code", "") or "").strip()
    if code:
        return code
    return (str(row.get("weld_no", "") or "").strip() + str(row.get("mark", "") or "").strip()).strip()


def _weld_change_label(row: dict[str, Any]) -> str:
    raw = _first_text(
        row.get("change_label"),
        row.get("change"),
        row.get("新增或修改"),
        row.get("op"),
        row.get("operation"),
        row.get("mark"),
        row.get("origin"),
    ).lower()
    if raw in {"新增", "新焊", "新增焊口", "new"}:
        return "新增"
    if raw in {"修改", "重焊", "原焊口重接", "裁切", "拆除不重焊", "existing", "r"}:
        return "修改"

    code_parts = _split_weld_code(_weld_code(row))
    suffix = code_parts.get("mark", "").lower()
    if suffix == "r":
        return "修改"
    if suffix:
        return "新增"
    try:
        if int(code_parts.get("weld_no", "") or "0") >= 1000:
            return "新增"
    except ValueError:
        pass
    if row.get("is_cut"):
        return "修改"
    return "新增"


def _weld_change_factor(label: Any) -> float | int | str:
    text = str(label or "").strip()
    if text == "新增":
        return 1
    if text == "修改":
        return 1.5
    return ""


def _split_weld_code(code: Any) -> dict[str, str]:
    text = str(code or "").strip()
    match = re.match(r"^(\d+)([A-Za-z]*)", text)
    if not match:
        return {"weld_no": text, "mark": ""}
    return {"weld_no": match.group(1), "mark": match.group(2)}


def _owner_weld_sort_key(row: dict[str, Any]) -> tuple[int, int, str, str]:
    code = _weld_code(row)
    match = re.match(r"^\s*(\d+)([A-Za-z]?)(.*)$", code)
    if not match:
        return (1, 0, code.lower(), "")
    number = int(match.group(1))
    mark = match.group(2).lower()
    tail = match.group(3).lower()
    mark_order = {"r": 0}
    if mark and mark not in mark_order:
        mark_order[mark] = ord(mark) - ord("a") + 1
    return (0, number, f"{mark_order.get(mark, 99):03d}-{mark}", tail)


def _first_text(*values: Any) -> str:
    for value in values:
        text = str(value or "").strip()
        if text:
            return text
    return ""


def _format_weld_size(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    normalized = text.replace("”", '"').replace("″", '"').replace("吋", '"')
    if re.search(r'("|DN|NPS|mm|MM)', normalized):
        return normalized
    if re.fullmatch(r"\d+(?:\.\d+)?", normalized):
        return f'{_trim_number(normalized)}"'
    if re.fullmatch(r"[0-9./ -]+", normalized):
        return f'{normalized}"'
    return normalized


def _format_db_text(value: Any, size_value: Any = "") -> str:
    explicit = str(value or "").strip()
    if explicit:
        if re.search(r"\bD[BI]\b", explicit, flags=re.IGNORECASE):
            return explicit
        return f"DB {explicit}"
    inferred = _infer_db_from_size(size_value)
    return f"DB {inferred}" if inferred else ""


def _format_budget_text(value: Any) -> str:
    text = str(value or "").strip()
    return f"預算 {text}" if text else ""


def _infer_db_from_size(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    dn_match = re.search(r"DN\s*(\d+)", text, flags=re.IGNORECASE)
    if dn_match:
        return _DN_TO_INCH.get(int(dn_match.group(1)), "")
    quote_match = re.search(r"(\d+(?:\.\d+)?)\s*(?:\"|吋)", text)
    if quote_match:
        return _trim_number(quote_match.group(1))
    numeric_match = re.fullmatch(r"\d+(?:\.\d+)?", text)
    if numeric_match:
        return _trim_number(text)
    return ""


def _trim_number(value: Any) -> str:
    text = str(value or "").strip()
    try:
        number = float(text)
    except ValueError:
        return text
    return str(int(number)) if number.is_integer() else f"{number:g}"


def _copy_photo_files(photos: list[dict[str, Any]], target_dir: Path, prefix: str) -> list[dict[str, str]]:
    copied = []
    for idx, photo in enumerate(photos, start=1):
        source = Path(str(photo.get("path", "") or ""))
        if not source.is_file():
            continue
        suffix = source.suffix.lower() or ".jpg"
        name = _safe_filename(f"{prefix}_{idx:02d}_{source.stem}", fallback=f"{prefix}_{idx:02d}") + suffix
        destination = target_dir / name
        shutil.copy2(source, destination)
        copied.append({
            "name": name,
            "source": str(source),
            "path": str(destination),
        })
    return copied


def _copy_pdf_file(pdf_info: dict[str, Any], target_dir: Path) -> list[dict[str, str]]:
    path = Path(str(pdf_info.get("path", "") or ""))
    if not path.is_file():
        return []
    name = _safe_filename(path.stem, fallback="drawing") + (path.suffix.lower() or ".pdf")
    destination = target_dir / name
    shutil.copy2(path, destination)
    return [{
        "name": name,
        "source": str(path),
        "path": str(destination),
    }]


def _write_cover_sheet(ws, report_set: dict[str, Any], entries: list[dict[str, Any]]) -> None:
    ws.title = "封面"
    navy = "0C2F5E"
    blue = "1860AB"
    light = "EEF4FC"
    border = _border()
    agg = report_set.get("aggregates", {}) or {}

    ws.cell(row=1, column=1, value="現場修改資料包索引")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws.cell(row=1, column=1).font = Font(name="Microsoft JhengHei UI", bold=True, size=18, color=navy)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.cell(row=2, column=1, value="業主查閱版")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws.cell(row=2, column=1).font = Font(name="Microsoft JhengHei UI", size=11, color="475569")

    project = report_set.get("project", {}) or {}
    rows = [
        ["工程資料夾", project.get("root", "")],
        ["資料包產生時間", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["索引統計時間", project.get("collected_at", "")],
        ["修改單數量", len(entries)],
        ["焊口總數", agg.get("weld_count", 0)],
        ["用料明細筆數", agg.get("material_row_count", 0)],
        ["照片張數", agg.get("photo_count", 0)],
        ["圖面 PDF", sum(1 for entry in entries if entry.get("pdf_files"))],
    ]
    start = 4
    for row_idx, row in enumerate(rows, start=start):
        for col, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = Font(name="Microsoft JhengHei UI", bold=(col == 1), size=10, color=navy if col == 1 else "1E293B")
            cell.fill = PatternFill(start_color=light if col == 1 else "FFFFFF", end_color=light if col == 1 else "FFFFFF", fill_type="solid")
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=6)

    guide_row = start + len(rows) + 2
    ws.cell(row=guide_row, column=1, value="使用說明")
    ws.merge_cells(start_row=guide_row, start_column=1, end_row=guide_row, end_column=6)
    ws.cell(row=guide_row, column=1).font = Font(name="Microsoft JhengHei UI", bold=True, size=12, color=blue)
    ws.cell(row=guide_row, column=1).fill = PatternFill(start_color=light, end_color=light, fill_type="solid")
    instructions = [
        "1. 每張修改單均有獨立資料夾，內含 before、after、pdf 子資料夾。",
        "2. 「資料索引」提供報告編號、日期、流水號、圖面資訊與資料夾連結。",
        "3. Excel 中 before / after / PDF 欄為預覽縮圖；正式檔案請以右側資料連結開啟。",
        "4. 本資料包使用相對路徑，整個 owner_data_report 資料夾可一起移動或交付。",
    ]
    for offset, text in enumerate(instructions, start=1):
        ws.cell(row=guide_row + offset, column=1, value=text)
        ws.merge_cells(start_row=guide_row + offset, start_column=1, end_row=guide_row + offset, end_column=6)
        ws.cell(row=guide_row + offset, column=1).font = Font(name="Microsoft JhengHei UI", size=10, color="334155")
        ws.cell(row=guide_row + offset, column=1).alignment = Alignment(wrap_text=True)

    for col, width in enumerate([18, 22, 20, 20, 20, 20], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_index_sheet(
    ws,
    report_set: dict[str, Any],
    entries: list[dict[str, Any]],
    *,
    media_dir: Path,
    base_dir: Path,
) -> None:
    title = f"{_owner_project_name(report_set)}-工務修改確認單"
    headers = [
        "項次", "工務修改確認單編號", "日期", "ISO流編", "圖號", "Line No.",
        "新增或修改說明", "新增修改焊口詳細", "材料新增或修改摘要",
        "修改前相片", "修改後相片", "相關圖說",
        "Before檔", "After檔", "圖面PDF",
    ]
    widths = [8, 24, 13, 11, 22, 18, 36, 42, 34, 40, 40, 40, 11, 11, 11]
    _prepare_titled_sheet(ws, title, headers, widths)

    for row_idx, entry in enumerate(entries, start=3):
        values = [
            row_idx - 2,
            entry.get("report_id", ""),
            entry.get("date", ""),
            entry.get("series", ""),
            entry.get("dwg_no", ""),
            entry.get("line_number", ""),
            entry.get("description", ""),
            _summary_to_cell_lines(entry.get("weld_summary", "")),
            _summary_to_cell_lines(entry.get("material_summary", "")),
            "",
            "",
            "",
            "開啟",
            "開啟",
            "開啟",
        ]
        _write_values_row(ws, row_idx, values)
        ws.row_dimensions[row_idx].height = PREVIEW_ROW_HEIGHT
        pdf_rel = _relpath(_first_path(entry.get("pdf_files", [])), base_dir)
        _link_cell(
            ws.cell(row=row_idx, column=2),
            _relpath(entry.get("report_dir", ""), base_dir),
            display=entry.get("report_id", "") or "開啟",
        )
        _link_cell(ws.cell(row=row_idx, column=13), _relpath(entry.get("before_dir", ""), base_dir), display="開啟")
        _link_cell(ws.cell(row=row_idx, column=14), _relpath(entry.get("after_dir", ""), base_dir), display="開啟")
        _link_cell(ws.cell(row=row_idx, column=15), pdf_rel, display="開啟")
        _center_cells(ws, row_idx, (1, 2, 13, 14, 15))
        _top_cells(ws, row_idx, (7, 8, 9))
        _add_preview(
            ws,
            row_idx,
            10,
            _first_path(entry.get("before_files", [])),
            media_dir,
            f"before_{row_idx}",
            title="Before",
        )
        _add_preview(
            ws,
            row_idx,
            11,
            _first_path(entry.get("after_files", [])),
            media_dir,
            f"after_{row_idx}",
            title="After",
        )
        _add_pdf_preview(
            ws,
            row_idx,
            12,
            _first_path(entry.get("pdf_files", [])),
            media_dir,
            f"pdf_{row_idx}",
            title="圖面 PDF",
        )

    _finish_table(ws, len(headers), max(3, len(entries) + 2))


def _write_photo_detail_sheet(
    ws,
    entries: list[dict[str, Any]],
    *,
    media_dir: Path,
    base_dir: Path,
) -> None:
    headers = ["報告編號", "日期", "組", "Before", "After", "before檔案", "after檔案", "說明"]
    widths = [18, 13, 8, 40, 40, 12, 12, 42]
    _prepare_titled_sheet(ws, "現場修改資料包 - 照片明細", headers, widths)
    row_idx = 3
    for entry in entries:
        before = entry.get("before_files", []) or []
        after = entry.get("after_files", []) or []
        pairs = max(len(before), len(after), 1)
        for pair_idx in range(pairs):
            before_item = before[pair_idx] if pair_idx < len(before) else {}
            after_item = after[pair_idx] if pair_idx < len(after) else {}
            _write_values_row(ws, row_idx, [
                entry.get("report_id", ""),
                entry.get("date", ""),
                pair_idx + 1,
                "",
                "",
                before_item.get("name", ""),
                after_item.get("name", ""),
                entry.get("description", ""),
            ])
            ws.row_dimensions[row_idx].height = PREVIEW_ROW_HEIGHT
            before_rel = _relpath(before_item.get("path", ""), base_dir)
            after_rel = _relpath(after_item.get("path", ""), base_dir)
            _link_cell(ws.cell(row=row_idx, column=6), before_rel, display="開啟")
            _link_cell(ws.cell(row=row_idx, column=7), after_rel, display="開啟")
            _center_cells(ws, row_idx, (6, 7))
            _add_preview(
                ws,
                row_idx,
                4,
                before_item.get("path", ""),
                media_dir,
                f"detail_before_{row_idx}",
                title="Before",
            )
            _add_preview(
                ws,
                row_idx,
                5,
                after_item.get("path", ""),
                media_dir,
                f"detail_after_{row_idx}",
                title="After",
            )
            row_idx += 1
    _finish_table(ws, len(headers), max(3, row_idx - 1))


def _write_weld_sheet(ws, report_set: dict[str, Any], *, weld_lookup: Any = None) -> None:
    headers = [
        "工務修改確認單編號",
        "日期",
        "ISO流編",
        "焊口編號",
        "尺寸",
        "材質",
        "厚度",
        "新增或修改",
        "新增或修改係數",
        "DB",
        "預算編號",
    ]
    widths = [24, 13, 11, 14, 12, 16, 14, 14, 16, 10, 14]
    title = f"{_owner_project_name(report_set)}-工務修改確認單 - 焊口統計"
    _prepare_titled_sheet(ws, title, headers, widths)
    row_idx = 3
    for report in report_set.get("reports", []) or []:
        info = report.get("report", {}) or {}
        label = _report_label(info)
        series = info.get("series", "")
        for weld in report.get("welds", {}).get("rows", []) or []:
            lookup_info = _lookup_weld_info(weld_lookup, series, weld)
            size_raw = _first_text(lookup_info.get("size"), weld.get("size"))
            material = _first_text(lookup_info.get("material"), weld.get("material"))
            thickness = _first_text(lookup_info.get("sch"), weld.get("thickness"), weld.get("sch"))
            db_text = _format_db_text(
                _first_text(
                    lookup_info.get("db"),
                    weld.get("db"),
                    weld.get("db_count"),
                    weld.get("DB數"),
                    weld.get("DB"),
                    weld.get("DI"),
                ),
                size_raw,
            ).replace("DB ", "", 1)
            budget_no = _first_text(
                lookup_info.get("budget_no"),
                weld.get("budget_no"),
                weld.get("預算編號"),
            )
            change_label = _weld_change_label(weld)
            _write_values_row(ws, row_idx, [
                label,
                info.get("date", ""),
                series,
                _weld_code(weld),
                _format_weld_size(size_raw),
                material,
                thickness,
                change_label,
                _weld_change_factor(change_label),
                db_text,
                budget_no,
            ])
            row_idx += 1
    _finish_table(ws, len(headers), max(3, row_idx - 1))


def _write_material_sheet(ws, report_set: dict[str, Any]) -> None:
    headers = ["報告編號", "日期", "流水號", "零件", "尺寸", "SCH", "材質", "數量", "單位", "備註"]
    widths = [18, 13, 11, 24, 14, 12, 18, 10, 10, 34]
    _prepare_titled_sheet(ws, f"{_owner_project_name(report_set)}-工務修改確認單 - 用料統計", headers, widths)
    row_idx = 3
    for report in report_set.get("reports", []) or []:
        info = report.get("report", {}) or {}
        label = _report_label(info)
        for material in report.get("materials", {}).get("rows", []) or []:
            _write_values_row(ws, row_idx, [
                label,
                info.get("date", ""),
                info.get("series", ""),
                material.get("component", ""),
                material.get("size", ""),
                material.get("sch", ""),
                material.get("material", ""),
                material.get("qty", ""),
                material.get("unit", ""),
                material.get("remark", ""),
            ])
            row_idx += 1
    _finish_table(ws, len(headers), max(3, row_idx - 1))


def _prepare_titled_sheet(ws, title: str, headers: list[str], widths: list[int]) -> None:
    navy = "0C2F5E"
    header_fill = PatternFill(start_color=navy, end_color=navy, fill_type="solid")
    border = _border()
    ws.cell(row=1, column=1, value=title)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1).font = Font(name="Microsoft JhengHei UI", bold=True, size=14, color=navy)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(name="Microsoft JhengHei UI", bold=True, size=10, color="FFFFFF")
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.sheet_view.showGridLines = False


def _write_values_row(ws, row_idx: int, values: list[Any]) -> None:
    stripe = PatternFill(start_color="EEF4FC", end_color="EEF4FC", fill_type="solid")
    border = _border()
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col, value=value)
        cell.font = Font(name="Microsoft JhengHei UI", size=9, color="1E293B")
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if (row_idx - 3) % 2 == 0:
            cell.fill = stripe


def _center_cells(ws, row_idx: int, columns: tuple[int, ...]) -> None:
    for col in columns:
        ws.cell(row=row_idx, column=col).alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )


def _top_cells(ws, row_idx: int, columns: tuple[int, ...]) -> None:
    for col in columns:
        ws.cell(row=row_idx, column=col).alignment = Alignment(
            horizontal="left",
            vertical="top",
            wrap_text=True,
        )


def _summary_to_cell_lines(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    count_text = ""
    count_match = re.search(r"([（(]共\d+口[)）])$", text)
    if count_match:
        count_text = count_match.group(1)
        text = text[:count_match.start()].rstrip("、;； \n")
    parts = [part.strip() for part in re.split(r"[、;；]+", text) if part.strip()]
    if not parts:
        return count_text or text
    if count_text:
        parts.append(count_text)
    return "\n".join(parts)


def _finish_table(ws, column_count: int, data_end: int) -> None:
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(column_count)}{data_end}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def _add_preview(
    ws,
    row_idx: int,
    col_idx: int,
    source: str,
    media_dir: Path,
    name: str,
    *,
    title: str,
) -> None:
    path = Path(str(source or ""))
    if not path.is_file():
        ws.cell(row=row_idx, column=col_idx, value="無檔案")
        return
    thumb = media_dir / f"{name}.png"
    if not _make_image_thumbnail(path, thumb, title=title):
        ws.cell(row=row_idx, column=col_idx, value="無法讀取")
        return
    image = XLImage(str(thumb))
    image.width = THUMB_WIDTH
    image.height = THUMB_HEIGHT
    ws.add_image(image, f"{get_column_letter(col_idx)}{row_idx}")


def _add_pdf_preview(
    ws,
    row_idx: int,
    col_idx: int,
    source: str,
    media_dir: Path,
    name: str,
    *,
    title: str,
) -> None:
    path = Path(str(source or ""))
    if not path.is_file():
        ws.cell(row=row_idx, column=col_idx, value="無PDF")
        return
    thumb = media_dir / f"{name}.png"
    if not _make_pdf_thumbnail(path, thumb, title=title):
        _make_placeholder_thumbnail(thumb, "PDF", path.name, title=title)
    image = XLImage(str(thumb))
    image.width = THUMB_WIDTH
    image.height = THUMB_HEIGHT
    ws.add_image(image, f"{get_column_letter(col_idx)}{row_idx}")


def _make_image_thumbnail(source: Path, output: Path, *, title: str) -> bool:
    try:
        from PIL import Image, ImageOps
        with Image.open(source) as img:
            img = ImageOps.exif_transpose(img).convert("RGB")
            _save_preview_card(output, img, title=title, subtitle=source.name)
        return True
    except Exception:
        return False


def _make_pdf_thumbnail(source: Path, output: Path, *, title: str) -> bool:
    try:
        img = _render_pdf_first_page(source)
        if img is None:
            return False
        _save_preview_card(output, img, title=title, subtitle=source.name)
        return True
    except Exception:
        return False


def _render_pdf_first_page(source: Path):
    rendered = _render_pdf_first_page_with_poppler(source)
    if rendered is not None:
        return rendered
    return _render_pdf_first_page_with_pymupdf(source)


def _render_pdf_first_page_with_poppler(source: Path):
    try:
        from PIL import Image, ImageOps

        with tempfile.TemporaryDirectory(prefix="iec_pdf_render_") as tmp:
            prefix = Path(tmp) / "page"
            command = [
                "pdftoppm",
                "-f",
                "1",
                "-l",
                "1",
                "-singlefile",
                "-png",
                "-r",
                "144",
                str(source),
                str(prefix),
            ]
            result = subprocess.run(command, capture_output=True, text=True, timeout=20)
            image_path = prefix.with_suffix(".png")
            if result.returncode != 0 or not image_path.is_file():
                return None
            with Image.open(image_path) as img:
                return ImageOps.exif_transpose(img).convert("RGB")
    except Exception:
        return None


def _render_pdf_first_page_with_pymupdf(source: Path):
    try:
        import fitz
        from io import BytesIO
        from PIL import Image, ImageOps

        doc = fitz.open(str(source))
        if doc.page_count < 1:
            doc.close()
            return None
        page = doc.load_page(0)
        pix = page.get_pixmap(matrix=fitz.Matrix(2.0, 2.0), alpha=False)
        raw = pix.tobytes("png")
        doc.close()
        with Image.open(BytesIO(raw)) as img:
            return ImageOps.exif_transpose(img).convert("RGB")
    except Exception:
        return None


def _save_preview_card(output: Path, image, *, title: str, subtitle: str) -> None:
    from PIL import Image, ImageDraw

    canvas = Image.new("RGB", (THUMB_WIDTH, THUMB_HEIGHT), "white")
    draw = ImageDraw.Draw(canvas)
    header_h = 26
    footer_h = 20
    draw.rounded_rectangle((0, 0, THUMB_WIDTH - 1, THUMB_HEIGHT - 1), radius=8, outline="#B4C6E7", width=2, fill="white")
    draw.rounded_rectangle((1, 1, THUMB_WIDTH - 2, header_h), radius=8, fill="#EEF4FC", outline="#EEF4FC")
    draw.rectangle((1, header_h - 8, THUMB_WIDTH - 2, header_h), fill="#EEF4FC")
    font_title = _preview_font(13, bold=True)
    font_sub = _preview_font(10)
    draw.text((12, 6), title, fill="#0C2F5E", font=font_title)
    work = image.copy()
    max_w = THUMB_WIDTH - 18
    max_h = THUMB_HEIGHT - header_h - footer_h - 12
    work.thumbnail((max_w, max_h), Image.Resampling.LANCZOS)
    x = (THUMB_WIDTH - work.width) // 2
    y = header_h + 6 + ((max_h - work.height) // 2)
    canvas.paste(work, (x, y))
    draw.rectangle((1, THUMB_HEIGHT - footer_h - 1, THUMB_WIDTH - 2, THUMB_HEIGHT - 2), fill="#F8FAFC")
    draw.text((12, THUMB_HEIGHT - footer_h + 4), subtitle[:38], fill="#475569", font=font_sub)
    canvas.save(output, "PNG")


def _make_placeholder_thumbnail(output: Path, badge: str, subtitle: str, *, title: str = "") -> None:
    from PIL import Image, ImageDraw

    canvas = Image.new("RGB", (THUMB_WIDTH, THUMB_HEIGHT), "white")
    draw = ImageDraw.Draw(canvas)
    header_h = 26
    draw.rounded_rectangle((0, 0, THUMB_WIDTH - 1, THUMB_HEIGHT - 1), radius=8, outline="#B4C6E7", width=2, fill="white")
    draw.rounded_rectangle((1, 1, THUMB_WIDTH - 2, header_h), radius=8, fill="#EEF4FC", outline="#EEF4FC")
    draw.rectangle((1, header_h - 8, THUMB_WIDTH - 2, header_h), fill="#EEF4FC")
    font_title = _preview_font(28, bold=True)
    font_head = _preview_font(13, bold=True)
    font_sub = _preview_font(12)
    draw.text((12, 6), title or "檔案預覽", fill="#0C2F5E", font=font_head)
    draw.rounded_rectangle((THUMB_WIDTH // 2 - 40, 58, THUMB_WIDTH // 2 + 40, 112), radius=8, fill="#EEF4FC", outline="#B4C6E7", width=2)
    draw.text((THUMB_WIDTH // 2, 84), badge, fill="#1860AB", font=font_title, anchor="mm")
    draw.text((THUMB_WIDTH // 2, 134), subtitle[:34], fill="#475569", font=font_sub, anchor="mm")
    canvas.save(output, "PNG")


def _preview_font(size: int, *, bold: bool = False):
    from PIL import ImageFont

    candidates = [
        r"C:\Windows\Fonts\msjhbd.ttc" if bold else r"C:\Windows\Fonts\msjh.ttc",
        r"C:\Windows\Fonts\mingliub.ttc" if bold else r"C:\Windows\Fonts\mingliu.ttc",
        "arialbd.ttf" if bold else "arial.ttf",
    ]
    for candidate in candidates:
        try:
            return ImageFont.truetype(candidate, size)
        except Exception:
            continue
    return ImageFont.load_default()


def _link_cell(cell, target: str, *, display: str | None = None) -> None:
    text = str(target or "").strip()
    if not text:
        cell.value = "無"
        return
    cell.value = (display or text).replace("\\", "/")
    cell.hyperlink = text.replace("\\", "/")
    cell.style = "Hyperlink"


def _relpath(path: str, base_dir: Path) -> str:
    if not path:
        return ""
    try:
        return os.path.relpath(path, base_dir).replace("\\", "/")
    except Exception:
        return str(path).replace("\\", "/")


def _first_path(items: list[dict[str, str]] | Any) -> str:
    if not items:
        return ""
    first = items[0] if isinstance(items, list) else {}
    return str(first.get("path", "") if isinstance(first, dict) else "")


def _report_label(info: dict[str, Any]) -> str:
    raw = str(info.get("report_id") or info.get("folder") or info.get("series") or "未編號").strip()
    return _clean_report_label(raw, info)


def _display_folder_label(info: dict[str, Any]) -> str:
    return _clean_report_label(str(info.get("folder", "") or "").strip(), info)


def _clean_report_label(value: str, info: dict[str, Any]) -> str:
    text = str(value or "").strip()
    if not text:
        return str(info.get("series") or "未編號").strip()
    date_raw = str(info.get("date_raw") or "").strip().replace("-", "")
    series = str(info.get("series") or "").strip().lstrip("0") or str(info.get("series") or "").strip()
    folder_text = str(info.get("folder", "") or "").strip()
    series_text = str(info.get("series", "") or "").strip()
    match = re.match(r"^(\d+)[_-](\d{8})[_-](\d+)$", text)
    if match:
        return f"CO-{match.group(1)}-{match.group(2)}-{int(match.group(3)):02d}"
    cleaned = re.sub(r"[_\s]+", "-", text)
    cleaned = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff.-]+", "-", cleaned)
    cleaned = re.sub(r"-{2,}", "-", cleaned).strip("-.")
    if text == folder_text and text != series_text and cleaned:
        return cleaned if cleaned.upper().startswith("CO-") else f"CO-{cleaned}"
    if date_raw and series and text == series_text:
        return f"CO-{series}-{date_raw}"
    return cleaned or "未編號"


def _safe_filename(value: str, *, fallback: str = "file") -> str:
    text = str(value or "").strip() or fallback
    invalid = '<>:"/\\|?*'
    cleaned = "".join("_" if char in invalid or ord(char) < 32 else char for char in text)
    cleaned = cleaned.strip(" .")
    return cleaned or fallback


def _unique_name(name: str, used: set[str]) -> str:
    candidate = name
    idx = 2
    while candidate.lower() in used:
        candidate = f"{name}_{idx}"
        idx += 1
    used.add(candidate.lower())
    return candidate


def _border() -> Border:
    side = Side(style="thin", color="B4C6E7")
    return Border(left=side, right=side, top=side, bottom=side)


def _apply_tab_colors(wb: Workbook) -> None:
    colors = {
        "封面": "0C2F5E",
        "資料索引": "1860AB",
        "照片明細": "1860AB",
        "焊口統計": "64748B",
        "用料統計": "64748B",
    }
    for sheet, color in colors.items():
        if sheet in wb.sheetnames:
            wb[sheet].sheet_properties.tabColor = color
