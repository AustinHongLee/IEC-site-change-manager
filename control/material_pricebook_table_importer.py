# -*- coding: utf-8 -*-
"""
material_pricebook_table_importer.py - Excel/CSV 材料價格表安全匯入核心

這個模組負責把公司合約價表匯入專案價目表：
- 可新增尚不存在的材料 key。
- 可把既有骨架的空白單價補上。
- 不覆蓋既有有價項目；價格不同時列為 conflict 並略過。
"""

from __future__ import annotations

import copy
import csv
import os
import re
from datetime import date, datetime
from typing import Any

from billing_calculator import parse_amount
from material_pricebook import load_material_pricebook, normalize_pricebook_items, save_material_pricebook
from material_pricebook_importer import material_key
from material_pricebook_validation import PricebookValidationReport, validate_pricebook_items


CANONICAL_FIELDS = (
    "id",
    "零件類型",
    "尺寸",
    "SCH",
    "材質",
    "類別",
    "單位",
    "單價",
    "來源",
    "生效日",
    "備註",
)

HEADER_ALIASES = {
    "id": ("id", "料號", "料號/id", "料號id", "material_id", "materialid"),
    "零件類型": ("零件類型", "零件", "零件名稱", "品項", "component", "name", "item"),
    "尺寸": ("尺寸", "size", "口徑", "管徑"),
    "SCH": ("sch", "SCH", "schedule", "厚度"),
    "材質": ("材質", "material", "管材", "材料"),
    "類別": ("類別", "category", "分類"),
    "單位": ("單位", "unit", "計價單位"),
    "單價": ("單價", "price", "unit_price", "unitprice", "合約單價", "未稅單價"),
    "來源": ("來源", "source", "price_source", "pricesource", "價格來源"),
    "生效日": ("生效日", "effective_date", "effectivefrom", "effective_from", "日期"),
    "備註": ("備註", "remark", "remarks", "note", "notes", "說明"),
}

HEADER_TO_FIELD: dict[str, str] = {}
for _field, _aliases in HEADER_ALIASES.items():
    for _alias in _aliases:
        HEADER_TO_FIELD[re.sub(r"[\s_／/\\\-:：()（）]+", "", str(_alias).strip().lower())] = _field


def load_price_table_items(path: str, *, sheet_name: str | None = None) -> list[dict[str, Any]]:
    """讀取 CSV/XLSX 價格表，回傳尚未正規化的 item dict。"""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        rows = _load_csv_rows(path)
    elif ext in (".xlsx", ".xlsm"):
        rows = _load_xlsx_rows(path, sheet_name=sheet_name)
    else:
        raise ValueError("價格表只支援 .csv / .xlsx / .xlsm")
    return _rows_to_items(rows)


def validate_price_table_items(items: list[dict[str, Any]]) -> PricebookValidationReport:
    """驗證價格表資料。單價允許合法數字或空白，但不可是不可解析文字。"""
    report = PricebookValidationReport()
    if not items:
        report.err("找不到可匯入的價格資料列")
        return report

    for idx, item in enumerate(items, start=1):
        price_text = str(item.get("單價", "")).strip()
        if price_text and parse_amount(price_text) is None:
            row = item.get("__source_row") or idx
            report.err(f"第 {row} 列：單價=「{price_text}」不是合法數字")

    normalized = normalize_pricebook_items(_strip_private_fields(items))
    normalized_report = validate_pricebook_items(normalized, allow_price=True)
    report.errors.extend(normalized_report.errors)
    report.warnings.extend(normalized_report.warnings)
    return report


def build_price_table_import_plan(
    table_items: list[dict[str, Any]],
    current_pricebook: dict[str, Any],
) -> dict[str, Any]:
    """建立匯入計畫；此函式不寫檔。"""
    current_items = normalize_pricebook_items(current_pricebook.get("items", []))
    candidates = normalize_pricebook_items(_strip_private_fields(table_items))
    merged = copy.deepcopy(current_items)
    index_by_key = {
        material_key(item): idx
        for idx, item in enumerate(merged)
        if item.get("零件類型")
    }

    added: list[dict[str, str]] = []
    updated: list[dict[str, dict[str, str]]] = []
    skipped: list[dict[str, Any]] = []
    conflicts: list[dict[str, Any]] = []

    for candidate in candidates:
        key = material_key(candidate)
        if not key[0]:
            skipped.append({"item": candidate, "reason": "缺少零件類型"})
            continue

        existing_idx = index_by_key.get(key)
        if existing_idx is None:
            merged.append(candidate)
            index_by_key[key] = len(merged) - 1
            added.append(candidate)
            continue

        existing = merged[existing_idx]
        existing_price = str(existing.get("單價", "")).strip()
        incoming_price = str(candidate.get("單價", "")).strip()

        if not existing_price and incoming_price:
            before = copy.deepcopy(existing)
            merged[existing_idx] = _merge_blank_price_item(existing, candidate)
            updated.append({"before": before, "after": copy.deepcopy(merged[existing_idx])})
            continue

        if existing_price and incoming_price and parse_amount(existing_price) != parse_amount(incoming_price):
            conflicts.append({
                "existing": copy.deepcopy(existing),
                "incoming": copy.deepcopy(candidate),
                "reason": "既有單價與匯入單價不同，未自動覆蓋",
            })
            continue

        skipped.append({"item": candidate, "reason": _skip_reason(existing_price, incoming_price)})

    return {
        "items": merged,
        "added": added,
        "updated": updated,
        "skipped": skipped,
        "conflicts": conflicts,
        "existing_count": len(current_items),
        "candidate_count": len(candidates),
    }


def format_price_table_import_summary(
    plan: dict[str, Any],
    *,
    apply: bool = False,
    target: str = "",
) -> str:
    mode = "APPLY" if apply else "DRY-RUN"
    lines = [f"[{mode}]"]
    if target:
        lines.append(f"目標價目表: {target}")
    lines.extend([
        f"既有項目: {plan['existing_count']}",
        f"價格表項目: {plan['candidate_count']}",
        f"將新增: {len(plan['added'])}",
        f"將補空白單價: {len(plan['updated'])}",
        f"已略過: {len(plan['skipped'])}",
        f"價格衝突略過: {len(plan['conflicts'])}",
    ])
    if plan.get("conflicts"):
        lines.extend(["", "價格衝突預覽（前 5 筆）："])
        for conflict in plan["conflicts"][:5]:
            existing = conflict.get("existing", {})
            incoming = conflict.get("incoming", {})
            lines.append(
                "- "
                + _item_label(existing)
                + f"：既有 {existing.get('單價', '') or '空白'}，匯入 {incoming.get('單價', '') or '空白'}"
            )
        if len(plan["conflicts"]) > 5:
            lines.append(f"...另有 {len(plan['conflicts']) - 5} 筆 conflict")
    if not apply:
        lines.append("尚未寫入；確認後才會更新目標價目表。")
    return "\n".join(lines)


def apply_price_table_import_plan(
    plan: dict[str, Any],
    current_pricebook: dict[str, Any],
    *,
    target_path: str | None = None,
) -> None:
    merged = dict(current_pricebook)
    merged["items"] = plan["items"]
    save_material_pricebook(merged, target_path)


def load_and_plan_price_table_import(
    price_table_path: str,
    *,
    target_path: str | None = None,
    sheet_name: str | None = None,
) -> tuple[list[dict[str, Any]], PricebookValidationReport, dict[str, Any], dict[str, Any]]:
    table_items = load_price_table_items(price_table_path, sheet_name=sheet_name)
    report = validate_price_table_items(table_items)
    current = load_material_pricebook(target_path)
    if report.ok:
        plan = build_price_table_import_plan(table_items, current)
    else:
        plan = {
            "items": current.get("items", []),
            "added": [],
            "updated": [],
            "skipped": [],
            "conflicts": [],
            "existing_count": len(current.get("items", [])),
            "candidate_count": len(table_items),
        }
    return table_items, report, plan, current


def _load_csv_rows(path: str) -> list[list[Any]]:
    last_error: UnicodeDecodeError | None = None
    for encoding in ("utf-8-sig", "cp950"):
        try:
            with open(path, "r", encoding=encoding, newline="") as f:
                sample = f.read(4096)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample)
                except csv.Error:
                    dialect = csv.excel
                return [list(row) for row in csv.reader(f, dialect)]
        except UnicodeDecodeError as exc:
            last_error = exc
            continue
    if last_error:
        raise last_error
    return []


def _load_xlsx_rows(path: str, *, sheet_name: str | None) -> list[list[Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError("缺少 openpyxl，無法讀取 Excel 價格表") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _rows_to_items(rows: list[list[Any]]) -> list[dict[str, Any]]:
    header_idx = _find_header_row(rows)
    if header_idx < 0:
        raise ValueError("找不到價格表表頭；至少需要零件類型與單價等欄位")

    headers = [_canonical_header(cell) for cell in rows[header_idx]]
    if not any(headers):
        raise ValueError("價格表表頭沒有可辨識欄位")

    items: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        if _row_blank(row):
            continue
        item: dict[str, Any] = {"__source_row": row_number}
        for col_idx, field in enumerate(headers):
            if not field or col_idx >= len(row):
                continue
            item[field] = _cell_to_text(row[col_idx])
        if any(str(item.get(field, "")).strip() for field in CANONICAL_FIELDS):
            items.append(item)
    return items


def _find_header_row(rows: list[list[Any]]) -> int:
    for idx, row in enumerate(rows):
        fields = {_canonical_header(cell) for cell in row}
        fields.discard("")
        if "零件類型" in fields and ("單價" in fields or "單位" in fields or "材質" in fields):
            return idx
        if len(fields.intersection({"id", "零件類型", "尺寸", "SCH", "材質", "單價"})) >= 3:
            return idx
    return -1


def _canonical_header(value: Any) -> str:
    key = re.sub(r"[\s_／/\\\-:：()（）]+", "", str(value or "").strip().lower().replace("\ufeff", ""))
    if key in HEADER_TO_FIELD:
        return HEADER_TO_FIELD[key]
    if "單價" in key or key == "price" or "unitprice" in key:
        return "單價"
    if "生效" in key and ("日" in key or "date" in key):
        return "生效日"
    if "備註" in key or "remark" in key or "note" in key:
        return "備註"
    return ""


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _row_blank(row: list[Any]) -> bool:
    return not any(str(cell or "").strip() for cell in row)


def _strip_private_fields(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {key: value for key, value in item.items() if not str(key).startswith("__")}
        for item in items
        if isinstance(item, dict)
    ]


def _merge_blank_price_item(existing: dict[str, str], candidate: dict[str, str]) -> dict[str, str]:
    merged = copy.deepcopy(existing)
    for key in CANONICAL_FIELDS:
        incoming = str(candidate.get(key, "")).strip()
        if not incoming:
            continue
        if key == "單價" or not str(merged.get(key, "")).strip():
            merged[key] = incoming
    return merged


def _skip_reason(existing_price: str, incoming_price: str) -> str:
    if not incoming_price:
        return "匯入列沒有單價可補"
    if existing_price:
        return "既有項目已有相同單價"
    return "已存在且無需更新"


def _item_label(item: dict[str, Any]) -> str:
    parts = [
        str(item.get("零件類型", "")).strip(),
        str(item.get("尺寸", "")).strip(),
        str(item.get("SCH", "")).strip(),
        str(item.get("材質", "")).strip(),
    ]
    return " / ".join(part for part in parts if part) or str(item.get("id", "")).strip() or "未命名項目"
