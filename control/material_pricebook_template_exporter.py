# -*- coding: utf-8 -*-
"""
material_pricebook_template_exporter.py - 材料補價表模板匯出

把目前專案價目表匯出成可用 Excel 填價的 CSV/XLSX。匯出的欄位與
material_pricebook_table_importer 相同，填好單價後可直接再匯入。
"""

from __future__ import annotations

import csv
import os
from typing import Any

from material_pricebook import load_material_pricebook, normalize_pricebook_items
from material_pricebook_table_importer import CANONICAL_FIELDS


PRICE_TABLE_TEMPLATE_HEADERS = list(CANONICAL_FIELDS)


def build_price_table_template_items(
    items: list[dict[str, Any]],
    *,
    only_unpriced: bool = True,
) -> list[dict[str, str]]:
    normalized = normalize_pricebook_items(items)
    if not only_unpriced:
        return normalized
    return [
        item for item in normalized
        if str(item.get("零件類型", "")).strip() and not str(item.get("單價", "")).strip()
    ]


def export_price_table_template(
    path: str,
    items: list[dict[str, Any]],
    *,
    only_unpriced: bool = True,
) -> dict[str, Any]:
    rows = build_price_table_template_items(items, only_unpriced=only_unpriced)
    _ensure_parent_dir(path)
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        _write_csv(path, rows)
    elif ext in (".xlsx", ".xlsm"):
        _write_xlsx(path, rows)
    else:
        raise ValueError("補價表只支援匯出 .csv / .xlsx / .xlsm")
    return {
        "path": path,
        "count": len(rows),
        "only_unpriced": only_unpriced,
        "headers": list(PRICE_TABLE_TEMPLATE_HEADERS),
    }


def export_pricebook_template_from_file(
    output_path: str,
    *,
    source_path: str | None = None,
    only_unpriced: bool = True,
) -> dict[str, Any]:
    pricebook = load_material_pricebook(source_path)
    return export_price_table_template(
        output_path,
        pricebook.get("items", []),
        only_unpriced=only_unpriced,
    )


def _write_csv(path: str, rows: list[dict[str, str]]) -> None:
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=PRICE_TABLE_TEMPLATE_HEADERS, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in PRICE_TABLE_TEMPLATE_HEADERS})


def _write_xlsx(path: str, rows: list[dict[str, str]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:
        raise ImportError("缺少 openpyxl，無法匯出 Excel 補價表") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "材料補價表"
    ws.append(PRICE_TABLE_TEMPLATE_HEADERS)
    for row in rows:
        ws.append([row.get(field, "") for field in PRICE_TABLE_TEMPLATE_HEADERS])

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, width in enumerate((22, 18, 10, 10, 18, 10, 8, 12, 12, 14, 28), start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width

    wb.save(path)


def _ensure_parent_dir(path: str) -> None:
    parent = os.path.dirname(os.path.abspath(path))
    if parent:
        os.makedirs(parent, exist_ok=True)
