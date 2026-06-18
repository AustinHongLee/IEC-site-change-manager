# -*- coding: utf-8 -*-
"""
gui_dialogs.py — 對話框集合（PyQt6 版）

包含以下七個對話框/類別：
- WeldBackfillDialog        焊口補登工具
- WeldSnapshotManager       焊口快照管理器（純邏輯，無 UI）
- WeldDuplicateCheckDialog  焊口重複性檢查
- SupplementInfoDialog      補充資料夾資訊
- RecordManagerDialog       記錄管理工具
- EditRecordDialog          編輯記錄
- WeldOrphanAuditDialog     孤兒焊口稽查工具
"""

import os
import re
import json
import shutil
from datetime import datetime

from PyQt6.QtWidgets import (
    QDialog, QWidget, QVBoxLayout, QHBoxLayout, QGridLayout, QFormLayout,
    QLabel, QPushButton, QLineEdit, QComboBox, QTextEdit, QGroupBox,
    QTreeWidget, QTreeWidgetItem, QHeaderView, QSplitter, QProgressBar,
    QMessageBox, QFileDialog, QScrollArea, QFrame, QSizePolicy,
    QApplication, QTableWidget, QTableWidgetItem,
)
from PyQt6.QtCore import Qt, QSize
from PyQt6.QtGui import QColor, QFont

from config import BASE_DIR, ATTACHMENTS_ROOT, OUTPUT_ROOT
from record_manager import _load_store, _save_store, auto_backup, RECORDS_JSON_PATH
from settings_manager import get_weld_control_config
from theme import Colors, Fonts, set_button_role, make_separator, make_hint_label
from utils import atomic_write_json, reentry_guard


# ---------------------------------------------------------------------------
#  Helpers
# ---------------------------------------------------------------------------
_FONT_TITLE = Fonts.heading()
_FONT_SUBTITLE = Fonts.subheading()
_FONT_CODE = Fonts.code(11)
_FONT_SMALL = Fonts.small()


def _make_hline() -> QFrame:
    """水平分隔線"""
    return make_separator()


def _show_reentry_notice(owner):
    QMessageBox.information(owner, "提示", "此動作正在執行中，請稍候。")


# =====================================================================
#  WeldBackfillDialog - 焊口補登工具
# =====================================================================
class WeldBackfillDialog(QDialog):
    """焊口補登工具對話框"""

    def __init__(self, parent: QWidget, settings_panel=None):
        super().__init__(parent)
        self.settings_panel = settings_panel
        self.missing_welds: list[dict] = []
        self.selected_indices: set[int] = set()

        self.setWindowTitle("📋 焊口補登工具 - 掃描未登錄焊口")
        self.resize(900, 650)
        self.setModal(True)

        self._build_ui()
        self._scan_missing_welds()

    # ---- UI ----
    def _build_ui(self):
        root = QVBoxLayout(self)

        # 標題
        root.setSpacing(10)
        header = QVBoxLayout()
        lbl = QLabel("掃描所有資料群，找出未登錄於焊口管制表的焊口")
        lbl.setFont(_FONT_SUBTITLE)
        lbl.setStyleSheet(f"color: {Colors.TEXT_SECONDARY}; border:none; background:transparent;")
        header.addWidget(lbl)
        self.stat_label = QLabel("正在掃描...")
        self.stat_label.setStyleSheet(f"color: {Colors.PRIMARY}; font-weight:bold; border:none; background:transparent;")
        header.addWidget(self.stat_label)
        root.addLayout(header)

        # 篩選區
        filter_grp = QGroupBox("🔍 篩選")
        flay = QHBoxLayout(filter_grp)
        flay.addWidget(QLabel("日期:"))
        self.date_combo = QComboBox()
        self.date_combo.addItem("全部")
        self.date_combo.currentTextChanged.connect(lambda: self._apply_filter())
        flay.addWidget(self.date_combo)

        flay.addSpacing(20)
        flay.addWidget(QLabel("流水號:"))
        self.serial_edit = QLineEdit()
        self.serial_edit.setFixedWidth(90)
        self.serial_edit.returnPressed.connect(self._apply_filter)
        flay.addWidget(self.serial_edit)

        btn_filter = QPushButton("篩選")
        btn_filter.clicked.connect(self._apply_filter)
        flay.addWidget(btn_filter)
        btn_rescan = QPushButton("重新掃描")
        btn_rescan.clicked.connect(self._scan_missing_welds)
        flay.addWidget(btn_rescan)
        flay.addStretch()
        root.addWidget(filter_grp)

        # 焊口列表
        list_grp = QGroupBox("📋 未登錄焊口")
        llay = QVBoxLayout(list_grp)
        self.tree = QTreeWidget()
        self.tree.setHeaderLabels(["☑", "日期", "資料夾", "流水號", "焊口編號", "尺寸", "代碼"])
        hdr = self.tree.header()
        hdr.resizeSection(0, 40)
        hdr.resizeSection(1, 90)
        hdr.resizeSection(2, 180)
        hdr.resizeSection(3, 80)
        hdr.resizeSection(4, 100)
        hdr.resizeSection(5, 60)
        hdr.resizeSection(6, 100)
        self.tree.setAlternatingRowColors(True)
        self.tree.itemClicked.connect(self._on_tree_click)
        self.tree.itemDoubleClicked.connect(self._on_tree_double_click)
        llay.addWidget(self.tree)
        root.addWidget(list_grp, stretch=1)

        # 選取工具
        sel_row = QHBoxLayout()
        for txt, fn in [("☑ 全選", self._select_all),
                        ("☐ 取消全選", self._deselect_all),
                        ("☑ 反選", self._invert_selection)]:
            b = QPushButton(txt)
            b.clicked.connect(fn)
            sel_row.addWidget(b)
        self.select_count_label = QLabel("已選取: 0 筆")
        self.select_count_label.setStyleSheet(f"color: {Colors.PRIMARY}; font-weight:bold; border:none; background:transparent;")
        sel_row.addSpacing(20)
        sel_row.addWidget(self.select_count_label)
        sel_row.addStretch()
        root.addLayout(sel_row)

        # 寫入設定
        write_grp = QGroupBox("📝 寫入設定")
        wlay = QHBoxLayout(write_grp)
        wlay.addWidget(QLabel("另存檔名:"))
        self.output_name_edit = QLineEdit("焊口管制表_補登")
        self.output_name_edit.setFixedWidth(200)
        wlay.addWidget(self.output_name_edit)
        wlay.addWidget(QLabel(f"_{datetime.now().strftime('%Y.%m.%d')}.xlsx"))
        wlay.addStretch()
        root.addWidget(write_grp)

        # 按鈕
        btn_row = QHBoxLayout()
        btn_preview = QPushButton("👁 預覽匯出")
        set_button_role(btn_preview, "flat")
        btn_preview.clicked.connect(self._preview_selected)
        btn_row.addWidget(btn_preview)
        btn_write = QPushButton("📝 寫入選取的焊口")
        set_button_role(btn_write, "primary")
        btn_write.clicked.connect(self._write_selected)
        btn_row.addWidget(btn_write)
        btn_open = QPushButton("📂 開啟資料夾")
        set_button_role(btn_open, "flat")
        btn_open.clicked.connect(self._open_selected_folder)
        btn_row.addWidget(btn_open)
        btn_row.addStretch()
        btn_close = QPushButton("關閉")
        btn_close.clicked.connect(self.close)
        btn_row.addWidget(btn_close)
        root.addLayout(btn_row)

    # ---- Logic ----
    def _scan_missing_welds(self):
        self.stat_label.setText("正在掃描資料群...")
        QApplication.processEvents()

        try:
            from weld_control import find_missing_welds, get_missing_welds_summary

            if self.settings_panel and hasattr(self.settings_panel, 'serial_format_var'):
                serial_format = self.settings_panel.serial_format_var
            else:
                config = get_weld_control_config()
                serial_format = config.get("serial_format", "{serial}")

            self.missing_welds = find_missing_welds(serial_format=serial_format)
            summary = get_missing_welds_summary(self.missing_welds)

            if self.missing_welds:
                dates = sorted(summary["by_date"].keys())
                self.date_combo.clear()
                self.date_combo.addItem("全部")
                self.date_combo.addItems(dates)

                self.stat_label.setText(
                    f"找到 {summary['total_missing']} 個未登錄焊口，"
                    f"涉及 {len(summary['by_serial'])} 個流水號，"
                    f"{len(summary['by_date'])} 個日期"
                )
            else:
                self.stat_label.setText("✅ 所有焊口皆已登錄，無需補登")

            self.selected_indices.clear()
            self._refresh_tree()

        except Exception as e:
            self.stat_label.setText(f"❌ 掃描失敗: {e}")
            import traceback; traceback.print_exc()

    def _refresh_tree(self, filtered_welds=None):
        self.tree.clear()
        welds = filtered_welds if filtered_welds is not None else self.missing_welds

        for i, w in enumerate(welds):
            orig_idx = self.missing_welds.index(w) if filtered_welds else i
            mark = "☑" if orig_idx in self.selected_indices else "☐"
            item = QTreeWidgetItem([
                mark,
                w["date"],
                w["folder_name"],
                w["serial_formatted"],
                w["weld_id"],
                f'{w["size"]}\"' if w.get("size") else "",
                w["code"],
            ])
            item.setData(0, Qt.ItemDataRole.UserRole, orig_idx)
            self.tree.addTopLevelItem(item)

        self._update_select_count()

    def _apply_filter(self):
        date_f = self.date_combo.currentText()
        serial_f = self.serial_edit.text().strip()
        filtered = []
        for w in self.missing_welds:
            if date_f != "全部" and w["date"] != date_f:
                continue
            if serial_f and serial_f not in w["serial"] and serial_f not in w["serial_formatted"]:
                continue
            filtered.append(w)
        self._refresh_tree(filtered)
        self.stat_label.setText(f"顯示 {len(filtered)} / {len(self.missing_welds)} 筆")

    def _on_tree_click(self, item: QTreeWidgetItem, column: int):
        if column == 0:
            idx = item.data(0, Qt.ItemDataRole.UserRole)
            if idx in self.selected_indices:
                self.selected_indices.remove(idx)
            else:
                self.selected_indices.add(idx)
            item.setText(0, "☑" if idx in self.selected_indices else "☐")
            self._update_select_count()

    def _on_tree_double_click(self, item: QTreeWidgetItem, column: int):
        idx = item.data(0, Qt.ItemDataRole.UserRole)
        if idx is not None and 0 <= idx < len(self.missing_welds):
            folder_path = self.missing_welds[idx]["folder_path"]
            if os.path.exists(folder_path):
                os.startfile(folder_path)

    def _select_all(self):
        for i in range(self.tree.topLevelItemCount()):
            it = self.tree.topLevelItem(i)
            idx = it.data(0, Qt.ItemDataRole.UserRole)
            self.selected_indices.add(idx)
            it.setText(0, "☑")
        self._update_select_count()

    def _deselect_all(self):
        for i in range(self.tree.topLevelItemCount()):
            it = self.tree.topLevelItem(i)
            idx = it.data(0, Qt.ItemDataRole.UserRole)
            self.selected_indices.discard(idx)
            it.setText(0, "☐")
        self._update_select_count()

    def _invert_selection(self):
        for i in range(self.tree.topLevelItemCount()):
            it = self.tree.topLevelItem(i)
            idx = it.data(0, Qt.ItemDataRole.UserRole)
            if idx in self.selected_indices:
                self.selected_indices.remove(idx)
            else:
                self.selected_indices.add(idx)
            it.setText(0, "☑" if idx in self.selected_indices else "☐")
        self._update_select_count()

    def _update_select_count(self):
        self.select_count_label.setText(f"已選取: {len(self.selected_indices)} 筆")

    def _open_selected_folder(self):
        items = self.tree.selectedItems()
        if not items:
            QMessageBox.information(self, "提示", "請先選擇一筆資料")
            return
        idx = items[0].data(0, Qt.ItemDataRole.UserRole)
        if idx is not None and 0 <= idx < len(self.missing_welds):
            p = self.missing_welds[idx]["folder_path"]
            if os.path.exists(p):
                os.startfile(p)

    # ---- 共用：建立匯出用資料列 ----
    def _build_export_rows(self):
        """
        根據 selected_indices 建立要匯出到 Excel 的資料列。

        Returns:
            (columns, rows, row_flags, warnings)
            columns:   [str, ...]  — 欄位名稱（與 Excel 對應）
            rows:      [[cell_value, ...], ...]  — 每列資料
            row_flags: [{"fallback_remark": bool}, ...]  — 每列標記
            warnings:  [str, ...]  — 潛在問題提示
        """
        from settings_manager import get_weld_control_table_path, get_weld_control_config
        from weld_control import build_report_id_lookup
        from utils import resolve_col_map

        config = get_weld_control_config()
        report_id_lookup = build_report_id_lookup()

        col_serial  = config.get("col_serial", "流水號")
        col_weld_no = config.get("col_weld_no", "焊口編號")
        col_date    = config.get("col_date", "登錄日期")

        # 固定欄位順序
        columns = [col_serial, col_weld_no, "SIZE", "材質", "厚度", col_date, "備註"]
        today = datetime.now().strftime("%Y/%m/%d")
        rows = []
        row_flags = []
        warnings = []

        for idx in sorted(self.selected_indices):
            if idx >= len(self.missing_welds):
                continue
            w = self.missing_welds[idx]

            serial_key = w["serial_formatted"].lstrip("0") or "0"
            report_id = report_id_lookup.get((serial_key, w["weld_id"]))
            is_fallback = False
            if not report_id:
                report_id = f"{w['date']}-{w['serial']}"
                is_fallback = True

            row = [
                w["serial_formatted"],
                w["weld_id"],
                w.get("size", ""),
                w.get("material", ""),
                w.get("thickness", ""),
                today,
                report_id,
            ]
            rows.append(row)
            row_flags.append({"fallback_remark": is_fallback})

            # 檢測潛在問題
            if is_fallback:
                warnings.append(
                    f"⚠ 流水號 {w['serial_formatted']} 焊口 {w['weld_id']}："
                    f"找不到對應報告編號，使用備用格式 \"{report_id}\""
                )
            if not w.get("size"):
                warnings.append(
                    f"⚠ 流水號 {w['serial_formatted']} 焊口 {w['weld_id']}："
                    f"尺寸(SIZE) 為空"
                )

        return columns, rows, row_flags, warnings

    # ---- 預覽匯出 ----
    def _preview_selected(self):
        if not self.selected_indices:
            QMessageBox.information(self, "提示", "請先選取要預覽的焊口")
            return
        try:
            columns, rows, row_flags, warnings = self._build_export_rows()
            dlg = BackfillPreviewDialog(self, columns, rows, row_flags, warnings)
            dlg.exec()
        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"產生預覽失敗: {e}")
            import traceback; traceback.print_exc()

    def _write_selected(self):
        if not self.selected_indices:
            QMessageBox.information(self, "提示", "請先選取要補登的焊口")
            return

        count = len(self.selected_indices)
        if QMessageBox.question(
            self, "確認",
            f"確定要補登 {count} 個焊口嗎？\n\n將另存為新檔案，不會覆蓋原檔。"
        ) != QMessageBox.StandardButton.Yes:
            return

        try:
            from settings_manager import get_weld_control_table_path, get_weld_control_config
            from openpyxl import load_workbook
            from weld_control import build_report_id_lookup

            table_path = get_weld_control_table_path()
            config = get_weld_control_config()
            report_id_lookup = build_report_id_lookup()

            wb = load_workbook(table_path)
            sheet_name = config.get("sheet_name", "焊口編號明細")
            if sheet_name not in wb.sheetnames:
                QMessageBox.critical(self, "錯誤", f"找不到工作表: {sheet_name}")
                return
            ws = wb[sheet_name]

            headers = [cell.value for cell in ws[1]]
            col_map = {h: i + 1 for i, h in enumerate(headers) if h}

            # 同義字欄位解析（焊↔銲 等）
            from utils import resolve_col_map
            col_serial = config.get("col_serial", "流水號")
            col_weld_no = config.get("col_weld_no", "焊口編號")
            col_date = config.get("col_date", "登錄日期")

            # 預先解析各欄位在 col_map 中的實際 column index
            _ci_serial  = resolve_col_map(col_map, col_serial)
            _ci_weld    = resolve_col_map(col_map, col_weld_no)
            _ci_date    = resolve_col_map(col_map, col_date)
            _ci_size    = resolve_col_map(col_map, "SIZE") or resolve_col_map(col_map, "尺寸")
            _ci_mat     = resolve_col_map(col_map, "材質") or resolve_col_map(col_map, "MATERIAL")
            _ci_thick   = resolve_col_map(col_map, "厚度") or resolve_col_map(col_map, "SCH")
            _ci_remark  = resolve_col_map(col_map, "備註")

            next_row = ws.max_row + 1
            today = datetime.now().strftime("%Y/%m/%d")
            written = 0
            found_report_id_count = 0

            for idx in sorted(self.selected_indices):
                if idx >= len(self.missing_welds):
                    continue
                w = self.missing_welds[idx]

                if _ci_serial is not None:
                    ws.cell(row=next_row, column=_ci_serial, value=w["serial_formatted"])
                if _ci_weld is not None:
                    ws.cell(row=next_row, column=_ci_weld, value=w["weld_id"])

                size_value = w.get("size", "")
                if size_value and _ci_size is not None:
                    ws.cell(row=next_row, column=_ci_size, value=size_value)

                material_value = w.get("material", "")
                if material_value and _ci_mat is not None:
                    ws.cell(row=next_row, column=_ci_mat, value=material_value)

                thickness_value = w.get("thickness", "")
                if thickness_value and _ci_thick is not None:
                    ws.cell(row=next_row, column=_ci_thick, value=thickness_value)

                if _ci_date is not None:
                    ws.cell(row=next_row, column=_ci_date, value=today)

                serial_key = w["serial_formatted"].lstrip("0") or "0"
                report_id = report_id_lookup.get((serial_key, w["weld_id"]))
                if report_id:
                    found_report_id_count += 1
                else:
                    report_id = f"{w['date']}-{w['serial']}"
                if _ci_remark is not None:
                    ws.cell(row=next_row, column=_ci_remark, value=report_id)

                next_row += 1
                written += 1

            output_name = self.output_name_edit.text().strip() or "焊口管制表_補登"
            date_str = datetime.now().strftime("%Y.%m.%d")
            base_dir = os.path.dirname(table_path)
            output_path = os.path.join(base_dir, f"{output_name}_{date_str}.xlsx")
            if os.path.exists(output_path):
                for i in range(1, 100):
                    output_path = os.path.join(base_dir, f"{output_name}_{date_str}_{i}.xlsx")
                    if not os.path.exists(output_path):
                        break

            wb.save(output_path)
            wb.close()

            try:
                from weld_control import get_weld_manager
                get_weld_manager().invalidate_cache()
            except Exception:
                pass

            QMessageBox.information(
                self, "成功",
                f"✅ 已補登 {written} 個焊口\n\n"
                "報告編號：\n"
                f"  • 從記錄清單查到: {found_report_id_count} 筆\n"
                f"  • 使用備用格式: {written - found_report_id_count} 筆\n\n"
                f"另存至:\n{output_path}\n\n"
                "請確認內容無誤後，取代原管制表檔案。"
            )

            if QMessageBox.question(
                self, "開啟檔案", "是否開啟補登後的檔案？"
            ) == QMessageBox.StandardButton.Yes:
                os.startfile(output_path)

            self._scan_missing_welds()

        except PermissionError:
            QMessageBox.critical(self, "錯誤", "焊口管制表被開啟中，請先關閉 Excel")
        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"寫入失敗: {e}")
            import traceback; traceback.print_exc()


# =====================================================================
#  BackfillPreviewDialog - 補登預覽表
# =====================================================================
class BackfillPreviewDialog(QDialog):
    """以表格形式預覽即將寫入 Excel 的補登資料，方便確認正確性。"""

    def __init__(self, parent: QWidget, columns: list[str],
                 rows: list[list], row_flags: list[dict], warnings: list[str]):
        super().__init__(parent)
        self.setWindowTitle("👁 補登預覽 — 即將寫入 Excel 的資料")
        self.resize(820, 520)
        self.setModal(True)

        root = QVBoxLayout(self)
        root.setSpacing(8)

        # 摘要
        summary = QLabel(f"共 {len(rows)} 筆資料，{len(warnings)} 個警告")
        summary.setFont(_FONT_SUBTITLE)
        summary.setStyleSheet(
            f"color: {Colors.WARNING if warnings else Colors.SUCCESS};"
            "font-weight:bold; border:none; background:transparent;"
        )
        root.addWidget(summary)

        # 表格
        table = QTableWidget(len(rows), len(columns))
        table.setHorizontalHeaderLabels(columns)
        table.setAlternatingRowColors(True)
        table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        table.horizontalHeader().setStretchLastSection(True)
        table.verticalHeader().setDefaultSectionSize(26)

        warn_bg = QColor(Colors.WARNING_BG)
        empty_bg = QColor(Colors.DANGER_BG)
        remark_col = len(columns) - 1

        for r, row_data in enumerate(rows):
            flags = row_flags[r] if r < len(row_flags) else {}
            for c, val in enumerate(row_data):
                cell_text = str(val) if val else ""
                item = QTableWidgetItem(cell_text)
                if not cell_text:
                    item.setBackground(empty_bg)
                elif c == remark_col and flags.get("fallback_remark"):
                    item.setBackground(warn_bg)
                table.setItem(r, c, item)

        table.resizeColumnsToContents()
        root.addWidget(table, stretch=1)

        # 警告區
        if warnings:
            warn_grp = QGroupBox(f"⚠ 警告 ({len(warnings)})")
            warn_grp.setStyleSheet(
                f"QGroupBox {{ color: {Colors.WARNING}; font-weight: bold; }}"
            )
            wlay = QVBoxLayout(warn_grp)
            warn_text = QTextEdit()
            warn_text.setReadOnly(True)
            warn_text.setMaximumHeight(120)
            warn_text.setPlainText("\n".join(warnings))
            warn_text.setStyleSheet(
                f"color: {Colors.WARNING}; background: {Colors.WARNING_BG};"
                "font-size: 12px;"
            )
            wlay.addWidget(warn_text)
            root.addWidget(warn_grp)

        # 按鈕
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_close = QPushButton("關閉")
        btn_close.clicked.connect(self.close)
        btn_row.addWidget(btn_close)
        root.addLayout(btn_row)


# =====================================================================
#  WeldSnapshotManager - 焊口快照管理器（純邏輯）
# =====================================================================
class WeldSnapshotManager:
    """焊口快照管理器 — 用於快速檢查焊口重複性"""

    SNAPSHOT_FILE = "weld_snapshot.json"

    def __init__(self, attachments_root: str | None = None):
        self.attachments_root = attachments_root or ATTACHMENTS_ROOT
        self.snapshot_path = os.path.join(
            os.path.dirname(self.attachments_root), "records", self.SNAPSHOT_FILE
        )
        self.snapshot: dict | None = None

    def _portable_attachments_root(self) -> str:
        project_root = os.path.dirname(os.path.abspath(self.attachments_root))
        try:
            rel = os.path.relpath(os.path.abspath(self.attachments_root), project_root)
            if not rel.startswith("..") and not os.path.isabs(rel):
                return rel.replace("\\", "/")
        except Exception:
            pass
        return self.attachments_root

    def load_snapshot(self) -> dict | None:
        if os.path.exists(self.snapshot_path):
            try:
                with open(self.snapshot_path, "r", encoding="utf-8") as f:
                    self.snapshot = json.load(f)
                    return self.snapshot
            except Exception:
                pass
        return None

    def build_snapshot(self, progress_callback=None) -> dict:
        pattern = re.compile(r"^(\d+)([rab])(.+)$")
        snapshot = {
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "attachments_root": self._portable_attachments_root(),
            "folder_count": 0,
            "weld_count": 0,
            "weld_index": {},
            "folders": {},
        }

        date_folders = []
        for name in os.listdir(self.attachments_root):
            path = os.path.join(self.attachments_root, name)
            if os.path.isdir(path) and re.match(r"^\d{8}$", name):
                date_folders.append((name, path))

        total_folders = 0
        for _, dp in date_folders:
            for sn in os.listdir(dp):
                sp = os.path.join(dp, sn)
                if os.path.isdir(sp) and not sn.startswith("_"):
                    total_folders += 1

        processed = 0
        for date_name, date_path in date_folders:
            for sub_name in os.listdir(date_path):
                sub_path = os.path.join(date_path, sub_name)
                if not os.path.isdir(sub_path) or sub_name.startswith("_"):
                    continue
                processed += 1
                if progress_callback:
                    progress_callback(processed, total_folders, sub_name)

                folder_key = f"{date_name}/{sub_name}"
                raw_serial = sub_name.split("_")[0] if "_" in sub_name else ""
                serial = raw_serial.lstrip("0") or raw_serial
                if not serial:
                    continue

                welds = self._extract_welds_from_folder(sub_path, sub_name, pattern)
                if not welds:
                    continue

                snapshot["folders"][folder_key] = {
                    "serial": serial,
                    "raw_serial": raw_serial,
                    "welds": welds,
                }
                snapshot["folder_count"] += 1

                for w in welds:
                    weld_key = f"{serial}_{w['no']}_{w['mark']}"
                    if weld_key not in snapshot["weld_index"]:
                        snapshot["weld_index"][weld_key] = []
                    snapshot["weld_index"][weld_key].append(folder_key)
                    snapshot["weld_count"] += 1

        self.snapshot = snapshot
        return snapshot

    def _extract_welds_from_folder(self, folder_path, folder_name, pattern) -> list:
        welds = []
        weld_info_path = os.path.join(folder_path, "weld_info.json")
        if os.path.exists(weld_info_path):
            try:
                with open(weld_info_path, "r", encoding="utf-8") as f:
                    info = json.load(f)
                    for w in info.get("welds", []):
                        wn = str(w.get("weld_no", ""))
                        mk = str(w.get("mark", ""))
                        m_sp = re.match(r'^(\d+)([rab])', wn)
                        if m_sp:
                            wn = m_sp.group(1)
                            mk = mk or m_sp.group(2)
                        welds.append({
                            "no": wn,
                            "mark": mk,
                            "size": str(w.get("size", "")),
                        })
                    if welds:
                        return welds
            except Exception:
                pass

        gw_path = os.path.join(folder_path, "GroupWeld.txt")
        if os.path.exists(gw_path):
            try:
                with open(gw_path, "r", encoding="utf-8") as f:
                    for line in f:
                        code = line.strip()
                        if code and not code.startswith("#"):
                            m = pattern.match(code)
                            if m:
                                welds.append({"no": m.group(1), "mark": m.group(2), "size": m.group(3)})
                if welds:
                    return welds
            except Exception:
                pass

        parts = folder_name.split("_")
        if len(parts) >= 2:
            for part in parts[1:]:
                m = pattern.match(part)
                if m:
                    welds.append({"no": m.group(1), "mark": m.group(2), "size": m.group(3)})
        return welds

    def save_snapshot(self) -> bool:
        if not self.snapshot:
            return False
        try:
            os.makedirs(os.path.dirname(self.snapshot_path), exist_ok=True)
            tmp = self.snapshot_path + ".tmp"
            with open(tmp, "w", encoding="utf-8") as f:
                json.dump(self.snapshot, f, ensure_ascii=False, indent=2)
            os.replace(tmp, self.snapshot_path)
            return True
        except Exception as e:
            print(f"儲存快照失敗: {e}")
            return False

    def find_duplicates(self) -> list:
        if not self.snapshot:
            self.load_snapshot()
        if not self.snapshot:
            return []
        duplicates = []
        for weld_key, folders in self.snapshot["weld_index"].items():
            if len(folders) > 1:
                parts = weld_key.split("_")
                if len(parts) >= 3:
                    serial, weld_no, mark = parts[0], parts[1], parts[2]
                else:
                    serial = parts[0] if len(parts) > 0 else ""
                    weld_no = parts[1] if len(parts) > 1 else ""
                    mark = ""
                duplicates.append({
                    "serial": serial,
                    "weld_no": weld_no,
                    "mark": mark,
                    "weld_key": weld_key,
                    "folders": folders,
                    "count": len(folders),
                })
        duplicates.sort(key=lambda x: (
            int(x["serial"]) if x["serial"].isdigit() else 0,
            int(x["weld_no"]) if x["weld_no"].isdigit() else 0,
            x["mark"],
        ))
        return duplicates

    def check_weld_exists(self, serial: str, weld_no: str, mark: str = "") -> list:
        if not self.snapshot:
            self.load_snapshot()
        if not self.snapshot:
            return []
        serial = serial.lstrip("0") or serial
        weld_key = f"{serial}_{weld_no}_{mark}"
        return self.snapshot["weld_index"].get(weld_key, [])

    def get_snapshot_info(self) -> dict | None:
        if not self.snapshot:
            self.load_snapshot()
        if not self.snapshot:
            return None
        return {
            "created_at": self.snapshot.get("created_at", "未知"),
            "folder_count": self.snapshot.get("folder_count", 0),
            "weld_count": self.snapshot.get("weld_count", 0),
            "unique_welds": len(self.snapshot.get("weld_index", {})),
        }


# =====================================================================
#  WeldDuplicateCheckDialog - 焊口重複性檢查
# =====================================================================
class WeldDuplicateCheckDialog(QDialog):
    """焊口重複性檢查對話框"""

    def __init__(self, parent: QWidget):
        super().__init__(parent)
        self.snapshot_manager = WeldSnapshotManager()

        self.setWindowTitle("🔍 焊口重複性檢查")
        self.resize(900, 600)
        self.setModal(True)

        self._build_ui()
        self._load_snapshot_info()

    def _build_ui(self):
        root = QVBoxLayout(self)

        # --- 快照資訊 ---
        info_grp = QGroupBox("📷 快照資訊")
        ilay = QVBoxLayout(info_grp)
        self.snapshot_info_label = QLabel("載入中...")
        ilay.addWidget(self.snapshot_info_label)

        btn_row = QHBoxLayout()
        btn_build = QPushButton("📷 建立/更新快照")
        set_button_role(btn_build, "primary")
        btn_build.clicked.connect(self._build_snapshot)
        btn_row.addWidget(btn_build)
        btn_check = QPushButton("🔍 檢查重複焊口")
        set_button_role(btn_check, "success")
        btn_check.clicked.connect(self._check_duplicates)
        btn_row.addWidget(btn_check)
        btn_row.addStretch()
        ilay.addLayout(btn_row)

        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.hide()
        ilay.addWidget(self.progress_bar)
        self.progress_label = QLabel("")
        self.progress_label.hide()
        ilay.addWidget(self.progress_label)
        root.addWidget(info_grp)

        # --- 結果 (上下分割) ---
        splitter = QSplitter(Qt.Orientation.Vertical)

        # 上: 重複焊口列表
        top_grp = QGroupBox("📊 重複焊口列表")
        tlay = QVBoxLayout(top_grp)
        self.tree = QTreeWidget()
        self.tree.setHeaderLabels(["流水號", "焊口號", "標記", "重複次數", "所在資料夾"])
        hdr = self.tree.header()
        hdr.resizeSection(0, 70)
        hdr.resizeSection(1, 70)
        hdr.resizeSection(2, 50)
        hdr.resizeSection(3, 70)
        hdr.setStretchLastSection(True)
        self.tree.setAlternatingRowColors(True)
        self.tree.itemSelectionChanged.connect(self._on_duplicate_select)
        tlay.addWidget(self.tree)
        splitter.addWidget(top_grp)

        # 下: 涉及資料夾
        bot_grp = QGroupBox("📁 涉及的資料夾（選擇上方項目查看）")
        blay = QVBoxLayout(bot_grp)
        self.folder_tree = QTreeWidget()
        self.folder_tree.setHeaderLabels(["日期", "資料夾名稱", "完整路徑"])
        fhdr = self.folder_tree.header()
        fhdr.resizeSection(0, 100)
        fhdr.resizeSection(1, 200)
        fhdr.setStretchLastSection(True)
        self.folder_tree.setAlternatingRowColors(True)
        blay.addWidget(self.folder_tree)

        action_row = QHBoxLayout()
        for txt, fn in [("📂 開啟資料夾", self._open_selected_folder),
                        ("📂 開啟 Output", self._open_selected_output)]:
            b = QPushButton(txt)
            b.clicked.connect(fn)
            action_row.addWidget(b)
        action_row.addWidget(_make_hline())
        for txt, fn in [("📦 封存選中資料夾", self._archive_selected_folder),
                        ("✏️ 編輯焊口資訊", self._edit_selected_folder)]:
            b = QPushButton(txt)
            b.clicked.connect(fn)
            action_row.addWidget(b)
        action_row.addStretch()
        blay.addLayout(action_row)
        splitter.addWidget(bot_grp)
        root.addWidget(splitter, stretch=1)

        # 統計
        self.stat_label = QLabel("")
        root.addWidget(self.stat_label)

        # 底部
        bottom = QHBoxLayout()
        btn_export = QPushButton("📋 匯出報告")
        btn_export.clicked.connect(self._export_report)
        bottom.addWidget(btn_export)
        bottom.addStretch()
        btn_close = QPushButton("關閉")
        btn_close.clicked.connect(self.close)
        bottom.addWidget(btn_close)
        root.addLayout(bottom)

    # ---- Logic ----
    def _load_snapshot_info(self):
        info = self.snapshot_manager.get_snapshot_info()
        if info:
            self.snapshot_info_label.setText(
                f"快照時間: {info['created_at']}\n"
                f"資料夾數: {info['folder_count']} 個\n"
                f"焊口總數: {info['weld_count']} 個\n"
                f"不重複焊口: {info['unique_welds']} 個"
            )
        else:
            self.snapshot_info_label.setText("⚠️ 尚未建立快照\n請點擊「建立/更新快照」按鈕")

    def _build_snapshot(self):
        if QMessageBox.question(
            self, "確認", "將掃描整個 attachments 目錄建立快照\n這可能需要一些時間，是否繼續？"
        ) != QMessageBox.StandardButton.Yes:
            return

        self.progress_bar.show()
        self.progress_label.show()
        self.progress_bar.setValue(0)

        def cb(current, total, name):
            pct = int(current / total * 100) if total else 0
            self.progress_bar.setValue(pct)
            self.progress_label.setText(f"掃描中: {name} ({current}/{total})")
            QApplication.processEvents()

        try:
            self.snapshot_manager.build_snapshot(cb)
            self.snapshot_manager.save_snapshot()
            self._load_snapshot_info()
            self.progress_label.setText("✅ 快照建立完成！")
            QMessageBox.information(self, "完成", "快照建立完成！")
        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"建立快照失敗: {e}")
            import traceback; traceback.print_exc()
        finally:
            self.progress_bar.hide()

    def _check_duplicates(self):
        self.tree.clear()
        if not self.snapshot_manager.snapshot:
            self.snapshot_manager.load_snapshot()
        if not self.snapshot_manager.snapshot:
            QMessageBox.warning(self, "提示", "請先建立快照")
            return

        duplicates = self.snapshot_manager.find_duplicates()
        if not duplicates:
            self.stat_label.setText("✅ 沒有發現重複的焊口！")
            self.stat_label.setStyleSheet(f"color: {Colors.SUCCESS}; font-weight:bold; border:none; background:transparent;")
            QMessageBox.information(self, "檢查結果", "✅ 沒有發現重複的焊口！")
            return

        for dup in duplicates:
            folders_str = ", ".join(dup["folders"])
            self.tree.addTopLevelItem(QTreeWidgetItem([
                dup["serial"], dup["weld_no"], dup["mark"],
                str(dup["count"]), folders_str,
            ]))

        self.stat_label.setText(f"⚠️ 發現 {len(duplicates)} 組重複焊口！")
        self.stat_label.setStyleSheet(f"color: {Colors.DANGER}; font-weight:bold; border:none; background:transparent;")

    def _export_report(self):
        duplicates = self.snapshot_manager.find_duplicates()
        if not duplicates:
            QMessageBox.information(self, "提示", "沒有重複焊口，無需匯出")
            return

        filepath, _ = QFileDialog.getSaveFileName(
            self, "匯出報告", "", "文字檔案 (*.txt);;所有檔案 (*.*)"
        )
        if not filepath:
            return

        try:
            with open(filepath, "w", encoding="utf-8") as f:
                f.write("焊口重複性檢查報告\n")
                f.write("=" * 60 + "\n")
                f.write(f"快照時間: {self.snapshot_manager.snapshot.get('created_at', '未知')}\n")
                f.write(f"重複焊口數: {len(duplicates)} 組\n")
                f.write("=" * 60 + "\n\n")
                for dup in duplicates:
                    mark_str = f", 標記: {dup['mark']}" if dup.get("mark") else ""
                    f.write(f"流水號: {dup['serial']}, 焊口號: {dup['weld_no']}{mark_str}\n")
                    f.write(f"  重複 {dup['count']} 次，出現在:\n")
                    for folder in dup["folders"]:
                        f.write(f"    - {folder}\n")
                    f.write("\n")
            QMessageBox.information(self, "完成", f"報告已匯出至:\n{filepath}")
        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"匯出失敗: {e}")

    def _on_duplicate_select(self):
        self.folder_tree.clear()
        items = self.tree.selectedItems()
        if not items:
            return
        it = items[0]
        folders_str = it.text(4)
        for fk in [x.strip() for x in folders_str.split(",")]:
            if "/" in fk:
                date, folder_name = fk.split("/", 1)
            else:
                date, folder_name = "", fk
            full_path = os.path.join(ATTACHMENTS_ROOT, fk)
            self.folder_tree.addTopLevelItem(QTreeWidgetItem([date, folder_name, full_path]))

    def _get_selected_folder_path(self) -> str | None:
        items = self.folder_tree.selectedItems()
        if not items:
            return None
        return items[0].text(2)

    def _open_selected_folder(self):
        p = self._get_selected_folder_path()
        if not p:
            QMessageBox.information(self, "提示", "請先在下方列表選擇一個資料夾")
            return
        if os.path.exists(p):
            os.startfile(p)
        else:
            QMessageBox.critical(self, "錯誤", f"資料夾不存在:\n{p}")

    def _open_selected_output(self):
        p = self._get_selected_folder_path()
        if not p:
            QMessageBox.information(self, "提示", "請先在下方列表選擇一個資料夾")
            return
        if ATTACHMENTS_ROOT in p:
            op = p.replace(ATTACHMENTS_ROOT, OUTPUT_ROOT)
            if os.path.exists(op):
                os.startfile(op)
            else:
                QMessageBox.information(self, "提示", f"Output 資料夾不存在:\n{op}\n\n可能尚未產出報告")
        else:
            QMessageBox.critical(self, "錯誤", "無法轉換路徑")

    def _archive_selected_folder(self):
        items = self.folder_tree.selectedItems()
        if not items:
            QMessageBox.information(self, "提示", "請先在下方列表選擇一個資料夾")
            return
        it = items[0]
        date, folder_name, full_path = it.text(0), it.text(1), it.text(2)

        if not os.path.exists(full_path):
            QMessageBox.critical(self, "錯誤", f"資料夾不存在:\n{full_path}")
            return

        if QMessageBox.question(
            self, "確認封存",
            f"確定要封存此資料夾嗎？\n\n📁 {folder_name}\n📅 日期: {date}\n\n封存後將移動到 _archived 資料夾"
        ) != QMessageBox.StandardButton.Yes:
            return

        try:
            from operation_journal import OperationJournal
            date_folder = os.path.dirname(full_path)
            archived_folder = os.path.join(date_folder, "_archived")
            os.makedirs(archived_folder, exist_ok=True)
            dest_path = os.path.join(archived_folder, folder_name)
            if os.path.exists(dest_path):
                QMessageBox.critical(self, "錯誤", f"封存目標已存在:\n{dest_path}")
                return
            with OperationJournal(BASE_DIR, "archive_selected_folder", {
                "folder": folder_name,
                "date": date,
            }) as journal:
                journal.step("move_attachment_folder", source=full_path, target=dest_path)
                shutil.move(full_path, dest_path)
            QMessageBox.information(self, "完成", f"✅ 已封存:\n{folder_name}\n\n到:\n{dest_path}")
            idx = self.folder_tree.indexOfTopLevelItem(it)
            self.folder_tree.takeTopLevelItem(idx)
            if QMessageBox.question(
                self, "更新快照", "封存後建議更新快照\n是否立即更新？"
            ) == QMessageBox.StandardButton.Yes:
                self._build_snapshot()
        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"封存失敗: {e}")
            import traceback; traceback.print_exc()

    def _edit_selected_folder(self):
        items = self.folder_tree.selectedItems()
        if not items:
            QMessageBox.information(self, "提示", "請先在下方列表選擇一個資料夾")
            return
        it = items[0]
        date, folder_name, full_path = it.text(0), it.text(1), it.text(2)
        if not os.path.exists(full_path):
            QMessageBox.critical(self, "錯誤", f"資料夾不存在:\n{full_path}")
            return

        parts = folder_name.split("_")
        serial = parts[0].lstrip("0") or parts[0] if parts else ""
        welds_str = "_".join(parts[1:]) if len(parts) > 1 else ""

        weld_info = None
        wip = os.path.join(full_path, "weld_info.json")
        if os.path.exists(wip):
            try:
                with open(wip, "r", encoding="utf-8") as f:
                    weld_info = json.load(f)
            except Exception:
                pass

        record = {
            "date": date,
            "folder_name": folder_name,
            "folder_path": full_path,
            "serial": serial,
            "welds_str": welds_str,
            "weld_info": weld_info,
        }

        def on_refresh():
            if QMessageBox.question(
                self, "更新快照", "編輯後建議更新快照\n是否立即更新？"
            ) == QMessageBox.StandardButton.Yes:
                self._build_snapshot()

        dlg = EditRecordDialog(self, record, on_refresh)
        dlg.exec()


# =====================================================================
#  SupplementInfoDialog - 補充資料夾資訊
# =====================================================================
class SupplementInfoDialog(QDialog):
    """補充資料夾資訊對話框（材質、厚度等）"""

    def __init__(self, parent: QWidget, folders_to_supplement: list[dict]):
        super().__init__(parent)
        self.folders = folders_to_supplement
        self.current_index = 0
        self.weld_entries: list[dict] = []
        self.ref_welds: list[dict] = []
        self.current_serial = ""

        self.weld_manager = None
        try:
            from weld_control import init_weld_manager_from_settings
            self.weld_manager = init_weld_manager_from_settings()
        except Exception:
            pass

        self.setWindowTitle("📝 補充資料夾資訊")
        self.resize(950, 700)
        self.setModal(True)

        self._build_ui()
        if self.folders:
            self._load_folder(0)

    def _build_ui(self):
        root = QVBoxLayout(self)

        # 進度 + 導航
        nav_row = QHBoxLayout()
        self.progress_label = QLabel("")
        nav_row.addWidget(self.progress_label)
        nav_row.addStretch()
        self.btn_prev = QPushButton("◀ 上一個")
        self.btn_prev.clicked.connect(self._prev_folder)
        nav_row.addWidget(self.btn_prev)
        self.btn_next = QPushButton("下一個 ▶")
        self.btn_next.clicked.connect(self._next_folder)
        nav_row.addWidget(self.btn_next)
        root.addLayout(nav_row)

        # 資料夾資訊
        info_grp = QGroupBox("📁 資料夾資訊")
        ilay = QVBoxLayout(info_grp)
        self.folder_label = QLabel("")
        self.folder_label.setFont(_FONT_CODE)
        ilay.addWidget(self.folder_label)
        self.date_label = QLabel("")
        ilay.addWidget(self.date_label)
        root.addWidget(info_grp)

        # 左右分割
        splitter = QSplitter(Qt.Orientation.Horizontal)

        # ---- 左側：補充資訊 ----
        left_w = QWidget()
        left_lay = QVBoxLayout(left_w)
        left_lay.setContentsMargins(0, 0, 0, 0)

        weld_grp = QGroupBox("🔧 補充焊口資訊")
        wg_lay = QVBoxLayout(weld_grp)

        # 表頭
        hdr_row = QHBoxLayout()
        for txt, w in [("焊口號", 60), ("標記", 40), ("尺寸", 50), ("材質", 100), ("厚度", 80)]:
            lb = QLabel(txt)
            lb.setFixedWidth(w)
            hdr_row.addWidget(lb)
        hdr_row.addStretch()
        wg_lay.addLayout(hdr_row)

        # 可捲動區域
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_content = QWidget()
        self.scroll_layout = QVBoxLayout(self.scroll_content)
        self.scroll_layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        self.scroll_area.setWidget(self.scroll_content)
        wg_lay.addWidget(self.scroll_area, stretch=1)

        # 快速填充
        qf_row = QHBoxLayout()
        qf_row.addWidget(QLabel("快速填充："))
        self.quick_material = QComboBox()
        self.quick_material.setEditable(True)
        self.quick_material.addItems(["", "A106-B", "A312-316L", "A358-316L", "A333-6", "A312-304", "A358-304"])
        self.quick_material.setFixedWidth(120)
        qf_row.addWidget(self.quick_material)
        self.quick_thickness = QComboBox()
        self.quick_thickness.setEditable(True)
        self.quick_thickness.addItems(["", "S10", "S20", "S40", "S80", "STD", "XS", "SCH40", "SCH80"])
        self.quick_thickness.setFixedWidth(100)
        qf_row.addWidget(self.quick_thickness)
        btn_qf = QPushButton("套用")
        btn_qf.clicked.connect(self._apply_quick_fill)
        qf_row.addWidget(btn_qf)
        qf_row.addStretch()
        wg_lay.addLayout(qf_row)

        left_lay.addWidget(weld_grp)
        splitter.addWidget(left_w)

        # ---- 右側：參考資訊 ----
        right_w = QWidget()
        right_lay = QVBoxLayout(right_w)
        right_lay.setContentsMargins(0, 0, 0, 0)

        ref_grp = QGroupBox("📊 參考資訊（焊口管制表）")
        ref_lay = QVBoxLayout(ref_grp)

        self.ref_status_label = QLabel("")
        ref_lay.addWidget(self.ref_status_label)

        self.ref_tree = QTreeWidget()
        self.ref_tree.setHeaderLabels(["焊口號", "尺寸", "材質", "厚度", "狀態"])
        rhdr = self.ref_tree.header()
        rhdr.resizeSection(0, 70)
        rhdr.resizeSection(1, 60)
        rhdr.resizeSection(2, 100)
        rhdr.resizeSection(3, 70)
        rhdr.resizeSection(4, 80)
        self.ref_tree.setAlternatingRowColors(True)
        ref_lay.addWidget(self.ref_tree, stretch=1)

        ref_btn_row = QHBoxLayout()
        btn_copy = QPushButton("📋 複製選中的材質/厚度")
        btn_copy.clicked.connect(self._copy_from_ref)
        ref_btn_row.addWidget(btn_copy)
        btn_copy_match = QPushButton("📥 複製全部相同尺寸")
        btn_copy_match.clicked.connect(self._copy_matching_size)
        ref_btn_row.addWidget(btn_copy_match)
        ref_btn_row.addStretch()
        ref_lay.addLayout(ref_btn_row)

        right_lay.addWidget(ref_grp)
        splitter.addWidget(right_w)

        root.addWidget(splitter, stretch=1)

        # 底部按鈕
        bot_row = QHBoxLayout()
        self.btn_save = QPushButton("💾 儲存此資料夾")
        set_button_role(self.btn_save, "primary")
        self.btn_save.clicked.connect(self._save_current)
        bot_row.addWidget(self.btn_save)
        btn_save_next = QPushButton("💾 儲存並下一個")
        set_button_role(btn_save_next, "success")
        btn_save_next.clicked.connect(self._save_and_next)
        bot_row.addWidget(btn_save_next)
        self.status_label = QLabel("")
        self.status_label.setStyleSheet(f"color: {Colors.SUCCESS}; font-weight:bold; border:none; background:transparent;")
        bot_row.addWidget(self.status_label)
        bot_row.addStretch()
        btn_close = QPushButton("關閉")
        btn_close.clicked.connect(self.close)
        bot_row.addWidget(btn_close)
        root.addLayout(bot_row)

    # ---- folder loading ----
    def _load_folder(self, index: int):
        if index < 0 or index >= len(self.folders):
            return
        self.current_index = index
        folder = self.folders[index]

        self.progress_label.setText(f"資料夾 {index + 1} / {len(self.folders)}")
        self.btn_prev.setEnabled(index > 0)
        self.btn_next.setEnabled(index < len(self.folders) - 1)

        self.folder_label.setText(folder["folder_name"])
        self.date_label.setText(f"日期: {folder['date']}")

        raw_serial = folder["folder_name"].split("_")[0] if "_" in folder["folder_name"] else ""
        self.current_serial = raw_serial.lstrip("0") or raw_serial

        # 清空
        while self.scroll_layout.count():
            child = self.scroll_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()
        self.weld_entries = []

        folder_path = folder["folder_path"]
        folder_name = folder["folder_name"]
        existing_welds = []

        weld_info_path = os.path.join(folder_path, "weld_info.json")
        if os.path.exists(weld_info_path):
            try:
                with open(weld_info_path, "r", encoding="utf-8") as f:
                    weld_info = json.load(f)
                    existing_welds = weld_info.get("welds", [])
            except Exception:
                pass

        if not existing_welds:
            pattern = re.compile(r"^(\d+)([rab])(.+)$")
            gw_path = os.path.join(folder_path, "GroupWeld.txt")
            if os.path.exists(gw_path):
                try:
                    with open(gw_path, "r", encoding="utf-8") as f:
                        for line in f:
                            code = line.strip()
                            if code and not code.startswith("#"):
                                m = pattern.match(code)
                                if m:
                                    existing_welds.append({
                                        "weld_no": m.group(1), "mark": m.group(2),
                                        "size": m.group(3), "material": "", "thickness": "",
                                    })
                except Exception:
                    pass
            if not existing_welds:
                parts = folder_name.split("_")
                if len(parts) >= 2:
                    for part in parts[1:]:
                        m = pattern.match(part)
                        if m:
                            existing_welds.append({
                                "weld_no": m.group(1), "mark": m.group(2),
                                "size": m.group(3), "material": "", "thickness": "",
                            })

        for w in existing_welds:
            wn = str(w.get("weld_no", ""))
            mk = str(w.get("mark", ""))
            m_sp = re.match(r'^(\d+)([rab])', wn)
            if m_sp:
                wn = m_sp.group(1)
                mk = mk or m_sp.group(2)
            self._add_weld_row(
                wn, mk, str(w.get("size", "")),
                str(w.get("material", "")), str(w.get("thickness", "")),
            )

        self.status_label.setText("")
        self._load_reference_data()

    def _add_weld_row(self, weld_no="", mark="", size="", material="", thickness=""):
        row_w = QWidget()
        row_lay = QHBoxLayout(row_w)
        row_lay.setContentsMargins(0, 2, 0, 2)

        weld_no_edit = QLineEdit(weld_no)
        weld_no_edit.setFixedWidth(60)
        weld_no_edit.setReadOnly(True)
        row_lay.addWidget(weld_no_edit)

        mark_edit = QLineEdit(mark)
        mark_edit.setFixedWidth(40)
        mark_edit.setReadOnly(True)
        row_lay.addWidget(mark_edit)

        size_edit = QLineEdit(size)
        size_edit.setFixedWidth(50)
        row_lay.addWidget(size_edit)

        mat_combo = QComboBox()
        mat_combo.setEditable(True)
        mat_combo.addItems(["", "A106-B", "A312-316L", "A358-316L", "A333-6", "A312-304", "A358-304"])
        mat_combo.setCurrentText(material)
        mat_combo.setFixedWidth(120)
        row_lay.addWidget(mat_combo)

        thk_combo = QComboBox()
        thk_combo.setEditable(True)
        thk_combo.addItems(["", "S10", "S20", "S40", "S80", "STD", "XS", "SCH40", "SCH80"])
        thk_combo.setCurrentText(thickness)
        thk_combo.setFixedWidth(100)
        row_lay.addWidget(thk_combo)

        row_lay.addStretch()
        self.scroll_layout.addWidget(row_w)

        self.weld_entries.append({
            "weld_no": weld_no_edit,
            "mark": mark_edit,
            "size": size_edit,
            "material": mat_combo,
            "thickness": thk_combo,
        })

    def _load_reference_data(self):
        self.ref_tree.clear()
        self.ref_welds = []
        if not self.weld_manager or not self.current_serial:
            self.ref_status_label.setText("⚠️ 未設定焊口管制表或無法解析流水號")
            return

        try:
            welds = self.weld_manager.get_all_welds_by_serial(self.current_serial)
            if not welds:
                self.ref_status_label.setText(f"📊 流水號 {self.current_serial}: 無資料")
                return

            self.ref_status_label.setText(f"📊 流水號 {self.current_serial}: 共 {len(welds)} 個焊口")

            weld_no_keys = ["銲口編號", "焊口編號", "焊口號", "Weld No.", "weld_no"]
            size_keys = ["尺寸", "Size", "size", "管徑"]
            material_keys = ["材質", "Material", "material", "材料"]
            thickness_keys = ["厚度", "Thickness", "thickness", "壁厚", "SCH"]
            status_keys = ["組銲完成日期", "狀態", "Status", "status", "進度"]

            def gv(row, keys):
                for k in keys:
                    if k in row and row[k]:
                        return str(row[k])
                return ""

            for weld in welds:
                wn = gv(weld, weld_no_keys)
                sz = gv(weld, size_keys)
                mt = gv(weld, material_keys)
                tk_ = gv(weld, thickness_keys)
                st = gv(weld, status_keys)
                self.ref_tree.addTopLevelItem(QTreeWidgetItem([wn, sz, mt, tk_, st]))
                self.ref_welds.append({"weld_no": wn, "size": sz, "material": mt, "thickness": tk_, "status": st})
        except Exception as e:
            self.ref_status_label.setText(f"❌ 載入失敗: {e}")
            import traceback; traceback.print_exc()

    def _copy_from_ref(self):
        items = self.ref_tree.selectedItems()
        if not items:
            QMessageBox.information(self, "提示", "請先在右側參考表格中選擇一個焊口")
            return
        it = items[0]
        material, thickness = it.text(2), it.text(3)
        self.quick_material.setCurrentText(material)
        self.quick_thickness.setCurrentText(thickness)
        self._apply_quick_fill()
        self.status_label.setText(f"✅ 已套用: {material}, {thickness}")

    def _copy_matching_size(self):
        if not self.ref_welds:
            QMessageBox.information(self, "提示", "沒有參考資料")
            return

        size_to_info: dict[str, dict] = {}
        for ref in self.ref_welds:
            sz = ref.get("size", "").strip()
            if sz and (ref.get("material") or ref.get("thickness")):
                if sz not in size_to_info:
                    size_to_info[sz] = {"material": ref.get("material", ""), "thickness": ref.get("thickness", "")}

        if not size_to_info:
            QMessageBox.information(self, "提示", "參考資料中沒有可用的材質/厚度資訊")
            return

        matched = 0
        for entry in self.weld_entries:
            sz = entry["size"].text().strip()
            if sz in size_to_info:
                info = size_to_info[sz]
                if info["material"]:
                    entry["material"].setCurrentText(info["material"])
                if info["thickness"]:
                    entry["thickness"].setCurrentText(info["thickness"])
                matched += 1

        if matched:
            self.status_label.setText(f"✅ 已套用 {matched} 個焊口")
        else:
            QMessageBox.information(self, "提示", "沒有找到尺寸匹配的焊口")

    def _apply_quick_fill(self):
        mat = self.quick_material.currentText().strip()
        thk = self.quick_thickness.currentText().strip()
        for entry in self.weld_entries:
            if mat:
                entry["material"].setCurrentText(mat)
            if thk:
                entry["thickness"].setCurrentText(thk)

    def _save_current(self) -> bool:
        folder = self.folders[self.current_index]
        folder_path = folder["folder_path"]
        welds = []
        for entry in self.weld_entries:
            wn = entry["weld_no"].text().strip()
            mk = entry["mark"].text().strip()
            sz = entry["size"].text().strip()
            mt = entry["material"].currentText().strip()
            tk_ = entry["thickness"].currentText().strip()
            if wn and mk and sz:
                welds.append({
                    "weld_no": wn, "mark": mk, "size": sz,
                    "material": mt, "thickness": tk_,
                    "code": f"{wn}{mk}{sz}",
                })
        if not welds:
            QMessageBox.critical(self, "錯誤", "沒有有效的焊口資訊")
            return False

        # 防空口編號重複
        _seen_ids: set[str] = set()
        _unique: list[dict] = []
        for w in welds:
            _wid = f"{w['weld_no']}{w['mark']}"
            if _wid not in _seen_ids:
                _seen_ids.add(_wid)
                _unique.append(w)
        welds = _unique

        raw_series = folder["folder_name"].split("_")[0] if "_" in folder["folder_name"] else ""
        series = raw_series.lstrip("0") or raw_series

        weld_info = {"series": series, "date": folder["date"], "welds": welds}
        try:
            wip = os.path.join(folder_path, "weld_info.json")
            with open(wip, "w", encoding="utf-8") as f:
                json.dump(weld_info, f, ensure_ascii=False, indent=2)
            self.status_label.setText("✅ 已儲存 weld_info.json")
            return True
        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"儲存失敗: {e}")
            return False

    def _save_and_next(self):
        if self._save_current():
            if self.current_index < len(self.folders) - 1:
                self._load_folder(self.current_index + 1)
            else:
                QMessageBox.information(self, "完成", "✅ 已處理完所有資料夾！")

    def _prev_folder(self):
        if self.current_index > 0:
            self._load_folder(self.current_index - 1)

    def _next_folder(self):
        if self.current_index < len(self.folders) - 1:
            self._load_folder(self.current_index + 1)


# =====================================================================
#  RecordManagerDialog - 記錄管理工具（⚠️ 已棄用，功能已合併至 RecordManagerPanel）
# =====================================================================
class RecordManagerDialog(QDialog):
    """記錄管理工具 — ⚠️ DEPRECATED: 功能已合併至 gui_panels.RecordManagerPanel 的右鍵選單"""

    def __init__(self, parent: QWidget):
        super().__init__(parent)
        self.all_records: list[dict] = []
        self.selected_record: dict | None = None

        self.setWindowTitle("📋 記錄管理工具")
        self.resize(1100, 700)
        self.setModal(True)

        self._build_ui()
        self._scan_all_records()

    def _build_ui(self):
        root = QVBoxLayout(self)

        # 標題
        lbl_title = QLabel("記錄管理工具")
        lbl_title.setFont(_FONT_TITLE)
        root.addWidget(lbl_title)
        lbl_sub = QLabel("管理 attachments 中的資料群：歸檔、修改、檢查產出狀態")
        lbl_sub.setStyleSheet(f"color: {Colors.TEXT_SECONDARY}; border:none; background:transparent;")
        root.addWidget(lbl_sub)

        # 主區域 (左右分割)
        splitter = QSplitter(Qt.Orientation.Horizontal)

        # ---- 左側：記錄列表 ----
        left_w = QWidget()
        left_lay = QVBoxLayout(left_w)
        left_lay.setContentsMargins(0, 0, 0, 0)

        # 篩選
        filter_grp = QGroupBox("🔍 篩選")
        flay = QHBoxLayout(filter_grp)
        flay.addWidget(QLabel("日期:"))
        self.filter_date = QComboBox()
        self.filter_date.setFixedWidth(120)
        self.filter_date.addItem("全部")
        self.filter_date.currentTextChanged.connect(lambda: self._apply_filter())
        flay.addWidget(self.filter_date)

        flay.addSpacing(10)
        flay.addWidget(QLabel("流水號:"))
        self.filter_serial = QLineEdit()
        self.filter_serial.setFixedWidth(80)
        self.filter_serial.returnPressed.connect(self._apply_filter)
        flay.addWidget(self.filter_serial)

        flay.addSpacing(10)
        flay.addWidget(QLabel("狀態:"))
        self.filter_status = QComboBox()
        self.filter_status.addItems(["全部", "已產出", "未產出", "已歸檔"])
        self.filter_status.setFixedWidth(100)
        self.filter_status.currentTextChanged.connect(lambda: self._apply_filter())
        flay.addWidget(self.filter_status)

        flay.addStretch()
        btn_rescan = QPushButton("🔄 重新掃描")
        btn_rescan.clicked.connect(self._scan_all_records)
        flay.addWidget(btn_rescan)
        left_lay.addWidget(filter_grp)

        # 列表
        list_grp = QGroupBox("📋 資料群列表")
        llay = QVBoxLayout(list_grp)
        self.tree = QTreeWidget()
        self.tree.setHeaderLabels(["日期", "資料夾", "流水號", "焊口", "狀態"])
        hdr = self.tree.header()
        hdr.resizeSection(0, 80)
        hdr.resizeSection(1, 180)
        hdr.resizeSection(2, 60)
        hdr.resizeSection(3, 120)
        hdr.resizeSection(4, 80)
        self.tree.itemSelectionChanged.connect(self._on_record_select)
        self.tree.setAlternatingRowColors(True)
        llay.addWidget(self.tree)
        left_lay.addWidget(list_grp, stretch=1)

        self.stats_label = QLabel("載入中...")
        self.stats_label.setStyleSheet(f"color: {Colors.PRIMARY}; border:none; background:transparent;")
        left_lay.addWidget(self.stats_label)

        splitter.addWidget(left_w)

        # ---- 右側：操作區 ----
        right_w = QWidget()
        right_lay = QVBoxLayout(right_w)
        right_lay.setContentsMargins(0, 0, 0, 0)

        # 詳細資訊
        detail_grp = QGroupBox("📄 詳細資訊")
        dlay = QVBoxLayout(detail_grp)
        self.detail_text = QTextEdit()
        self.detail_text.setReadOnly(True)
        dlay.addWidget(self.detail_text)
        right_lay.addWidget(detail_grp)

        # 操作按鈕
        action_grp = QGroupBox("🔧 操作")
        alay = QVBoxLayout(action_grp)

        for txt, desc, fn in [
            ("📦 歸檔選中記錄", "將選中的資料群移至 _archived 資料夾", self._archive_record),
            ("✏️ 修改焊口/尺寸", "修改焊口編號、尺寸，同步更新資料夾", self._edit_record),
            ("🔍 檢查未產出報告", "列出 attachments 中尚未產出報告的資料群", self._check_unproduced),
            ("📂 開啟資料夾", "", self._open_folder),
        ]:
            btn = QPushButton(txt)
            btn.setFixedWidth(220)
            btn.clicked.connect(fn)
            alay.addWidget(btn, alignment=Qt.AlignmentFlag.AlignHCenter)
            if desc:
                lb = QLabel(desc)
                lb.setFont(_FONT_SMALL)
                lb.setStyleSheet(f"color: {Colors.TEXT_MUTED}; border:none; background:transparent;")
                lb.setAlignment(Qt.AlignmentFlag.AlignCenter)
                alay.addWidget(lb)
            if fn != self._open_folder:
                alay.addWidget(_make_hline())
        right_lay.addWidget(action_grp)

        splitter.addWidget(right_w)

        root.addWidget(splitter, stretch=1)

        # 底部
        bot = QHBoxLayout()
        bot.addStretch()
        btn_close = QPushButton("關閉")
        btn_close.setFixedWidth(120)
        btn_close.clicked.connect(self.close)
        bot.addWidget(btn_close)
        root.addLayout(bot)

    # ---- Logic ----
    def _scan_all_records(self):
        self.all_records.clear()
        try:
            from config import ATTACHMENTS_ROOT, OUTPUT_ROOT
            if not os.path.exists(ATTACHMENTS_ROOT):
                QMessageBox.critical(self, "錯誤", f"attachments 資料夾不存在:\n{ATTACHMENTS_ROOT}")
                return

            dates = set()
            for date_dir in sorted(os.listdir(ATTACHMENTS_ROOT)):
                date_path = os.path.join(ATTACHMENTS_ROOT, date_dir)
                if not os.path.isdir(date_path) or date_dir.startswith("_"):
                    continue
                dates.add(date_dir)

                for fn in sorted(os.listdir(date_path)):
                    fp = os.path.join(date_path, fn)
                    if not os.path.isdir(fp):
                        continue
                    parts = fn.split("_")
                    if not parts or not parts[0].isdigit():
                        continue
                    serial = parts[0]
                    welds_str = "_".join(parts[1:]) if len(parts) > 1 else ""

                    output_dp = os.path.join(OUTPUT_ROOT, date_dir)
                    has_output = False
                    output_file = None
                    if os.path.exists(output_dp):
                        for of in os.listdir(output_dp):
                            if of.startswith(f"管線修改單_{serial.zfill(4)}_") and of.endswith(".xlsm"):
                                has_output = True
                                output_file = of
                                break

                    weld_info = None
                    wip = os.path.join(fp, "weld_info.json")
                    if os.path.exists(wip):
                        try:
                            with open(wip, "r", encoding="utf-8") as f:
                                weld_info = json.load(f)
                        except Exception:
                            pass

                    self.all_records.append({
                        "date": date_dir, "folder_name": fn, "folder_path": fp,
                        "serial": serial, "welds_str": welds_str,
                        "has_output": has_output, "output_file": output_file,
                        "weld_info": weld_info, "is_archived": False,
                    })

            # 掃描 _archived
            archived_root = os.path.join(ATTACHMENTS_ROOT, "_archived")
            if os.path.exists(archived_root):
                for date_dir in sorted(os.listdir(archived_root)):
                    date_path = os.path.join(archived_root, date_dir)
                    if not os.path.isdir(date_path):
                        continue
                    for fn in sorted(os.listdir(date_path)):
                        fp = os.path.join(date_path, fn)
                        if not os.path.isdir(fp):
                            continue
                        parts = fn.split("_")
                        if not parts or not parts[0].isdigit():
                            continue
                        serial = parts[0]
                        welds_str = "_".join(parts[1:]) if len(parts) > 1 else ""
                        self.all_records.append({
                            "date": date_dir, "folder_name": fn, "folder_path": fp,
                            "serial": serial, "welds_str": welds_str,
                            "has_output": False, "output_file": None,
                            "weld_info": None, "is_archived": True,
                        })

            self.filter_date.clear()
            self.filter_date.addItem("全部")
            self.filter_date.addItems(sorted(dates, reverse=True))
            self._apply_filter()

        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"掃描失敗: {e}")
            import traceback; traceback.print_exc()

    def _apply_filter(self):
        self.tree.clear()
        d_f = self.filter_date.currentText()
        s_f = self.filter_serial.text().strip()
        st_f = self.filter_status.currentText()

        filtered = []
        for r in self.all_records:
            if d_f and d_f != "全部" and r["date"] != d_f:
                continue
            if s_f and s_f not in r["serial"]:
                continue
            if st_f == "已產出" and not r["has_output"]:
                continue
            elif st_f == "未產出" and (r["has_output"] or r["is_archived"]):
                continue
            elif st_f == "已歸檔" and not r["is_archived"]:
                continue
            filtered.append(r)

        for r in filtered:
            status = "📦 已歸檔" if r["is_archived"] else ("✅ 已產出" if r["has_output"] else "⏳ 未產出")
            self.tree.addTopLevelItem(QTreeWidgetItem([
                r["date"], r["folder_name"], r["serial"], r["welds_str"], status,
            ]))

        total = len(self.all_records)
        produced = sum(1 for r in self.all_records if r["has_output"])
        unproduced = sum(1 for r in self.all_records if not r["has_output"] and not r["is_archived"])
        archived = sum(1 for r in self.all_records if r["is_archived"])
        self.stats_label.setText(
            f"總計: {total} 筆 | 已產出: {produced} | 未產出: {unproduced} | 已歸檔: {archived} | 篩選結果: {len(filtered)}"
        )

    def _on_record_select(self):
        items = self.tree.selectedItems()
        if not items:
            self.selected_record = None
            return
        it = items[0]
        date, folder_name = it.text(0), it.text(1)
        for r in self.all_records:
            if r["date"] == date and r["folder_name"] == folder_name:
                self.selected_record = r
                self._show_record_detail(r)
                break

    def _show_record_detail(self, record: dict):
        info = (
            f"📁 資料夾: {record['folder_name']}\n"
            f"📅 日期: {record['date']}\n"
            f"🔢 流水號: {record['serial']}\n"
            f"🔧 焊口: {record['welds_str']}\n"
            f"\n📍 路徑:\n{record['folder_path']}\n"
        )
        if record["has_output"]:
            info += f"\n✅ 已產出: {record['output_file']}\n"
        elif record["is_archived"]:
            info += "\n📦 已歸檔\n"
        else:
            info += "\n⏳ 尚未產出報告\n"

        if record.get("weld_info"):
            info += "\n📋 焊口詳細 (weld_info.json):\n"
            for w in record["weld_info"].get("welds", []):
                info += f"  • {w.get('code', '')} "
                if w.get("material"):
                    info += f"[{w['material']}]"
                if w.get("thickness"):
                    info += f"[{w['thickness']}]"
                info += "\n"
        self.detail_text.setPlainText(info)

    def _archive_record(self):
        if not self.selected_record:
            QMessageBox.information(self, "提示", "請先選擇一筆記錄")
            return
        record = self.selected_record

        if record["is_archived"]:
            if QMessageBox.question(
                self, "還原記錄",
                f"確定要還原以下記錄嗎？\n\n📁 {record['folder_name']}\n📅 {record['date']}\n\n將從 _archived 移回原位置"
            ) == QMessageBox.StandardButton.Yes:
                self._restore_record(record)
        else:
            if QMessageBox.question(
                self, "歸檔記錄",
                f"確定要歸檔以下記錄嗎？\n\n📁 {record['folder_name']}\n📅 {record['date']}\n\n"
                f"將移至 _archived/{record['date']}/ 資料夾"
            ) == QMessageBox.StandardButton.Yes:
                self._do_archive_record(record)

    @reentry_guard("_record_move_in_progress", _show_reentry_notice)
    def _do_archive_record(self, record: dict):
        try:
            from config import ATTACHMENTS_ROOT
            from operation_journal import OperationJournal
            archived_date_dir = os.path.join(ATTACHMENTS_ROOT, "_archived", record["date"])
            os.makedirs(archived_date_dir, exist_ok=True)
            src = record["folder_path"]
            dst = os.path.join(archived_date_dir, record["folder_name"])
            if os.path.exists(dst):
                ts = datetime.now().strftime("%H%M%S")
                dst = os.path.join(archived_date_dir, f"{record['folder_name']}_{ts}")
            with OperationJournal(BASE_DIR, "archive_record_dialog", {
                "folder": record["folder_name"],
                "date": record["date"],
            }) as journal:
                journal.step("move_attachment_folder", source=src, target=dst)
                shutil.move(src, dst)
            QMessageBox.information(self, "成功", f"✅ 已歸檔至:\n{dst}")
            self._scan_all_records()
        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"歸檔失敗: {e}")

    @reentry_guard("_record_move_in_progress", _show_reentry_notice)
    def _restore_record(self, record: dict):
        try:
            from config import ATTACHMENTS_ROOT
            from operation_journal import OperationJournal
            target_dir = os.path.join(ATTACHMENTS_ROOT, record["date"])
            os.makedirs(target_dir, exist_ok=True)
            src = record["folder_path"]
            dst = os.path.join(target_dir, record["folder_name"])
            if os.path.exists(dst):
                QMessageBox.critical(self, "錯誤", f"目標位置已存在同名資料夾:\n{dst}")
                return
            with OperationJournal(BASE_DIR, "restore_record_dialog", {
                "folder": record["folder_name"],
                "date": record["date"],
            }) as journal:
                journal.step("move_attachment_folder", source=src, target=dst)
                shutil.move(src, dst)
            QMessageBox.information(self, "成功", f"✅ 已還原至:\n{dst}")
            self._scan_all_records()
        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"還原失敗: {e}")

    def _edit_record(self):
        if not self.selected_record:
            QMessageBox.information(self, "提示", "請先選擇一筆記錄")
            return
        if self.selected_record["is_archived"]:
            QMessageBox.warning(self, "提示", "已歸檔的記錄不能編輯，請先還原")
            return
        dlg = EditRecordDialog(self, self.selected_record, self._scan_all_records)
        dlg.exec()

    def _check_unproduced(self):
        unproduced = [r for r in self.all_records if not r["has_output"] and not r["is_archived"]]
        if not unproduced:
            QMessageBox.information(self, "檢查結果", "✅ 所有資料群都已產出報告！")
            return

        by_date: dict[str, list] = {}
        for r in unproduced:
            by_date.setdefault(r["date"], []).append(r)

        result = f"⏳ 共有 {len(unproduced)} 個資料群尚未產出報告：\n\n"
        for date in sorted(by_date.keys(), reverse=True):
            result += f"📅 {date}:\n"
            for r in by_date[date]:
                result += f"   • {r['folder_name']}\n"
            result += "\n"

        dlg = QDialog(self)
        dlg.setWindowTitle("未產出報告清單")
        dlg.resize(500, 400)
        dlay = QVBoxLayout(dlg)
        te = QTextEdit()
        te.setPlainText(result)
        te.setReadOnly(True)
        dlay.addWidget(te)

        brow = QHBoxLayout()
        btn_filter = QPushButton("篩選顯示未產出")

        def do_filter():
            self.filter_status.setCurrentText("未產出")
            self._apply_filter()
            dlg.close()

        btn_filter.clicked.connect(do_filter)
        brow.addWidget(btn_filter)
        btn_c = QPushButton("關閉")
        btn_c.clicked.connect(dlg.close)
        brow.addWidget(btn_c)
        dlay.addLayout(brow)
        dlg.exec()

    def _open_folder(self):
        if not self.selected_record:
            QMessageBox.information(self, "提示", "請先選擇一筆記錄")
            return
        p = self.selected_record["folder_path"]
        if os.path.exists(p):
            os.startfile(p)
        else:
            QMessageBox.critical(self, "錯誤", f"資料夾不存在:\n{p}")


# =====================================================================
#  WeldSyncConflictDialog — 焊口資料來源衝突解決
# =====================================================================
class WeldSyncConflictDialog(QDialog):
    """
    比對多來源焊口資料（folder/weld_info.json/records.json details），
    空值自動補入、差異值列出讓使用者選擇。

    resolved_welds: 解決後的焊口清單 list[dict]
        每項 = {weld_no, mark, size, material, thickness, code}
    """

    def __init__(self, parent: QWidget, conflicts: list[dict]):
        """
        conflicts: list of dict, each =
            {field, weld_label, source_a_name, source_a_val, source_b_name, source_b_val}
        """
        super().__init__(parent)
        self.conflicts = conflicts
        self.choices: dict[int, str] = {}  # idx → chosen value

        self.setWindowTitle("⚠️ 焊口資料差異 — 請選擇要保留的值")
        self.resize(700, min(400, 150 + 50 * len(conflicts)))
        self.setModal(True)
        self._build_ui()

    def _build_ui(self):
        root = QVBoxLayout(self)

        hint = QLabel(
            "以下欄位在不同來源之間有差異，請選擇要保留的值：\n"
            "（空白值已自動從另一來源補入，此處僅列出雙方都有值但不同的項目）"
        )
        hint.setWordWrap(True)
        hint.setStyleSheet(f"color: {Colors.TEXT_SECONDARY}; border:none; background:transparent;")
        root.addWidget(hint)

        tree = QTreeWidget()
        tree.setHeaderLabels(["焊口", "欄位", "來源 A", "值 A", "來源 B", "值 B", "選擇"])
        for i, w in enumerate([80, 60, 80, 100, 80, 100, 80]):
            tree.setColumnWidth(i, w)
        tree.setAlternatingRowColors(True)

        self._combos: list[QComboBox] = []
        for idx, c in enumerate(self.conflicts):
            item = QTreeWidgetItem(tree, [
                c["weld_label"], c["field"],
                c["source_a_name"], str(c["source_a_val"]),
                c["source_b_name"], str(c["source_b_val"]),
                "",
            ])
            combo = QComboBox()
            combo.addItem(f"用 {c['source_a_name']}", c["source_a_val"])
            combo.addItem(f"用 {c['source_b_name']}", c["source_b_val"])
            tree.setItemWidget(item, 6, combo)
            self._combos.append(combo)

        root.addWidget(tree, stretch=1)

        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_ok = QPushButton("✅ 確認選擇")
        set_button_role(btn_ok, "primary")
        btn_ok.clicked.connect(self.accept)
        btn_row.addWidget(btn_ok)
        btn_cancel = QPushButton("取消")
        btn_cancel.clicked.connect(self.reject)
        btn_row.addWidget(btn_cancel)
        root.addLayout(btn_row)

    def get_choices(self) -> dict[int, str]:
        """回傳 {conflict_index: chosen_value}"""
        result = {}
        for idx, combo in enumerate(self._combos):
            result[idx] = combo.currentData()
        return result


def _collect_and_merge_weld_sources(
    folder_path: str,
    folder_name: str,
    report_id: str,
    new_welds: list[dict] | None = None,
) -> tuple[list[dict], list[dict]]:
    """
    收集三個來源的焊口資料並合併。

    來源優先級：
      new_welds (本次編輯)  >  weld_info.json  >  records.json details

    回傳 (merged_welds, conflicts)
        merged_welds = [{weld_no, mark, size, material, thickness, code}, ...]
        conflicts = [{field, weld_label, source_a_name, source_a_val, source_b_name, source_b_val}, ...]
    """
    # ── 來源 1: weld_info.json ──
    wi_welds = []
    wip = os.path.join(folder_path, "weld_info.json") if os.path.isdir(folder_path) else ""
    if wip and os.path.isfile(wip):
        try:
            with open(wip, "r", encoding="utf-8") as f:
                wi_data = json.load(f)
            for w in wi_data.get("welds", []):
                wn = str(w.get("weld_no", ""))
                mk = str(w.get("mark", ""))
                # 若 weld_no 含內嵌標記 (如 "1001a")，拆分出純數字 + 標記
                m_split = re.match(r'^(\d+)([rab])', wn)
                if m_split:
                    wn = m_split.group(1)
                    if not mk:
                        mk = m_split.group(2)
                wi_welds.append({
                    "weld_no": wn,
                    "mark": mk,
                    "size": str(w.get("size", "")),
                    "material": str(w.get("material", "")),
                    "thickness": str(w.get("thickness", "")),
                })
        except Exception:
            pass

    # ── 來源 2: records.json details ──
    rj_welds = []
    if report_id:
        try:
            store = _load_store()
            for det in store.get("details", []):
                if str(det.get("紀錄編號", "")) == report_id:
                    code = str(det.get("焊口編號", ""))
                    size = str(det.get("焊口尺寸", ""))
                    # 嘗試拆解 code → weld_no + mark
                    m = re.match(r"^(\d+)([A-Za-z])?(.*)$", code)
                    weld_no = m.group(1) if m else code
                    mark = m.group(2) if m and m.group(2) else ""
                    rj_welds.append({
                        "weld_no": weld_no,
                        "mark": mark.lower(),
                        "size": size,
                        "material": "",
                        "thickness": "",
                    })
        except Exception:
            pass

    # ── 決定基底 ──
    if new_welds is not None:
        base = [{
            "weld_no": str(w.get("weld_no", "")),
            "mark": str(w.get("mark", "")),
            "size": str(w.get("size", "")),
            "material": str(w.get("material", "")),
            "thickness": str(w.get("thickness", "")),
        } for w in new_welds]
        other_source = wi_welds if wi_welds else rj_welds
        other_name = "weld_info.json" if wi_welds else "records.json"
        base_name = "本次編輯"
    elif wi_welds:
        base = wi_welds
        other_source = rj_welds
        other_name = "records.json"
        base_name = "weld_info.json"
    else:
        base = rj_welds
        other_source = []
        other_name = ""
        base_name = "records.json"

    conflicts = []
    merged = []

    for i, bw in enumerate(base):
        ow = other_source[i] if i < len(other_source) else {}
        result = dict(bw)
        for field in ("weld_no", "mark", "size", "material", "thickness"):
            bv = str(bw.get(field, "")).strip()
            ov = str(ow.get(field, "")).strip()
            if not bv and ov:
                result[field] = ov  # 空值補入
            elif bv and ov and bv != ov and other_name:
                conflicts.append({
                    "field": field,
                    "weld_label": f"#{i+1} ({bw.get('weld_no', '?')})",
                    "source_a_name": base_name,
                    "source_a_val": bv,
                    "source_b_name": other_name,
                    "source_b_val": ov,
                    "_index": i,
                    "_field": field,
                })
        result["code"] = f"{result['weld_no']}{result['mark']}{result['size']}"
        merged.append(result)

    return merged, conflicts


def _apply_conflict_choices(merged: list[dict], conflicts: list[dict], choices: dict[int, str]):
    """將使用者的選擇套用到 merged 上"""
    for idx, val in choices.items():
        c = conflicts[idx]
        i = c["_index"]
        field = c["_field"]
        merged[i][field] = val
        merged[i]["code"] = f"{merged[i]['weld_no']}{merged[i]['mark']}{merged[i]['size']}"


def _sync_records_json(
    date_str: str,
    old_folder_name: str,
    new_folder_name: str,
    new_serial: str,
    merged_welds: list[dict],
):
    """
    將合併後的焊口資料同步寫入 records.json（主表 + 明細）。
    """
    store = _load_store()

    # ── 更新 records[] 主表 ──
    for rec in store["records"]:
        if rec.get("日期") == date_str and rec.get("資料夾名") == old_folder_name:
            codes = [w["code"] for w in merged_welds]
            weld_list_str = "、".join(
                f"{w['weld_no']}{w['mark']}" for w in merged_welds
            )
            weld_size_str = "；".join(
                f"{w['weld_no']}{w['mark']}={w['size']}" for w in merged_welds
            )
            rec["資料夾名"] = new_folder_name
            rec["Series NO"] = new_serial.zfill(4) if new_serial.isdigit() else new_serial
            rec["焊口清單"] = weld_list_str
            rec["焊口與尺寸"] = weld_size_str
            break

    # ── 更新 details[] ──
    report_id = ""
    for rec in store["records"]:
        if rec.get("日期") == date_str and rec.get("資料夾名") == new_folder_name:
            report_id = rec.get("報告編號", "")
            break

    if report_id:
        # 移除舊明細
        store["details"] = [
            d for d in store["details"]
            if str(d.get("紀錄編號", "")) != report_id
        ]
        # 寫入新明細
        for idx, w in enumerate(merged_welds):
            store["details"].append({
                "項目": str(idx + 1),
                "紀錄編號": report_id,
                "修改日期": date_str,
                "修改原因敘述": "",
                "Series NO": new_serial.zfill(4) if new_serial.isdigit() else new_serial,
                "DWG NO": "",
                "焊口編號": f"{w['weld_no']}{w['mark']}",
                "焊口尺寸": w["size"],
                "係數": "",
                "單價/DB": "",
                "金額": "",
                "備註": "",
            })

    auto_backup(RECORDS_JSON_PATH)
    _save_store(store)


def _sync_records_json_preserve(
    date_str: str,
    old_folder_name: str,
    new_folder_name: str,
    new_serial: str,
    merged_welds: list[dict],
):
    """
    同步 records.json，但保留明細中已有的非焊口欄位（係數/單價/金額/備註等）。
    """
    store = _load_store()

    # ── 更新 records[] 主表 ──
    for rec in store["records"]:
        if rec.get("日期") == date_str and rec.get("資料夾名") == old_folder_name:
            weld_list_str = "、".join(
                f"{w['weld_no']}{w['mark']}" for w in merged_welds
            )
            weld_size_str = "；".join(
                f"{w['weld_no']}{w['mark']}={w['size']}" for w in merged_welds
            )
            rec["資料夾名"] = new_folder_name
            rec["Series NO"] = new_serial.zfill(4) if new_serial.isdigit() else new_serial
            rec["焊口清單"] = weld_list_str
            rec["焊口與尺寸"] = weld_size_str
            break

    # ── 更新 details[]（保留非焊口欄位）──
    report_id = ""
    for rec in store["records"]:
        if rec.get("日期") == date_str and rec.get("資料夾名") == new_folder_name:
            report_id = rec.get("報告編號", "")
            break

    if report_id:
        old_details = [
            d for d in store["details"]
            if str(d.get("紀錄編號", "")) == report_id
        ]
        store["details"] = [
            d for d in store["details"]
            if str(d.get("紀錄編號", "")) != report_id
        ]
        for idx, w in enumerate(merged_welds):
            old = old_details[idx] if idx < len(old_details) else {}
            store["details"].append({
                "項目": str(idx + 1),
                "紀錄編號": report_id,
                "修改日期": old.get("修改日期", date_str),
                "修改原因敘述": old.get("修改原因敘述", ""),
                "Series NO": new_serial.zfill(4) if new_serial.isdigit() else new_serial,
                "DWG NO": old.get("DWG NO", ""),
                "焊口編號": f"{w['weld_no']}{w['mark']}",
                "焊口尺寸": w["size"],
                "係數": old.get("係數", ""),
                "單價/DB": old.get("單價/DB", ""),
                "金額": old.get("金額", ""),
                "備註": old.get("備註", ""),
            })

    auto_backup(RECORDS_JSON_PATH)
    _save_store(store)


# =====================================================================
#  EditRecordDialog - 編輯記錄
# =====================================================================
class EditRecordDialog(QDialog):
    """編輯記錄對話框"""

    def __init__(self, parent: QWidget, record: dict, refresh_callback=None):
        super().__init__(parent)
        self.record = record
        self.refresh_callback = refresh_callback
        self.weld_entries: list[dict] = []

        self.setWindowTitle(f"✏️ 編輯記錄 - {record['folder_name']}")
        self.resize(600, 500)
        self.setModal(True)

        self._build_ui()

    def _build_ui(self):
        root = QVBoxLayout(self)

        # 基本資訊
        info_grp = QGroupBox("📋 基本資訊")
        ilay = QVBoxLayout(info_grp)
        ilay.addWidget(QLabel(f"日期: {self.record['date']}"))
        ilay.addWidget(QLabel(f"原資料夾: {self.record['folder_name']}"))

        serial_row = QHBoxLayout()
        serial_row.addWidget(QLabel("流水號:"))
        self.serial_edit = QLineEdit(self.record["serial"])
        self.serial_edit.setFixedWidth(80)
        self.serial_edit.textChanged.connect(self._update_preview)
        serial_row.addWidget(self.serial_edit)
        serial_row.addStretch()
        ilay.addLayout(serial_row)
        root.addWidget(info_grp)

        # 焊口清單
        weld_grp = QGroupBox("🔧 焊口清單")
        wg_lay = QVBoxLayout(weld_grp)

        # 表頭
        hdr_row = QHBoxLayout()
        for txt, w in [("焊口號", 60), ("標記", 45), ("尺寸", 60), ("材質", 80), ("厚度", 80)]:
            lb = QLabel(txt)
            lb.setFixedWidth(w)
            hdr_row.addWidget(lb)
        hdr_row.addStretch()
        wg_lay.addLayout(hdr_row)

        # 捲動區
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_content = QWidget()
        self.scroll_layout = QVBoxLayout(self.scroll_content)
        self.scroll_layout.setAlignment(Qt.AlignmentFlag.AlignTop)
        self.scroll_area.setWidget(self.scroll_content)
        wg_lay.addWidget(self.scroll_area, stretch=1)

        # 解析現有焊口
        welds_str = self.record["welds_str"]
        welds = []

        if self.record.get("weld_info") and self.record["weld_info"].get("welds"):
            for w in self.record["weld_info"]["welds"]:
                wn = str(w.get("weld_no", ""))
                mk = str(w.get("mark", ""))
                # 一律拆分：若 weld_no 尾部帶字母（如 "1001a" 或 "6r"），拆出數字+標記
                m_wn = re.match(r'^(\d+)([rab])', wn, re.IGNORECASE)
                if m_wn:
                    wn = m_wn.group(1)
                    mk = mk or m_wn.group(2).lower()
                welds.append({
                    "weld_no": wn,
                    "mark": mk,
                    "size": w.get("size", ""),
                    "material": w.get("material", ""),
                    "thickness": w.get("thickness", ""),
                })
        else:
            pattern = re.compile(r"^(\d+)([rab])(.+)$")
            for part in welds_str.split("_"):
                m = pattern.match(part)
                if m:
                    welds.append({
                        "weld_no": m.group(1), "mark": m.group(2), "size": m.group(3),
                        "material": "", "thickness": "",
                    })

        for w in welds:
            self._add_weld_row(w["weld_no"], w["mark"], w["size"], w["material"], w["thickness"])

        # 新增按鈕
        add_btn_row = QHBoxLayout()
        btn_add = QPushButton("➕ 新增焊口")
        btn_add.clicked.connect(lambda: self._add_weld_row())
        add_btn_row.addWidget(btn_add)
        add_btn_row.addStretch()
        wg_lay.addLayout(add_btn_row)
        root.addWidget(weld_grp, stretch=1)

        # 預覽
        preview_grp = QGroupBox("👁️ 預覽新資料夾名稱")
        play = QVBoxLayout(preview_grp)
        self.preview_label = QLabel("(請填寫完整資訊)")
        self.preview_label.setFont(_FONT_CODE)
        play.addWidget(self.preview_label)
        root.addWidget(preview_grp)

        self._update_preview()

        # 按鈕
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_cancel = QPushButton("取消")
        btn_cancel.clicked.connect(self.close)
        btn_row.addWidget(btn_cancel)
        btn_save = QPushButton("💾 儲存變更")
        set_button_role(btn_save, "primary")
        btn_save.clicked.connect(self._save_changes)
        btn_row.addWidget(btn_save)
        root.addLayout(btn_row)

    def _add_weld_row(self, weld_no="", mark="r", size="", material="", thickness=""):
        row_w = QWidget()
        row_lay = QHBoxLayout(row_w)
        row_lay.setContentsMargins(0, 2, 0, 2)

        weld_no_edit = QLineEdit(str(weld_no))
        weld_no_edit.setFixedWidth(60)
        weld_no_edit.textChanged.connect(self._update_preview)
        row_lay.addWidget(weld_no_edit)

        mark_combo = QComboBox()
        mark_combo.addItems(["r", "a", "b"])
        mark_combo.setCurrentText(str(mark))
        mark_combo.setFixedWidth(45)
        mark_combo.currentTextChanged.connect(self._update_preview)
        row_lay.addWidget(mark_combo)

        size_edit = QLineEdit(str(size))
        size_edit.setFixedWidth(60)
        size_edit.textChanged.connect(self._update_preview)
        row_lay.addWidget(size_edit)

        mat_edit = QLineEdit(str(material))
        mat_edit.setFixedWidth(80)
        row_lay.addWidget(mat_edit)

        thk_edit = QLineEdit(str(thickness))
        thk_edit.setFixedWidth(80)
        row_lay.addWidget(thk_edit)

        row_lay.addStretch()
        self.scroll_layout.addWidget(row_w)

        self.weld_entries.append({
            "weld_no": weld_no_edit,
            "mark": mark_combo,
            "size": size_edit,
            "material": mat_edit,
            "thickness": thk_edit,
        })

    def _update_preview(self):
        serial = self.serial_edit.text().strip()
        codes = []
        for entry in self.weld_entries:
            wn = entry["weld_no"].text().strip()
            mk = entry["mark"].currentText().strip()
            sz = entry["size"].text().strip()
            if wn and mk and sz:
                codes.append(f"{wn}{mk}{sz}")
        if serial and codes:
            self.preview_label.setText(f"{serial}_{'_'.join(codes)}")
        else:
            self.preview_label.setText("(請填寫完整資訊)")

    @reentry_guard("_edit_record_save_in_progress", _show_reentry_notice)
    def _save_changes(self):
        new_serial = self.serial_edit.text().strip()
        if not new_serial or not new_serial.isdigit():
            QMessageBox.critical(self, "錯誤", "流水號必須是數字")
            return

        new_welds = []
        for entry in self.weld_entries:
            wn = entry["weld_no"].text().strip()
            mk = entry["mark"].currentText().strip()
            sz = entry["size"].text().strip()
            mt = entry["material"].text().strip()
            tk_ = entry["thickness"].text().strip()
            if wn and mk and sz:
                new_welds.append({
                    "weld_no": wn, "mark": mk, "size": sz,
                    "material": mt, "thickness": tk_,
                    "code": f"{wn}{mk}{sz}",
                })

        if not new_welds:
            QMessageBox.critical(self, "錯誤", "至少需要一個焊口")
            return

        codes = [w["code"] for w in new_welds]
        new_folder_name = f"{new_serial}_{'_'.join(codes)}"
        old_path = self.record["folder_path"]
        old_folder_name = self.record["folder_name"]
        new_path = os.path.join(os.path.dirname(old_path), new_folder_name)

        if new_folder_name != old_folder_name:
            if os.path.exists(new_path):
                QMessageBox.critical(self, "錯誤", f"目標資料夾已存在:\n{new_path}")
                return
            if QMessageBox.question(
                self, "確認變更",
                f"將進行以下變更：\n\n📁 原資料夾: {old_folder_name}\n📁 新資料夾: {new_folder_name}\n\n確定要繼續嗎？"
            ) != QMessageBox.StandardButton.Yes:
                return

        # ── 衝突偵測：比對 records.json 中的既有資料 ──
        report_id = self.record.get("report_id", "")
        merged, conflicts = _collect_and_merge_weld_sources(
            old_path, old_folder_name, report_id, new_welds,
        )

        if conflicts:
            dlg = WeldSyncConflictDialog(self, conflicts)
            if dlg.exec() != QDialog.DialogCode.Accepted:
                return
            _apply_conflict_choices(merged, conflicts, dlg.get_choices())

        try:
            # 1) weld_info.json
            _deduped_welds: list[dict] = []
            _seen_wids: set[str] = set()
            for w in merged:
                _wid = f"{w['weld_no']}{w['mark']}"
                if _wid not in _seen_wids:
                    _seen_wids.add(_wid)
                    _deduped_welds.append(w)

            weld_info = {"series": new_serial, "date": self.record["date"], "welds": [
                {k: w[k] for k in ("weld_no", "mark", "size", "material", "thickness")}
                for w in _deduped_welds
            ]}
            wip = os.path.join(old_path, "weld_info.json")
            atomic_write_json(wip, weld_info)

            # 2) GroupWeld.txt
            gw_path = os.path.join(old_path, "GroupWeld.txt")
            if os.path.exists(gw_path) or len(merged) > 6:
                with open(gw_path, "w", encoding="utf-8") as f:
                    for w in merged:
                        f.write(w["code"] + "\n")

            # 3) rename folder
            if new_folder_name != old_folder_name:
                os.rename(old_path, new_path)

            # 4) sync records.json（主表 + 明細，保留既有欄位）
            _sync_records_json_preserve(
                self.record["date"], old_folder_name,
                new_folder_name, new_serial, merged,
            )

            QMessageBox.information(self, "成功",
                f"✅ 已更新記錄並同步所有資料\n\n新資料夾: {new_folder_name}")

            if self.refresh_callback:
                self.refresh_callback()

            self.close()

        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"儲存失敗: {e}")
            import traceback; traceback.print_exc()


# =====================================================================
#  WeldOrphanAuditDialog - 孤兒焊口稽查工具
# =====================================================================
class WeldOrphanAuditDialog(QDialog):
    """掃描焊口管制表，找出修改/新增焊口中的孤兒與重複"""

    _STATUS_COLORS = {
        "orphan": "#FF6B6B",
        "duplicate": "#FFD93D",
        "matched": "#6BCB77",
    }
    _STATUS_LABELS = {
        "orphan": "🔴 孤兒",
        "duplicate": "🟡 重複",
        "matched": "🟢 已對應",
    }

    def __init__(self, parent: QWidget):
        super().__init__(parent)
        self.audit_results: list[dict] = []

        self.setWindowTitle("🔍 孤兒焊口稽查 - 管制表 vs 報告比對")
        self.resize(950, 680)
        self.setModal(True)

        self._build_ui()
        self._run_audit()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setSpacing(10)

        # 標題
        header = QVBoxLayout()
        lbl = QLabel("掃描焊口管制表中的修改口(r)/新增口(a)/1000+號，"
                     "比對 records.json 報告紀錄")
        lbl.setFont(_FONT_SUBTITLE)
        lbl.setStyleSheet(f"color: {Colors.TEXT_SECONDARY}; border:none; background:transparent;")
        lbl.setWordWrap(True)
        header.addWidget(lbl)
        self.stat_label = QLabel("正在掃描...")
        self.stat_label.setStyleSheet(
            f"color: {Colors.PRIMARY}; font-weight:bold; border:none; background:transparent;")
        header.addWidget(self.stat_label)
        root.addLayout(header)

        # 篩選區
        filter_grp = QGroupBox("🔍 篩選")
        flay = QHBoxLayout(filter_grp)
        flay.addWidget(QLabel("狀態:"))
        self.status_combo = QComboBox()
        self.status_combo.addItems(["全部", "🔴 孤兒", "🟡 重複", "🟢 已對應"])
        self.status_combo.currentTextChanged.connect(self._apply_filter)
        flay.addWidget(self.status_combo)

        flay.addSpacing(20)
        flay.addWidget(QLabel("流水號:"))
        self.serial_edit = QLineEdit()
        self.serial_edit.setFixedWidth(90)
        self.serial_edit.setPlaceholderText("輸入篩選")
        self.serial_edit.returnPressed.connect(self._apply_filter)
        flay.addWidget(self.serial_edit)

        btn_filter = QPushButton("篩選")
        btn_filter.clicked.connect(self._apply_filter)
        flay.addWidget(btn_filter)
        btn_rescan = QPushButton("重新掃描")
        btn_rescan.clicked.connect(self._run_audit)
        flay.addWidget(btn_rescan)
        flay.addStretch()
        root.addWidget(filter_grp)

        # 結果列表
        list_grp = QGroupBox("📋 稽查結果")
        llay = QVBoxLayout(list_grp)
        self.tree = QTreeWidget()
        self.tree.setHeaderLabels(["狀態", "流水號", "焊口編號", "對應報告", "備註"])
        hdr = self.tree.header()
        hdr.resizeSection(0, 80)
        hdr.resizeSection(1, 80)
        hdr.resizeSection(2, 100)
        hdr.resizeSection(3, 160)
        hdr.resizeSection(4, 400)
        self.tree.setAlternatingRowColors(True)
        llay.addWidget(self.tree)
        root.addWidget(list_grp, stretch=1)

        # 統計
        stat_row = QHBoxLayout()
        self.orphan_label = QLabel("")
        self.orphan_label.setStyleSheet(
            f"color: #FF6B6B; font-weight:bold; border:none; background:transparent;")
        stat_row.addWidget(self.orphan_label)
        self.dup_label = QLabel("")
        self.dup_label.setStyleSheet(
            f"color: #FFD93D; font-weight:bold; border:none; background:transparent;")
        stat_row.addWidget(self.dup_label)
        self.match_label = QLabel("")
        self.match_label.setStyleSheet(
            f"color: #6BCB77; font-weight:bold; border:none; background:transparent;")
        stat_row.addWidget(self.match_label)
        stat_row.addStretch()
        root.addLayout(stat_row)

        # 按鈕
        btn_row = QHBoxLayout()
        btn_export = QPushButton("📋 複製結果到剪貼簿")
        btn_export.clicked.connect(self._copy_to_clipboard)
        btn_row.addWidget(btn_export)
        btn_row.addStretch()
        btn_close = QPushButton("關閉")
        btn_close.clicked.connect(self.close)
        btn_row.addWidget(btn_close)
        root.addLayout(btn_row)

    def _run_audit(self):
        """執行稽查"""
        self.tree.clear()
        self.stat_label.setText("正在掃描...")
        QApplication.processEvents()

        from weld_control import audit_orphan_welds
        self.audit_results = audit_orphan_welds()

        # 統計
        counts = {"orphan": 0, "duplicate": 0, "matched": 0}
        for r in self.audit_results:
            counts[r["status"]] = counts.get(r["status"], 0) + 1

        total = len(self.audit_results)
        self.stat_label.setText(
            f"共掃描 {total} 筆修改/新增焊口")
        self.orphan_label.setText(f"🔴 孤兒: {counts['orphan']}")
        self.dup_label.setText(f"🟡 重複: {counts['duplicate']}")
        self.match_label.setText(f"🟢 已對應: {counts['matched']}")

        self._populate_tree(self.audit_results)

    def _populate_tree(self, data: list[dict]):
        """填入樹狀列表"""
        self.tree.clear()
        from PyQt6.QtGui import QBrush, QColor

        for r in data:
            status = r["status"]
            report_str = ", ".join(r["report_ids"]) if r["report_ids"] else "—"
            item = QTreeWidgetItem([
                self._STATUS_LABELS.get(status, status),
                r["serial"],
                r["weld_no"],
                report_str,
                r["remark"],
            ])
            color = self._STATUS_COLORS.get(status, "#FFFFFF")
            item.setForeground(0, QBrush(QColor(color)))
            self.tree.addTopLevelItem(item)

    def _apply_filter(self):
        """篩選"""
        status_text = self.status_combo.currentText()
        serial_filter = self.serial_edit.text().strip()

        status_map = {
            "🔴 孤兒": "orphan",
            "🟡 重複": "duplicate",
            "🟢 已對應": "matched",
        }
        target_status = status_map.get(status_text)

        filtered = []
        for r in self.audit_results:
            if target_status and r["status"] != target_status:
                continue
            if serial_filter:
                if serial_filter not in r["serial"]:
                    continue
            filtered.append(r)

        self._populate_tree(filtered)

    def _copy_to_clipboard(self):
        """複製結果到剪貼簿（TSV 格式）"""
        lines = ["狀態\t流水號\t焊口編號\t對應報告\t備註"]
        for r in self.audit_results:
            report_str = ", ".join(r["report_ids"]) if r["report_ids"] else "—"
            lines.append(
                f"{self._STATUS_LABELS.get(r['status'], r['status'])}\t"
                f"{r['serial']}\t{r['weld_no']}\t{report_str}\t{r['remark']}"
            )
        QApplication.clipboard().setText("\n".join(lines))
        QMessageBox.information(self, "已複製",
                                f"✅ 已複製 {len(self.audit_results)} 筆結果到剪貼簿\n"
                                "可貼到 Excel 或記事本")
