# -*- coding: utf-8 -*-
"""co_main_bridge.py — 新版主介面的前後端橋（transport-agnostic，先做「讀」路徑）

定位同 co_bridge：**不認 UI、不認傳輸層**，JSON 進 JSON 出，每個對外方法回
``{"ok": bool, "data": ..., "error": str|None}`` 信封；pywebview 的 js_api 直接拿它
（見 co_main_app.py），未來 FastAPI 也能無痛包成 HTTP route。

第一刀只做唯讀讀取，先把「橋 → pywebview → 前端」這條鏈打通：
    - ``pricebook()``：讀 records/material_pricebook.json（舊系統真實料表，442 筆），
      對映成主介面前端 material 管理表的形狀。
    - ``records()``：讀舊 store records.json（目前空），對映成前端記錄形狀。

不放任何業務邏輯；之後寫入 / 產出 / 中央查價再逐刀加。
"""
from __future__ import annotations

import functools
import json
import os
import subprocess
import sys
import traceback
import re
from pathlib import Path
from typing import Any, Callable

from material_taxonomy import (
    category_for_part,
    enrich_material_item,
    load_taxonomy,
    material_match_key,
    normalize_material,
    normalize_schedule,
    normalize_size,
    taxonomy_options,
)
from material_catalog_rules import all_catalog_rows, build_frontend_item, catalog_summary, query_catalog, rows_by_ids
from support_bom import analyze_support_bom
from utils import atomic_write_json

API_VERSION = "main-0.1"
PDF_HEALTH_SCAN_FILE_LIMIT = 2000
PDF_HEALTH_SCAN_DIR_LIMIT = 300


SOURCE_SCHEMA_ROLES = {
    "dwg": [
        {
            "key": "serial_column",
            "setting_key": "col_serial",
            "label": "ISO流編",
            "default": "NO",
            "aliases": ["流水號", "管線序號", "ISO流編", "Series NO", "NO", "Serial", "Series"],
            "required": True,
            "note": "用來把 DWG LIST 資料掛回修改單流水號。",
        },
        {
            "key": "dwg_no_column",
            "setting_key": "col_dwg_no",
            "label": "圖號",
            "default": "DWG NO",
            "aliases": ["圖號", "圖面編號", "DWG NO", "DRAWING NO", "DRAWING"],
            "required": False,
            "note": "業主資料包的圖號 / DWG 欄位。",
        },
        {
            "key": "line_no_column",
            "setting_key": "col_line_no",
            "label": "Line No.",
            "default": "LINE NUMBER",
            "aliases": ["Line No.", "LINE NO", "LINE NUMBER", "LINE   NUMBER", "管線編號"],
            "required": False,
            "note": "業主資料包的 Line No. 欄位。",
        },
        {
            "key": "date_column",
            "setting_key": "col_date",
            "label": "日期",
            "default": "日期",
            "aliases": ["日期", "DATE", "施工日期", "配管完成日期"],
            "required": False,
            "note": "若來源表有日期，可作為報表索引參考。",
        },
    ],
    "weld": [
        {
            "key": "serial_column",
            "setting_key": "col_serial",
            "label": "ISO流編",
            "default": "流水號",
            "aliases": ["流水號", "ISO流編", "Series NO", "NO", "Serial", "管線序號"],
            "required": True,
            "note": "用來查同一張 ISO / 流水號底下的焊口。",
        },
        {
            "key": "weld_no_column",
            "setting_key": "col_weld_no",
            "label": "焊口編號",
            "default": "焊口編號",
            "aliases": ["焊口編號", "銲口編號", "焊口碼", "Weld No", "WELD NO"],
            "required": True,
            "note": "修改單精靈與業主資料包的焊口主鍵。",
        },
        {
            "key": "size_column",
            "setting_key": "col_size",
            "label": "尺寸",
            "default": "尺寸",
            "aliases": ["尺寸", "SIZE", "口徑", "管徑"],
            "required": False,
            "note": "焊口摘要與焊口統計的尺寸欄。",
        },
        {
            "key": "material_column",
            "setting_key": "col_material",
            "label": "材質",
            "default": "材質",
            "aliases": ["材質", "MATERIAL", "鋼種"],
            "required": False,
            "note": "焊口摘要與焊口統計的材質欄。",
        },
        {
            "key": "thickness_column",
            "setting_key": "col_thickness",
            "label": "厚度/SCH",
            "default": "厚度",
            "aliases": ["厚度", "SCH", "Schedule", "管厚"],
            "required": False,
            "note": "焊口摘要與焊口統計的厚度 / SCH 欄。",
        },
        {
            "key": "weld_type_column",
            "setting_key": "col_weld_type",
            "label": "焊接型式",
            "default": "銲接型式",
            "aliases": ["銲接型式", "焊接型式", "焊口型式", "WELD TYPE"],
            "required": False,
            "note": "修改單精靈帶入 BW / SW / FSW 等型式。",
        },
        {
            "key": "db_column",
            "setting_key": "col_db",
            "label": "DB",
            "default": "DB數",
            "aliases": ["DB數", "DB", "D.B.", "DI", "D.I.", "Dia-Inch", "DIA INCH", "管徑吋數"],
            "required": False,
            "note": "業主焊口統計的 DB 數，保留 0.5 / 0.75 等小數。",
        },
        {
            "key": "inside_diameter_column",
            "setting_key": "col_inside_diameter",
            "label": "I.D",
            "default": "I.D",
            "aliases": ["I.D", "I.D.", "ID", "內徑"],
            "required": False,
            "note": "需要內徑資訊時保留來源欄位。",
        },
        {
            "key": "budget_no_column",
            "setting_key": "col_budget_no",
            "label": "預算編號",
            "default": "預算編號",
            "aliases": ["預算編號", "Budget No", "BudgetNo", "Budget", "預算"],
            "required": False,
            "note": "業主焊口統計與摘要的預算編號欄。",
        },
        {
            "key": "attribute_1_column",
            "setting_key": "col_attribute_1",
            "label": "屬性.1",
            "default": "屬性.1",
            "aliases": ["屬性.1", "屬性1", "焊口屬性", "屬性"],
            "required": False,
            "note": "用來過濾實際焊口列，避免把閥件/法蘭安裝列算進去。",
        },
        {
            "key": "attribute_2_column",
            "setting_key": "col_attribute_2",
            "label": "屬性.2",
            "default": "屬性.2",
            "aliases": ["屬性.2", "屬性2", "分類"],
            "required": False,
            "note": "保留第二層屬性/分類，供後續稽核與輸出使用。",
        },
    ],
}


def _wizard_launch_command(root: Path) -> tuple[list[str], str]:
    if getattr(sys, "frozen", False):
        return [sys.executable, "--wizard"], "frozen"

    app = root / "control" / "co_wizard_app.py"
    if not app.exists():
        raise FileNotFoundError(f"找不到精靈啟動器：{app}")
    return [sys.executable, str(app)], "source"


def _enveloped(fn: Callable) -> Callable:
    @functools.wraps(fn)
    def wrapper(self, *args, **kwargs):
        try:
            return {"ok": True, "data": fn(self, *args, **kwargs), "error": None}
        except Exception as exc:  # 橋不可把例外漏給前端
            return {
                "ok": False,
                "data": None,
                "error": f"{type(exc).__name__}: {exc}",
                "trace": traceback.format_exc(),
            }
    return wrapper


def _num(v: Any) -> Any:
    """單價等欄位：'100' → 100、'' → 0；無法轉的原樣回（不炸）。"""
    if v is None:
        return 0
    try:
        s = str(v).strip().replace(",", "")
        if s == "":
            return 0
        f = float(s)
        return int(f) if f == int(f) else f
    except (TypeError, ValueError):
        return v


def _read_json(path: Path) -> Any:
    """容錯讀 JSON（utf-8-sig 吃 BOM）；不存在 / 壞檔回 None。"""
    try:
        return json.loads(path.read_text(encoding="utf-8-sig"))
    except Exception:
        return None


def _settings_path_value(root: Path, key: str) -> str:
    data = _read_json(root / "settings.json")
    if not isinstance(data, dict):
        return ""
    paths = data.get("paths")
    if not isinstance(paths, dict):
        return ""
    return str(paths.get(key) or "")


def _settings_project_name(root: Path) -> str:
    data = _read_json(root / "settings.json")
    if not isinstance(data, dict):
        return ""
    project = data.get("project")
    if not isinstance(project, dict):
        return ""
    return str(project.get("name") or project.get("project_name") or project.get("title") or "").strip()


def _settings_section(root: Path, section: str) -> dict:
    data = _read_json(root / "settings.json")
    if not isinstance(data, dict):
        return {}
    value = data.get(section)
    return value if isinstance(value, dict) else {}


def _resolved_setting_path(root: Path, value: Any) -> Path:
    path = Path(str(value or "").strip())
    return path if path.is_absolute() else root / path


def _norm_header(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip()).replace("銲", "焊").upper()


def _norm_sheet_name(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip()).lower()


def _sheet_candidate_sort_key(sheet_name: str, desired: str, order: int) -> tuple[int, int]:
    name_norm = _norm_sheet_name(sheet_name)
    desired_norm = _norm_sheet_name(desired)
    score = 0
    if sheet_name == desired:
        score += 1000
    if desired_norm and name_norm == desired_norm:
        score += 900
    if desired_norm and name_norm.startswith(desired_norm):
        score += 500
    if desired_norm and desired_norm in name_norm:
        score += 250
    lower_name = sheet_name.lower()
    if "new" in lower_name or "新版" in sheet_name or sheet_name.endswith("-NEW"):
        score += 100
    if "old" in lower_name or "舊" in sheet_name or sheet_name.endswith("-OLD"):
        score -= 100
    return (-score, order)


def _excel_sheet_candidates(wb: Any, sheet_name: str) -> list[str]:
    return [
        name
        for _, name in sorted(
            enumerate(wb.sheetnames),
            key=lambda item: _sheet_candidate_sort_key(item[1], sheet_name, item[0]),
        )
    ]


def _source_problem(label: str, message: str) -> dict:
    return {
        "state": "err",
        "label": label,
        "count": 0,
        "summary": message,
        "message": message,
    }


def _required_options(value: Any) -> list[str]:
    if isinstance(value, (list, tuple, set)):
        return [str(v) for v in value if str(v or "").strip()]
    return [str(value)] if str(value or "").strip() else []


def _source_role_defs(kind: str, cfg: dict | None = None) -> list[dict]:
    cfg = cfg if isinstance(cfg, dict) else {}
    roles = []
    for role in SOURCE_SCHEMA_ROLES.get(kind, []):
        value = str(cfg.get(role["setting_key"]) or "").strip()
        default = str(role.get("default") or "").strip()
        aliases = []
        for item in [value, default, *role.get("aliases", [])]:
            text = str(item or "").strip()
            if text and text not in aliases:
                aliases.append(text)
        roles.append({
            **role,
            "value": value or default,
            "aliases": aliases,
        })
    return roles


def _source_required_options(roles: list[dict]) -> list[list[str]]:
    return [role.get("aliases", []) for role in roles if role.get("required")]


def _resolve_required_indexes(headers: list[Any], required: list[Any]) -> tuple[list[int], list[str], list[dict]]:
    header_map = {_norm_header(h): idx for idx, h in enumerate(headers) if str(h or "").strip()}
    header_text = {idx: str(h or "").strip() for idx, h in enumerate(headers) if str(h or "").strip()}
    indexes: list[int] = []
    missing: list[str] = []
    matched: list[dict] = []
    for item in required:
        options = _required_options(item)
        found = None
        matched_alias = ""
        for opt in options:
            idx = header_map.get(_norm_header(opt))
            if idx is not None:
                found = idx
                matched_alias = opt
                break
        if found is None:
            missing.append(options[0] if options else "")
        else:
            indexes.append(found)
            matched.append({
                "expected": options[0] if options else "",
                "alias": matched_alias,
                "actual": header_text.get(found, ""),
                "column": found + 1,
            })
    return indexes, missing, matched


def _resolve_role_fields(headers: list[Any], roles: list[dict]) -> list[dict]:
    fields = []
    for role in roles:
        _, missing, matched = _resolve_required_indexes(headers, [role.get("aliases", [])])
        hit = matched[0] if matched else {}
        actual = str(hit.get("actual") or "").strip()
        fields.append({
            "key": role.get("key", ""),
            "label": role.get("label", ""),
            "setting_key": role.get("setting_key", ""),
            "expected": role.get("value") or role.get("default") or "",
            "alias": hit.get("alias", ""),
            "actual": actual,
            "column": hit.get("column"),
            "required": bool(role.get("required")),
            "missing": bool(missing),
            "note": role.get("note", ""),
        })
    return fields


def _find_excel_header(ws: Any, required: list[Any], *, scan_rows: int = 12) -> dict | None:
    max_row = min(int(ws.max_row or scan_rows), scan_rows)
    best_missing: list[str] = []
    for row_no, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, values_only=True), start=1):
        headers = list(row or [])
        indexes, missing, matched = _resolve_required_indexes(headers, required)
        if not missing:
            return {
                "row": row_no,
                "headers": headers,
                "indexes": indexes,
                "matched": matched,
            }
        if not best_missing or len(missing) < len(best_missing):
            best_missing = missing
    return {"missing": best_missing} if best_missing else None


def _excel_source_health(
    root: Path,
    path_value: Any,
    *,
    sheet_name: str,
    required: list[Any],
    roles: list[dict] | None = None,
) -> dict:
    raw = str(path_value or "").strip()
    if not raw:
        return _source_problem("未設定", "尚未設定路徑")
    path = _resolved_setting_path(root, raw)
    if not path.is_file():
        return _source_problem("找不到", "找不到檔案")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet_candidates = _excel_sheet_candidates(wb, sheet_name)
            chosen = None
            chosen_headers: list[Any] = []
            chosen_indexes: list[int] = []
            chosen_header_row = 1
            chosen_matched: list[dict] = []
            last_missing: list[str] = []
            for candidate in sheet_candidates:
                ws = wb[candidate]
                header_info = _find_excel_header(ws, required)
                if header_info and not header_info.get("missing"):
                    chosen = candidate
                    chosen_headers = list(header_info.get("headers") or [])
                    chosen_indexes = list(header_info.get("indexes") or [])
                    chosen_header_row = int(header_info.get("row") or 1)
                    chosen_matched = list(header_info.get("matched") or [])
                    break
                if header_info and header_info.get("missing"):
                    last_missing = list(header_info.get("missing") or [])
            if not chosen:
                if sheet_name not in wb.sheetnames:
                    available = "、".join(wb.sheetnames[:4])
                    return _source_problem("需檢查", f"找不到工作表：{sheet_name}；可用：{available}")
                return _source_problem("需檢查", "缺少欄位：" + "、".join(last_missing))
            ws = wb[chosen]
            role_fields = _resolve_role_fields(chosen_headers, roles or [])
            count = 0
            for row in ws.iter_rows(min_row=chosen_header_row + 1, values_only=True):
                if all(idx < len(row) and str(row[idx] or "").strip() for idx in chosen_indexes):
                    count += 1
            field_count = len([h for h in chosen_headers if str(h or "").strip()])
            matched_text = "、".join(
                f"{m.get('expected')}→{m.get('actual')}"
                for m in chosen_matched
                if m.get("expected") and m.get("actual")
            )
            message = f"{chosen} 第 {chosen_header_row} 列 · {field_count} 個欄位"
            if matched_text:
                message += f" · {matched_text}"
            if chosen != sheet_name:
                message = f"自動辨識 {message}（設定值：{sheet_name}）"
            return {
                "state": "ok",
                "label": "已讀取",
                "count": count,
                "summary": f"有效資料 {count} 筆",
                "message": message,
                "sheet": chosen,
                "header_row": chosen_header_row,
                "fields": role_fields or chosen_matched,
                "missing_optional": [
                    item for item in role_fields if item.get("missing") and not item.get("required")
                ],
            }
        finally:
            wb.close()
    except PermissionError:
        return _source_problem("被占用", "檔案可能正由 Excel 開啟")
    except Exception as exc:
        return _source_problem("讀取失敗", str(exc))


def _cell_preview_value(value: Any) -> str:
    if value is None:
        return ""
    try:
        import datetime
        if isinstance(value, (datetime.datetime, datetime.date)):
            return value.strftime("%Y-%m-%d")
    except Exception:
        pass
    text = str(value)
    return text if len(text) <= 120 else text[:117] + "..."


def _excel_source_preview(
    root: Path,
    path_value: Any,
    *,
    sheet_name: str,
    required: list[Any],
    roles: list[dict] | None = None,
    max_rows: int = 18,
    max_cols: int = 18,
) -> dict:
    raw = str(path_value or "").strip()
    if not raw:
        raise FileNotFoundError("尚未設定 Excel 來源路徑")
    path = _resolved_setting_path(root, raw)
    if not path.is_file():
        raise FileNotFoundError(f"找不到檔案：{path}")
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_candidates = _excel_sheet_candidates(wb, sheet_name)
        chosen = sheet_candidates[0] if sheet_candidates else wb.sheetnames[0]
        header_info: dict | None = None
        for candidate in sheet_candidates:
            found = _find_excel_header(wb[candidate], required)
            if found and not found.get("missing"):
                chosen = candidate
                header_info = found
                break
            if candidate == chosen and found:
                header_info = found
        ws = wb[chosen]
        header_row = int((header_info or {}).get("row") or 1)
        matched = list((header_info or {}).get("matched") or [])
        required_indexes = list((header_info or {}).get("indexes") or [])
        max_index = max(required_indexes, default=-1) + 1
        col_count = min(max(int(ws.max_column or 1), max_index), max_cols)
        row_count = min(int(ws.max_row or 1), max(header_row + max_rows, max_rows))

        header_values = []
        try:
            header_values = [cell for cell in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
        except StopIteration:
            header_values = []
        columns = []
        for idx in range(1, col_count + 1):
            header = header_values[idx - 1] if idx - 1 < len(header_values) else ""
            columns.append({
                "index": idx,
                "letter": get_column_letter(idx),
                "header": _cell_preview_value(header),
            })

        rows = []
        for row_no, row in enumerate(ws.iter_rows(min_row=1, max_row=row_count, max_col=col_count, values_only=True), start=1):
            rows.append({
                "number": row_no,
                "is_header": row_no == header_row,
                "cells": [_cell_preview_value(v) for v in row],
            })
        return {
            "path": str(path),
            "sheets": list(wb.sheetnames),
            "sheet": chosen,
            "configured_sheet": sheet_name,
            "header_row": header_row,
            "columns": columns,
            "rows": rows,
            "roles": roles or [],
            "fields": _resolve_role_fields(header_values, roles or []) or matched,
            "missing": list((header_info or {}).get("missing") or []),
        }
    finally:
        wb.close()


def _serial_tokens(value: Any) -> set[str]:
    raw = str(value or "").strip()
    digits = re.sub(r"\D", "", raw)
    tokens = {raw.lower()} if raw else set()
    if digits:
        tokens.add(digits)
        tokens.add(digits.lstrip("0") or "0")
        tokens.add(digits.zfill(3))
        tokens.add(digits.zfill(4))
    return {t for t in tokens if t}


def _pdf_source_health(root: Path, path_value: Any, serials: set[str]) -> dict:
    raw = str(path_value or "").strip()
    if not raw:
        return _source_problem("未設定", "尚未設定路徑")
    path = _resolved_setting_path(root, raw)
    if not path.is_dir():
        return _source_problem("找不到", "找不到資料夾")
    pdfs, partial, dirs_scanned, errors = _scan_pdf_files(path)
    if errors and not pdfs:
        return _source_problem("讀取失敗", errors[0])
    serial_tokens = set()
    for serial in serials:
        serial_tokens.update(_serial_tokens(serial))
    matched = 0
    for pdf in pdfs:
        stem = pdf.stem.lower()
        if any(re.search(rf"(?<!\d){re.escape(token)}(?!\d)", stem) for token in serial_tokens):
            matched += 1
    if serial_tokens and pdfs and matched == 0:
        return {
            "state": "warn",
            "label": "部分讀取" if partial else "未匹配",
            "count": 0,
            "total": len(pdfs),
            "matched": 0,
            "summary": f"匹配 0 / {len(pdfs)} 張 PDF",
            "message": f"目前有 {len(serials)} 個流水號可比對",
            "partial": partial,
            "scanned_dirs": dirs_scanned,
            "scan_limit": PDF_HEALTH_SCAN_FILE_LIMIT,
            "errors": errors[:3],
        }
    return {
        "state": "warn" if partial else "ok",
        "label": "部分讀取" if partial else "已讀取",
        "count": matched,
        "total": len(pdfs),
        "matched": matched,
        "summary": f"匹配 {matched} / {len(pdfs)} 張 PDF" if serial_tokens else f"PDF {len(pdfs)} 張",
        "message": f"目前有 {len(serials)} 個流水號可比對" if serial_tokens else "目前沒有紀錄可比對",
        "partial": partial,
        "scanned_dirs": dirs_scanned,
        "scan_limit": PDF_HEALTH_SCAN_FILE_LIMIT,
        "errors": errors[:3],
    }


def _scan_pdf_files(path: Path) -> tuple[list[Path], bool, int, list[str]]:
    """Bounded recursive PDF scan for settings health checks."""
    pdfs: list[Path] = []
    stack = [path]
    dirs_scanned = 0
    partial = False
    errors: list[str] = []
    seen: set[str] = set()

    while stack:
        if len(pdfs) >= PDF_HEALTH_SCAN_FILE_LIMIT or dirs_scanned >= PDF_HEALTH_SCAN_DIR_LIMIT:
            partial = True
            break
        current = stack.pop()
        key = str(current.absolute())
        if key in seen:
            continue
        seen.add(key)
        dirs_scanned += 1
        try:
            with os.scandir(current) as entries:
                for entry in entries:
                    if len(pdfs) >= PDF_HEALTH_SCAN_FILE_LIMIT:
                        partial = True
                        break
                    try:
                        if entry.is_file(follow_symlinks=False) and entry.name.lower().endswith(".pdf"):
                            pdfs.append(Path(entry.path))
                        elif entry.is_dir(follow_symlinks=False):
                            stack.append(Path(entry.path))
                    except OSError as exc:
                        errors.append(f"{entry.path}: {exc}")
        except OSError as exc:
            errors.append(f"{current}: {exc}")
    if stack:
        partial = True
    return pdfs, partial, dirs_scanned, errors


def _photo_label(role: Any) -> str:
    value = str(role or "").strip().lower()
    if value == "before":
        return "修改前"
    if value == "after":
        return "修改後"
    return str(role or "")


def _date_key(value: Any) -> str:
    """Normalize UI/report dates to YYYYMMDD so the output tab can group records."""
    digits = re.sub(r"\D", "", str(value or ""))
    return digits[:8] if len(digits) >= 8 else ""


def _status_key(value: Any) -> str:
    raw = str(value or "").strip()
    return {"完整": "done", "待補": "pending", "草稿": "pending"}.get(raw, raw or "pending")


def _weld_kind_label(value: Any, origin: Any = None) -> str:
    raw = str(value or "").strip()
    origin_text = str(origin or "").strip()
    if raw in {"新焊", "新增焊口"}:
        return "新焊"
    if raw in {"重焊", "原焊口重接", "拆除不重焊"}:
        return "重焊"
    if raw in {"裁切", "加長", "縮短"}:
        return "新焊" if origin_text == "new" else "重焊"
    if origin_text == "new":
        return "新焊"
    if origin_text == "existing":
        return "重焊"
    return raw


def _weld_identity_from_code(
    code: Any,
    mark: Any = None,
    origin: Any = None,
    *,
    code_changed: bool = True,
) -> dict[str, Any]:
    text = str(code or "").strip()
    try:
        from weld_codec import parse as parse_weld_code

        parsed = parse_weld_code(text)
    except Exception:
        parsed = None
    if parsed is not None and getattr(parsed, "parsed", False):
        if getattr(parsed, "is_new", False):
            fallback_op = _weld_kind_label(mark, origin) or ""
            if not code_changed and fallback_op == "重焊":
                return {"base": str(getattr(parsed, "raw", text) or text), "origin": "existing", "op": "重焊"}
            return {"base": None, "origin": "new", "op": "新焊"}
        base = getattr(parsed, "base", None)
        if base:
            return {"base": str(base), "origin": "existing", "op": "重焊"}

    op = _weld_kind_label(mark, origin) or ""
    origin_text = str(origin or "").strip()
    if origin_text not in {"existing", "new"}:
        origin_text = "new" if op == "新焊" else "existing" if op == "重焊" else origin_text
    match = re.match(r"^\s*(\d+)", text)
    return {
        "base": match.group(1) if match and origin_text == "existing" else None,
        "origin": origin_text or None,
        "op": op or None,
    }


_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".bmp", ".gif", ".webp"}
_IMAGE_MIME = {
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".bmp": "image/bmp",
    ".gif": "image/gif",
    ".webp": "image/webp",
}


class MainBridge:
    """主介面橋。對外方法皆回信封；`project_root` 下的 records/ 放舊系統資料。"""

    def __init__(self, project_root: Any, *, pick_file_fn: Optional[Callable[[str], Optional[str]]] = None):
        self.root = Path(project_root)
        self.records_dir = self.root / "records"
        self._pick_file_fn = pick_file_fn  # 由 launcher 注入原生檔案對話框
        self._save_file_fn = None          # 由 launcher 注入原生存檔對話框
        self._pick_folder_fn = None        # 由 launcher 注入原生資料夾對話框
        self._open_path_fn = None          # 測試可注入；桌面版預設用系統開啟
        self._pricebook_cache_key = None
        self._pricebook_cache_data = None
        self._catalog_summary_cache = None

    # ---- 對外 API（皆回信封） -------------------------------------------- #
    @_enveloped
    def info(self) -> dict:
        return {
            "api_version": API_VERSION,
            "root": str(self.root),
            "records_dir": str(self.records_dir),
            "pricebook_exists": (self.records_dir / "material_pricebook.json").exists(),
        }

    @_enveloped
    def pricebook(self) -> list:
        """讀真實料表 → 對映成前端 PRICE 形狀；缺檔回 []。"""
        path = self.records_dir / "material_pricebook.json"
        try:
            stat = path.stat()
            project_path = self._project_parts_path()
            try:
                project_stat = project_path.stat()
                project_cache_key = (project_stat.st_mtime_ns, project_stat.st_size)
            except OSError:
                project_cache_key = None
            cache_key = (stat.st_mtime_ns, stat.st_size, project_cache_key)
        except OSError:
            cache_key = None
        if cache_key is not None and cache_key == self._pricebook_cache_key:
            return self._pricebook_cache_data or []

        data = _read_json(path)
        items = data.get("items") if isinstance(data, dict) else data
        items = list(items or []) + self._read_custom_project_parts()
        taxonomy = load_taxonomy(str(self.root))
        out = []
        for it in (items or []):
            if not isinstance(it, dict):
                continue
            out.append(self._frontend_material_row(it, taxonomy))
        self._pricebook_cache_key = cache_key
        self._pricebook_cache_data = out
        return out

    def _frontend_material_row(self, it: dict, taxonomy: dict | None = None) -> dict:
        """Normalize one material row to the frontend shape used by PRICE."""
        taxonomy = taxonomy or load_taxonomy(str(self.root))
        row = enrich_material_item(it, taxonomy)
        return {
            "id": it.get("id") or "",
            "part": row.get("零件類型") or "",
            "size": row.get("尺寸") or "",
            "sch": row.get("SCH") or "",
            "mat": row.get("材質") or "",
            "cat": row.get("類別") or "",
            "unit": row.get("單位") or "",
            "src": it.get("來源") or row.get("來源") or "",
            "remark": it.get("備註") or row.get("備註") or "",
            "type": it.get("Type") or row.get("Type") or "",
            "level": it.get("支撐級別") or row.get("支撐級別") or "",
            "spec": it.get("規格") or row.get("規格") or "",
            "icon": it.get("icon") or row.get("icon") or "",
            "material_family": row.get("material_family") or "",
            "match_key": row.get("match_key") or "",
            "project_only": bool(it.get("project_only")),
            "source_designation": it.get("source_designation") or "",
        }

    def _legacy_material_rows(self) -> list:
        doc = _read_json(self.records_dir / "material_pricebook.json")
        items = doc.get("items") if isinstance(doc, dict) else doc
        return [it for it in (items or []) if isinstance(it, dict)]

    def _material_items_for_ids(self, ids) -> list:
        wanted = {str(x) for x in (ids or []) if str(x or "").strip()}
        if not wanted:
            return []
        items_by_id: dict[str, dict] = {str(row.get("id")): row for row in rows_by_ids(self.root, wanted)}
        missing = wanted - set(items_by_id)
        if missing:
            taxonomy = load_taxonomy(str(self.root))
            for row in self._legacy_material_rows() + self._read_custom_project_parts():
                rid = str(row.get("id") or "")
                if rid in missing and rid not in items_by_id:
                    items_by_id[rid] = self._frontend_material_row(row, taxonomy)
        return [items_by_id[x] for x in sorted(items_by_id)]

    @_enveloped
    def material_catalog_summary(self) -> dict:
        """Compact material catalog summary; frontend uses this before loading rows."""
        if self._catalog_summary_cache is None:
            self._catalog_summary_cache = catalog_summary(self.root)
        return self._catalog_summary_cache

    @_enveloped
    def material_catalog_query(self, filters: dict | None = None, offset: int = 0, limit: int = 200) -> dict:
        """Query the compact material rules and expand only the requested page."""
        return query_catalog(self.root, filters or {}, offset=offset, limit=limit)

    @_enveloped
    def material_catalog_items(self, ids) -> dict:
        """Resolve material ids from the compact rule catalog, with legacy fallback."""
        return {"items": self._material_items_for_ids(ids)}

    @_enveloped
    def build_project_part(self, spec: dict) -> dict:
        """Build one controlled material spec and register it into project parts."""
        item = build_frontend_item(self.root, spec or {})
        mid = str(item.get("id") or "").strip()
        if not mid:
            raise ValueError("材料規格未產生料號")
        cur = self._read_registered()
        cur.add(mid)
        self._write_registered(cur)
        return {"item": item, "registered": sorted(cur)}

    @_enveloped
    def material_taxonomy(self) -> dict:
        """材料分類 / 規格軸 / 圖示對照。前端用它覆蓋硬編碼 fallback。"""
        taxonomy = load_taxonomy(str(self.root))
        return {**taxonomy, "options": taxonomy_options(taxonomy)}

    @_enveloped
    def support_bom(self, designation: str, overrides: dict | None = None) -> dict:
        """展開管架 Type 編碼為 BOM 與本系統材料列。"""
        return analyze_support_bom(designation, project_root=self.root, overrides=overrides or {})

    def _attachments_root(self) -> Path:
        """精靈出單的根目錄（＝專案 attachments/，與預設 config.ATTACHMENTS_ROOT 同位置、可測）。"""
        return self.root / "attachments"

    def _read_change_orders(self) -> list:
        """掃 attachments 下所有 change_order.json（精靈出的單）→ 前端 record 形狀，日期新到舊。"""
        root = self._attachments_root()
        if not root.exists():
            return []
        out = []
        for cop in root.rglob("change_order.json"):
            data = _read_json(cop)
            if not isinstance(data, dict):
                continue
            out.append(self._record_from_change_order(cop, data))
        out.sort(key=lambda r: (r.get("date") or "", r.get("id") or ""), reverse=True)
        return out

    def _record_from_change_order(self, cop: Path, data: dict) -> dict:
        folder = cop.parent
        welds = []
        for w in (data.get("welds") or []):
            if not isinstance(w, dict):
                continue
            sp = w.get("spec") or {}
            welds.append({
                "code": w.get("code") or "",
                "mark": _weld_kind_label(w.get("op"), w.get("origin")),
                "size": sp.get("size") or "",
                "mat": sp.get("material") or "",
                "sch": sp.get("sch") or "",
                "coef": "",
            })
        mats = []
        for m in (data.get("materials") or []):
            if not isinstance(m, dict):
                continue
            mats.append({
                "id": m.get("component_id") or "",
                "part": m.get("component") or "",
                "size": m.get("size") or "",
                "sch": m.get("schedule") or "",
                "mat": m.get("material") or "",
                "qty": "" if m.get("qty") is None else m.get("qty"),
                "unit": m.get("unit") or "",
                "remark": m.get("remark") or "",
            })
        photos = []
        for ph in (data.get("photos") or []):
            if not isinstance(ph, dict):
                continue
            src = self._attachment_url(folder, ph.get("file") or "")
            photos.append({"src": src, "label": _photo_label(ph.get("role")), "file": ph.get("file") or ""})
        return {
            "id": data.get("id") or folder.name,
            "date": data.get("date") or "",
            "series": str(data.get("series") or ""),
            "status": _status_key(data.get("status")),
            "reason": data.get("reason") or "",
            "welds": welds,
            "mats": mats,
            "photos": photos,
            "folder": str(folder),
        }

    def _attachment_url(self, folder: Path, value: str) -> str:
        text = str(value or "").strip()
        if not text or text.lower().startswith("data:"):
            return text
        path = self._path_from_file_value(text)
        candidate = path if path.is_absolute() else folder / path
        try:
            if candidate.is_file():
                return str(candidate.resolve())
        except OSError:
            pass
        return text

    def _find_change_order(self, record_id: str) -> tuple[Path, dict]:
        target = str(record_id or "").strip()
        if not target:
            raise ValueError("缺少記錄編號")
        for path in self._attachments_root().rglob("change_order.json"):
            data = _read_json(path)
            if not isinstance(data, dict):
                continue
            if str(data.get("id") or path.parent.name) == target:
                return path, data
        raise FileNotFoundError(f"找不到修改單記錄：{record_id}")

    def _write_change_order(self, path: Path, data: dict) -> None:
        atomic_write_json(str(path), data)

    def _photo_at(self, data: dict, index: int) -> dict:
        photos = data.get("photos")
        if not isinstance(photos, list):
            raise ValueError("此記錄沒有照片清單")
        if index < 0 or index >= len(photos) or not isinstance(photos[index], dict):
            raise IndexError("照片序號超出範圍")
        return photos[index]

    def _photo_response(self, folder: Path, photo: dict) -> dict:
        return {
            "src": self._attachment_url(folder, photo.get("file") or ""),
            "label": _photo_label(photo.get("role")),
            "file": photo.get("file") or "",
        }

    def _path_from_file_value(self, value: str) -> Path:
        text = str(value or "").strip()
        if text.lower().startswith("file:"):
            from urllib.parse import unquote, urlparse
            parsed = urlparse(text)
            raw = unquote(parsed.path or "")
            if parsed.netloc:
                raw = f"//{parsed.netloc}{raw}"
            if len(raw) >= 4 and raw[0] == "/" and raw[2] == ":":
                raw = raw[1:]
            return Path(raw)
        return Path(text)

    def _resolved_photo_path(self, folder: Path, value: str) -> Path:
        path = self._path_from_file_value(value)
        return path if path.is_absolute() else folder / path

    @_enveloped
    def image_data_url(self, file_path: str) -> dict:
        """把本機圖片讀成 data URL，給燈箱與縮圖使用。"""
        text = str(file_path or "").strip()
        if not text:
            raise ValueError("缺少圖片路徑")
        if text.lower().startswith("data:"):
            return {"name": "", "path": text, "url": text}
        path = self._path_from_file_value(text)
        if not path.is_absolute():
            path = self.root / path
        if not path.is_file():
            raise FileNotFoundError(f"找不到圖片：{file_path}")
        suffix = path.suffix.lower()
        if suffix not in _IMAGE_EXTS:
            raise ValueError(f"不支援的圖片格式：{suffix or path.name}")
        import base64
        import mimetypes
        mime = _IMAGE_MIME.get(suffix) or mimetypes.guess_type(path.name)[0] or "application/octet-stream"
        encoded = base64.b64encode(path.read_bytes()).decode("ascii")
        return {"name": path.name, "path": str(path.resolve()), "url": f"data:{mime};base64,{encoded}"}

    @_enveloped
    def replace_photo(self, record_id: str, photo_index: int, source_path: str) -> dict:
        """取代某張照片：複製新檔到該修改單附件資料夾，並更新 change_order.json。"""
        record_path, data = self._find_change_order(record_id)
        folder = record_path.parent
        photo = self._photo_at(data, int(photo_index))
        source = self._path_from_file_value(source_path)
        if not source.is_file():
            raise FileNotFoundError(f"找不到取代照片：{source_path}")
        suffix = source.suffix.lower()
        if suffix not in _IMAGE_EXTS:
            raise ValueError(f"不支援的圖片格式：{suffix or source.name}")
        current = str(photo.get("file") or "").strip()
        role = str(photo.get("role") or "photo").strip().lower() or "photo"
        old_path = self._resolved_photo_path(folder, current) if current else None
        stem = old_path.stem if old_path is not None and old_path.name else role
        destination = folder / f"{stem}{suffix}"
        if source.resolve() != destination.resolve():
            import shutil
            shutil.copy2(source, destination)
        photo["file"] = destination.name
        self._write_change_order(record_path, data)
        return {"record": str(record_path), "photo": self._photo_response(folder, photo)}

    @_enveloped
    def save_photo_annotation(self, record_id: str, photo_index: int, data_url: str) -> dict:
        """儲存燈箱 canvas 合成後的標註圖，並把該照片指向標註檔。"""
        payload = str(data_url or "")
        if "," in payload:
            header, payload = payload.split(",", 1)
            if "base64" not in header.lower():
                raise ValueError("標註資料必須是 base64 data URL")
        import base64
        raw = base64.b64decode(payload, validate=True)
        record_path, data = self._find_change_order(record_id)
        folder = record_path.parent
        photo = self._photo_at(data, int(photo_index))
        current = str(photo.get("file") or "").strip()
        current_path = self._resolved_photo_path(folder, current) if current else folder / "photo.png"
        base = current_path.stem[:-10] if current_path.stem.endswith("_annotated") else current_path.stem
        destination = folder / f"{base}_annotated.png"
        destination.write_bytes(raw)
        photo["file"] = destination.name
        self._write_change_order(record_path, data)
        return {"record": str(record_path), "photo": self._photo_response(folder, photo)}

    @_enveloped
    def delete_photo(self, record_id: str, photo_index: int) -> dict:
        """從 change_order.json 移除照片引用；檔案保留在附件資料夾中。"""
        record_path, data = self._find_change_order(record_id)
        photos = data.get("photos")
        if not isinstance(photos, list):
            raise ValueError("此記錄沒有照片清單")
        removed = photos.pop(int(photo_index))
        self._write_change_order(record_path, data)
        return {"record": str(record_path), "removed": removed, "count": len(photos)}

    @_enveloped
    def records(self) -> list:
        """讀精靈出的 change_order.json（attachments/）→ 前端記錄形狀；無則 []。"""
        return self._read_change_orders()


    @_enveloped
    def dates(self) -> list:
        """產出報告左欄：優先彙整新版草稿 change_order.json，再補舊 weld_snapshot。"""
        by_date: dict[str, list] = {}
        seen: set[tuple[str, str]] = set()
        for rec in self._read_change_orders():
            date = _date_key(rec.get("date"))
            series = str(rec.get("series") or "")
            if not date or not series:
                continue
            seen.add((date, series))
            welds = rec.get("welds") or []
            weld_codes = [
                str(w.get("code") or "").strip()
                for w in welds
                if isinstance(w, dict) and str(w.get("code") or "").strip()
            ]
            by_date.setdefault(date, []).append({
                "series": series,
                "welds": len(welds),
                "weld_codes": weld_codes,
                "status": rec.get("status") or "pending",
                "sel": False,
                "folder": rec.get("folder") or "",
                "record_id": rec.get("id") or "",
                "source": "change_order",
            })

        snap = _read_json(self.records_dir / "weld_snapshot.json") or {}
        folders = snap.get("folders") if isinstance(snap, dict) else None
        for key, info in (folders or {}).items():
            if not isinstance(info, dict):
                continue
            date = _date_key(str(key).split("/", 1)[0])
            serial = info.get("serial") or info.get("raw_serial") or ""
            serial = str(serial)
            if not date or not serial or (date, serial) in seen:
                continue
            snap_welds = info.get("welds") or []
            weld_codes = []
            for w in snap_welds:
                code = w.get("code") if isinstance(w, dict) else w
                code = str(code or "").strip()
                if code:
                    weld_codes.append(code)
            by_date.setdefault(date, []).append({
                "series": serial,
                "welds": len(snap_welds),
                "weld_codes": weld_codes,
                "status": "pending",
                "sel": False,
                "folder": str(key),
                "record_id": "",
                "source": "weld_snapshot",
            })
        out = []
        for date in sorted(by_date.keys(), reverse=True):
            items = sorted(by_date[date], key=lambda it: str(it.get("series") or ""))
            out.append({"date": date, "open": False, "items": items})
        if out:
            out[0]["open"] = True
        return out

    @_enveloped
    def billing(self) -> dict:
        """請款：每筆記錄一列（rec 對映 records()），狀態預設未請款、可從 billing.json 讀回。"""
        recs = self._read_change_orders()
        saved = _read_json(self.records_dir / "billing.json") or {}
        by_id = saved.get("byId") if isinstance(saved, dict) else {}
        by_id = by_id if isinstance(by_id, dict) else {}
        rows = []
        for i, r in enumerate(recs):
            st = by_id.get(r["id"]) or {}
            rows.append({
                "rec": i,
                "status": st.get("status") or "未請款",
                "billDate": st.get("billDate") or "",
                "remark": st.get("remark") or "",
            })
        batches_doc = _read_json(self.records_dir / "billing_batches.json") or {}
        meta = batches_doc.get("meta") or {}
        batches = [b for b in (batches_doc.get("batches") or []) if isinstance(b, dict)]
        return {"rows": rows, "batches": batches, "tax_rate": meta.get("tax_rate", "5%")}

    @_enveloped
    def save_billing(self, rows: list) -> dict:
        """儲存請款追蹤狀態；只存狀態/日期/備註，不涉及金額。"""
        recs = self._read_change_orders()
        by_id = {}
        for row in rows or []:
            if not isinstance(row, dict):
                continue
            try:
                idx = int(row.get("rec"))
            except (TypeError, ValueError):
                continue
            if idx < 0 or idx >= len(recs):
                continue
            rid = recs[idx].get("id")
            if not rid:
                continue
            by_id[str(rid)] = {
                "status": row.get("status") or "未請款",
                "billDate": row.get("billDate") or "",
                "remark": row.get("remark") or "",
            }
        self.records_dir.mkdir(parents=True, exist_ok=True)
        atomic_write_json(str(self.records_dir / "billing.json"), {"byId": by_id})
        return {"saved": True, "count": len(by_id)}

    def _open_path(self, path: Path) -> None:
        target = Path(path)
        if not target.exists():
            raise FileNotFoundError(f"找不到路徑：{target}")
        if self._open_path_fn is not None:
            self._open_path_fn(str(target))
            return
        import os
        import subprocess
        import sys
        if sys.platform.startswith("win"):
            os.startfile(str(target))  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            subprocess.Popen(["open", str(target)])
        else:
            subprocess.Popen(["xdg-open", str(target)])

    def _drawing_pdf_path(self, folder: Path, data: dict) -> Path:
        drawing = data.get("drawing_pdf")
        if isinstance(drawing, dict) and drawing.get("file"):
            candidate = self._resolved_photo_path(folder, str(drawing.get("file") or ""))
            if candidate.is_file():
                return candidate
        pdfs = sorted(p for p in folder.glob("*.pdf") if p.is_file())
        if pdfs:
            return pdfs[0]
        raise FileNotFoundError("這筆記錄沒有可開啟的 PDF")

    @_enveloped
    def open_record_folder(self, record_id: str) -> dict:
        """用系統檔案總管開啟修改單附件資料夾。"""
        record_path, _data = self._find_change_order(record_id)
        folder = record_path.parent
        self._open_path(folder)
        return {"path": str(folder)}

    @_enveloped
    def open_record_pdf(self, record_id: str) -> dict:
        """開啟這筆修改單附件資料夾中的圖面 PDF。"""
        record_path, data = self._find_change_order(record_id)
        pdf = self._drawing_pdf_path(record_path.parent, data)
        self._open_path(pdf)
        return {"path": str(pdf)}

    @_enveloped
    def save_record(self, record: dict) -> dict:
        """把紀錄管理中可編輯的焊口/材料明細寫回 change_order.json。"""
        if not isinstance(record, dict):
            raise ValueError("紀錄資料格式錯誤")
        record_path, data = self._find_change_order(str(record.get("id") or ""))
        if "date" in record:
            data["date"] = record.get("date") or data.get("date")
        if "series" in record:
            data["series"] = str(record.get("series") or data.get("series") or "")
        if "reason" in record:
            data["reason"] = record.get("reason") or None

        old_welds = data.get("welds") if isinstance(data.get("welds"), list) else []
        new_welds = []
        for i, row in enumerate(record.get("welds") or []):
            if not isinstance(row, dict):
                continue
            old = dict(old_welds[i]) if i < len(old_welds) and isinstance(old_welds[i], dict) else {
                "joint_type": "焊口",
                "origin": "manual",
            }
            spec = dict(old.get("spec") or {})
            code = str(row.get("code") or "")
            old_code = str(old.get("code") or "")
            identity = _weld_identity_from_code(
                code,
                row.get("mark"),
                old.get("origin"),
                code_changed=code != old_code,
            )
            old["code"] = code
            old["base"] = identity.get("base")
            old["origin"] = identity.get("origin") or old.get("origin")
            old["op"] = identity.get("op") or _weld_kind_label(row.get("mark"), old.get("origin"))
            spec["size"] = str(row.get("size") or "")
            spec["material"] = str(row.get("mat") or "")
            spec["sch"] = str(row.get("sch") or "")
            old["spec"] = spec
            new_welds.append(old)
        data["welds"] = new_welds

        new_mats = []
        for row in record.get("mats") or []:
            if not isinstance(row, dict):
                continue
            qty = row.get("qty")
            try:
                qn = float(str(qty).replace(",", ""))
                qty = int(qn) if qn == int(qn) else qn
            except (TypeError, ValueError):
                qty = qty if qty not in (None, "") else 1
            new_mats.append({
                "component_id": str(row.get("id") or ""),
                "component": str(row.get("part") or ""),
                "size": str(row.get("size") or ""),
                "schedule": str(row.get("sch") or ""),
                "material": str(row.get("mat") or ""),
                "qty": qty,
                "unit": str(row.get("unit") or ""),
                "remark": str(row.get("remark") or ""),
            })
        data["materials"] = new_mats

        import datetime
        audit = data.get("audit") if isinstance(data.get("audit"), dict) else {}
        history = audit.get("history") if isinstance(audit.get("history"), list) else []
        history.append({
            "who": None,
            "when": datetime.datetime.now().isoformat(timespec="seconds"),
            "action": "saved_from_main",
            "detail": None,
        })
        audit["history"] = history
        data["audit"] = audit
        self._write_change_order(record_path, data)
        return self._record_from_change_order(record_path, data)

    def _export_rows(self, path: str, default_name: str, sheet_name: str, columns: list[str], rows: list[list[Any]]) -> dict:
        if not path:
            if self._save_file_fn is None:
                raise RuntimeError("存檔對話框未注入（此環境不支援匯出）")
            path = self._save_file_fn("excel", default_name)
            if not path:
                return {"cancelled": True}
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(columns)
        for row in rows:
            ws.append(row)
        ws.freeze_panes = "A2"
        for col in ws.columns:
            width = max(len(str(cell.value or "")) for cell in col[:60])
            ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 10), 42)
        wb.save(path)
        return {"path": str(path), "count": len(rows)}

    @_enveloped
    def export_records(self, path: str = "") -> dict:
        """匯出目前 change_order 記錄清單成 Excel。"""
        rows = []
        for r in self._read_change_orders():
            rows.append([
                r.get("id", ""),
                r.get("date", ""),
                r.get("series", ""),
                r.get("status", ""),
                r.get("reason", ""),
                len(r.get("welds") or []),
                len(r.get("mats") or []),
                len(r.get("photos") or []),
                r.get("folder", ""),
            ])
        return self._export_rows(
            path,
            "修改單紀錄清單.xlsx",
            "紀錄清單",
            ["報告編號", "日期", "流水號", "狀態", "說明", "焊口數", "材料數", "照片數", "資料夾"],
            rows,
        )

    @_enveloped
    def export_record_materials(self, path: str = "") -> dict:
        """匯出所有修改單材料明細彙總；只記品項與量，不帶價格。"""
        rows = []
        for r in self._read_change_orders():
            for m in r.get("mats") or []:
                rows.append([
                    r.get("id", ""),
                    r.get("date", ""),
                    r.get("series", ""),
                    m.get("id", ""),
                    m.get("part", ""),
                    m.get("size", ""),
                    m.get("sch", ""),
                    m.get("mat", ""),
                    m.get("qty", ""),
                    m.get("unit", ""),
                    m.get("remark", ""),
                ])
        return self._export_rows(
            path,
            "修改單材料彙總.xlsx",
            "材料彙總",
            ["報告編號", "日期", "流水號", "料號", "零件", "尺寸", "SCH", "材質", "數量", "單位", "備註"],
            rows,
        )

    def _include_keys_from_output_items(self, items: Any) -> list[tuple[str, str]] | None:
        keys: list[tuple[str, str]] = []
        seen: set[tuple[str, str]] = set()
        attachments_root = self._attachments_root().resolve()
        for item in items or []:
            if not isinstance(item, dict):
                continue
            date = _date_key(item.get("date"))
            folder = str(item.get("folder") or "").strip().replace("\\", "/")
            if folder:
                raw_path = Path(folder)
                if raw_path.is_absolute():
                    try:
                        rel = raw_path.resolve().relative_to(attachments_root)
                        parts = rel.parts
                        if len(parts) >= 2 and (not date or _date_key(parts[0]) == date):
                            date = _date_key(parts[0])
                            folder = "/".join(parts[1:])
                        elif parts:
                            folder = parts[-1]
                    except Exception:
                        folder = raw_path.name or folder
                elif "/" in folder:
                    first, rest = folder.split("/", 1)
                    if not date or _date_key(first) == date:
                        date = _date_key(first)
                        folder = rest.strip("/")
            if not folder:
                folder = str(item.get("record_id") or item.get("series") or "").strip()
            key = (date, folder)
            if not date or not folder or key in seen:
                continue
            seen.add(key)
            keys.append(key)
        return keys or None

    @_enveloped
    def export_output_center(self, report_type: str = "owner-data", selected: list | None = None) -> dict:
        """建立輸出中心；新版 GUI 預設產出可交付的業主資料包。"""
        from site_output_center import run_site_output_center

        output_dir = self.root / "staging" / "site_output_center_web"
        return run_site_output_center(
            output_dir,
            project_root=self.root,
            attachments_root=self._attachments_root(),
            include_report_keys=self._include_keys_from_output_items(selected),
            overwrite=True,
            render_pdf=True,
            render_png=False,
            render_statistics=True,
            render_summary_pdf=True,
            render_photo_grid_pdf=True,
            report_type=report_type or "owner-data",
        )

    @_enveloped
    def open_output_folder(self, kind: str = "output") -> dict:
        """開啟設定中的 output/pdf 資料夾；未設定時使用專案內 output/ 或 pdf/。"""
        kind_text = str(kind or "").strip().lower()
        if kind_text in {"owner", "owner-data", "owner_data", "owner_data_report"}:
            target = self.root / "staging" / "site_output_center_web" / "owner_data_report"
            if not target.exists():
                target = self.root / "staging" / "site_output_center_web"
            target.mkdir(parents=True, exist_ok=True)
            self._open_path(target)
            return {"path": str(target)}
        if kind_text in {"site", "site-output", "site_output", "site_output_center", "developer"}:
            target = self.root / "staging" / "site_output_center_web"
            target.mkdir(parents=True, exist_ok=True)
            self._open_path(target)
            return {"path": str(target)}
        settings = self.app_settings()
        data = settings.get("data") if isinstance(settings, dict) else {}
        key = "pdf_dir" if kind_text == "pdf" else "output_dir"
        fallback = "pdf" if key == "pdf_dir" else "output"
        raw = str((data or {}).get(key) or "").strip()
        target = Path(raw) if raw else self.root / fallback
        if not target.is_absolute():
            target = self.root / target
        target.mkdir(parents=True, exist_ok=True)
        self._open_path(target)
        return {"path": str(target)}

    @_enveloped
    def open_wizard(self) -> dict:
        """開修改單精靈（subprocess 另開原生視窗、獨立行程；接真橋、能出單寫 change_order.json）。"""
        command, mode = _wizard_launch_command(self.root)
        subprocess.Popen(command, cwd=str(self.root))
        return {"launched": True, "mode": mode, "command": command}

    # ---- 設定：路徑（config 預設 + records/app_settings.json 覆蓋）---------- #
    @_enveloped
    def app_settings(self) -> dict:
        """回目前路徑設定：有存過就用存的，否則讀專案 settings/config 預設。"""
        saved = _read_json(self.records_dir / "app_settings.json")
        saved = saved if isinstance(saved, dict) else {}
        dwg = out = pdf = ""
        try:
            import config as _C
            dwg = str(getattr(_C, "DRAWING_LIST_PATH", "") or "")
            out = str(getattr(_C, "OUTPUT_ROOT", "") or "")
            pdf = str(getattr(_C, "PDF_OUTPUT_DIR", "") or "")
        except Exception:
            pass
        weld = _settings_path_value(self.root, "weld_control_table")
        prefab = _settings_path_value(self.root, "prefab_drawing_dir")
        data = {
            "project_name": saved.get("project_name") or _settings_project_name(self.root),
            "dwg_list": saved.get("dwg_list") or _settings_path_value(self.root, "drawing_list") or dwg,
            "weld_control_table": saved.get("weld_control_table") or weld,
            "prefab_drawing_dir": saved.get("prefab_drawing_dir") or prefab,
            "output_dir": saved.get("output_dir") or out,
            "pdf_dir": saved.get("pdf_dir") or pdf,
        }
        data["source_schema"] = self._source_schema_settings()
        data["source_health"] = self._source_health(data)
        return data

    def _source_schema_settings(self) -> dict:
        weld_cfg = _settings_section(self.root, "weld_control")
        dwg_cfg = _settings_section(self.root, "dwg_list")
        dwg = {"sheet_name": str(dwg_cfg.get("sheet_name") or "DRAWING LIST")}
        for role in _source_role_defs("dwg", dwg_cfg):
            dwg[role["key"]] = str(dwg_cfg.get(role["setting_key"]) or role.get("default") or "")
        weld = {"sheet_name": str(weld_cfg.get("sheet_name") or "焊口編號明細")}
        for role in _source_role_defs("weld", weld_cfg):
            weld[role["key"]] = str(weld_cfg.get(role["setting_key"]) or role.get("default") or "")
        return {"dwg": dwg, "weld": weld}

    def _source_health(self, settings: dict) -> dict:
        weld_cfg = _settings_section(self.root, "weld_control")
        dwg_cfg = _settings_section(self.root, "dwg_list")
        serials = {
            str(rec.get("series") or "").strip()
            for rec in self._read_change_orders()
            if isinstance(rec, dict) and str(rec.get("series") or "").strip()
        }
        return {
            "dwg": _excel_source_health(
                self.root,
                settings.get("dwg_list"),
                sheet_name=str(dwg_cfg.get("sheet_name") or "DRAWING LIST"),
                required=_source_required_options(_source_role_defs("dwg", dwg_cfg)),
                roles=_source_role_defs("dwg", dwg_cfg),
            ),
            "weld": _excel_source_health(
                self.root,
                settings.get("weld_control_table"),
                sheet_name=str(weld_cfg.get("sheet_name") or "焊口編號明細"),
                required=_source_required_options(_source_role_defs("weld", weld_cfg)),
                roles=_source_role_defs("weld", weld_cfg),
            ),
            "drawingpdf": _pdf_source_health(
                self.root,
                settings.get("prefab_drawing_dir"),
                serials,
            ),
        }

    @_enveloped
    def save_setting(self, key: str, value: Any) -> dict:
        """存單一設定；路徑同時寫入 settings.json，讓修改單精靈可讀到焊口表。"""
        p = self.records_dir / "app_settings.json"
        saved = _read_json(p)
        saved = saved if isinstance(saved, dict) else {}
        key = str(key)
        saved[key] = value
        self.records_dir.mkdir(parents=True, exist_ok=True)
        atomic_write_json(str(p), saved)
        if key in {"dwg_list", "weld_control_table", "prefab_drawing_dir"}:
            settings_key = "drawing_list" if key == "dwg_list" else key
            self._write_settings_path(settings_key, value)
        if key == "project_name":
            self._write_settings_project_name(value)
        return {"saved": True, "key": key, "value": value}

    @_enveloped
    def save_source_schema(self, kind: str, config: dict) -> dict:
        """存 DWG LIST / 焊口表的格式設定，供來源健康檢查與查詢流程共用。"""
        kind = str(kind or "").strip().lower()
        config = config if isinstance(config, dict) else {}
        if kind == "dwg":
            roles = _source_role_defs("dwg")
            updates = {"sheet_name": str(config.get("sheet_name") or "").strip() or "DRAWING LIST"}
            for role in roles:
                updates[role["setting_key"]] = (
                    str(config.get(role["key"]) or config.get(role["setting_key"]) or "").strip()
                    or str(role.get("default") or "")
                )
            self._write_settings_section("dwg_list", updates)
        elif kind == "weld":
            roles = _source_role_defs("weld")
            updates = {"sheet_name": str(config.get("sheet_name") or "").strip() or "焊口編號明細"}
            for role in roles:
                updates[role["setting_key"]] = (
                    str(config.get(role["key"]) or config.get(role["setting_key"]) or "").strip()
                    or str(role.get("default") or "")
                )
            self._write_settings_section("weld_control", updates)
        else:
            raise ValueError(f"不支援的來源格式設定：{kind}")
        return {"saved": True, "kind": kind, "settings": self._source_schema_settings().get(kind, {})}

    @_enveloped
    def source_excel_preview(self, kind: str) -> dict:
        """把來源 Excel 映成前端可視化欄位設定預覽，不要求使用者背欄位常數。"""
        kind = str(kind or "").strip().lower()
        settings_res = self.app_settings()
        settings = settings_res.get("data") if isinstance(settings_res, dict) else {}
        schema = self._source_schema_settings()
        if kind == "dwg":
            cfg = schema["dwg"]
            roles = _source_role_defs("dwg", _settings_section(self.root, "dwg_list"))
            return _excel_source_preview(
                self.root,
                settings.get("dwg_list"),
                sheet_name=cfg["sheet_name"],
                required=_source_required_options(roles),
                roles=roles,
            )
        if kind == "weld":
            cfg = schema["weld"]
            roles = _source_role_defs("weld", _settings_section(self.root, "weld_control"))
            return _excel_source_preview(
                self.root,
                settings.get("weld_control_table"),
                sheet_name=cfg["sheet_name"],
                required=_source_required_options(roles),
                roles=roles,
            )
        raise ValueError(f"不支援的來源預覽：{kind}")

    def _write_settings_section(self, section_name: str, updates: dict) -> None:
        p = self.root / "settings.json"
        data = _read_json(p)
        data = data if isinstance(data, dict) else {}
        section = data.get(section_name)
        if not isinstance(section, dict):
            section = {}
            data[section_name] = section
        for key, value in updates.items():
            section[str(key)] = value
        meta = data.get("meta")
        if not isinstance(meta, dict):
            meta = {}
            data["meta"] = meta
        try:
            import datetime
            meta["last_modified"] = datetime.datetime.now().isoformat()
        except Exception:
            pass
        atomic_write_json(str(p), data)

    def _write_settings_project_name(self, value: Any) -> None:
        p = self.root / "settings.json"
        data = _read_json(p)
        data = data if isinstance(data, dict) else {}
        project = data.get("project")
        if not isinstance(project, dict):
            project = {}
            data["project"] = project
        project["name"] = "" if value is None else str(value)
        meta = data.get("meta")
        if not isinstance(meta, dict):
            meta = {}
            data["meta"] = meta
        try:
            import datetime
            meta["last_modified"] = datetime.datetime.now().isoformat()
        except Exception:
            pass
        atomic_write_json(str(p), data)

    def _write_settings_path(self, key: str, value: Any) -> None:
        p = self.root / "settings.json"
        data = _read_json(p)
        data = data if isinstance(data, dict) else {}
        paths = data.get("paths")
        if not isinstance(paths, dict):
            paths = {}
            data["paths"] = paths
        paths[str(key)] = "" if value is None else str(value)
        meta = data.get("meta")
        if not isinstance(meta, dict):
            meta = {}
            data["meta"] = meta
        try:
            import datetime
            meta["last_modified"] = datetime.datetime.now().isoformat()
        except Exception:
            pass
        atomic_write_json(str(p), data)

    @_enveloped
    def pick_path(self, kind: str = "output") -> dict:
        """開原生對話框選路徑：kind='dwg'/'weld' 選 Excel 檔，'output'/'pdf' 選資料夾。"""
        if kind in {"dwg", "weld"}:
            if self._pick_file_fn is None:
                raise RuntimeError("檔案對話框未注入")
            return {"path": self._pick_file_fn("excel")}
        if self._pick_folder_fn is None:
            raise RuntimeError("資料夾對話框未注入")
        return {"path": self._pick_folder_fn()}

    @_enveloped
    def health(self) -> dict:
        """健康：跑既有 integrity_audit / project_guard，回計數 + 問題 + 狀態。"""
        counts: dict = {}
        issues: list = []
        try:
            from integrity_audit import audit_integrity
            a = audit_integrity(str(self.root))
            counts = dict(getattr(a, "counts", {}) or {})
            for it in (getattr(a, "issues", []) or []):
                issues.append(_issue_dict("資料稽核", it))
        except Exception:
            pass
        try:
            from project_guard import inspect_project
            g = inspect_project(str(self.root))
            for it in (getattr(g, "issues", []) or []):
                issues.append(_issue_dict("啟動守門", it))
        except Exception:
            pass
        errs = sum(1 for i in issues if i["level"] == "error")
        warns = sum(1 for i in issues if i["level"] == "warning")
        infos = sum(1 for i in issues if i["level"] == "info")
        if errs:
            status, label = "err", f"需要人工確認：{errs} 個錯誤"
        elif warns or infos:
            status, label = "warn", f"可使用，有 {warns} 個提醒"
        else:
            status, label = "ok", "正常：未發現問題"
        return {
            "status": status, "label": label, "counts": counts, "issues": issues,
            "root": str(self.root), "errors": errs, "warnings": warns, "infos": infos,
        }

    @_enveloped
    def pick_file(self, kind: str = "excel") -> dict:
        """開原生檔案對話框（由 launcher 注入）；純環境未注入則回明確錯。"""
        if self._pick_file_fn is None:
            raise RuntimeError("檔案對話框未注入（此環境不支援選檔）")
        return {"path": self._pick_file_fn(kind)}

    @_enveloped
    def import_material_excel(self, path: str) -> dict:
        """匯入管制單 Excel（自動辨識多格式）→ 併入總庫（去重、不覆蓋、只加新品項）。"""
        new_items = _parse_material_xlsx(path)
        taxonomy = load_taxonomy(str(self.root))
        clean_new_items = []
        for it in new_items:
            row = enrich_material_item(it, taxonomy)
            clean_new_items.append({
                **it,
                "零件類型": row.get("零件類型") or "",
                "尺寸": row.get("尺寸") or "",
                "SCH": row.get("SCH") or "",
                "材質": row.get("材質") or "",
                "類別": row.get("類別") or "",
                "單位": row.get("單位") or it.get("單位") or "",
            })
        new_items = clean_new_items
        self.records_dir.mkdir(parents=True, exist_ok=True)
        p = self.records_dir / "material_pricebook.json"
        doc = _read_json(p)
        existing = (doc.get("items") if isinstance(doc, dict) else doc) or []

        def _key(it):
            return material_match_key(it, taxonomy)   # 顯示鍵：畫面看到的四欄相同即視為同一料

        seen = {_key(it) for it in existing}
        added = []
        for it in new_items:
            k = _key(it)
            if k in seen:
                continue
            seen.add(k)
            added.append(it)
        merged = list(existing) + added
        for i, it in enumerate(merged):
            it["id"] = f"{i + 1:04d}"
        if p.exists():                            # 備份用時間戳，絕不蓋掉 442 的 .backup.json
            import shutil
            import datetime
            stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            shutil.copy(p, self.records_dir / f"material_pricebook.{stamp}.bak.json")
        atomic_write_json(
            str(p),
            {
                "items": merged,
                "history": [],
                "meta": {
                    "version": "3.1",
                    "kind": "總庫",
                    "count": len(merged),
                    "priced": False,
                },
            },
        )
        return {"added": len(added), "total": len(merged)}

    # ---- 本案配件登記（總庫 → 本案子集，只記品項不記數量）------------------ #
    def _project_parts_path(self) -> Path:
        return self.records_dir / "project_parts.json"

    def _read_project_parts_doc(self) -> dict:
        data = _read_json(self._project_parts_path())
        return data if isinstance(data, dict) else {}

    def _read_registered(self) -> set:
        data = self._read_project_parts_doc()
        reg = data.get("registered") if isinstance(data, dict) else None
        return set(str(x) for x in (reg or []))

    def _read_custom_project_parts(self) -> list:
        data = self._read_project_parts_doc()
        custom = data.get("custom") if isinstance(data, dict) else None
        return [x for x in (custom or []) if isinstance(x, dict) and x.get("id")]

    def _write_project_parts_doc(self, reg, custom=None) -> None:
        self.records_dir.mkdir(parents=True, exist_ok=True)
        ids = sorted(set(str(x) for x in reg))
        if custom is None:
            custom = self._read_custom_project_parts()
        atomic_write_json(
            str(self._project_parts_path()),
            {
                "registered": ids,
                "custom": list(custom or []),
                "meta": {"count": len(ids), "custom_count": len(custom or [])},
            },
        )

    def _write_registered(self, reg) -> None:
        self._write_project_parts_doc(reg)

    def _project_material_from_frontend(self, row: dict) -> dict:
        mid = str(row.get("id") or "").strip()
        if not mid:
            raise ValueError("project material missing id")
        return {
            "id": mid,
            "零件類型": str(row.get("part") or row.get("零件類型") or ""),
            "尺寸": str(row.get("size") or row.get("尺寸") or ""),
            "SCH": str(row.get("sch") or row.get("SCH") or ""),
            "材質": str(row.get("mat") or row.get("材質") or ""),
            "類別": str(row.get("cat") or row.get("類別") or ""),
            "單位": str(row.get("unit") or row.get("單位") or ""),
            "來源": str(row.get("src") or row.get("來源") or "管架展開"),
            "規格": str(row.get("spec") or row.get("規格") or ""),
            "備註": str(row.get("remark") or row.get("備註") or ""),
            "Type": str(row.get("type") or row.get("Type") or ""),
            "支撐級別": str(row.get("level") or row.get("支撐級別") or "管架展開"),
            "project_only": True,
            "source_designation": str(row.get("source_designation") or ""),
        }

    @_enveloped
    def project_parts(self) -> dict:
        """本案已登記的料號清單（前端據此把總庫過濾成本案配件）。"""
        cur = self._read_registered()
        items = self._material_items_for_ids(cur)
        registered = {str(row.get("id")) for row in items if row.get("id")}
        dropped = cur - registered
        if dropped:
            self._write_registered(registered)
        return {
            "registered": sorted(registered),
            "dropped": sorted(dropped),
            "items": items,
            "custom": self._read_custom_project_parts(),
        }

    @_enveloped
    def register_parts(self, ids) -> dict:
        """把料號加入本案配件（勾選登記）。"""
        cur = self._read_registered()
        requested = [str(x) for x in (ids or [])]
        valid_requested = {str(row.get("id")) for row in self._material_items_for_ids(requested) if row.get("id")}
        add = [x for x in requested if x in valid_requested]
        cur.update(add)
        self._write_registered(cur)
        return {"registered": sorted(cur), "added": add, "ignored": [x for x in requested if x not in add]}

    @_enveloped
    def unregister_parts(self, ids) -> dict:
        """把料號移出本案配件（取消登記）。"""
        cur = self._read_registered()
        rem = [str(x) for x in (ids or [])]
        cur.difference_update(rem)
        self._write_registered(cur)
        return {"registered": sorted(cur), "removed": rem}

    @_enveloped
    def upsert_project_parts(self, items) -> dict:
        """新增/更新本案自訂材料，並自動登記到本案配件。"""
        requested = [x for x in (items or []) if isinstance(x, dict)]
        custom = {str(x.get("id")): x for x in self._read_custom_project_parts() if isinstance(x, dict) and x.get("id")}
        added: list[str] = []
        for row in requested:
            item = self._project_material_from_frontend(row)
            custom[item["id"]] = item
            added.append(item["id"])
        cur = self._read_registered()
        cur.update(added)
        self._write_project_parts_doc(cur, custom.values())
        self._pricebook_cache_key = None
        self._pricebook_cache_data = None
        return {"registered": sorted(cur), "added": added, "custom_count": len(custom)}

    # ---- 匯出（總庫 / 本案配件 → Excel，交回採購/收料）--------------------- #
    def _all_materials(self) -> list:
        return all_catalog_rows(self.root)

    def _export_materials(self, items, path, sheet_name, default_stem) -> dict:
        if not path:
            if self._save_file_fn is None:
                raise RuntimeError("存檔對話框未注入（此環境不支援匯出）")
            path = self._save_file_fn("excel", f"{default_stem}.xlsx")
            if not path:
                return {"cancelled": True}
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        cols = ["料號", "零件類型", "尺寸", "SCH", "材質", "料別", "單位", "來源", "規格", "備註"]
        keys = ["id", "零件類型", "尺寸", "SCH", "材質", "類別", "單位", "來源", "規格", "備註"]
        ws.append(cols)
        for it in items:
            ws.append([it.get(k, "") for k in keys])
        wb.save(path)
        return {"path": str(path), "count": len(items)}

    @_enveloped
    def export_master(self, path: str = "") -> dict:
        """匯出整個總庫成 Excel。"""
        return self._export_materials(self._all_materials(), path, "總庫", "總庫材料")

    @_enveloped
    def export_project_parts(self, path: str = "") -> dict:
        """匯出本案已登記配件成 Excel（交回採購/收料）。"""
        reg = self._read_registered()
        items = []
        for row in self._material_items_for_ids(reg):
            items.append({
                "id": row.get("id") or "",
                "零件類型": row.get("part") or "",
                "尺寸": row.get("size") or "",
                "SCH": row.get("sch") or "",
                "材質": row.get("mat") or "",
                "類別": row.get("cat") or "",
                "單位": row.get("unit") or "",
                "來源": row.get("src") or "",
                "規格": row.get("spec") or "",
                "備註": row.get("remark") or "",
            })
        return self._export_materials(items, path, "本案配件", "本案配件")


def _pm_clean(s):
    return " ".join(str(s or "").split())


_INCH_DN = {
    '1/8"': 'DN6', '1/4"': 'DN8', '3/8"': 'DN10', '1/2"': 'DN15', '3/4"': 'DN20',
    '1"': 'DN25', '1.1/4"': 'DN32', '1.1/2"': 'DN40', '2"': 'DN50', '2.1/2"': 'DN65',
    '3"': 'DN80', '3.1/2"': 'DN90', '4"': 'DN100', '5"': 'DN125', '6"': 'DN150',
    '8"': 'DN200', '10"': 'DN250', '12"': 'DN300', '14"': 'DN350', '16"': 'DN400',
    '18"': 'DN450', '20"': 'DN500', '22"': 'DN550', '24"': 'DN600', '26"': 'DN650',
    '28"': 'DN700', '30"': 'DN750', '32"': 'DN800', '36"': 'DN900', '40"': 'DN1000',
    '42"': 'DN1050', '48"': 'DN1200',
}


def _canon_size(s):
    """尺寸統一成 DN（吋 → DN；DN 保留；認不出原樣回）。"""
    return normalize_size(s)


def _canon_mat(m):
    """材質正規化：收空白、去尾點、A182-F304→A182 F304（合併空格/連字號的重複值）。"""
    return normalize_material(m)


def _pm_size(d):
    import re
    m = re.search(r'(\d[\d./]*)\s*"', str(d))
    if m:
        return m.group(1) + '"'
    for t in str(d).split(","):               # 退而求其次抓 DN 代號
        t = t.strip()
        if t.upper().startswith("DN"):
            return t
    return ""


def _pm_sch(d):
    import re
    s = str(d)
    for t in s.split(","):                 # 逗號式規格：整段保留（SCH40S、SCH40XSCH80、150#）不截斷
        t = t.strip()
        if t.upper().startswith("SCH") or t.endswith("#"):
            return t.upper().replace(" ", "")
    m = re.search(r'(SCH\s*\d+[A-Z]*|\d+#)', s.upper())   # 描述式：抓 SCHxx(含尾字母) 或 xxx#
    return m.group(1).replace(" ", "") if m else ""


def _pm_mat(s, fb=""):
    import re
    m = re.search(r'(A\d{2,3}\s*GR\.?\s*[A-Z0-9]+|A182[^ ]*|A105|A234[^ ]*|A106[^ ]*|SUS\s?\d{3}[A-Z]?|SS\d{3}|WPB)', str(s).upper())
    return m.group(1).strip() if m else fb


def _pm_type(d):
    """從英文規格描述抽出乾淨的中文零件類型；認不出回 ''（呼叫端保留原字串）。"""
    import re
    u = str(d).upper()
    ang = "90°" if re.search(r'X90|-90|\b90\b', u) else ("45°" if re.search(r'X45|-45|\b45\b', u) else "")
    if 'ELBOW' in u or re.search(r'\bEB-', u):
        return ang + "彎頭"
    if 'OLET' in u:
        return "OLET"
    if 'REDUCING TEE' in u:
        return "異徑三通"
    if 'TEE' in u or re.search(r'\bTE-', u):
        return "三通"
    if 'REDUCER' in u or re.search(r'\bRE-|\bCR-|\bER-', u):
        return "大小頭"
    if 'CAP' in u:
        return "管帽"
    if 'BLIND' in u:
        return "盲法蘭"
    if 'FLANGE' in u or re.search(r'\bFLG|\bSOF-|\bWNF-|\bSWF-', u):
        return "法蘭"
    if 'GASKET' in u or re.search(r'\bGR-', u):
        return "墊片"
    if 'COUPLING' in u or 'CPLG' in u:
        return "COUPLING"
    if 'NIPPLE' in u or re.search(r'\bNIP', u):
        return "短節"
    if 'UNION' in u:
        return "由令"
    if 'BUSHING' in u:
        return "補心"
    if 'PLUG' in u:
        return "管塞"
    if 'VALVE' in u or re.search(r'\bVA-', u):
        if 'GATE' in u:
            return "閘閥"
        if 'GLOBE' in u:
            return "球心閥"
        if 'BALL' in u:
            return "球閥"
        if 'CHECK' in u:
            return "止回閥"
        if 'BUTTERFLY' in u:
            return "蝶閥"
        if 'KNIFE' in u:
            return "刀閘閥"
        return "閥"
    if 'BOLT' in u or 'STUD' in u or re.search(r'\bNUT\b', u) or re.match(r'^M\d', u):
        return "螺栓"
    if 'PIPE' in u or 'SMLS' in u:
        return "鋼管"
    return ""


def _pm_cat(n):
    return category_for_part(n)


def _parse_material_xlsx(path: Any) -> list:
    """讀管制單 Excel → 正規化成總庫品項（無單價）。自動辨識 4 種格式：

    - 0408 品名式：項次/品名/規格型號/材質/單位（名稱已是乾淨中文 → 原樣保留）
    - 消防 規格式：項次/規格型號/材質/單位（無品名 → 名稱＝規格，抽中文類型）
    - 工業級：項次/品名及規格/尺寸/單位（名稱是英文描述 → 抽中文類型）
    - GASKET：Item/品名/Size/Description/Unit
    以表頭欄名定位欄位、以「有無單位」跳過分節列；規格全文保留供比對/匯出。
    """
    import openpyxl
    import os
    import re
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))

    stem = os.path.splitext(os.path.basename(str(path)))[0]      # 來源標記＝檔名括號內容
    paren = re.findall(r"[（(]([^）)]+)[）)]", stem)
    src = _pm_clean(paren[-1]) if paren else _pm_clean(stem)[:16]

    hi, header = None, []
    for idx, r in enumerate(rows[:6]):
        cells = [_pm_clean(c) for c in (r or [])]
        if any(("項次" in c) or (c.lower() == "item") for c in cells):
            hi, header = idx, cells
            break
    if hi is None:
        hi = 2
        header = [_pm_clean(c) for c in (rows[2] if len(rows) > 2 else [])]

    def find(*names):
        for i, c in enumerate(header):
            cl = c.replace(" ", "").lower()          # 表頭常含全形/半形空白（品  名、材   質）
            if any(n.replace(" ", "").lower() in cl for n in names):
                return i
        return None

    i_pname = find("品名")
    i_spec = find("規格型號")
    i_size = find("尺寸", "size")
    i_desc = find("description", "描述")
    i_mat = find("材質")
    i_unit = find("單位", "unit")

    name_i = i_pname if i_pname is not None else i_spec
    if name_i is None:
        name_i = 1
    spec_i = i_spec if i_spec is not None else i_size
    if spec_i is None:
        spec_i = name_i
    # 只有「同時有 品名 與 規格型號」的 0408 式才原樣保留名稱；其餘名稱是英文/規格 → 正規化
    derive = not (i_pname is not None and i_spec is not None)

    items, seen = [], set()
    for r in rows[hi + 1:]:
        if not r:
            continue
        row = [_pm_clean(c) for c in r] + [""] * 12
        raw = row[name_i]
        unit = row[i_unit] if i_unit is not None else ""
        if not raw or not unit:               # 分節列/空列（無單位）跳過
            continue
        spec = row[spec_i] if spec_i is not None else ""
        if i_desc is not None:
            spec = _pm_clean(spec + " " + row[i_desc])
        matcol = row[i_mat] if i_mat is not None else ""
        k = f"{raw}|{spec}|{matcol}".lower()
        if k in seen:
            continue
        seen.add(k)
        blob = raw + " " + spec
        typ = (_pm_type(blob) or raw) if derive else raw
        items.append({
            "零件類型": typ, "尺寸": _canon_size(_pm_size(blob)), "SCH": normalize_schedule(_pm_sch(blob)),
            "材質": _canon_mat(matcol or _pm_mat(blob)), "類別": _pm_cat(typ),
            "單位": unit, "規格": spec, "來源": src, "備註": "",
        })
    return items


def _issue_dict(source: str, it: Any) -> dict:
    sev = getattr(it, "severity", None) or (it.get("severity") if isinstance(it, dict) else None) or "info"
    sev = str(sev).lower()
    if sev not in ("error", "warning", "info"):
        sev = "info"
    g = lambda *names: next((getattr(it, n, None) or (it.get(n) if isinstance(it, dict) else None)
                             for n in names if (getattr(it, n, None) or (isinstance(it, dict) and it.get(n)))), "")
    return {
        "source": source,
        "level": sev,
        "title": g("title", "name", "code") or "",
        "message": g("message", "detail", "content") or "",
        "ref": g("path", "ref", "reference") or "",
    }


__all__ = ["MainBridge", "API_VERSION"]
