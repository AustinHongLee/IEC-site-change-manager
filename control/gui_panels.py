# -*- coding: utf-8 -*-
"""
gui_panels.py — 紀錄管理面板 + 請款追蹤面板 (PyQt6)

從 gui.py 拆分而來，降低單檔複雜度。
"""

import os
import re
import shutil
from datetime import datetime
from decimal import Decimal

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGroupBox, QLabel, QPushButton,
    QLineEdit, QTreeWidget, QTreeWidgetItem, QTabWidget, QComboBox,
    QMessageBox, QFileDialog, QHeaderView, QGridLayout, QAbstractItemView,
    QFrame, QSizePolicy, QMenu, QInputDialog, QCheckBox, QDialog,
    QDialogButtonBox, QTextEdit,
)
from PyQt6.QtCore import Qt, QSize
from PyQt6.QtGui import QBrush, QColor, QFont, QImage, QPixmap

# PDF 縮圖渲染
try:
    import fitz as _fitz
    _FITZ_OK = True
except ImportError:
    _fitz = None
    _FITZ_OK = False

from config import BASE_DIR, RECORD_XLSX_PATH, ATTACHMENTS_ROOT, PDF_OUTPUT_DIR
from billing_calculator import (
    BILLING_CURRENCY,
    BILLING_ROUNDING_RULE,
    BILLING_TAX_MODE,
    amount_to_text,
    build_billing_rows,
    calculate_tax_amount,
    money_to_text,
    parse_amount as parse_billing_amount,
    tax_rate_to_text,
)
from billing_audit import append_billing_audit, build_billing_change_events
from billing_batch import (
    BATCH_STATUS_OPTIONS,
    BillingBatchError,
    active_batch_index,
    batch_report_ids,
    create_billing_batch,
    is_active_batch_status,
    load_billing_batches,
    save_billing_batches,
    update_billing_batch_status,
)
from billing_status import (
    BILLING_STATUS_OPTIONS,
    normalize_billing_status,
    validate_billing_status_changes,
)
from material_pricebook import (
    PRICEBOOK_JSON_PATH,
    load_material_pricebook,
    normalize_pricebook_items,
    save_material_pricebook,
    unresolved_material_counts_by_report,
)
from material_pricebook_importer import (
    apply_import_plan,
    format_import_plan_summary,
    load_and_plan_seed_import,
)
from material_pricebook_table_importer import (
    apply_price_table_import_plan,
    format_price_table_import_summary,
    load_and_plan_price_table_import,
)
from material_pricebook_template_exporter import (
    build_price_table_template_items,
    export_price_table_template,
)
from material_constants import load_material_constants, material_default_unit
from material_repricing import (
    apply_project_reprice_plan,
    build_project_reprice_plan,
    format_reprice_summary,
)
from record_manager import (
    RECORDS_JSON_PATH, BILLING_JSON_PATH,
    _load_store, _save_store, auto_backup, export_records_to_excel,
)
from record_rebuild_queue import (
    build_rebuild_queue,
    export_rebuild_queue_csv,
    format_rebuild_queue_summary,
)
from site_output_center import run_site_output_center
from site_statistics_exporter import export_site_statistics_workbook
from workbook_pdf_converter import convert_workbook_to_pdf
from utils import atomic_save_wb, atomic_write_json, reentry_guard
from theme import Colors, Fonts, set_button_role, make_separator, make_stat_card, make_hint_label

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def _show_reentry_notice(owner):
    QMessageBox.information(owner, "提示", "此動作正在執行中，請稍候。")


# ──────────────── hover-to-zoom 圖片預覽 ────────────────
class _ZoomLabel(QLabel):
    """滑鼠懸停時顯示的放大預覽浮窗"""
    _ZOOM = 380

    def __init__(self):
        super().__init__(None, Qt.WindowType.ToolTip | Qt.WindowType.FramelessWindowHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setStyleSheet(
            "QLabel { background: white; border: 2px solid #2563eb; border-radius: 8px; padding: 4px; }"
        )
        self.setFixedSize(self._ZOOM + 8, self._ZOOM + 8)
        self.hide()

    def show_image(self, pixmap, gpos):
        if pixmap and not pixmap.isNull():
            scaled = pixmap.scaled(
                self._ZOOM, self._ZOOM,
                Qt.AspectRatioMode.KeepAspectRatio,
                Qt.TransformationMode.SmoothTransformation,
            )
            self.setPixmap(scaled)
            self.move(gpos.x() + 16, gpos.y() + 16)
            self.show()


_zoom_inst = None

def _get_zoom():
    global _zoom_inst
    if _zoom_inst is None:
        _zoom_inst = _ZoomLabel()
    return _zoom_inst


class _HoverThumb(QLabel):
    """可 hover 放大的縮圖 QLabel"""

    def __init__(self, full_pixmap, thumb_size: int, tag: str = "", parent=None):
        super().__init__(parent)
        self._full = full_pixmap
        self.setFixedSize(thumb_size, thumb_size)
        self.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.setStyleSheet(
            f"border: 1px solid {Colors.BORDER}; border-radius: 4px; background: white;"
        )
        self.setToolTip(tag)
        if full_pixmap and not full_pixmap.isNull():
            self.setPixmap(full_pixmap.scaled(
                thumb_size, thumb_size,
                Qt.AspectRatioMode.KeepAspectRatio,
                Qt.TransformationMode.SmoothTransformation,
            ))
        else:
            self.setText(tag or "—")
            self.setFont(QFont("Segoe UI", 10))
        self.setMouseTracking(True)

    def enterEvent(self, ev):
        super().enterEvent(ev)
        self.setStyleSheet(
            f"border: 2px solid {Colors.PRIMARY}; border-radius: 4px; background: white;"
        )
        if self._full and not self._full.isNull():
            _get_zoom().show_image(self._full, self.mapToGlobal(self.rect().topRight()))

    def leaveEvent(self, ev):
        super().leaveEvent(ev)
        self.setStyleSheet(
            f"border: 1px solid {Colors.BORDER}; border-radius: 4px; background: white;"
        )
        _get_zoom().hide()

    def mouseMoveEvent(self, ev):
        super().mouseMoveEvent(ev)
        z = _get_zoom()
        if z.isVisible() and self._full:
            z.move(ev.globalPosition().toPoint().x() + 16,
                   ev.globalPosition().toPoint().y() + 16)


# ──────────────── helper: inline cell edit ────────────────
class _CellEditor:
    """在 QTreeWidget 上方疊加編輯 widget，模擬 tkinter 行內編輯。"""

    def __init__(self):
        self._widget = None

    def destroy(self):
        if self._widget:
            self._widget.setParent(None)
            self._widget.deleteLater()
            self._widget = None

    def start_entry(self, tree: QTreeWidget, item: QTreeWidgetItem,
                    col: int, callback):
        """文字輸入編輯"""
        self.destroy()
        rect = tree.visualItemRect(item)
        header = tree.header()
        x = header.sectionPosition(col) - tree.horizontalScrollBar().value()
        w = header.sectionSize(col)

        edit = QLineEdit(tree.viewport())
        current = item.text(col) or ""
        clean = current.replace("$", "").replace(",", "")
        edit.setText(clean)
        edit.selectAll()
        edit.setGeometry(x, rect.y(), w, rect.height())
        edit.show()
        edit.setFocus()

        def finish():
            val = edit.text().strip()
            self.destroy()
            callback(item, col, val)

        edit.returnPressed.connect(finish)
        edit.editingFinished.connect(finish)
        self._widget = edit

    def start_combo(self, tree: QTreeWidget, item: QTreeWidgetItem,
                    col: int, options: list, callback):
        """下拉選單編輯"""
        self.destroy()
        rect = tree.visualItemRect(item)
        header = tree.header()
        x = header.sectionPosition(col) - tree.horizontalScrollBar().value()
        w = header.sectionSize(col)

        combo = QComboBox(tree.viewport())
        combo.addItems(options)
        cur = item.text(col)
        if cur in options:
            combo.setCurrentText(cur)
        combo.setGeometry(x, rect.y(), w, rect.height())
        combo.show()
        combo.setFocus()
        combo.showPopup()

        def on_select(_idx):
            val = combo.currentText()
            self.destroy()
            callback(item, col, val)

        combo.activated.connect(on_select)
        self._widget = combo


# ========= 紀錄管理面板 =========
class RecordManagerPanel(QWidget):
    """紀錄管理面板 - 顯示與編輯紀錄清單 (PyQt6)"""

    def __init__(self, parent: QWidget = None):
        super().__init__(parent)
        self.records = []
        self.details = []
        self.materials = []
        self.details_modified = False
        self.materials_modified = False
        self._editor = _CellEditor()
        self._build_ui()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(8, 8, 8, 8)
        root.setSpacing(10)

        # ── 篩選工具列 ──
        filter_group = QGroupBox("🔍 篩選")
        fr = QHBoxLayout(filter_group)
        fr.setSpacing(8)
        fr.addWidget(QLabel("日期："))
        self.date_from_edit = QLineEdit()
        self.date_from_edit.setPlaceholderText("YYYYMMDD")
        self.date_from_edit.setFixedWidth(95)
        fr.addWidget(self.date_from_edit)
        sep_lbl = QLabel("–")
        sep_lbl.setStyleSheet(f"color: {Colors.TEXT_MUTED}; border:none; background:transparent;")
        fr.addWidget(sep_lbl)
        self.date_to_edit = QLineEdit()
        self.date_to_edit.setPlaceholderText("YYYYMMDD")
        self.date_to_edit.setFixedWidth(95)
        fr.addWidget(self.date_to_edit)
        fr.addSpacing(12)
        fr.addWidget(QLabel("🔎"))
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("搜尋編號 / 說明...")
        self.search_edit.setFixedWidth(180)
        fr.addWidget(self.search_edit)
        # 狀態篩選
        fr.addSpacing(8)
        fr.addWidget(QLabel("狀態："))
        self.status_combo = QComboBox()
        self.status_combo.addItems(["全部", "已產出", "需重產", "未產出", "已歸檔"])
        self.status_combo.setFixedWidth(90)
        self.status_combo.currentIndexChanged.connect(lambda: self.load_records())
        fr.addWidget(self.status_combo)
        btn_reload = QPushButton("🔄 載入")
        btn_reload.clicked.connect(self.load_records)
        fr.addWidget(btn_reload)
        btn_excel = QPushButton("� 匯出 Excel")
        btn_excel.setProperty("role", "flat")
        btn_excel.setToolTip("從 JSON 資料匯出為 Excel 紀錄清單")
        btn_excel.clicked.connect(self._open_excel)
        fr.addWidget(btn_excel)
        btn_site_stats = QPushButton("現場統計單")
        btn_site_stats.setToolTip("匯出總覽、修改單清單、焊口統計、照片索引、用料統計與問題清單")
        btn_site_stats.clicked.connect(self._export_site_statistics)
        fr.addWidget(btn_site_stats)
        btn_site_stats_pdf = QPushButton("統計PDF")
        btn_site_stats_pdf.setToolTip("匯出現場統計單 xlsx 後，用 LibreOffice 轉出 PDF")
        btn_site_stats_pdf.clicked.connect(lambda checked=False: self._export_site_statistics(with_pdf=True))
        fr.addWidget(btn_site_stats_pdf)
        btn_output_center = QPushButton("輸出中心")
        btn_output_center.setToolTip("用目前 attachments 產出現場統計單、summary PDF 或 before/after 照片 PDF")
        btn_output_center.clicked.connect(self._export_site_output_center)
        fr.addWidget(btn_output_center)
        btn_rebuild_filter = QPushButton("需重產")
        btn_rebuild_filter.setToolTip("快速顯示補價或資料異動後需要重新產出的修改單")
        btn_rebuild_filter.clicked.connect(self._show_rebuild_queue)
        fr.addWidget(btn_rebuild_filter)
        btn_rebuild_export = QPushButton("匯出重產清單")
        btn_rebuild_export.setToolTip("匯出目前需重產修改單清單，方便安排重產與複查")
        btn_rebuild_export.clicked.connect(self._export_rebuild_queue)
        fr.addWidget(btn_rebuild_export)
        fr.addStretch()
        root.addWidget(filter_group)

        # ── 主紀錄 + 明細 ──
        main_h = QHBoxLayout()
        main_h.setSpacing(10)

        # 左側：主紀錄
        record_group = QGroupBox("📋 報告紀錄")
        rg_layout = QVBoxLayout(record_group)
        self.record_tree = QTreeWidget()
        self.record_tree.setAlternatingRowColors(True)
        self.record_tree.setSortingEnabled(True)
        self.record_tree.setHeaderLabels(["報告編號", "日期", "Series", "焊口清單", "變更類型", "說明", "狀態", "材料提醒"])
        for i, w in enumerate([100, 80, 60, 150, 80, 180, 70, 95]):
            self.record_tree.setColumnWidth(i, w)
        self.record_tree.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.record_tree.itemSelectionChanged.connect(self._on_record_select)
        self.record_tree.itemDoubleClicked.connect(lambda: self._open_folder())
        self.record_tree.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.record_tree.customContextMenuRequested.connect(self._show_context_menu)
        rg_layout.addWidget(self.record_tree)
        main_h.addWidget(record_group, 1)

        # 右側：明細 notebook
        detail_nb = QTabWidget()

        # 焊口明細
        detail_w = QWidget()
        dv = QVBoxLayout(detail_w)
        dv.setContentsMargins(6, 6, 6, 6)
        self.detail_tree = QTreeWidget()
        self.detail_tree.setAlternatingRowColors(True)
        self.detail_tree.setHeaderLabels(
            ["焊口編號", "標記", "尺寸", "材質", "厚度", "係數", "單價", "金額"]
        )
        for i, w in enumerate([70, 45, 55, 70, 60, 55, 70, 70]):
            self.detail_tree.setColumnWidth(i, w)
        self.detail_tree.itemDoubleClicked.connect(self._on_detail_double_click)
        dv.addWidget(self.detail_tree)
        detail_nb.addTab(detail_w, "🔧 焊口明細")

        # 材料明細
        mat_w = QWidget()
        mv = QVBoxLayout(mat_w)
        mv.setContentsMargins(6, 6, 6, 6)
        self.material_tree = QTreeWidget()
        self.material_tree.setAlternatingRowColors(True)
        self.material_tree.setHeaderLabels(["零件類型", "尺寸", "SCH", "材質", "數量", "單價", "金額"])
        for i, w in enumerate([100, 60, 55, 80, 60, 80, 80]):
            self.material_tree.setColumnWidth(i, w)
        self.material_tree.itemDoubleClicked.connect(self._on_material_double_click)
        mv.addWidget(self.material_tree)
        detail_nb.addTab(mat_w, "📦 材料明細")

        # 圖片明細
        img_w = QWidget()
        iv = QVBoxLayout(img_w)
        iv.setContentsMargins(6, 6, 6, 6)
        iv.setSpacing(8)
        self._img_hint = QLabel("選取左側紀錄以顯示圖片")
        self._img_hint.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._img_hint.setStyleSheet(
            f"color: {Colors.TEXT_MUTED}; border:none; background:transparent; padding: 30px;"
        )
        iv.addWidget(self._img_hint)
        # 圖片容器
        self._img_container = QWidget()
        self._img_layout = QHBoxLayout(self._img_container)
        self._img_layout.setContentsMargins(0, 0, 0, 0)
        self._img_layout.setSpacing(12)
        self._img_layout.addStretch()
        self._img_container.hide()
        iv.addWidget(self._img_container, stretch=1)
        detail_nb.addTab(img_w, "🖼️ 圖片明細")

        main_h.addWidget(detail_nb, 1)
        root.addLayout(main_h, 1)

        # ── 底部按鈕 ──
        bot = QHBoxLayout()
        bot.setSpacing(8)
        self.stat_label = QLabel("共 0 筆紀錄")
        self.stat_label.setFont(Fonts.small())
        self.stat_label.setStyleSheet(f"color: {Colors.TEXT_MUTED}; border:none; background:transparent;")
        bot.addWidget(self.stat_label)
        self.modified_label = QLabel("")
        self.modified_label.setStyleSheet(f"color: {Colors.DANGER}; font-weight:bold; border:none; background:transparent;")
        bot.addWidget(self.modified_label)
        bot.addStretch()
        for txt, slot, role in [
            ("💾 儲存", self._save_changes, "primary"),
            ("� 匯出", self._open_excel, "success"),
            ("�📄 PDF", self._open_pdf, ""),
            ("📂 資料夾", self._open_folder, ""),
            (" 補登", self._open_backfill_tool, ""),
            ("🔍 稽查", self._open_orphan_audit, ""),
        ]:
            b = QPushButton(txt)
            if role:
                b.setProperty("role", role)
            b.clicked.connect(slot)
            bot.addWidget(b)
        root.addLayout(bot)

    # ────────────────── 資料載入 ──────────────────
    def load_records(self):
        self.record_tree.setSortingEnabled(False)   # 批次插入時關閉排序
        self.record_tree.clear()
        self.records.clear()

        try:
            store = _load_store()
            search_text = self.search_edit.text().lower()
            status_filter = self.status_combo.currentText()
            unresolved_counts = unresolved_material_counts_by_report(store.get("materials", []))

            # ── 1) 從 JSON 載入已產出紀錄 ──
            json_folders = set()          # (日期, 資料夾名) — 追蹤已有紀錄的
            for rec in store["records"]:
                date_val = rec.get("日期", "")
                folder_name = rec.get("資料夾名", "")
                json_folders.add((date_val, folder_name))
                status = "需重產" if str(rec.get("需重產", "")).strip() == "1" else "已產出"
                record = {
                    "report_id": rec.get("報告編號", ""),
                    "date": date_val,
                    "series": rec.get("Series NO", ""),
                    "welds": rec.get("焊口清單", ""),
                    "change_type": rec.get("變更類型", ""),
                    "desc": rec.get("說明", ""),
                    "folder": folder_name,
                    "status": status,
                    "needs_rebuild": str(rec.get("需重產", "")).strip() == "1",
                    "rebuild_reason": rec.get("需重產原因", ""),
                    "rebuild_at": rec.get("需重產時間", ""),
                    "unresolved_materials": unresolved_counts.get(str(rec.get("報告編號", "")).strip(), {}),
                    "folder_path": os.path.join(ATTACHMENTS_ROOT, date_val, folder_name),
                    "is_archived": False,
                }
                self.records.append(record)

            # ── 2) 掃描 attachments/ 找未產出的 ──
            if os.path.isdir(ATTACHMENTS_ROOT):
                for date_dir in sorted(os.listdir(ATTACHMENTS_ROOT)):
                    date_path = os.path.join(ATTACHMENTS_ROOT, date_dir)
                    if not os.path.isdir(date_path) or date_dir.startswith("_"):
                        continue
                    if not re.match(r"^\d{8}$", date_dir):
                        continue
                    for fn in sorted(os.listdir(date_path)):
                        fp = os.path.join(date_path, fn)
                        if not os.path.isdir(fp) or fn.startswith("_"):
                            continue
                        if (date_dir, fn) in json_folders:
                            continue  # 已在 JSON 裡
                        parts = fn.split("_")
                        serial = parts[0] if parts and parts[0].isdigit() else ""
                        welds_str = "_".join(parts[1:]) if len(parts) > 1 else ""
                        self.records.append({
                            "report_id": "",
                            "date": date_dir,
                            "series": serial,
                            "welds": welds_str,
                            "change_type": "",
                            "desc": "",
                            "folder": fn,
                            "status": "未產出",
                            "needs_rebuild": False,
                            "rebuild_reason": "",
                            "rebuild_at": "",
                            "unresolved_materials": {},
                            "folder_path": fp,
                            "is_archived": False,
                        })

            # ── 3) 掃描 _archived/ 找已歸檔的 ──
            archived_root = os.path.join(ATTACHMENTS_ROOT, "_archived")
            if os.path.isdir(archived_root):
                for date_dir in sorted(os.listdir(archived_root)):
                    date_path = os.path.join(archived_root, date_dir)
                    if not os.path.isdir(date_path):
                        continue
                    for fn in sorted(os.listdir(date_path)):
                        fp = os.path.join(date_path, fn)
                        if not os.path.isdir(fp):
                            continue
                        parts = fn.split("_")
                        serial = parts[0] if parts and parts[0].isdigit() else ""
                        welds_str = "_".join(parts[1:]) if len(parts) > 1 else ""
                        self.records.append({
                            "report_id": "",
                            "date": date_dir,
                            "series": serial,
                            "welds": welds_str,
                            "change_type": "",
                            "desc": "",
                            "folder": fn,
                            "status": "已歸檔",
                            "needs_rebuild": False,
                            "rebuild_reason": "",
                            "rebuild_at": "",
                            "unresolved_materials": {},
                            "folder_path": fp,
                            "is_archived": True,
                        })

            # ── 4) 篩選 + 新增到 tree ──
            visible_count = 0
            for rec_idx, record in enumerate(self.records):
                # 狀態篩選
                if status_filter != "全部" and record["status"] != status_filter:
                    continue
                # 文字搜尋
                if search_text:
                    searchable = (
                        f"{record['report_id']} {record['series']} "
                        f"{record['welds']} {record['desc']} {record['folder']}"
                    ).lower()
                    if search_text not in searchable:
                        continue

                welds_s = str(record["welds"])
                desc_s = str(record["desc"])
                status_s = record["status"]
                material_s = self._format_material_warning(record.get("unresolved_materials", {}))
                item = QTreeWidgetItem(self.record_tree, [
                    str(record["report_id"]),
                    str(record["date"]),
                    str(record["series"]),
                    welds_s[:30] + "..." if len(welds_s) > 30 else welds_s,
                    str(record["change_type"]),
                    desc_s[:30] + "..." if len(desc_s) > 30 else desc_s,
                    status_s,
                    material_s,
                ])
                item.setData(0, Qt.ItemDataRole.UserRole, rec_idx)
                # 狀態顏色
                if status_s == "未產出":
                    item.setForeground(6, QBrush(QColor("#d97706")))  # 橘
                elif status_s == "需重產":
                    item.setForeground(6, QBrush(QColor(Colors.WARNING)))
                    item.setToolTip(6, self._format_rebuild_tooltip(record))
                elif status_s == "已歸檔":
                    item.setForeground(6, QBrush(QColor("#9ca3af")))  # 灰
                    for col in range(6):
                        item.setForeground(col, QBrush(QColor("#9ca3af")))
                else:
                    item.setForeground(6, QBrush(QColor("#16a34a")))  # 綠
                if material_s:
                    item.setForeground(7, QBrush(QColor(Colors.DANGER)))
                    item.setToolTip(7, self._format_material_warning_tooltip(record.get("unresolved_materials", {})))
                visible_count += 1

            # 明細
            self.details.clear()
            for det in store["details"]:
                self.details.append({
                    "report_id": det.get("紀錄編號", ""),
                    "weld_code": det.get("焊口編號", ""),
                    "size": det.get("焊口尺寸", ""),
                    "coefficient": det.get("係數", ""),
                    "price": det.get("單價/DB", ""),
                    "amount": det.get("金額", ""),
                })

            # 材料明細（若有）
            self.materials.clear()
            for mat in store.get("materials", []):
                self.materials.append({
                    "report_id": mat.get("報告編號", ""),
                    "component": mat.get("零件類型", ""),
                    "size": mat.get("尺寸", ""),
                    "sch": mat.get("SCH", ""),
                    "material": mat.get("材質", ""),
                    "category": mat.get("類別", "材料"),
                    "qty": mat.get("數量", ""),
                    "unit": mat.get("單位", ""),
                    "price": mat.get("單價", ""),
                    "amount": mat.get("金額", ""),
                    "price_source": mat.get("單價來源", ""),
                    "amount_source": mat.get("金額來源", ""),
                    "pricebook_id": mat.get("價目表ID", ""),
                    "pricebook_source": mat.get("價目來源", ""),
                    "pricebook_effective_date": mat.get("價目生效日", ""),
                    "pricing_status": mat.get("配價狀態", ""),
                })

            produced = sum(1 for r in self.records if r["status"] == "已產出")
            needs_rebuild = sum(1 for r in self.records if r["status"] == "需重產")
            unproduced = sum(1 for r in self.records if r["status"] == "未產出")
            archived = sum(1 for r in self.records if r["status"] == "已歸檔")
            unresolved_total = sum(int((r.get("unresolved_materials") or {}).get("total", 0)) for r in self.records)
            self.stat_label.setText(
                f"總計 {len(self.records)} 筆 │ "
                f"✅ 已產出 {produced}  ⚠️ 需重產 {needs_rebuild}  ⏳ 未產出 {unproduced}  📦 已歸檔 {archived}"
                + (f"  │ 材料未定價 {unresolved_total}" if unresolved_total else "")
                + (f"  │ 篩選 {visible_count}" if status_filter != "全部" or search_text else "")
            )

        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"載入紀錄失敗：{e}")
        finally:
            self.record_tree.setSortingEnabled(True)  # 重新啟用欄位擊點排序
            self.record_tree.sortByColumn(1, Qt.SortOrder.AscendingOrder)  # 預設日期升冪

    # ────────────────── 選擇 / 雙擊 ──────────────────
    @staticmethod
    def _format_material_warning(counts: dict) -> str:
        total = int((counts or {}).get("total", 0))
        if total <= 0:
            return ""
        missing_price = int((counts or {}).get("missing_price", 0))
        missing_pricebook = int((counts or {}).get("missing_pricebook", 0))
        parts = []
        if missing_price:
            parts.append(f"待補價 {missing_price}")
        if missing_pricebook:
            parts.append(f"待建料 {missing_pricebook}")
        return "、".join(parts) or f"未定價 {total}"

    @staticmethod
    def _format_material_warning_tooltip(counts: dict) -> str:
        total = int((counts or {}).get("total", 0))
        missing_price = int((counts or {}).get("missing_price", 0))
        missing_pricebook = int((counts or {}).get("missing_pricebook", 0))
        lines = [f"此修改單有 {total} 筆材料尚未可請款。"]
        if missing_price:
            lines.append(f"未定價（待補價）：{missing_price} 筆")
        if missing_pricebook:
            lines.append(f"查無價目（待建料）：{missing_pricebook} 筆")
        lines.append("請到材料價目表補價或建料後，再套用補價。")
        return "\n".join(lines)

    @staticmethod
    def _format_rebuild_tooltip(record: dict) -> str:
        reason = str(record.get("rebuild_reason", "")).strip() or "資料已變更"
        at = str(record.get("rebuild_at", "")).strip()
        text = f"此修改單已產出過，但 {reason}，建議重新產出 PDF/Excel。"
        if at:
            text += f"\n標記時間：{at}"
        return text

    @staticmethod
    def _format_rebuild_queue_export_message(rows: list[dict[str, str]], path: str) -> str:
        return f"已匯出 {len(rows)} 張需重產修改單：\n{path}"

    def _on_record_select(self):
        items = self.record_tree.selectedItems()
        if not items:
            return
        report_id = items[0].text(0)
        record = self._find_record_by_item(items[0])

        # ── 焊口明細 ──
        self.detail_tree.clear()

        # 從 weld_info.json 載入材質/厚度（如果有）
        weld_extra = {}  # key: weld_code → {mark, material, thickness}
        if record:
            import json as _json
            wip = os.path.join(record.get("folder_path", ""), "weld_info.json")
            if os.path.isfile(wip):
                try:
                    with open(wip, "r", encoding="utf-8") as f:
                        wi = _json.load(f)
                    for w in wi.get("welds", []):
                        wn = str(w.get("weld_no", ""))
                        mk = str(w.get("mark", ""))
                        m_sp = re.match(r'^(\d+)([rab])', wn)
                        if m_sp:
                            wn = m_sp.group(1)
                            mk = mk or m_sp.group(2)
                        # key = weld_no + mark 以匹配 detail 的焊口編號
                        code = f"{wn}{mk}"
                        weld_extra[code] = {
                            "mark": mk,
                            "material": w.get("material", ""),
                            "thickness": w.get("thickness", ""),
                        }
                except Exception:
                    pass

        for d in self.details:
            if str(d["report_id"]) == report_id:
                wc = str(d["weld_code"])
                extra = weld_extra.get(wc, {})
                # 合併到 detail 供存檔用
                d["mark"] = d.get("mark") or extra.get("mark", "")
                d["material"] = d.get("material") or extra.get("material", "")
                d["thickness"] = d.get("thickness") or extra.get("thickness", "")
                QTreeWidgetItem(self.detail_tree, [
                    wc,
                    str(d["mark"]),
                    str(d["size"]),
                    str(d["material"]),
                    str(d["thickness"]),
                    str(d["coefficient"]),
                    str(d["price"]),
                    str(d["amount"]),
                ])

        # ── 材料明細 ──
        self.material_tree.clear()
        for m in self.materials:
            if str(m["report_id"]) == report_id:
                qty_str = f"{m['qty']} {m.get('unit', '')}".strip()
                price_s = str(m["price"])
                amount_s = str(m["amount"])
                if m.get("pricing_status") == "missing_pricebook":
                    price_s = "未配價"
                    amount_s = "未配價"
                elif m.get("pricing_status") == "missing_price":
                    price_s = "未定價"
                    amount_s = "未定價"
                item = QTreeWidgetItem(self.material_tree, [
                    str(m["component"]), str(m["size"]),
                    str(m.get("sch", "")), str(m["material"]),
                    qty_str, price_s, amount_s,
                ])
                if m.get("price_source") == "pricebook":
                    item.setForeground(5, QBrush(QColor(Colors.INFO)))
                    item.setToolTip(5, f"由價目表帶入：{m.get('pricebook_id', '')}")
                elif m.get("pricing_status") == "missing_pricebook":
                    for col in (5, 6):
                        item.setForeground(col, QBrush(QColor(Colors.DANGER)))
                        item.setToolTip(col, "材料價目表找不到對應單價，請先到材料價目補價或手動輸入單價")
                elif m.get("pricing_status") == "missing_price":
                    for col in (5, 6):
                        item.setForeground(col, QBrush(QColor(Colors.WARNING)))
                        item.setToolTip(col, "材料價目表已有此材料，但尚未填單價，請到材料價目補價或手動輸入單價")

        # ── 圖片明細 ──
        self._update_image_detail(record)

    # ── 焊口明細欄位對照 ──
    _DETAIL_COL_MAP = {
        0: "weld_code",    # 焊口編號
        1: "mark",         # 標記 (r/a/b)
        2: "size",         # 尺寸
        3: "material",     # 材質
        4: "thickness",    # 厚度
        5: "coefficient",  # 係數
        6: "price",        # 單價
        7: "amount",       # 金額
    }

    def _on_detail_double_click(self, item: QTreeWidgetItem, col: int):
        if col not in self._DETAIL_COL_MAP:
            return

        def cb(it, c, val):
            it.setText(c, val)
            field = self._DETAIL_COL_MAP[c]
            self._update_detail_data(it, field, val)
            self.details_modified = True
            self._update_modified_label()

        # 標記欄用下拉選單
        if col == 1:
            self._editor.start_combo(
                self.detail_tree, item, col, ["r", "a", "b"], cb
            )
        else:
            self._editor.start_entry(self.detail_tree, item, col, cb)

    def _on_material_double_click(self, item: QTreeWidgetItem, col: int):
        editable = {5: "price", 6: "amount"}
        if col not in editable:
            return

        def cb(it, c, val):
            if str(val).strip() in ("未配價", "未定價"):
                val = ""
            it.setText(c, val)
            self._update_material_data(it, editable[c], val)
            self.materials_modified = True
            self._update_modified_label()

        self._editor.start_entry(self.material_tree, item, col, cb)

    def _update_detail_data(self, item: QTreeWidgetItem, field: str, val: str):
        """更新 self.details 中對應項目的欄位值（用 tree 行索引精確匹配）"""
        sel = self.record_tree.selectedItems()
        if not sel:
            return
        report_id = sel[0].text(0)
        # 用行索引而非焊口編號匹配（因為焊口編號本身可被修改）
        row_idx = self.detail_tree.indexOfTopLevelItem(item)
        matching = [d for d in self.details if str(d["report_id"]) == report_id]
        if 0 <= row_idx < len(matching):
            matching[row_idx][field] = val

    def _update_material_data(self, item: QTreeWidgetItem, field: str, val: str):
        val = self._normalize_material_edit_value(val)
        comp = item.text(0)
        size = item.text(1)
        sel = self.record_tree.selectedItems()
        if not sel:
            return
        report_id = sel[0].text(0)
        for m in self.materials:
            if str(m["report_id"]) == report_id and str(m["component"]) == comp and str(m["size"]) == size:
                m[field] = val
                if field == "price":
                    m["price_source"] = "manual" if str(val).strip() else ""
                    if str(val).strip():
                        m["pricebook_id"] = ""
                        m["pricebook_source"] = ""
                        m["pricebook_effective_date"] = ""
                        m["pricing_status"] = "manual"
                    else:
                        m["pricebook_id"] = ""
                        m["pricebook_source"] = ""
                        m["pricebook_effective_date"] = ""
                        m["pricing_status"] = ""
                elif field == "amount":
                    m["amount_source"] = "manual" if str(val).strip() else ""
                break

    @staticmethod
    def _normalize_material_edit_value(val: str) -> str:
        """畫面提示文字不可寫回資料層。"""
        text = str(val).strip()
        if text in ("未配價", "未定價"):
            return ""
        return text

    def _update_modified_label(self):
        parts = []
        if self.details_modified:
            parts.append("焊口明細")
        if self.materials_modified:
            parts.append("材料明細")
        self.modified_label.setText(f"⚠️ 已修改: {', '.join(parts)}" if parts else "")

    # ────────────────── 圖片明細 ──────────────────

    def _update_image_detail(self, record: dict | None):
        """根據選取的紀錄更新圖片明細 tab"""
        # 清除舊卡片
        while self._img_layout.count() > 1:
            item = self._img_layout.takeAt(0)
            w = item.widget()
            if w:
                w.deleteLater()

        if not record:
            self._img_hint.setText("此紀錄沒有資料夾資訊")
            self._img_hint.show()
            self._img_container.hide()
            return

        folder_path = record.get("folder_path", "")
        if not folder_path or not os.path.isdir(folder_path):
            self._img_hint.setText(f"資料夾不存在:\n{folder_path}")
            self._img_hint.show()
            self._img_container.hide()
            return

        self._img_hint.hide()
        self._img_container.show()

        # before / after（支援 single + group 模式）
        for img_path, label in self._find_images(folder_path):
            card = self._make_image_card(img_path, label, folder_path=folder_path)
            self._img_layout.insertWidget(self._img_layout.count() - 1, card)

        # PDF
        pdf_files = [f for f in os.listdir(folder_path) if f.lower().endswith(".pdf")]
        for pdf_name in pdf_files[:2]:
            pdf_path = os.path.join(folder_path, pdf_name)
            card = self._make_image_card(pdf_path, f"📄 {pdf_name}", is_pdf=True)
            self._img_layout.insertWidget(self._img_layout.count() - 1, card)

        # ➕ 新增圖片 按鈕卡
        add_card = self._make_add_image_card(folder_path)
        self._img_layout.insertWidget(self._img_layout.count() - 1, add_card)

    def _make_image_card(self, path: str, label: str, is_pdf: bool = False,
                         folder_path: str = "") -> QFrame:
        """建立一張圖片預覽卡片（含縮圖 + 標籤，hover 放大）"""
        card = QFrame()
        card.setStyleSheet(
            f"QFrame {{ background: {Colors.BG_WHITE}; border: 1px solid {Colors.BORDER_LIGHT};"
            f" border-radius: 6px; }}"
        )
        card.setFixedWidth(160)
        clay = QVBoxLayout(card)
        clay.setContentsMargins(6, 6, 6, 6)
        clay.setSpacing(4)
        clay.setAlignment(Qt.AlignmentFlag.AlignTop)

        THUMB_SIZE = 140
        full_pm = None

        if is_pdf:
            full_pm = self._render_pdf_thumb(path, 400)
        elif os.path.exists(path):
            full_pm = QPixmap(path)

        thumb_lbl = _HoverThumb(full_pm, THUMB_SIZE, label)
        clay.addWidget(thumb_lbl, alignment=Qt.AlignmentFlag.AlignCenter)

        # 標籤
        txt = QLabel(label)
        txt.setFont(QFont("Segoe UI", 8, QFont.Weight.Bold))
        txt.setAlignment(Qt.AlignmentFlag.AlignCenter)
        txt.setStyleSheet(f"color: {Colors.TEXT}; border:none; background:transparent;")
        clay.addWidget(txt)

        # 存在/不存在狀態
        exists = (full_pm is not None and not full_pm.isNull())
        if not exists and not is_pdf:
            st = QLabel("(檔案不存在)")
            st.setFont(QFont("Segoe UI", 7))
            st.setAlignment(Qt.AlignmentFlag.AlignCenter)
            st.setStyleSheet(f"color: {Colors.TEXT_MUTED}; border:none; background:transparent;")
            clay.addWidget(st)

        # PDF 可點擊開啟
        if is_pdf and os.path.exists(path):
            btn = QPushButton("📂 開啟 PDF")
            btn.setFixedHeight(24)
            btn.setFont(QFont("Segoe UI", 7))
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.clicked.connect(lambda _, p=path: os.startfile(p))
            clay.addWidget(btn)

        # 🔄 替換圖片 按鈕（非 PDF 才顯示）
        if not is_pdf:
            btn_replace = QPushButton("🔄 替換圖片")
            btn_replace.setFixedHeight(24)
            btn_replace.setFont(QFont("Segoe UI", 7))
            btn_replace.setCursor(Qt.CursorShape.PointingHandCursor)
            btn_replace.clicked.connect(lambda _, p=path: self._replace_image(p))
            clay.addWidget(btn_replace)

        # ✏️ 標註（圖片和 PDF 都可使用）
        if os.path.exists(path):
            btn_annotate = QPushButton("✏️ 標註")
            btn_annotate.setFixedHeight(24)
            btn_annotate.setFont(QFont("Segoe UI", 7))
            btn_annotate.setCursor(Qt.CursorShape.PointingHandCursor)
            btn_annotate.clicked.connect(
                lambda _, p=path, pdf=is_pdf: self._open_annotator(p, pdf)
            )
            clay.addWidget(btn_annotate)

        return card

    def _make_add_image_card(self, folder_path: str) -> QFrame:
        """建立 ➕ 新增圖片 卡片"""
        card = QFrame()
        card.setStyleSheet(
            f"QFrame {{ background: {Colors.BG_WHITE}; border: 2px dashed {Colors.BORDER_LIGHT};"
            f" border-radius: 6px; }}"
        )
        card.setFixedWidth(160)
        card.setFixedHeight(160)
        clay = QVBoxLayout(card)
        clay.setAlignment(Qt.AlignmentFlag.AlignCenter)

        icon_lbl = QLabel("➕")
        icon_lbl.setFont(QFont("Segoe UI", 28))
        icon_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        icon_lbl.setStyleSheet("border:none; background:transparent;")
        clay.addWidget(icon_lbl)

        txt = QLabel("新增圖片")
        txt.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))
        txt.setAlignment(Qt.AlignmentFlag.AlignCenter)
        txt.setStyleSheet(f"color: {Colors.TEXT_MUTED}; border:none; background:transparent;")
        clay.addWidget(txt)

        card.setCursor(Qt.CursorShape.PointingHandCursor)
        card.mousePressEvent = lambda e, fp=folder_path: self._add_image(fp)
        return card

    def _replace_image(self, target_path: str):
        """以檔案對話框選取圖片替換指定路徑"""
        from PyQt6.QtWidgets import QFileDialog
        src, _ = QFileDialog.getOpenFileName(
            self, "選擇替換圖片", "",
            "圖片 (*.jpg *.jpeg *.png *.bmp);;所有檔案 (*)"
        )
        if not src:
            return
        import shutil
        shutil.copy2(src, target_path)
        # 重新整理
        items = self.record_tree.selectedItems()
        if items:
            record = self._find_record_by_item(items[0])
            self._update_image_detail(record)

    def _add_image(self, folder_path: str, preferred_prefix: str = "") -> bool:
        """新增一張圖片到資料夾（自動命名 before_N / after_N）"""
        return self._add_image_to_folder(folder_path, preferred_prefix=preferred_prefix)

    def _add_image_to_folder(self, folder_path: str, *, preferred_prefix: str = "", parent=None) -> bool:
        from PyQt6.QtWidgets import QFileDialog, QInputDialog
        src, _ = QFileDialog.getOpenFileName(
            parent or self, "選擇圖片", "",
            "圖片 (*.jpg *.jpeg *.png *.bmp);;所有檔案 (*)"
        )
        if not src:
            return False
        prefix = self._normalize_output_center_photo_prefix(preferred_prefix)
        if not prefix:
            choices = ["修改前 (before)", "修改後 (after)"]
            choice, ok = QInputDialog.getItem(
                parent or self, "圖片類型", "此圖片屬於：", choices, 0, False
            )
            if not ok:
                return False
            prefix = "before" if "before" in choice else "after"
        if not os.path.isdir(folder_path):
            QMessageBox.warning(parent or self, "資料夾不存在", f"找不到資料夾：\n{folder_path}")
            return False
        dest = self._next_output_center_photo_path(folder_path, prefix)
        if not dest:
            QMessageBox.warning(parent or self, "圖片類型錯誤", f"不支援的圖片類型：{preferred_prefix}")
            return False
        shutil.copy2(src, dest)
        # 重新整理
        items = self.record_tree.selectedItems()
        if items:
            record = self._find_record_by_item(items[0])
            self._update_image_detail(record)
        return True

    @staticmethod
    def _normalize_output_center_photo_prefix(value: str) -> str:
        text = str(value or "").strip().lower()
        if text in ("before", "修改前"):
            return "before"
        if text in ("after", "修改後"):
            return "after"
        return ""

    @staticmethod
    def _next_output_center_photo_path(folder_path: str, prefix: str) -> str:
        prefix = RecordManagerPanel._normalize_output_center_photo_prefix(prefix)
        if not prefix:
            return ""
        existing = set(os.listdir(folder_path)) if os.path.isdir(folder_path) else set()
        base_name = f"{prefix}.jpg"
        if base_name not in existing:
            return os.path.join(folder_path, base_name)
        idx = 1
        while f"{prefix}_{idx}.jpg" in existing:
            idx += 1
        return os.path.join(folder_path, f"{prefix}_{idx}.jpg")

    def _open_annotator(self, path: str, is_pdf: bool):
        """開啟標註對話框"""
        from gui_annotator import AnnotationDialog
        dlg = AnnotationDialog(path, is_pdf=is_pdf, parent=self)
        if not dlg._load_ok:
            QMessageBox.warning(self, "無法載入", f"無法開啟標註工具:\n{path}")
            return
        dlg.exec()
        if dlg.was_saved:
            items = self.record_tree.selectedItems()
            if items:
                record = self._find_record_by_item(items[0])
                self._update_image_detail(record)

    @staticmethod
    def _find_images(folder_path: str) -> list:
        """找出 before/after 圖片（支援 single + group 模式）

        Returns: [(path, label), ...]
        """
        result = []
        files = set(os.listdir(folder_path)) if os.path.isdir(folder_path) else set()
        # single: before.jpg / after.jpg
        if "before.jpg" in files:
            result.append((os.path.join(folder_path, "before.jpg"), "修改前"))
        if "after.jpg" in files:
            result.append((os.path.join(folder_path, "after.jpg"), "修改後"))
        # group: before_1.jpg, before_2.jpg, ... / after_1.jpg, after_2.jpg, ...
        idx = 1
        while True:
            bf = f"before_{idx}.jpg"
            af = f"after_{idx}.jpg"
            found = False
            if bf in files:
                result.append((os.path.join(folder_path, bf), f"修改前 {idx}"))
                found = True
            if af in files:
                result.append((os.path.join(folder_path, af), f"修改後 {idx}"))
                found = True
            if not found:
                break
            idx += 1
        return result

    @staticmethod
    def _render_pdf_thumb(pdf_path: str, size: int = 400):
        """用 PyMuPDF 渲染 PDF 第一頁為 QPixmap"""
        if not _FITZ_OK or not os.path.exists(pdf_path):
            return None
        try:
            doc = _fitz.open(pdf_path)
            page = doc[0]
            rect = page.rect
            zoom = size / min(rect.width, rect.height)
            mat = _fitz.Matrix(zoom, zoom)
            pix = page.get_pixmap(matrix=mat, alpha=False)
            img = QImage(pix.samples, pix.width, pix.height, pix.stride, QImage.Format.Format_RGB888)
            pm = QPixmap.fromImage(img)
            doc.close()
            return pm
        except Exception:
            return None

    # ────────────────── 右鍵選單 ──────────────────
    def _show_context_menu(self, pos):
        item = self.record_tree.itemAt(pos)
        if not item:
            return
        record = self._find_record_by_item(item)
        if not record:
            return

        menu = QMenu(self)
        is_produced = record["status"] in ("已產出", "需重產")
        is_archived = record["status"] == "已歸檔"

        act_folder = menu.addAction("📂 開啟資料夾")
        act_pdf = menu.addAction("📄 開啟 PDF")
        act_pdf.setEnabled(is_produced)
        menu.addSeparator()

        act_edit = menu.addAction("✏️ 修改焊口 / 尺寸")
        act_edit.setEnabled(not is_archived)

        if is_archived:
            act_archive = menu.addAction("📤 還原（取消歸檔）")
        else:
            act_archive = menu.addAction("📦 歸檔")

        menu.addSeparator()
        act_backfill = menu.addAction("🔧 補登")
        act_audit = menu.addAction("🔍 孤兒稽查")

        chosen = menu.exec(self.record_tree.viewport().mapToGlobal(pos))
        if chosen == act_folder:
            self._open_folder_for(record)
        elif chosen == act_pdf:
            self._open_pdf_for(record)
        elif chosen == act_edit:
            self._edit_record(record)
        elif chosen == act_archive:
            if is_archived:
                self._restore_record(record)
            else:
                self._archive_record(record)
        elif chosen == act_backfill:
            self._open_backfill_tool()
        elif chosen == act_audit:
            self._open_orphan_audit()

    def _find_record_by_item(self, item: QTreeWidgetItem) -> dict | None:
        """根據 tree item 在 self.records 中找到對應的 record"""
        rec_idx = item.data(0, Qt.ItemDataRole.UserRole)
        if rec_idx is not None and 0 <= rec_idx < len(self.records):
            return self.records[rec_idx]
        # fallback: 先用 report_id 精準比對
        report_id = item.text(0)
        if report_id:
            for r in self.records:
                if str(r["report_id"]) == report_id:
                    return r
        # 再用 date + 焊口清單 比對
        date_val = item.text(1)
        folder_text = item.text(3)
        for r in self.records:
            if str(r["date"]) == date_val:
                welds_s = str(r["welds"])
                display = welds_s[:30] + "..." if len(welds_s) > 30 else welds_s
                if display == folder_text:
                    return r
        return None

    # ────────────────── 歸檔 / 還原 ──────────────────
    @reentry_guard("_record_move_in_progress", _show_reentry_notice)
    def _archive_record(self, record: dict):
        folder = record.get("folder", "")
        date_val = record.get("date", "")
        if QMessageBox.question(
            self, "歸檔記錄",
            f"確定要歸檔以下記錄嗎？\n\n📁 {folder}\n📅 {date_val}\n\n"
            f"將移至 _archived/{date_val}/ 資料夾",
        ) != QMessageBox.StandardButton.Yes:
            return
        try:
            from operation_journal import OperationJournal
            src = record["folder_path"]
            archived_date_dir = os.path.join(ATTACHMENTS_ROOT, "_archived", date_val)
            os.makedirs(archived_date_dir, exist_ok=True)
            dst = os.path.join(archived_date_dir, folder)
            if os.path.exists(dst):
                ts = datetime.now().strftime("%H%M%S")
                dst = os.path.join(archived_date_dir, f"{folder}_{ts}")
            with OperationJournal(BASE_DIR, "archive_record", {
                "folder": folder,
                "date": date_val,
            }) as journal:
                journal.step("move_attachment_folder", source=src, target=dst)
                shutil.move(src, dst)
            QMessageBox.information(self, "成功", f"✅ 已歸檔至:\n{dst}")
            self.load_records()
        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"歸檔失敗: {e}")

    @reentry_guard("_record_move_in_progress", _show_reentry_notice)
    def _restore_record(self, record: dict):
        folder = record.get("folder", "")
        date_val = record.get("date", "")
        if QMessageBox.question(
            self, "還原記錄",
            f"確定要還原以下記錄嗎？\n\n📁 {folder}\n📅 {date_val}\n\n"
            f"將從 _archived 移回 attachments/{date_val}/",
        ) != QMessageBox.StandardButton.Yes:
            return
        try:
            from operation_journal import OperationJournal
            src = record["folder_path"]
            target_dir = os.path.join(ATTACHMENTS_ROOT, date_val)
            os.makedirs(target_dir, exist_ok=True)
            dst = os.path.join(target_dir, folder)
            if os.path.exists(dst):
                QMessageBox.critical(self, "錯誤", f"目標位置已存在同名資料夾:\n{dst}")
                return
            with OperationJournal(BASE_DIR, "restore_record", {
                "folder": folder,
                "date": date_val,
            }) as journal:
                journal.step("move_attachment_folder", source=src, target=dst)
                shutil.move(src, dst)
            QMessageBox.information(self, "成功", f"✅ 已還原至:\n{dst}")
            self.load_records()
        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"還原失敗: {e}")

    # ────────────────── 編輯焊口 ──────────────────
    def _edit_record(self, record: dict):
        """開啟 EditRecordDialog 來編輯焊口/尺寸/資料夾名稱"""
        import json
        folder_path = record["folder_path"]
        if not os.path.isdir(folder_path):
            QMessageBox.warning(self, "提示", f"資料夾不存在:\n{folder_path}")
            return

        # 組裝 EditRecordDialog 需要的 record 格式
        parts = record["folder"].split("_")
        serial = parts[0] if parts and parts[0].isdigit() else ""
        welds_str = "_".join(parts[1:]) if len(parts) > 1 else ""

        weld_info = None
        wip = os.path.join(folder_path, "weld_info.json")
        if os.path.exists(wip):
            try:
                with open(wip, "r", encoding="utf-8") as f:
                    weld_info = json.load(f)
            except Exception:
                pass

        dialog_record = {
            "date": record["date"],
            "folder_name": record["folder"],
            "folder_path": folder_path,
            "serial": serial,
            "welds_str": welds_str,
            "weld_info": weld_info,
        }
        from gui_dialogs import EditRecordDialog
        EditRecordDialog(self, dialog_record, self.load_records).exec()

    # ────────────────── 動作（通用）──────────────────
    def _open_folder_for(self, record: dict):
        fp = record.get("folder_path", "")
        if fp and os.path.exists(fp):
            os.startfile(fp)
        else:
            QMessageBox.warning(self, "提示", f"資料夾不存在：{fp}")

    def _open_pdf_for(self, record: dict):
        report_id = record.get("report_id", "")
        if not report_id:
            QMessageBox.warning(self, "提示", "此項目尚無報告編號")
            return
        pdf = os.path.join(PDF_OUTPUT_DIR, f"{report_id}.pdf")
        if os.path.exists(pdf):
            os.startfile(pdf)
        else:
            QMessageBox.warning(self, "提示", f"PDF 不存在：{pdf}")

    # ────────────────── 動作（按鈕列）──────────────────

    def _open_backfill_tool(self):
        from gui_dialogs import WeldBackfillDialog
        WeldBackfillDialog(self, None).exec()

    def _open_orphan_audit(self):
        from gui_dialogs import WeldOrphanAuditDialog
        WeldOrphanAuditDialog(self).exec()

    def _open_folder(self):
        items = self.record_tree.selectedItems()
        if not items:
            return
        record = self._find_record_by_item(items[0])
        if record:
            self._open_folder_for(record)

    def _open_pdf(self):
        items = self.record_tree.selectedItems()
        if not items:
            return
        record = self._find_record_by_item(items[0])
        if record:
            self._open_pdf_for(record)

    @reentry_guard("_record_export_in_progress", _show_reentry_notice)
    def _open_excel(self):
        """匯出並開啟 Excel 紀錄清單"""
        try:
            path = export_records_to_excel()
            os.startfile(path)
        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"匯出失敗：{e}")

    @reentry_guard("_site_statistics_export_in_progress", _show_reentry_notice)
    def _export_site_statistics(self, with_pdf: bool = False):
        try:
            path = export_site_statistics_workbook()
            if with_pdf:
                pdf_path = os.path.splitext(path)[0] + ".pdf"
                result = convert_workbook_to_pdf(path, pdf_path)
                if result.get("ok"):
                    os.startfile(result.get("path") or pdf_path)
                    return
                os.startfile(path)
                QMessageBox.warning(
                    self,
                    "PDF 轉檔失敗",
                    "已匯出並開啟 Excel 統計單，但 PDF 轉檔未完成。\n\n"
                    + self._format_pdf_conversion_issues(result),
                )
                return
            os.startfile(path)
        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"匯出現場統計單失敗：{e}")

    def _format_pdf_conversion_issues(self, result: dict) -> str:
        issues = result.get("issues", []) or []
        if not issues:
            return "沒有取得詳細錯誤。"
        messages = [f"- {item.get('message', '')}" for item in issues if item.get("message")]
        return "\n".join(messages) if messages else "沒有取得詳細錯誤。"

    @reentry_guard("_site_output_center_in_progress", _show_reentry_notice)
    def _export_site_output_center(self):
        default_output_dir = os.path.join(BASE_DIR, "staging", "site_output_center_gui")
        scope = self._choose_output_center_scope(default_output_dir)
        if not scope:
            return
        output_dir = scope["output_dir"]

        try:
            result = run_site_output_center(
                output_dir,
                project_root=BASE_DIR,
                attachments_root=ATTACHMENTS_ROOT,
                include_report_keys=scope["include_report_keys"],
                overwrite=True,
                render_pdf=True,
                render_png=False,
                render_statistics=scope["content"]["statistics_xlsx"],
                render_summary_pdf=scope["content"]["summary_pdf"],
                render_photo_grid_pdf=scope["content"]["photo_grid_pdf"],
            )
            self._show_output_center_result_dialog(result)
        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"輸出中心產出失敗：{e}")

    def _show_output_center_result_dialog(self, result: dict) -> None:
        dlg = QDialog(self)
        dlg.setWindowTitle("輸出中心結果")
        dlg.resize(820, 460)
        layout = QVBoxLayout(dlg)
        layout.setSpacing(10)

        summary = QLabel(self._format_output_center_export_message(result), dlg)
        summary.setWordWrap(True)
        summary.setStyleSheet(f"color: {Colors.TEXT_SECONDARY};")
        layout.addWidget(summary)

        tree = QTreeWidget(dlg)
        tree.setHeaderLabels(["項目", "修改單", "狀態", "檔案 / 說明"])
        tree.setColumnWidth(0, 160)
        tree.setColumnWidth(1, 120)
        tree.setColumnWidth(2, 110)
        tree.setColumnWidth(3, 410)
        tree.setAlternatingRowColors(True)
        for group in self._output_center_output_groups(result):
            group_title = group.get("title", "")
            group_status = group.get("status", "")
            if group_status:
                group_title = f"{group_title}（{group_status}）"
            group_item = QTreeWidgetItem(tree, [
                group_title,
                "",
                "",
                "",
            ])
            group_item.setFirstColumnSpanned(True)
            font = group_item.font(0)
            font.setBold(True)
            group_item.setFont(0, font)
            group_item.setForeground(0, QBrush(QColor(Colors.PRIMARY)))
            group_item.setData(0, Qt.ItemDataRole.UserRole, "")
            for item_data in group.get("items", []):
                item = QTreeWidgetItem(group_item, [
                    item_data.get("kind", ""),
                    item_data.get("report", ""),
                    item_data.get("status", item_data.get("pages", "")),
                    item_data.get("path", "") or item_data.get("message", ""),
                ])
                item.setData(0, Qt.ItemDataRole.UserRole, item_data.get("path", ""))
                item.setData(1, Qt.ItemDataRole.UserRole, item_data.get("record_ref", {}))
                item.setData(2, Qt.ItemDataRole.UserRole, item_data.get("issue_action", {}))
                tooltip = item_data.get("tooltip", "") or item_data.get("message", "")
                if tooltip:
                    item.setToolTip(3, tooltip)
                if item_data.get("severity") == "warning" or not item_data.get("exists", True):
                    item.setForeground(2, QBrush(QColor(Colors.WARNING)))
                    item.setForeground(3, QBrush(QColor(Colors.WARNING)))
                elif item_data.get("ok") is False:
                    item.setForeground(2, QBrush(QColor(Colors.DANGER)))
                    item.setForeground(3, QBrush(QColor(Colors.DANGER)))
        tree.expandAll()
        layout.addWidget(tree, 1)
        for row in range(tree.topLevelItemCount()):
            group_item = tree.topLevelItem(row)
            if group_item.childCount():
                tree.setCurrentItem(group_item.child(0))
                break

        btn_row = QHBoxLayout()
        btn_open = QPushButton("開啟選取檔案", dlg)
        btn_focus_record = QPushButton("定位修改單", dlg)
        btn_fix_issue = QPushButton("處理提醒", dlg)
        btn_open_folder = QPushButton("開啟輸出資料夾", dlg)
        btn_close = QPushButton("關閉", dlg)
        btn_row.addWidget(btn_open)
        btn_row.addWidget(btn_focus_record)
        btn_row.addWidget(btn_fix_issue)
        btn_row.addWidget(btn_open_folder)
        btn_row.addStretch()
        btn_row.addWidget(btn_close)
        layout.addLayout(btn_row)

        def open_selected_file():
            selected = tree.selectedItems()
            if not selected:
                QMessageBox.information(dlg, "提示", "請先選取要開啟的檔案。")
                return
            path = selected[0].data(0, Qt.ItemDataRole.UserRole)
            if not path:
                QMessageBox.information(dlg, "提示", "這列沒有可開啟的檔案。")
                return
            if path and os.path.exists(path):
                os.startfile(path)
                return
            QMessageBox.warning(dlg, "檔案不存在", f"找不到檔案：\n{path or '<空白>'}")

        def focus_selected_record():
            selected = tree.selectedItems()
            if not selected:
                QMessageBox.information(dlg, "提示", "請先選取資料提醒或修改單輸出列。")
                return
            record_ref = selected[0].data(1, Qt.ItemDataRole.UserRole) or {}
            if self._focus_output_center_record(record_ref, parent=dlg):
                dlg.accept()

        def handle_selected_issue_action():
            selected = tree.selectedItems()
            if not selected:
                QMessageBox.information(dlg, "提示", "請先選取資料提醒列。")
                return
            action = selected[0].data(2, Qt.ItemDataRole.UserRole) or {}
            record_ref = selected[0].data(1, Qt.ItemDataRole.UserRole) or {}
            if self._handle_output_center_issue_action(action, record_ref, parent=dlg):
                dlg.accept()

        def activate_selected_row():
            selected = tree.selectedItems()
            if not selected:
                return
            if selected[0].data(0, Qt.ItemDataRole.UserRole):
                open_selected_file()
                return
            if selected[0].data(1, Qt.ItemDataRole.UserRole):
                focus_selected_record()
                return

        def open_output_folder():
            root = result.get("output_center") or result.get("showcase", "")
            folder = os.path.join(root, "output")
            if not os.path.isdir(folder):
                folder = root
            if folder and os.path.isdir(folder):
                os.startfile(folder)
                return
            QMessageBox.warning(dlg, "資料夾不存在", f"找不到輸出資料夾：\n{folder or '<空白>'}")

        btn_open.clicked.connect(open_selected_file)
        btn_focus_record.clicked.connect(focus_selected_record)
        btn_fix_issue.clicked.connect(handle_selected_issue_action)
        btn_open_folder.clicked.connect(open_output_folder)
        btn_close.clicked.connect(dlg.accept)
        tree.itemDoubleClicked.connect(lambda _item, _col: activate_selected_row())
        dlg.exec()

    def _choose_output_center_scope(self, default_output_dir: str) -> dict | None:
        all_keys = self._output_center_report_keys(self.records)
        visible_keys = self._output_center_report_keys(self._visible_records())
        selected_keys = self._output_center_report_keys(self._selected_records())
        options = self._output_center_scope_options(
            selected_count=len(selected_keys),
            visible_count=len(visible_keys),
            total_count=len(all_keys),
        )
        if not options:
            QMessageBox.information(self, "沒有可輸出的資料", "目前沒有可輸出的 attachments 修改單。")
            return None

        dlg = QDialog(self)
        dlg.setWindowTitle("輸出中心")
        dlg.resize(520, 220)
        layout = QVBoxLayout(dlg)
        layout.setSpacing(10)

        scope_combo = QComboBox(dlg)
        key_by_mode = {
            "all": None,
            "visible": visible_keys,
            "selected": selected_keys,
        }
        for option in options:
            scope_combo.addItem(option["label"], option)
        layout.addWidget(QLabel("輸出範圍："))
        layout.addWidget(scope_combo)

        content_group = QGroupBox("輸出內容", dlg)
        content_layout = QVBoxLayout(content_group)
        chk_statistics = QCheckBox("現場統計單 Excel", content_group)
        chk_statistics.setChecked(True)
        chk_summary_pdf = QCheckBox("summary PDF", content_group)
        chk_summary_pdf.setChecked(True)
        chk_photo_grid_pdf = QCheckBox("before/after 照片 PDF", content_group)
        chk_photo_grid_pdf.setChecked(True)
        content_layout.addWidget(chk_statistics)
        content_layout.addWidget(chk_summary_pdf)
        content_layout.addWidget(chk_photo_grid_pdf)
        layout.addWidget(content_group)

        output_group = QGroupBox("輸出位置", dlg)
        output_layout = QHBoxLayout(output_group)
        output_edit = QLineEdit(default_output_dir, output_group)
        output_edit.setPlaceholderText("選擇輸出資料夾")
        output_layout.addWidget(output_edit, 1)
        btn_browse_output = QPushButton("瀏覽", output_group)
        output_layout.addWidget(btn_browse_output)
        layout.addWidget(output_group)

        def browse_output_dir():
            current = self._normalize_output_center_output_dir(output_edit.text(), default_output_dir)
            initial = current if os.path.isdir(current) else os.path.dirname(current)
            chosen = QFileDialog.getExistingDirectory(self, "選擇輸出資料夾", initial)
            if chosen:
                output_edit.setText(chosen)

        btn_browse_output.clicked.connect(browse_output_dir)

        message_label = QLabel(dlg)
        message_label.setWordWrap(True)
        message_label.setStyleSheet(f"color: {Colors.TEXT_SECONDARY};")
        layout.addWidget(message_label)

        def refresh_message():
            option = scope_combo.currentData() or options[0]
            message_label.setText(
                self._format_output_center_export_confirmation(
                    self._normalize_output_center_output_dir(output_edit.text(), default_output_dir),
                    option.get("label", ""),
                    option.get("count", 0),
                    self._format_output_center_content_label({
                        "statistics_xlsx": chk_statistics.isChecked(),
                        "summary_pdf": chk_summary_pdf.isChecked(),
                        "photo_grid_pdf": chk_photo_grid_pdf.isChecked(),
                    }),
                )
            )

        scope_combo.currentIndexChanged.connect(lambda _idx: refresh_message())
        chk_statistics.stateChanged.connect(lambda _state: refresh_message())
        chk_summary_pdf.stateChanged.connect(lambda _state: refresh_message())
        chk_photo_grid_pdf.stateChanged.connect(lambda _state: refresh_message())
        output_edit.textChanged.connect(lambda _text: refresh_message())
        refresh_message()

        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel,
            parent=dlg,
        )
        buttons.accepted.connect(dlg.accept)
        buttons.rejected.connect(dlg.reject)
        layout.addWidget(buttons)

        if dlg.exec() != QDialog.DialogCode.Accepted:
            return None

        option = scope_combo.currentData() or options[0]
        content = {
            "statistics_xlsx": chk_statistics.isChecked(),
            "summary_pdf": chk_summary_pdf.isChecked(),
            "photo_grid_pdf": chk_photo_grid_pdf.isChecked(),
        }
        if not any(content.values()):
            QMessageBox.information(self, "沒有選擇輸出內容", "請至少選擇一種輸出內容。")
            return None
        output_dir = self._normalize_output_center_output_dir(output_edit.text(), default_output_dir)
        if not output_dir:
            QMessageBox.information(self, "沒有輸出位置", "請指定輸出資料夾。")
            return None
        return {
            "mode": option["mode"],
            "label": option["label"],
            "include_report_keys": key_by_mode.get(option["mode"]),
            "content": content,
            "output_dir": output_dir,
        }

    def _visible_records(self) -> list[dict]:
        records = []
        for idx in range(self.record_tree.topLevelItemCount()):
            record = self._find_record_by_item(self.record_tree.topLevelItem(idx))
            if record:
                records.append(record)
        return records

    def _selected_records(self) -> list[dict]:
        records = []
        for item in self.record_tree.selectedItems():
            record = self._find_record_by_item(item)
            if record:
                records.append(record)
        return records

    def _focus_output_center_record(self, record_ref: dict, *, parent=None) -> bool:
        if not record_ref:
            QMessageBox.information(parent or self, "無法定位", "這列沒有對應的修改單定位資料。")
            return False
        item = self._find_record_tree_item_by_output_ref(record_ref)
        filters_reset = False
        if not item and self._reset_record_filters_for_output_focus():
            filters_reset = True
            item = self._find_record_tree_item_by_output_ref(record_ref)
        if not item:
            QMessageBox.information(
                parent or self,
                "找不到修改單",
                self._format_output_center_focus_missing_message(record_ref, filters_reset=filters_reset),
            )
            return False
        self._select_record_tree_item(item)
        return True

    def _select_record_tree_item(self, item: QTreeWidgetItem) -> None:
        self.record_tree.clearSelection()
        item.setSelected(True)
        self.record_tree.setCurrentItem(item)
        self.record_tree.scrollToItem(item)
        self._on_record_select()

    def _find_record_tree_item_by_output_ref(self, record_ref: dict) -> QTreeWidgetItem | None:
        for row in range(self.record_tree.topLevelItemCount()):
            item = self.record_tree.topLevelItem(row)
            record = self._find_record_by_item(item)
            if self._record_matches_output_ref(record, record_ref):
                return item
        return None

    def _handle_output_center_issue_action(self, action: dict, record_ref: dict, *, parent=None) -> bool:
        if not action:
            QMessageBox.information(parent or self, "無法處理", "這列沒有可用的修正動作。")
            return False
        kind = str(action.get("kind", "") or "focus_record")
        if kind == "open_folder":
            return self._open_output_center_record_folder(record_ref, parent=parent)
        if kind == "add_photo":
            return self._add_output_center_photo(record_ref, str(action.get("prefix", "") or ""), parent=parent)
        if kind == "edit_note":
            return self._edit_output_center_note(record_ref, parent=parent)
        if kind == "focus_record":
            return self._focus_output_center_record(record_ref, parent=parent)
        QMessageBox.information(parent or self, "尚未支援", str(action.get("message", "") or "此提醒尚未設定處理動作。"))
        return False

    def _open_output_center_record_folder(self, record_ref: dict, *, parent=None) -> bool:
        item = self._find_record_tree_item_by_output_ref(record_ref)
        if not item and self._reset_record_filters_for_output_focus():
            item = self._find_record_tree_item_by_output_ref(record_ref)
        if item:
            self._select_record_tree_item(item)
            record = self._find_record_by_item(item)
            if record:
                folder = record.get("folder_path", "")
                if folder and os.path.isdir(folder):
                    os.startfile(folder)
                    return True

        folder_path = self._output_center_folder_path_from_ref(record_ref)
        if folder_path and os.path.isdir(folder_path):
            os.startfile(folder_path)
            return True
        QMessageBox.warning(
            parent or self,
            "找不到資料夾",
            f"找不到對應 attachments 資料夾：\n{folder_path or '<無定位資料>'}",
        )
        return False

    def _add_output_center_photo(self, record_ref: dict, prefix: str, *, parent=None) -> bool:
        folder_path = ""
        item = self._find_record_tree_item_by_output_ref(record_ref)
        if not item and self._reset_record_filters_for_output_focus():
            item = self._find_record_tree_item_by_output_ref(record_ref)
        if item:
            self._select_record_tree_item(item)
            record = self._find_record_by_item(item)
            if record:
                folder_path = record.get("folder_path", "")
        if not folder_path:
            folder_path = self._output_center_folder_path_from_ref(record_ref)
        if not folder_path or not os.path.isdir(folder_path):
            QMessageBox.warning(
                parent or self,
                "找不到資料夾",
                f"找不到對應 attachments 資料夾：\n{folder_path or '<無定位資料>'}",
            )
            return False
        return self._add_image_to_folder(folder_path, preferred_prefix=prefix, parent=parent)

    @staticmethod
    def _output_center_folder_path_from_ref(record_ref: dict) -> str:
        date = str(record_ref.get("date", "") or "").strip()
        folder = str(record_ref.get("folder", "") or "").strip()
        if date and folder:
            return os.path.join(ATTACHMENTS_ROOT, date, folder)
        if not folder:
            return ""
        if not os.path.isdir(ATTACHMENTS_ROOT):
            return ""
        for date_name in sorted(os.listdir(ATTACHMENTS_ROOT)):
            if not re.match(r"^\d{8}$", date_name):
                continue
            candidate = os.path.join(ATTACHMENTS_ROOT, date_name, folder)
            if os.path.isdir(candidate):
                return candidate
        return os.path.join(ATTACHMENTS_ROOT, folder)

    def _edit_output_center_note(self, record_ref: dict, *, parent=None) -> bool:
        folder_path = self._output_center_folder_path_from_ref(record_ref)
        if not folder_path or not os.path.isdir(folder_path):
            QMessageBox.warning(
                parent or self,
                "找不到資料夾",
                f"找不到對應 attachments 資料夾：\n{folder_path or '<無定位資料>'}",
            )
            return False

        note_path = os.path.join(folder_path, "note.txt")
        try:
            current = self._read_output_center_note_text(note_path)
        except Exception as exc:
            QMessageBox.warning(parent or self, "讀取 note 失敗", f"無法讀取 note.txt：\n{exc}")
            return False

        dlg = QDialog(parent or self)
        dlg.setWindowTitle("編輯現場說明 note.txt")
        dlg.resize(560, 360)
        layout = QVBoxLayout(dlg)
        layout.setSpacing(10)

        title = QLabel(f"資料夾：{folder_path}", dlg)
        title.setWordWrap(True)
        title.setStyleSheet(f"color: {Colors.TEXT_SECONDARY};")
        layout.addWidget(title)

        editor = QTextEdit(dlg)
        editor.setPlainText(current)
        editor.setPlaceholderText("請輸入現場修改原因、施工說明或需要留給請款/統計的人看的內容。")
        layout.addWidget(editor, 1)

        hint = QLabel("儲存後會覆寫此資料夾的 note.txt；原始 attachments 內其他檔案不會被修改。", dlg)
        hint.setWordWrap(True)
        hint.setStyleSheet(f"color: {Colors.TEXT_MUTED};")
        layout.addWidget(hint)

        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel,
            parent=dlg,
        )
        buttons.accepted.connect(dlg.accept)
        buttons.rejected.connect(dlg.reject)
        layout.addWidget(buttons)

        if dlg.exec() != QDialog.DialogCode.Accepted:
            return False

        text = editor.toPlainText().strip()
        if not self._output_center_note_text_is_valid(text):
            QMessageBox.information(parent or self, "note 仍不完整", "請填入實際現場說明，不要留空或保留「請填寫」樣板文字。")
            return False
        try:
            self._write_output_center_note_text(note_path, text)
        except Exception as exc:
            QMessageBox.critical(parent or self, "儲存 note 失敗", f"無法儲存 note.txt：\n{exc}")
            return False
        QMessageBox.information(parent or self, "已儲存", f"已更新：\n{note_path}")
        return True

    @staticmethod
    def _read_output_center_note_text(note_path: str) -> str:
        if not os.path.exists(note_path):
            return ""
        with open(note_path, "r", encoding="utf-8") as f:
            return f.read()

    @staticmethod
    def _write_output_center_note_text(note_path: str, text: str) -> None:
        folder = os.path.dirname(note_path)
        if folder:
            os.makedirs(folder, exist_ok=True)
        tmp = note_path + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            f.write(str(text or "").strip())
            f.write("\n")
        os.replace(tmp, note_path)

    @staticmethod
    def _output_center_note_text_is_valid(text: str) -> bool:
        normalized = str(text or "").strip()
        if not normalized:
            return False
        compact = normalized.replace(" ", "")
        return not ("請填寫" in compact or compact.startswith("#"))

    def _reset_record_filters_for_output_focus(self) -> bool:
        current_status = self.status_combo.currentText() if hasattr(self, "status_combo") else "全部"
        search_text = self.search_edit.text() if hasattr(self, "search_edit") else ""
        date_from = self.date_from_edit.text() if hasattr(self, "date_from_edit") else ""
        date_to = self.date_to_edit.text() if hasattr(self, "date_to_edit") else ""
        if not self._output_center_filters_are_narrowed(current_status, search_text, date_from, date_to):
            return False

        if hasattr(self, "status_combo"):
            idx = self.status_combo.findText("全部")
            if idx >= 0:
                old = self.status_combo.blockSignals(True)
                self.status_combo.setCurrentIndex(idx)
                self.status_combo.blockSignals(old)
        for edit_name in ("search_edit", "date_from_edit", "date_to_edit"):
            edit = getattr(self, edit_name, None)
            if edit is not None and edit.text():
                edit.clear()
        self.load_records()
        return True

    @staticmethod
    def _output_center_filters_are_narrowed(
        status_text: str,
        search_text: str = "",
        date_from: str = "",
        date_to: str = "",
    ) -> bool:
        return (
            str(status_text or "").strip() not in ("", "全部")
            or bool(str(search_text or "").strip())
            or bool(str(date_from or "").strip())
            or bool(str(date_to or "").strip())
        )

    @staticmethod
    def _record_matches_output_ref(record: dict | None, record_ref: dict) -> bool:
        if not record or not record_ref:
            return False
        ref_report_id = str(record_ref.get("report_id", "") or "").strip()
        ref_report = str(record_ref.get("report", "") or "").strip()
        ref_date = str(record_ref.get("date", "") or "").strip()
        ref_folder = str(record_ref.get("folder", "") or "").strip()
        record_report_id = str(record.get("report_id", "") or "").strip()
        record_date = str(record.get("date", "") or "").strip()
        record_folder = str(record.get("folder", "") or "").strip()

        if ref_report_id and record_report_id == ref_report_id:
            return True
        if ref_date and ref_folder and record_date == ref_date and record_folder == ref_folder:
            return True
        if ref_folder and record_folder == ref_folder and (not ref_date or record_date == ref_date):
            return True
        if ref_report and ref_report in {record_report_id, record_folder}:
            return True
        return False

    @staticmethod
    def _format_output_center_focus_missing_message(record_ref: dict, *, filters_reset: bool = False) -> str:
        label = RecordManagerPanel._format_output_center_record_ref_label(record_ref)
        prefix = "已切回全部狀態並清空搜尋，仍找不到這筆修改單。" if filters_reset else "目前清單找不到這筆修改單。"
        return (
            f"{prefix}\n\n"
            f"修改單：{label}\n\n"
            "可能原因：目前篩選條件把它藏起來，或這筆資料只存在 attachments/ 尚未寫入 records.json。"
        )

    @staticmethod
    def _output_center_report_keys(records: list[dict]) -> list[tuple[str, str]]:
        keys = []
        seen = set()
        for record in records or []:
            if record.get("is_archived"):
                continue
            date_str = str(record.get("date", "") or "").strip()
            folder = str(record.get("folder", "") or "").strip()
            key = (date_str, folder)
            if not date_str or not folder or key in seen:
                continue
            seen.add(key)
            keys.append(key)
        return keys

    @staticmethod
    def _output_center_scope_options(*, selected_count: int, visible_count: int, total_count: int) -> list[dict]:
        options = []
        if total_count > 0:
            options.append({"mode": "all", "label": f"全部 attachments ({total_count})", "count": total_count})
        if visible_count > 0:
            options.append({"mode": "visible", "label": f"目前篩選結果 ({visible_count})", "count": visible_count})
        if selected_count > 0:
            options.append({"mode": "selected", "label": f"選取修改單 ({selected_count})", "count": selected_count})
        return options

    @staticmethod
    def _format_output_center_content_label(content: dict) -> str:
        labels = []
        if content.get("statistics_xlsx"):
            labels.append("現場統計單 Excel")
        if content.get("summary_pdf"):
            labels.append("summary PDF")
        if content.get("photo_grid_pdf"):
            labels.append("照片 PDF")
        return "、".join(labels) if labels else "未選擇"

    @staticmethod
    def _normalize_output_center_output_dir(value: str, default_output_dir: str) -> str:
        text = str(value or "").strip().strip('"').strip("'")
        if not text:
            text = str(default_output_dir or "").strip()
        if not text:
            return ""
        return os.path.abspath(os.path.expanduser(text))

    @staticmethod
    def _output_center_output_items(result: dict) -> list[dict]:
        items = []
        files = result.get("files", {}) or {}
        static_outputs = [
            ("資料檔", "資料 JSON", "", "", files.get("report_set", "")),
            ("主要輸出", "現場統計單 Excel", "", "", files.get("statistics_xlsx", "")),
            ("資料檔", "摘要 JSON", "", "", files.get("summary", "")),
        ]
        for group, kind, report, pages, path in static_outputs:
            if path:
                exists = os.path.exists(path)
                items.append({
                    "group": group,
                    "kind": kind,
                    "report": report,
                    "pages": pages,
                    "path": path,
                    "exists": exists,
                    "ok": exists,
                    "status": "可開啟" if exists else "找不到檔案",
                    "tooltip": "" if exists else "檔案路徑已記錄，但目前找不到檔案",
                })

        template_labels = {
            "summary": "summary PDF",
            "photo_grid": "照片 PDF",
        }
        for render in result.get("renders", []) or []:
            path = render.get("path", "")
            if not path and render.get("ok", True) is not False:
                continue
            template = render.get("template", "")
            folder = str(render.get("folder", "") or "")
            exists = bool(path and os.path.exists(path))
            ok = bool(render.get("ok", True)) and exists
            issue_codes = [str(code) for code in render.get("issue_codes", []) or [] if code]
            items.append({
                "group": "PDF",
                "kind": template_labels.get(template, template or "PDF"),
                "report": folder,
                "pages": str(render.get("pages", "") or ""),
                "path": path,
                "exists": exists,
                "ok": ok,
                "status": RecordManagerPanel._format_output_center_render_status(render, exists=exists),
                "tooltip": "；".join(issue_codes),
                "record_ref": {"report": folder, "report_id": "", "date": "", "folder": folder} if folder else {},
            })
        return items

    @staticmethod
    def _output_center_output_groups(result: dict) -> list[dict]:
        items = RecordManagerPanel._output_center_output_items(result)
        groups = []
        for title in ("主要輸出", "PDF", "資料檔"):
            group_items = [item for item in items if item.get("group") == title]
            if group_items:
                groups.append({
                    "title": title,
                    "status": RecordManagerPanel._format_output_center_group_status(group_items),
                    "items": group_items,
                })

        issue_items = RecordManagerPanel._output_center_issue_items(result)
        if issue_items:
            groups.append({
                "title": "資料提醒",
                "status": f"{len(issue_items)} 項",
                "items": issue_items,
            })
        return groups

    @staticmethod
    def _output_center_issue_items(result: dict) -> list[dict]:
        items = []
        for issue in result.get("issues", []) or []:
            severity = str(issue.get("severity", "") or "warning")
            code = str(issue.get("code", "") or "資料提醒")
            message = str(issue.get("message", "") or "")
            record_ref = RecordManagerPanel._output_center_issue_record_ref(issue)
            issue_action = RecordManagerPanel._output_center_issue_action(issue)
            items.append({
                "group": "資料提醒",
                "kind": code,
                "report": RecordManagerPanel._format_output_center_record_ref_label(record_ref),
                "pages": "",
                "path": "",
                "message": message,
                "exists": True,
                "ok": severity not in ("error", "warning"),
                "severity": severity,
                "status": "提醒" if severity == "warning" else severity,
                "tooltip": RecordManagerPanel._format_output_center_issue_tooltip(message, issue_action),
                "record_ref": record_ref,
                "issue_action": issue_action,
            })
        return items

    @staticmethod
    def _output_center_issue_action(issue: dict) -> dict[str, str]:
        code = str(issue.get("code", "") or "").strip()
        actions = {
            "note": {
                "kind": "edit_note",
                "label": "編輯 note.txt",
                "message": "請補齊 note.txt 的現場說明。",
            },
            "before_photo": {
                "kind": "add_photo",
                "prefix": "before",
                "label": "新增 before 照片",
                "message": "請選擇要加入此附件資料夾的 before 照片。",
            },
            "after_photo": {
                "kind": "add_photo",
                "prefix": "after",
                "label": "新增 after 照片",
                "message": "請選擇要加入此附件資料夾的 after 照片。",
            },
            "parse_error": {
                "kind": "open_folder",
                "label": "開啟資料夾檢查文字檔",
                "message": "請檢查 GroupWeld.txt、materials.txt 或資料夾命名是否可解析。",
            },
            "weld_or_material": {
                "kind": "focus_record",
                "label": "定位修改單檢查焊口/材料",
                "message": "請檢查此修改單是否缺少焊口或用料資料。",
            },
        }
        return actions.get(code, {
            "kind": "focus_record",
            "label": "定位修改單",
            "message": "請回到修改單檢查此資料提醒。",
        })

    @staticmethod
    def _format_output_center_issue_tooltip(message: str, issue_action: dict) -> str:
        label = str(issue_action.get("label", "") or "").strip()
        action_message = str(issue_action.get("message", "") or "").strip()
        parts = [str(message or "").strip()]
        if label:
            parts.append(f"處理：{label}")
        if action_message:
            parts.append(action_message)
        return "\n".join(part for part in parts if part)

    @staticmethod
    def _output_center_issue_record_ref(issue: dict) -> dict[str, str]:
        report = str(issue.get("report", "") or "").strip()
        report_id = str(issue.get("report_id", "") or "").strip()
        date = str(issue.get("date", "") or issue.get("date_raw", "") or "").strip()
        folder = str(issue.get("folder", "") or "").strip()
        if not folder and report and not report_id:
            folder = report
        return {
            "report": report,
            "report_id": report_id,
            "date": date,
            "folder": folder,
        }

    @staticmethod
    def _format_output_center_record_ref_label(record_ref: dict) -> str:
        report_id = str(record_ref.get("report_id", "") or "").strip()
        date = str(record_ref.get("date", "") or "").strip()
        folder = str(record_ref.get("folder", "") or "").strip()
        report = str(record_ref.get("report", "") or "").strip()
        if report_id:
            return report_id
        if date and folder:
            return f"{date}/{folder}"
        return folder or report or ""

    @staticmethod
    def _format_output_center_render_status(render: dict, *, exists: bool) -> str:
        if render.get("ok", True) is False:
            codes = [str(code) for code in render.get("issue_codes", []) or [] if code]
            return f"失敗：{', '.join(codes)}" if codes else "失敗"
        if not exists:
            return "找不到檔案"
        pages = str(render.get("pages", "") or "").strip()
        return f"OK，{pages} 頁" if pages else "OK"

    @staticmethod
    def _format_output_center_group_status(items: list[dict]) -> str:
        needs_attention = sum(
            1 for item in items
            if item.get("ok") is False or not item.get("exists", True)
        )
        if needs_attention:
            return f"{len(items)} 項，{needs_attention} 項需注意"
        return f"{len(items)} 項"

    @staticmethod
    def _format_output_center_export_confirmation(
        output_dir: str,
        scope_label: str = "全部 attachments",
        report_count: int = 0,
        content_label: str = "現場統計單 Excel、summary PDF、照片 PDF",
    ) -> str:
        return (
            "將用目前 attachments/ 建立現場輸出中心資料。\n\n"
            f"範圍：{scope_label}\n"
            f"預計修改單：{report_count} 張\n"
            f"輸出內容：{content_label}\n"
            f"輸出資料夾：{output_dir}\n\n"
            "會重新產生該輸出中心資料夾中的檔案；原始 attachments/ 不會被修改。"
        )

    @staticmethod
    def _format_output_center_export_message(result: dict) -> str:
        renders = result.get("renders", []) or []
        ok_renders = [item for item in renders if item.get("ok") and item.get("path")]
        by_template = {}
        for item in ok_renders:
            key = item.get("template") or "pdf"
            by_template[key] = by_template.get(key, 0) + 1
        template_labels = {
            "summary": "統計 PDF",
            "photo_grid": "照片 PDF",
            "pdf": "PDF",
        }
        render_parts = [
            f"{template_labels.get(key, key)} {count} 份"
            for key, count in sorted(by_template.items())
        ]
        aggregates = result.get("aggregates", {}) or {}
        files = result.get("files", {}) or {}
        lines = [
            "現場輸出中心已產出。" if result.get("ok") else "現場輸出中心產出未完全成功。",
            "",
            f"修改單：{result.get('report_count', 0)} 張",
            f"焊口：{aggregates.get('weld_count', 0)} 口",
            f"材料列：{aggregates.get('material_row_count', 0)} 筆",
            f"照片：{aggregates.get('photo_count', 0)} 張",
            f"PDF：{'、'.join(render_parts) if render_parts else '0 份'}",
            f"輸出資料夾：{result.get('output_center') or result.get('showcase', '')}",
        ]
        if files.get("statistics_xlsx"):
            lines.append(f"統計單：{files['statistics_xlsx']}")
        issues = result.get("issues", []) or []
        if issues:
            lines.extend(["", "資料提醒："])
            for issue in issues[:6]:
                lines.append(f"- {issue.get('report', '')}: {issue.get('message', '')}")
            if len(issues) > 6:
                lines.append(f"...還有 {len(issues) - 6} 筆")
        return "\n".join(lines)

    def _show_rebuild_queue(self):
        rows = build_rebuild_queue(_load_store())
        if not rows:
            QMessageBox.information(self, "需重產清單", format_rebuild_queue_summary(rows))
            return
        idx = self.status_combo.findText("需重產")
        if idx >= 0:
            self.status_combo.setCurrentIndex(idx)
        self.load_records()
        QMessageBox.information(self, "需重產清單", format_rebuild_queue_summary(rows))

    @reentry_guard("_record_rebuild_export_in_progress", _show_reentry_notice)
    def _export_rebuild_queue(self):
        try:
            rows = build_rebuild_queue(_load_store())
        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"讀取需重產清單失敗：{e}")
            return
        if not rows:
            QMessageBox.information(self, "需重產清單", format_rebuild_queue_summary(rows))
            return

        default_path = os.path.join(BASE_DIR, f"需重產清單_{datetime.now().strftime('%Y%m%d')}.csv")
        path, _ = QFileDialog.getSaveFileName(
            self,
            "匯出需重產清單",
            default_path,
            "CSV Files (*.csv);;All Files (*)",
        )
        if not path:
            return
        if not os.path.splitext(path)[1]:
            path += ".csv"
        try:
            export_rebuild_queue_csv(path, rows)
            QMessageBox.information(self, "匯出完成", self._format_rebuild_queue_export_message(rows, path))
        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"匯出需重產清單失敗：{e}")

    # ────────────────── 儲存 ──────────────────
    @reentry_guard("_record_save_in_progress", _show_reentry_notice)
    def _save_changes(self):
        if not self.details_modified and not self.materials_modified:
            QMessageBox.information(self, "提示", "沒有需要儲存的變更")
            return
        journal = None
        try:
            import json as _json
            from collections import defaultdict
            from gui_dialogs import (
                _collect_and_merge_weld_sources, _apply_conflict_choices,
                WeldSyncConflictDialog,
            )
            from operation_journal import OperationJournal
            from record_manager import auto_backup, RECORDS_JSON_PATH

            store = _load_store()
            renamed_folders = []   # 記錄需 rename 的資料夾
            journal = OperationJournal(
                BASE_DIR,
                "record_manager_save_changes",
                {
                    "details_modified": self.details_modified,
                    "materials_modified": self.materials_modified,
                },
            ).begin()

            if self.details_modified:
                # 按 report_id 分組 self.details
                mod_groups: dict[str, list] = defaultdict(list)
                for d in self.details:
                    mod_groups[d["report_id"]].append(d)

                # 依索引位置更新 store["details"]（焊口編號本身可能被修改）
                for rid, mods in mod_groups.items():
                    store_group = [
                        det for det in store["details"]
                        if det.get("紀錄編號", "") == rid
                    ]
                    for i, mod in enumerate(mods):
                        if i >= len(store_group):
                            break
                        det = store_group[i]
                        det["焊口編號"] = mod.get("weld_code", "")
                        det["焊口尺寸"] = mod.get("size", "")
                        det["係數"] = mod.get("coefficient", "")
                        det["單價/DB"] = mod.get("price", "")
                        det["金額"] = mod.get("amount", "")

                # ── 同步 weld_info.json ──
                for record in self.records:
                    rid = str(record.get("report_id", ""))
                    if rid not in mod_groups:
                        continue
                    folder_path = record.get("folder_path", "")
                    if not folder_path or not os.path.isdir(folder_path):
                        continue

                    wip = os.path.join(folder_path, "weld_info.json")
                    wi_data = {}
                    if os.path.isfile(wip):
                        try:
                            with open(wip, "r", encoding="utf-8") as f:
                                wi_data = _json.load(f)
                        except Exception:
                            wi_data = {}

                    old_welds = wi_data.get("welds", [])
                    new_welds = []
                    for i, mod in enumerate(mod_groups[rid]):
                        w = dict(old_welds[i]) if i < len(old_welds) else {}
                        weld_code = str(mod.get("weld_code", ""))
                        mark = str(mod.get("mark", ""))
                        # 一律 regex 拆分: "1001a" → weld_no="1001", mark="a"
                        m_wn = re.match(r'^(\d+)([rab])', weld_code, re.IGNORECASE)
                        if m_wn:
                            weld_code = m_wn.group(1)
                            mark = mark or m_wn.group(2).lower()
                        w["weld_no"] = weld_code
                        w["mark"] = mark
                        w["size"] = mod.get("size", "")
                        w["material"] = mod.get("material", "")
                        w["thickness"] = mod.get("thickness", "")
                        new_welds.append(w)
                    # 防空口編號重複
                    _seen_wids: set[str] = set()
                    _unique_welds: list[dict] = []
                    for _w in new_welds:
                        _wid = f"{_w.get('weld_no','')}{_w.get('mark','')}"
                        if _wid not in _seen_wids:
                            _seen_wids.add(_wid)
                            _unique_welds.append(_w)
                    wi_data["welds"] = _unique_welds

                    journal.step("write_weld_info", report_id=rid, path=wip)
                    atomic_write_json(wip, wi_data)

                    # ── 同步 records[] 主表的焊口欄位 ──
                    merged_for_main = []
                    for _mod in mod_groups[rid]:
                        _wn = str(_mod.get("weld_code", ""))
                        _mk = str(_mod.get("mark", ""))
                        _m_sp = re.match(r'^(\d+)([rab])', _wn, re.IGNORECASE)
                        if _m_sp:
                            _wn = _m_sp.group(1)
                            _mk = _mk or _m_sp.group(2).lower()
                        merged_for_main.append({
                            "weld_no": _wn,
                            "mark": _mk,
                            "size": _mod.get("size", ""),
                        })

                    for rec in store["records"]:
                        if rec.get("報告編號", "") == rid:
                            weld_list_str = "、".join(
                                f"{w['weld_no']}{w['mark']}" for w in merged_for_main
                            )
                            weld_size_str = "；".join(
                                f"{w['weld_no']}{w['mark']}={w['size']}" for w in merged_for_main
                            )
                            rec["焊口清單"] = weld_list_str
                            rec["焊口與尺寸"] = weld_size_str

                            # ── 檢查是否需要改資料夾名 ──
                            old_folder = rec.get("資料夾名", "")
                            serial = record.get("series", "")
                            codes = [f"{w['weld_no']}{w['mark']}{w['size']}"
                                     for w in merged_for_main if w['weld_no'] and w['mark'] and w['size']]
                            if serial and codes:
                                new_folder = f"{serial}_{'_'.join(codes)}"
                                if new_folder != old_folder:
                                    old_fp = os.path.join(ATTACHMENTS_ROOT, record["date"], old_folder)
                                    new_fp = os.path.join(ATTACHMENTS_ROOT, record["date"], new_folder)
                                    renamed_folders.append((old_fp, new_fp, old_folder, new_folder, rec))
                            break

            if self.materials_modified:
                for mat in store.get("materials", []):
                    rid = mat.get("報告編號", "")
                    comp = mat.get("零件類型", "")
                    sz = mat.get("尺寸", "")
                    for m in self.materials:
                        if m["report_id"] == rid and m["component"] == comp and m["size"] == sz:
                            mat["單價"] = m.get("price", "")
                            mat["金額"] = m.get("amount", "")
                            mat["類別"] = m.get("category", "材料")
                            mat["單價來源"] = m.get("price_source", "")
                            mat["金額來源"] = m.get("amount_source", "")
                            mat["價目表ID"] = m.get("pricebook_id", "")
                            mat["價目來源"] = m.get("pricebook_source", "")
                            mat["價目生效日"] = m.get("pricebook_effective_date", "")
                            mat["配價狀態"] = m.get("pricing_status", "")
                            break

            # ── 處理資料夾重新命名 ──
            for old_fp, new_fp, old_fn, new_fn, rec in renamed_folders:
                if os.path.isdir(old_fp) and not os.path.exists(new_fp):
                    answer = QMessageBox.question(
                        self, "資料夾名稱同步",
                        f"焊口已修改，是否同步更新資料夾名稱？\n\n"
                        f"📁 {old_fn}\n→ {new_fn}",
                    )
                    if answer == QMessageBox.StandardButton.Yes:
                        journal.step("rename_attachment_folder", source=old_fp, target=new_fp)
                        os.rename(old_fp, new_fp)
                        rec["資料夾名"] = new_fn

            auto_backup(RECORDS_JSON_PATH)
            journal.step("save_records_json", path=RECORDS_JSON_PATH)
            _save_store(store)
            journal.complete()
            journal = None
            self.details_modified = False
            self.materials_modified = False
            self._update_modified_label()
            QMessageBox.information(self, "成功", "變更已儲存（含主表 + 明細同步）")

        except Exception as e:
            if journal is not None:
                journal.fail(str(e))
            QMessageBox.critical(self, "錯誤", f"儲存失敗：{e}")


# ========= 材料價目表面板 =========
class MaterialPricebookPanel(QWidget):
    """材料價目表最小編輯 UI。"""

    PRICE_SOURCE_OPTIONS = ["合約", "報價", "手動", "參考"]

    COLUMNS = [
        ("id", "料號/ID", 130),
        ("零件類型", "零件類型", 120),
        ("尺寸", "尺寸", 70),
        ("SCH", "SCH", 70),
        ("材質", "材質", 80),
        ("類別", "類別", 80),
        ("單位", "單位", 60),
        ("單價", "單價", 90),
        ("來源", "來源", 80),
        ("生效日", "生效日", 90),
        ("備註", "備註", 220),
    ]

    def __init__(self, parent: QWidget = None):
        super().__init__(parent)
        self.items: list[dict] = []
        self.modified = False
        self._editor = _CellEditor()
        self._build_ui()
        self.load_data()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(8, 8, 8, 8)
        root.setSpacing(10)

        toolbar = QHBoxLayout()
        toolbar.setSpacing(8)

        toolbar.addWidget(QLabel("搜尋："))
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("零件類型 / 尺寸 / 材質 / 料號")
        self.search_edit.textChanged.connect(self._apply_filter)
        toolbar.addWidget(self.search_edit, 1)

        self.only_unpriced_checkbox = QCheckBox("只看未定價")
        self.only_unpriced_checkbox.toggled.connect(self._apply_filter)
        toolbar.addWidget(self.only_unpriced_checkbox)

        toolbar.addWidget(QLabel("零件："))
        self.component_filter_combo = QComboBox()
        self.component_filter_combo.addItems(self._filter_options_for_key("零件類型", "全部零件"))
        self.component_filter_combo.setFixedWidth(130)
        self.component_filter_combo.currentIndexChanged.connect(self._apply_filter)
        toolbar.addWidget(self.component_filter_combo)

        toolbar.addWidget(QLabel("材質："))
        self.material_filter_combo = QComboBox()
        self.material_filter_combo.addItems(self._filter_options_for_key("材質", "全部材質"))
        self.material_filter_combo.setFixedWidth(150)
        self.material_filter_combo.currentIndexChanged.connect(self._apply_filter)
        toolbar.addWidget(self.material_filter_combo)

        btn_reload = QPushButton("載入")
        btn_reload.clicked.connect(self.load_data)
        toolbar.addWidget(btn_reload)

        self.btn_import_seed = QPushButton("匯入骨架")
        self.btn_import_seed.clicked.connect(self._import_seed)
        toolbar.addWidget(self.btn_import_seed)

        self.btn_import_price_table = QPushButton("匯入價格表")
        self.btn_import_price_table.clicked.connect(self._import_price_table)
        toolbar.addWidget(self.btn_import_price_table)

        self.btn_export_price_table = QPushButton("匯出補價表")
        self.btn_export_price_table.clicked.connect(self._export_price_table_template)
        toolbar.addWidget(self.btn_export_price_table)

        btn_select_visible = QPushButton("選取顯示")
        btn_select_visible.clicked.connect(self._select_visible_items)
        toolbar.addWidget(btn_select_visible)

        self.btn_batch_fill_price = QPushButton("批次填價")
        self.btn_batch_fill_price.clicked.connect(self._batch_fill_price)
        toolbar.addWidget(self.btn_batch_fill_price)

        btn_add = QPushButton("新增")
        btn_add.setProperty("role", "primary")
        btn_add.clicked.connect(self._add_item)
        toolbar.addWidget(btn_add)

        btn_copy = QPushButton("複製")
        btn_copy.clicked.connect(self._copy_item)
        toolbar.addWidget(btn_copy)

        btn_delete = QPushButton("刪除")
        btn_delete.setProperty("role", "danger")
        btn_delete.clicked.connect(self._delete_item)
        toolbar.addWidget(btn_delete)

        btn_save = QPushButton("儲存")
        btn_save.setProperty("role", "success")
        btn_save.clicked.connect(self._save_changes)
        toolbar.addWidget(btn_save)

        self.btn_reprice_materials = QPushButton("套用補價")
        self.btn_reprice_materials.clicked.connect(self._reprice_materials)
        toolbar.addWidget(self.btn_reprice_materials)

        root.addLayout(toolbar)

        self.tree = QTreeWidget()
        self.tree.setAlternatingRowColors(True)
        self.tree.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.tree.setHeaderLabels([label for _, label, _ in self.COLUMNS])
        for idx, (_, _, width) in enumerate(self.COLUMNS):
            self.tree.setColumnWidth(idx, width)
        self.tree.header().setSectionResizeMode(len(self.COLUMNS) - 1, QHeaderView.ResizeMode.Stretch)
        self.tree.itemDoubleClicked.connect(self._on_double_click)
        root.addWidget(self.tree, 1)

        bottom = QHBoxLayout()
        self.status_label = QLabel("")
        self.status_label.setFont(Fonts.small())
        self.status_label.setStyleSheet(f"color: {Colors.TEXT_MUTED}; border:none; background:transparent;")
        bottom.addWidget(self.status_label, 1)
        self.modified_label = QLabel("")
        self.modified_label.setStyleSheet(f"color: {Colors.DANGER}; font-weight:bold; border:none; background:transparent;")
        bottom.addWidget(self.modified_label)
        root.addLayout(bottom)

    def load_data(self):
        try:
            data = load_material_pricebook()
            self.items = normalize_pricebook_items(data.get("items", []))
            self.modified = False
            self._apply_filter()
            self._update_labels()
        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"載入價目表失敗：{e}")

    @reentry_guard("_pricebook_import_in_progress", _show_reentry_notice)
    def _import_seed(self):
        if self.modified:
            QMessageBox.warning(self, "尚有未儲存變更", "請先儲存或重新載入價目表，再匯入骨架。")
            return

        path, _ = QFileDialog.getOpenFileName(
            self,
            "匯入材料價目骨架",
            BASE_DIR,
            "JSON Files (*.json);;All Files (*)",
        )
        if not path:
            return

        try:
            _seed_items, report, plan, current = load_and_plan_seed_import(path)
        except Exception as e:
            QMessageBox.critical(self, "匯入失敗", f"讀取或驗證 seed 失敗：\n{e}")
            return

        if report.errors:
            QMessageBox.critical(self, "驗證未通過", self._format_import_validation_message(report))
            return

        if not plan["added"]:
            QMessageBox.information(
                self,
                "沒有可匯入項目",
                self._format_import_seed_confirmation(path, plan, report.warnings, dry_run_only=True),
            )
            return

        if QMessageBox.question(
            self,
            "匯入材料價目骨架",
            self._format_import_seed_confirmation(path, plan, report.warnings),
        ) != QMessageBox.StandardButton.Yes:
            return

        try:
            apply_import_plan(plan, current)
            self.items = plan["items"]
            self.modified = False
            self._apply_filter()
            self._update_labels()
            QMessageBox.information(
                self,
                "匯入完成",
                f"已新增 {len(plan['added'])} 筆材料價目骨架；已存在略過 {len(plan['skipped'])} 筆。",
            )
        except Exception as e:
            QMessageBox.critical(self, "匯入失敗", f"寫入價目表失敗：\n{e}")

    @reentry_guard("_pricebook_table_import_in_progress", _show_reentry_notice)
    def _import_price_table(self):
        if self.modified:
            QMessageBox.warning(self, "尚有未儲存變更", "請先儲存或重新載入價目表，再匯入價格表。")
            return

        path, _ = QFileDialog.getOpenFileName(
            self,
            "匯入材料價格表",
            BASE_DIR,
            "Excel/CSV (*.xlsx *.xlsm *.csv);;All Files (*)",
        )
        if not path:
            return

        try:
            _table_items, report, plan, current = load_and_plan_price_table_import(path)
        except Exception as e:
            QMessageBox.critical(self, "匯入失敗", f"讀取或驗證價格表失敗：\n{e}")
            return

        if report.errors:
            QMessageBox.critical(self, "驗證未通過", self._format_import_validation_message(report))
            return

        if not plan["added"] and not plan["updated"]:
            QMessageBox.information(
                self,
                "沒有可匯入價格",
                self._format_import_price_table_confirmation(path, plan, report.warnings, dry_run_only=True),
            )
            return

        if QMessageBox.question(
            self,
            "匯入材料價格表",
            self._format_import_price_table_confirmation(path, plan, report.warnings),
        ) != QMessageBox.StandardButton.Yes:
            return

        try:
            apply_price_table_import_plan(plan, current)
            self.items = plan["items"]
            self.modified = False
            self._apply_filter()
            self._update_labels()
            QMessageBox.information(
                self,
                "匯入完成",
                (
                    f"已新增 {len(plan['added'])} 筆；"
                    f"補空白單價 {len(plan['updated'])} 筆；"
                    f"價格衝突略過 {len(plan['conflicts'])} 筆。"
                ),
            )
        except Exception as e:
            QMessageBox.critical(self, "匯入失敗", f"寫入價目表失敗：\n{e}")

    @reentry_guard("_pricebook_template_export_in_progress", _show_reentry_notice)
    def _export_price_table_template(self):
        if self.modified and QMessageBox.question(
            self,
            "尚有未儲存變更",
            "目前有未儲存變更；匯出的補價表會包含畫面上的暫存內容。要繼續嗎？",
        ) != QMessageBox.StandardButton.Yes:
            return

        items, scope = self._template_export_items_and_scope()
        rows = build_price_table_template_items(items, only_unpriced=True)
        if not rows:
            QMessageBox.information(self, "沒有可匯出項目", f"{scope} 沒有未定價材料。")
            return

        default_path = os.path.join(BASE_DIR, f"材料補價表_{datetime.now().strftime('%Y%m%d')}.xlsx")
        path, _ = QFileDialog.getSaveFileName(
            self,
            "匯出材料補價表",
            default_path,
            "Excel Files (*.xlsx);;CSV Files (*.csv);;All Files (*)",
        )
        if not path:
            return
        if not os.path.splitext(path)[1]:
            path += ".xlsx"

        try:
            result = export_price_table_template(path, rows, only_unpriced=False)
            QMessageBox.information(
                self,
                "匯出完成",
                f"已匯出 {result['count']} 筆未定價材料。\n來源範圍：{scope}\n{result['path']}",
            )
        except Exception as e:
            QMessageBox.critical(self, "匯出失敗", f"寫出補價表失敗：\n{e}")

    @reentry_guard("_material_reprice_in_progress", _show_reentry_notice)
    def _reprice_materials(self):
        if self.modified:
            QMessageBox.warning(self, "尚有未儲存變更", "請先儲存價目表，再套用補價。")
            return
        try:
            plan = build_project_reprice_plan()
        except Exception as e:
            QMessageBox.critical(self, "重配價失敗", f"讀取材料明細失敗：\n{e}")
            return

        summary = plan["summary"]
        affected_report_ids = plan.get("affected_report_ids", [])
        if summary.get("candidates", 0) == 0:
            QMessageBox.information(
                self,
                "沒有未定價材料",
                self._format_reprice_message(summary, apply=False, affected_report_ids=affected_report_ids),
            )
            return
        if summary.get("matched", 0) == 0 and summary.get("updated", 0) == 0:
            QMessageBox.information(
                self,
                "沒有可套用補價",
                self._format_reprice_message(summary, apply=False, affected_report_ids=affected_report_ids),
            )
            return
        if QMessageBox.question(
            self,
            "套用材料補價",
            self._format_reprice_message(summary, apply=True, affected_report_ids=affected_report_ids),
        ) != QMessageBox.StandardButton.Yes:
            return
        try:
            apply_project_reprice_plan(plan)
            QMessageBox.information(
                self,
                "套用完成",
                self._format_reprice_message(summary, apply=False, affected_report_ids=affected_report_ids),
            )
        except Exception as e:
            QMessageBox.critical(self, "重配價失敗", f"寫入材料明細失敗：\n{e}")

    @staticmethod
    def _format_reprice_message(
        summary: dict,
        *,
        apply: bool,
        affected_report_ids: list[str] | None = None,
    ) -> str:
        lines = [format_reprice_summary(summary)]
        affected_report_ids = affected_report_ids or []
        if affected_report_ids:
            lines.extend(["", "將標記需重產："])
            for report_id in affected_report_ids[:12]:
                lines.append(f"- {report_id}")
            if len(affected_report_ids) > 12:
                lines.append(f"...另有 {len(affected_report_ids) - 12} 張")
        if apply:
            lines.extend([
                "",
                "確認後只會更新未定價材料；手動價與已請款修改單不會被覆蓋。",
                "有金額變更的修改單會在紀錄管理標記為「需重產」。",
            ])
        return "\n".join(lines)

    @staticmethod
    def _format_import_validation_message(report) -> str:
        lines = [
            f"ERROR: {len(report.errors)}    WARNING: {len(report.warnings)}",
            "",
        ]
        for msg in report.errors[:8]:
            lines.append(f"- {msg}")
        if len(report.errors) > 8:
            lines.append(f"...另有 {len(report.errors) - 8} 個 error")
        if report.warnings:
            lines.append("")
            lines.append("WARNING:")
            for msg in report.warnings[:5]:
                lines.append(f"- {msg}")
            if len(report.warnings) > 5:
                lines.append(f"...另有 {len(report.warnings) - 5} 個 warning")
        return "\n".join(lines)

    @staticmethod
    def _format_import_seed_confirmation(
        path: str,
        plan: dict,
        warnings: list[str],
        *,
        dry_run_only: bool = False,
    ) -> str:
        lines = [
            f"來源：{path}",
            "",
            format_import_plan_summary(plan, apply=not dry_run_only),
        ]
        if warnings:
            lines.extend(["", f"WARNING: {len(warnings)}"])
            for msg in warnings[:5]:
                lines.append(f"- {msg}")
            if len(warnings) > 5:
                lines.append(f"...另有 {len(warnings) - 5} 個 warning")
        if not dry_run_only:
            lines.extend(["", "確認後會寫入專案材料價目表；既有項目不會被覆蓋。"])
        return "\n".join(lines)

    @staticmethod
    def _format_import_price_table_confirmation(
        path: str,
        plan: dict,
        warnings: list[str],
        *,
        dry_run_only: bool = False,
    ) -> str:
        lines = [
            f"來源：{path}",
            "",
            format_price_table_import_summary(plan, apply=not dry_run_only),
        ]
        if warnings:
            lines.extend(["", f"WARNING: {len(warnings)}"])
            for msg in warnings[:5]:
                lines.append(f"- {msg}")
            if len(warnings) > 5:
                lines.append(f"...另有 {len(warnings) - 5} 個 warning")
        if plan.get("conflicts"):
            lines.extend(["", "價格衝突會被略過，不會覆蓋既有有價項目。"])
        if not dry_run_only:
            lines.extend([
                "",
                "確認後會寫入專案材料價目表；只新增新項目或補空白單價。",
                "寫入後仍需按「套用補價」才會更新既有未定價材料明細。",
            ])
        return "\n".join(lines)

    def _apply_filter(self):
        self.tree.clear()
        keyword = self.search_edit.text().strip().lower()
        only_unpriced = self.only_unpriced_checkbox.isChecked()
        component_filter = self.component_filter_combo.currentText()
        material_filter = self.material_filter_combo.currentText()
        visible = 0
        unpriced = 0
        for idx, item in enumerate(self.items):
            if self._is_unpriced_item(item):
                unpriced += 1
            if not self._matches_pricebook_filters(
                item,
                keyword=keyword,
                component_filter=component_filter,
                material_filter=material_filter,
                only_unpriced=only_unpriced,
            ):
                continue
            tree_item = QTreeWidgetItem(self.tree, [
                str(item.get(key, "")) for key, _, _ in self.COLUMNS
            ])
            tree_item.setData(0, Qt.ItemDataRole.UserRole, idx)
            if self._is_unpriced_item(item):
                price_col = self._column_index("單價")
                if price_col >= 0:
                    tree_item.setText(price_col, "未定價")
                    tree_item.setForeground(price_col, QBrush(QColor(Colors.WARNING)))
                    tree_item.setToolTip(price_col, "價目表已有材料骨架，但尚未填單價")
            visible += 1
        self.status_label.setText(
            f"共 {len(self.items)} 筆價目"
            + f"，未定價 {unpriced} 筆"
            + (f"，顯示 {visible} 筆" if self._pricebook_filter_active(keyword, component_filter, material_filter, only_unpriced) else "")
            + f"  |  {PRICEBOOK_JSON_PATH}"
        )

    def _select_visible_items(self):
        self.tree.clearSelection()
        for i in range(self.tree.topLevelItemCount()):
            self.tree.topLevelItem(i).setSelected(True)

    @reentry_guard("_pricebook_batch_price_in_progress", _show_reentry_notice)
    def _batch_fill_price(self):
        indices = self._selected_pricebook_indices()
        if not indices:
            QMessageBox.information(self, "提示", "請先選取要填價的價目列。")
            return

        price_text, ok = QInputDialog.getText(
            self,
            "批次填價",
            f"選取 {len(indices)} 筆價目，請輸入單價：",
            text="",
        )
        if not ok:
            return
        price = parse_billing_amount(price_text)
        if price is None:
            QMessageBox.warning(self, "格式錯誤", "單價必須是數字。")
            return

        source, ok = QInputDialog.getItem(
            self,
            "批次填價",
            "價目來源：",
            self.PRICE_SOURCE_OPTIONS,
            0,
            False,
        )
        if not ok:
            return

        effective_date, ok = QInputDialog.getText(
            self,
            "批次填價",
            "生效日（YYYY-MM-DD，可空白）：",
            text=datetime.now().strftime("%Y-%m-%d"),
        )
        if not ok:
            return
        effective_date = effective_date.strip()
        if effective_date and not re.match(r"^\d{4}-\d{2}-\d{2}$", effective_date):
            QMessageBox.warning(self, "格式錯誤", "生效日格式需為 YYYY-MM-DD。")
            return

        if QMessageBox.question(
            self,
            "確認批次填價",
            self._format_batch_fill_price_confirmation(indices, amount_to_text(price), source, effective_date),
        ) != QMessageBox.StandardButton.Yes:
            return

        self._apply_batch_price(indices, amount_to_text(price), source, effective_date)
        self.modified = True
        self._apply_filter()
        self._update_labels()

    def _selected_pricebook_indices(self) -> list[int]:
        indices: set[int] = set()
        for item in self.tree.selectedItems():
            idx = item.data(0, Qt.ItemDataRole.UserRole)
            if idx is None:
                continue
            try:
                idx_int = int(idx)
            except (TypeError, ValueError):
                continue
            if 0 <= idx_int < len(self.items):
                indices.add(idx_int)
        return sorted(indices)

    def _visible_pricebook_indices(self) -> list[int]:
        indices: list[int] = []
        for i in range(self.tree.topLevelItemCount()):
            idx = self.tree.topLevelItem(i).data(0, Qt.ItemDataRole.UserRole)
            if idx is None:
                continue
            try:
                idx_int = int(idx)
            except (TypeError, ValueError):
                continue
            if 0 <= idx_int < len(self.items):
                indices.append(idx_int)
        return indices

    def _template_export_items_and_scope(self) -> tuple[list[dict], str]:
        selected = self._selected_pricebook_indices()
        if selected:
            return [self.items[idx] for idx in selected], f"已選取 {len(selected)} 筆"

        keyword = self.search_edit.text().strip().lower()
        component_filter = self.component_filter_combo.currentText()
        material_filter = self.material_filter_combo.currentText()
        only_unpriced = self.only_unpriced_checkbox.isChecked()
        if self._pricebook_filter_active(keyword, component_filter, material_filter, only_unpriced):
            visible = self._visible_pricebook_indices()
            return [self.items[idx] for idx in visible], f"目前顯示 {len(visible)} 筆"

        return list(self.items), "全部價目"

    def _apply_batch_price(self, indices: list[int], price: str, source: str, effective_date: str) -> None:
        for idx in indices:
            if not (0 <= idx < len(self.items)):
                continue
            self.items[idx]["單價"] = price
            self.items[idx]["來源"] = source
            self.items[idx]["生效日"] = effective_date

    def _format_batch_fill_price_confirmation(
        self,
        indices: list[int],
        price: str,
        source: str,
        effective_date: str,
    ) -> str:
        lines = [
            f"即將批次填價 {len(indices)} 筆價目。",
            f"單價：{price}",
            f"來源：{source or '空白'}",
            f"生效日：{effective_date or '空白'}",
            "",
        ]
        for idx in indices[:10]:
            item = self.items[idx] if 0 <= idx < len(self.items) else {}
            lines.append(
                "- "
                + " / ".join(
                    part for part in [
                        str(item.get("零件類型", "")).strip(),
                        str(item.get("尺寸", "")).strip(),
                        str(item.get("SCH", "")).strip(),
                        str(item.get("材質", "")).strip(),
                    ]
                    if part
                )
            )
        if len(indices) > 10:
            lines.append(f"...還有 {len(indices) - 10} 筆")
        lines.append("")
        lines.append("此動作只修改目前選取的價目列，尚需按「儲存」才會寫入檔案。")
        return "\n".join(lines)

    def _add_item(self):
        self.items.append({
            "id": "",
            "零件類型": "",
            "尺寸": "",
            "SCH": "",
            "材質": "",
            "類別": "材料",
            "單位": "個",
            "單價": "",
            "來源": "合約",
            "生效日": datetime.now().strftime("%Y-%m-%d"),
            "備註": "",
        })
        self.modified = True
        self._apply_filter()
        self._update_labels()

    def _copy_item(self):
        item, idx = self._selected_item()
        if item is None:
            QMessageBox.information(self, "提示", "請先選擇一筆價目")
            return
        copied = dict(self.items[idx])
        copied["id"] = f"{copied.get('id', '')}_copy".strip("_")
        self.items.append(copied)
        self.modified = True
        self._apply_filter()
        self._update_labels()

    def _delete_item(self):
        item, idx = self._selected_item()
        if item is None:
            QMessageBox.information(self, "提示", "請先選擇一筆價目")
            return
        if QMessageBox.question(
            self,
            "刪除價目",
            f"確定刪除此價目？\n\n{item.text(1)} {item.text(2)} {item.text(4)}",
        ) != QMessageBox.StandardButton.Yes:
            return
        self.items.pop(idx)
        self.modified = True
        self._apply_filter()
        self._update_labels()

    def _on_double_click(self, tree_item: QTreeWidgetItem, col: int):
        if col < 0 or col >= len(self.COLUMNS):
            return

        def cb(it, c, val):
            idx = it.data(0, Qt.ItemDataRole.UserRole)
            if idx is None or not (0 <= idx < len(self.items)):
                return
            key = self.COLUMNS[c][0]
            if key == "單價" and val.strip() == "未定價":
                val = ""
            if key == "單價" and val.strip() and parse_billing_amount(val) is None:
                QMessageBox.warning(self, "格式錯誤", "單價必須是數字")
                return
            if key == "生效日" and val.strip() and not re.match(r"^\d{4}-\d{2}-\d{2}$", val.strip()):
                QMessageBox.warning(self, "格式錯誤", "生效日格式需為 YYYY-MM-DD")
                return
            it.setText(c, val)
            self.items[idx][key] = val.strip()
            if key == "零件類型":
                unit_key = "單位"
                unit_col = self._column_index(unit_key)
                if unit_col >= 0 and not str(self.items[idx].get(unit_key, "")).strip():
                    unit = material_default_unit(val)
                    self.items[idx][unit_key] = unit
                    it.setText(unit_col, unit)
            self.modified = True
            self._update_labels()

        key = self.COLUMNS[col][0]
        options = self._controlled_options_for_key(key)
        if options is not None:
            self._editor.start_combo(self.tree, tree_item, col, options, cb)
        elif key == "類別":
            self._editor.start_combo(self.tree, tree_item, col, ["材料", "耗材", "工資", "雜項"], cb)
        elif key == "來源":
            self._editor.start_combo(self.tree, tree_item, col, ["合約", "報價", "手動", "參考"], cb)
        else:
            self._editor.start_entry(self.tree, tree_item, col, cb)

    def _controlled_options_for_key(self, key: str) -> list[str] | None:
        if key not in {"零件類型", "尺寸", "SCH", "材質"}:
            return None
        try:
            constants = load_material_constants()
        except Exception:
            return None
        if key == "零件類型":
            return [""] + list(constants.components)
        if key == "尺寸":
            return [""] + list(constants.sizes)
        if key == "SCH":
            return [""] + list(constants.schedules)
        if key == "材質":
            return [""] + list(constants.materials)
        return None

    def _filter_options_for_key(self, key: str, all_label: str) -> list[str]:
        options = self._controlled_options_for_key(key) or [""]
        return [all_label] + [option for option in options if str(option).strip()]

    def _column_index(self, key: str) -> int:
        for idx, (col_key, _, _) in enumerate(self.COLUMNS):
            if col_key == key:
                return idx
        return -1

    @staticmethod
    def _is_unpriced_item(item: dict) -> bool:
        return bool(str(item.get("零件類型", "")).strip()) and not str(item.get("單價", "")).strip()

    @classmethod
    def _matches_pricebook_filters(
        cls,
        item: dict,
        *,
        keyword: str,
        component_filter: str,
        material_filter: str,
        only_unpriced: bool,
    ) -> bool:
        if only_unpriced and not cls._is_unpriced_item(item):
            return False
        if component_filter and component_filter != "全部零件":
            if str(item.get("零件類型", "")).strip() != component_filter:
                return False
        if material_filter and material_filter != "全部材質":
            if str(item.get("材質", "")).strip() != material_filter:
                return False
        if keyword:
            haystack = " ".join(str(item.get(key, "")) for key, _, _ in cls.COLUMNS).lower()
            if keyword not in haystack:
                return False
        return True

    @staticmethod
    def _pricebook_filter_active(
        keyword: str,
        component_filter: str,
        material_filter: str,
        only_unpriced: bool,
    ) -> bool:
        return bool(
            keyword
            or only_unpriced
            or (component_filter and component_filter != "全部零件")
            or (material_filter and material_filter != "全部材質")
        )

    @reentry_guard("_pricebook_save_in_progress", _show_reentry_notice)
    def _save_changes(self):
        valid_items = []
        seen_ids = set()
        for idx, item in enumerate(normalize_pricebook_items(self.items), start=1):
            if not item.get("零件類型") and not item.get("單價"):
                continue
            if not item.get("零件類型"):
                QMessageBox.warning(self, "資料不足", f"第 {idx} 筆缺少零件類型")
                return
            if item.get("單價") and parse_billing_amount(item.get("單價")) is None:
                QMessageBox.warning(self, "格式錯誤", f"第 {idx} 筆單價不是有效數字")
                return
            if not item.get("單位"):
                QMessageBox.warning(self, "資料不足", f"第 {idx} 筆缺少單位")
                return
            item_id = item.get("id", "")
            if item_id in seen_ids:
                QMessageBox.warning(self, "資料重複", f"料號/ID 重複：{item_id}")
                return
            seen_ids.add(item_id)
            valid_items.append(item)

        try:
            save_material_pricebook({"items": valid_items})
            self.items = valid_items
            self.modified = False
            self._apply_filter()
            self._update_labels()
            QMessageBox.information(self, "成功", "材料價目表已儲存")
        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"儲存價目表失敗：{e}")

    def _selected_item(self) -> tuple[QTreeWidgetItem | None, int]:
        selected = self.tree.selectedItems()
        if not selected:
            return None, -1
        item = selected[0]
        idx = item.data(0, Qt.ItemDataRole.UserRole)
        if idx is None or not (0 <= idx < len(self.items)):
            return None, -1
        return item, idx

    def _update_labels(self):
        self.modified_label.setText("有未儲存變更" if self.modified else "")


# ========= 健康檢查面板 =========
class HealthCheckPanel(QWidget):
    """專案資料夾健康檢查最小 UI。"""

    def __init__(self, parent: QWidget = None):
        super().__init__(parent)
        self._last_guard = None
        self._last_audit = None
        self._build_ui()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(8, 8, 8, 8)
        root.setSpacing(10)

        top_row = QHBoxLayout()
        top_row.setSpacing(8)

        self.summary_label = QLabel("尚未檢查")
        self.summary_label.setFont(Fonts.subheading(11))
        self.summary_label.setStyleSheet(f"color: {Colors.TEXT}; border:none; background:transparent;")
        top_row.addWidget(self.summary_label, 1)

        btn_check = QPushButton("重新檢查")
        btn_check.clicked.connect(self.run_check)
        top_row.addWidget(btn_check)

        self.btn_repair = QPushButton("修復可自動修復項")
        self.btn_repair.setProperty("role", "primary")
        self.btn_repair.setEnabled(False)
        self.btn_repair.clicked.connect(self.repair_project)
        top_row.addWidget(self.btn_repair)

        btn_support = QPushButton("支援診斷包")
        btn_support.setToolTip("產生可交給工程端排查問題的診斷 zip")
        btn_support.clicked.connect(self.create_support_bundle)
        top_row.addWidget(btn_support)

        btn_probe_support = QPushButton("深度診斷包")
        btn_probe_support.setToolTip("產生診斷 zip，並探測 Excel COM / LibreOffice 實際狀態")
        btn_probe_support.clicked.connect(lambda: self.create_support_bundle(probe=True))
        top_row.addWidget(btn_probe_support)

        btn_about = QPushButton("版本資訊")
        btn_about.clicked.connect(self.show_app_info)
        top_row.addWidget(btn_about)

        root.addLayout(top_row)

        self.counts_label = QLabel("")
        self.counts_label.setFont(Fonts.small())
        self.counts_label.setStyleSheet(f"color: {Colors.TEXT_MUTED}; border:none; background:transparent;")
        root.addWidget(self.counts_label)

        self.issue_tree = QTreeWidget()
        self.issue_tree.setAlternatingRowColors(True)
        self.issue_tree.setHeaderLabels(["來源", "等級", "項目", "內容", "參考"])
        for i, width in enumerate([90, 70, 180, 360, 260]):
            self.issue_tree.setColumnWidth(i, width)
        self.issue_tree.header().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        root.addWidget(self.issue_tree, 1)

        self.hint_label = QLabel("健康檢查只會讀取資料；只有按下修復按鈕時，才會建立缺少的必要資料夾或預設設定。")
        self.hint_label.setWordWrap(True)
        self.hint_label.setFont(Fonts.small())
        self.hint_label.setStyleSheet(f"color: {Colors.TEXT_MUTED}; border:none; background:transparent;")
        root.addWidget(self.hint_label)

    def run_check(self):
        from integrity_audit import audit_integrity
        from project_guard import inspect_project

        self.issue_tree.clear()
        try:
            guard = inspect_project(BASE_DIR)
            audit = audit_integrity(BASE_DIR)
        except Exception as exc:
            QMessageBox.critical(self, "健康檢查失敗", str(exc))
            return

        self._last_guard = guard
        self._last_audit = audit

        guard_counts = self._count_guard_severity(guard.issues)
        audit_counts = audit.count_by_severity()
        total_errors = guard_counts["error"] + audit_counts["error"]
        total_warnings = guard_counts["warning"] + audit_counts["warning"]
        total_infos = guard_counts["info"] + audit_counts["info"]

        if total_errors:
            status_text = f"需要人工確認：{total_errors} 個錯誤"
            color = Colors.DANGER
        elif total_warnings or total_infos:
            status_text = f"可使用，有 {total_warnings} 個提醒"
            color = Colors.WARNING
        else:
            status_text = "正常：未發現問題"
            color = Colors.SUCCESS

        self.summary_label.setText(status_text)
        self.summary_label.setStyleSheet(
            f"color: {color}; font-weight:bold; border:none; background:transparent;"
        )
        self.counts_label.setText(
            f"專案：{guard.root}  |  "
            f"records={audit.counts.get('records', 0)}  "
            f"details={audit.counts.get('details', 0)}  "
            f"materials={audit.counts.get('materials', 0)}  "
            f"attachments={audit.counts.get('attachment_folders', 0)}  |  "
            f"error={total_errors} warning={total_warnings} info={total_infos}"
        )

        self.btn_repair.setEnabled(guard.can_auto_repair)

        for issue in guard.issues:
            self._add_issue(
                "啟動守門",
                issue.severity,
                issue.title,
                issue.message,
                issue.path,
                issue.auto_fixable,
            )
        for issue in audit.issues:
            refs = "\n".join(issue.refs[:8])
            if len(issue.refs) > 8:
                refs += f"\n...還有 {len(issue.refs) - 8} 筆"
            self._add_issue("一致性稽核", issue.severity, issue.title, issue.message, refs)

        if not guard.issues and not audit.issues:
            self._add_issue("健康檢查", "info", "正常", "目前沒有發現專案結構或資料一致性問題。", "")

    def repair_project(self):
        from project_guard import inspect_project, repair_project

        guard = self._last_guard or inspect_project(BASE_DIR)
        if not guard.can_auto_repair:
            QMessageBox.information(self, "提示", "目前沒有可自動修復的項目。")
            self.run_check()
            return

        reply = QMessageBox.question(
            self,
            "確認修復",
            "將建立缺少的必要資料夾、預設設定檔或專案識別檔。\n\n是否繼續？",
        )
        if reply != QMessageBox.StandardButton.Yes:
            return

        result = repair_project(BASE_DIR)
        repaired = "\n".join(result.repaired) if result.repaired else "沒有實際變更"
        QMessageBox.information(self, "修復完成", repaired)
        self.run_check()

    def create_support_bundle(self, probe: bool = False):
        from diagnostics import collect_support_bundle

        try:
            result = collect_support_bundle(
                BASE_DIR,
                probe_com_application=probe,
                probe_libreoffice_version=probe,
            )
        except Exception as exc:
            QMessageBox.critical(self, "支援診斷包失敗", str(exc))
            return

        title = "深度診斷包完成" if probe else "支援診斷包完成"
        QMessageBox.information(
            self,
            title,
            "已產生支援診斷包：\n"
            f"{result.get('bundle_path', '')}\n\n"
            f"啟動判斷：{result.get('startup_action', '')}",
        )

    def show_app_info(self):
        from app_info import format_app_identity

        QMessageBox.information(self, "版本資訊", format_app_identity())

    @staticmethod
    def _count_guard_severity(issues) -> dict[str, int]:
        counts = {"error": 0, "warning": 0, "info": 0}
        for issue in issues:
            if issue.severity in counts:
                counts[issue.severity] += 1
        return counts

    def _add_issue(
        self,
        source: str,
        severity: str,
        title: str,
        message: str,
        refs: str,
        auto_fixable: bool = False,
    ):
        label = {
            "error": "錯誤",
            "warning": "提醒",
            "info": "資訊",
        }.get(severity, severity)
        if auto_fixable:
            label += " / 可修"

        item = QTreeWidgetItem(self.issue_tree, [source, label, title, message, refs])
        color = {
            "error": QColor(Colors.DANGER),
            "warning": QColor(Colors.WARNING),
            "info": QColor(Colors.TEXT_MUTED),
        }.get(severity, QColor(Colors.TEXT))
        for col in range(5):
            item.setForeground(col, QBrush(color))


# ========= 請款追蹤面板 =========
class BillingPanel(QWidget):
    """請款追蹤面板 - 管理請款狀態、彙總計算、匯出功能 (PyQt6)"""

    BILLING_STATUS = list(BILLING_STATUS_OPTIONS)

    def __init__(self, parent: QWidget = None):
        super().__init__(parent)
        self.records = []
        self.modified = False
        self._editor = _CellEditor()
        self._batch_store = {"batches": []}
        self._build_ui()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(8, 8, 8, 8)
        root.setSpacing(10)

        # ── 工具列 ──
        toolbar = QHBoxLayout()
        toolbar.setSpacing(10)

        filter_group = QGroupBox("🔍 篩選")
        fg = QHBoxLayout(filter_group)
        fg.setSpacing(8)
        fg.addWidget(QLabel("日期："))
        self.date_from_edit = QLineEdit()
        self.date_from_edit.setPlaceholderText("YYYYMMDD")
        self.date_from_edit.setFixedWidth(95)
        fg.addWidget(self.date_from_edit)
        sep_lbl = QLabel("–")
        sep_lbl.setStyleSheet(f"color: {Colors.TEXT_MUTED}; border:none; background:transparent;")
        fg.addWidget(sep_lbl)
        self.date_to_edit = QLineEdit()
        self.date_to_edit.setPlaceholderText("YYYYMMDD")
        self.date_to_edit.setFixedWidth(95)
        fg.addWidget(self.date_to_edit)
        fg.addSpacing(8)
        fg.addWidget(QLabel("狀態："))
        self.status_filter_combo = QComboBox()
        self.status_filter_combo.addItems(["全部"] + self.BILLING_STATUS)
        self.status_filter_combo.setFixedWidth(90)
        self.status_filter_combo.currentIndexChanged.connect(lambda: self._apply_filter())
        fg.addWidget(self.status_filter_combo)
        btn_load = QPushButton("🔄 載入")
        btn_load.clicked.connect(self.load_data)
        fg.addWidget(btn_load)
        btn_filter = QPushButton("🔍 篩選")
        btn_filter.clicked.connect(self._apply_filter)
        fg.addWidget(btn_filter)
        toolbar.addWidget(filter_group, 1)

        action_group = QGroupBox("📤 操作")
        ag = QHBoxLayout(action_group)
        ag.setSpacing(6)
        btn_save = QPushButton("💾 儲存")
        btn_save.setProperty("role", "primary")
        btn_save.clicked.connect(self._save_changes)
        ag.addWidget(btn_save)
        btn_rpt = QPushButton("📊 匯出報表")
        btn_rpt.clicked.connect(self._export_report)
        ag.addWidget(btn_rpt)
        btn_bill = QPushButton("📋 請款單")
        btn_bill.setProperty("role", "success")
        btn_bill.clicked.connect(self._export_billing)
        ag.addWidget(btn_bill)
        btn_batch = QPushButton("建立批次")
        btn_batch.setProperty("role", "primary")
        btn_batch.setToolTip("將目前選取的修改單建立為請款批次；已在活躍批次中的單會被阻擋")
        btn_batch.clicked.connect(self._create_batch_from_selection)
        ag.addWidget(btn_batch)
        toolbar.addWidget(action_group)
        root.addLayout(toolbar)

        # ── 主列表 ──
        self.tree = QTreeWidget()
        self.tree.setAlternatingRowColors(True)
        self.tree.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        col_config = [
            ("報告編號", 90), ("修改日期", 80), ("Series", 55), ("說明", 150),
            ("請款狀態", 70), ("請款日期", 80), ("焊口金額", 80),
            ("材料金額", 80), ("未稅小計", 85), ("稅額(5%)", 80),
            ("含稅總額", 90), ("備註", 120),
        ]
        self.tree.setHeaderLabels([h for h, _ in col_config])
        for i, (_, w) in enumerate(col_config):
            self.tree.setColumnWidth(i, w)
        self.tree.itemDoubleClicked.connect(self._on_double_click)
        root.addWidget(self.tree, 1)

        # ── 請款批次清單 ──
        batch_group = QGroupBox("請款批次")
        bg = QVBoxLayout(batch_group)
        bg.setSpacing(6)

        batch_toolbar = QHBoxLayout()
        batch_toolbar.setSpacing(8)
        self.batch_summary_label = QLabel("尚未載入批次")
        self.batch_summary_label.setFont(Fonts.small())
        self.batch_summary_label.setStyleSheet(f"color: {Colors.TEXT_MUTED}; border:none; background:transparent;")
        batch_toolbar.addWidget(self.batch_summary_label, 1)

        btn_batch_reload = QPushButton("重新載入批次")
        btn_batch_reload.clicked.connect(self._refresh_batch_tree)
        batch_toolbar.addWidget(btn_batch_reload)

        btn_batch_status = QPushButton("更新狀態")
        btn_batch_status.setProperty("role", "primary")
        btn_batch_status.clicked.connect(self._update_selected_batch_status)
        batch_toolbar.addWidget(btn_batch_status)
        bg.addLayout(batch_toolbar)

        self.batch_tree = QTreeWidget()
        self.batch_tree.setAlternatingRowColors(True)
        self.batch_tree.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        batch_cols = [
            ("批次編號", 170), ("狀態", 80), ("期數", 90), ("業主", 120),
            ("張數", 55), ("建立時間", 140), ("更新時間", 140),
        ]
        self.batch_tree.setHeaderLabels([h for h, _ in batch_cols])
        for i, (_, w) in enumerate(batch_cols):
            self.batch_tree.setColumnWidth(i, w)
        self.batch_tree.itemDoubleClicked.connect(self._on_batch_double_click)
        self.batch_tree.setMaximumHeight(150)
        bg.addWidget(self.batch_tree)
        root.addWidget(batch_group)

        # ── 統計 ──
        stat_group = QGroupBox("📈 彙總統計")
        sg = QVBoxLayout(stat_group)
        sg.setSpacing(8)

        # 第一行：統計卡片
        cards_row = QHBoxLayout()
        cards_row.setSpacing(10)

        self._stat_total_card = make_stat_card("總筆數", "0", Colors.TEXT)
        cards_row.addWidget(self._stat_total_card)
        self._stat_pending_card = make_stat_card("未請款", "0", Colors.DANGER)
        cards_row.addWidget(self._stat_pending_card)
        self._stat_billed_card = make_stat_card("已請款", "0", Colors.SUCCESS)
        cards_row.addWidget(self._stat_billed_card)
        self._stat_closed_card = make_stat_card("已結案", "0", Colors.TEXT_MUTED)
        cards_row.addWidget(self._stat_closed_card)
        sg.addLayout(cards_row)

        # 第二行：金額
        amt_row = QHBoxLayout()
        amt_row.setSpacing(10)
        self._stat_weld_card = make_stat_card("焊口總額", "$0", Colors.PRIMARY)
        amt_row.addWidget(self._stat_weld_card)
        self._stat_mat_card = make_stat_card("材料總額", "$0", Colors.INFO)
        amt_row.addWidget(self._stat_mat_card)
        self._stat_pend_amt_card = make_stat_card("未請款額", "$0", Colors.DANGER)
        amt_row.addWidget(self._stat_pend_amt_card)
        self._stat_bill_amt_card = make_stat_card("已請款額", "$0", Colors.SUCCESS)
        amt_row.addWidget(self._stat_bill_amt_card)
        sg.addLayout(amt_row)

        mod_row = QHBoxLayout()
        mod_row.addStretch()
        self.modified_label = QLabel("")
        self.modified_label.setStyleSheet(f"color: {Colors.DANGER}; font-weight:bold; border:none; background:transparent;")
        mod_row.addWidget(self.modified_label)
        sg.addLayout(mod_row)
        root.addWidget(stat_group)

    # ────────────────── 載入 ──────────────────
    def load_data(self):
        self.records.clear()
        self.tree.clear()
        try:
            store = _load_store()
            billing = self._load_billing()
            self.records.extend(build_billing_rows(store, billing))
            unresolved_counts = unresolved_material_counts_by_report(store.get("materials", []))
            for row in self.records:
                counts = unresolved_counts.get(str(row.get("report_id", "")).strip(), {})
                row["unresolved_material_total"] = str(counts.get("total", 0))
                row["missing_price_count"] = str(counts.get("missing_price", 0))
                row["missing_pricebook_count"] = str(counts.get("missing_pricebook", 0))
            self._apply_filter()
            self._refresh_batch_tree()
            self._update_statistics()
        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"載入失敗：{e}")

    def _load_billing(self) -> dict:
        """載入 billing.json，回傳 {report_id: {...}}"""
        if not os.path.exists(BILLING_JSON_PATH):
            # 嘗試從舊 Excel 遷移
            self._migrate_billing_from_excel()
        if os.path.exists(BILLING_JSON_PATH):
            import json
            with open(BILLING_JSON_PATH, "r", encoding="utf-8") as f:
                data = json.load(f)
            return data.get("billing", {})
        return {}

    def _migrate_billing_from_excel(self):
        """從舊 Excel 的請款欄位遷移到 billing.json"""
        if not os.path.exists(RECORD_XLSX_PATH) or not OPENPYXL_AVAILABLE:
            return
        try:
            wb = load_workbook(RECORD_XLSX_PATH, read_only=True, data_only=True)
            if 'record' not in wb.sheetnames:
                wb.close()
                return
            ws = wb['record']
            headers = [c.value for c in ws[1]]
            col_map = {h: i for i, h in enumerate(headers) if h}
            # 需要有請款欄位才遷移
            if "請款狀態" not in col_map:
                wb.close()
                return
            billing = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                rid = str(row[col_map.get("報告編號", 1)] or "")
                if not rid:
                    continue
                status = row[col_map.get("請款狀態")] or ""
                if status or row[col_map.get("焊口金額", 17)]:
                    billing[rid] = {
                        "status": str(status),
                        "billing_date": str(row[col_map.get("請款日期", 16)] or ""),
                        "weld_amount": str(row[col_map.get("焊口金額", 17)] or ""),
                        "material_amount": str(row[col_map.get("材料金額", 18)] or ""),
                        "total": str(row[col_map.get("總金額", 19)] or ""),
                        "remark": str(row[col_map.get("備註", 20)] or ""),
                    }
            wb.close()
            if billing:
                self._save_billing_json(billing)
        except Exception:
            pass

    def _load_active_batch_index(self) -> dict[str, str]:
        try:
            return active_batch_index(load_billing_batches())
        except Exception:
            return {}

    def _load_batch_store(self) -> dict:
        self._batch_store = load_billing_batches()
        return self._batch_store

    def _save_billing_json(self, billing: dict):
        """儲存 billing.json"""
        os.makedirs(os.path.dirname(BILLING_JSON_PATH), exist_ok=True)
        auto_backup(BILLING_JSON_PATH, max_backups=20)
        data = {
            "billing": billing,
            "meta": {
                "version": "1.0",
                "last_modified": datetime.now().isoformat(),
                "currency": BILLING_CURRENCY,
                "tax_mode": BILLING_TAX_MODE,
                "tax_rate": tax_rate_to_text(),
                "rounding_rule": BILLING_ROUNDING_RULE,
            }
        }
        atomic_write_json(BILLING_JSON_PATH, data)

    # ────────────────── 篩選 ──────────────────
    def _apply_filter(self):
        self.tree.clear()
        date_from = self.date_from_edit.text().strip()
        date_to = self.date_to_edit.text().strip()
        status_filter = self.status_filter_combo.currentText()
        active_batches = self._load_active_batch_index()

        for r in self.records:
            if date_from and str(r["date"]) < date_from:
                continue
            if date_to and str(r["date"]) > date_to:
                continue
            row_status = normalize_billing_status(r.get("status", ""))
            if status_filter != "全部":
                if row_status != status_filter:
                    continue
            desc_s = str(r["desc"])
            item = QTreeWidgetItem(self.tree, [
                str(r["report_id"]), str(r["date"]), str(r["series"]),
                desc_s[:20] + "..." if len(desc_s) > 20 else desc_s,
                row_status, str(r["billing_date"]),
                self._format_amount(r["weld_amount"]),
                self._format_amount(r["material_amount"]),
                self._format_amount(r["total"]),
                self._format_amount(r.get("tax_amount", "")),
                self._format_amount(r.get("grand_total", "")),
                str(r["remark"]),
            ])
            active_batch_id = active_batches.get(str(r["report_id"]))
            if active_batch_id:
                item.setForeground(0, QBrush(QColor(Colors.WARNING)))
                item.setToolTip(0, f"已在活躍請款批次：{active_batch_id}")
                item.setToolTip(4, f"此修改單已在活躍請款批次 {active_batch_id}，不可重複加入新批次")
            if int(str(r.get("unresolved_material_total", "0") or "0")):
                tooltip = self._format_unresolved_material_tooltip(r)
                item.setForeground(0, QBrush(QColor(Colors.DANGER)))
                item.setForeground(7, QBrush(QColor(Colors.DANGER)))
                item.setToolTip(0, tooltip)
                item.setToolTip(7, tooltip)
            for col, source_field in [
                (6, "weld_amount_source"),
                (7, "material_amount_source"),
            ]:
                source = r.get(source_field, "")
                if source == "calculated":
                    item.setForeground(col, QBrush(QColor(Colors.INFO)))
                    item.setToolTip(col, "由焊口/材料明細自動計算；雙擊可手動覆蓋")
                elif source == "manual":
                    item.setToolTip(col, "手動覆蓋金額")
            if r.get("total_mismatch") == "1":
                item.setForeground(8, QBrush(QColor(Colors.WARNING)))
                diff = r.get("total_mismatch_amount", "")
                manual_total = r.get("manual_total", "")
                tip = f"未稅小計由焊口金額 + 材料金額計算；舊手填總額 {manual_total or '空白'} 已不再使用"
                if diff:
                    tip += f"，差額 {diff}"
                item.setToolTip(8, tip)
            elif r.get("total_source") == "calculated":
                item.setForeground(8, QBrush(QColor(Colors.INFO)))
                item.setToolTip(8, "未稅小計由焊口金額 + 材料金額自動計算，不可手動覆蓋")
            item.setForeground(9, QBrush(QColor(Colors.INFO)))
            item.setToolTip(9, "稅額依未稅小計 x 5% 外加計算，四捨五入到元")
            item.setForeground(10, QBrush(QColor(Colors.INFO)))
            item.setToolTip(10, "含稅總額 = 未稅小計 + 稅額")
        self._update_statistics()

    # ────────────────── 統計 ──────────────────
    @staticmethod
    def _format_amount(val):
        if not val:
            return ""
        text = money_to_text(val)
        if text:
            return f"${int(text):,}"
        if parse_billing_amount(val) == Decimal("0"):
            return ""
        return str(val)

    @staticmethod
    def _parse_amount(val):
        amount = parse_billing_amount(val)
        return float(amount) if amount is not None else 0

    def _update_statistics(self):
        tc = pc = bc = cc = 0
        wt = mt = pa = ba = 0.0
        for i in range(self.tree.topLevelItemCount()):
            it = self.tree.topLevelItem(i)
            tc += 1
            status = normalize_billing_status(it.text(4))
            wa = self._parse_amount(it.text(6))
            ma = self._parse_amount(it.text(7))
            ta = self._parse_amount(it.text(10)) or self._parse_amount(it.text(8)) or (wa + ma)
            wt += wa
            mt += ma
            if status in ["未請款", "暫緩", "退回", "作廢"]:
                pc += 1; pa += ta
            elif status in ["請款中", "已請款", "部分付款", "已付款", "補件中"]:
                bc += 1; ba += ta
            elif status == "已結案":
                cc += 1; ba += ta

        # 更新統計卡片
        def _set_card_value(card, value_text):
            v_lbl = card.layout().itemAt(0).widget()
            if v_lbl:
                v_lbl.setText(value_text)

        _set_card_value(self._stat_total_card, str(tc))
        _set_card_value(self._stat_pending_card, str(pc))
        _set_card_value(self._stat_billed_card, str(bc))
        _set_card_value(self._stat_closed_card, str(cc))
        _set_card_value(self._stat_weld_card, f"${wt:,.0f}")
        _set_card_value(self._stat_mat_card, f"${mt:,.0f}")
        _set_card_value(self._stat_pend_amt_card, f"${pa:,.0f}")
        _set_card_value(self._stat_bill_amt_card, f"${ba:,.0f}")

    # ────────────────── 雙擊編輯 ──────────────────
    def _on_double_click(self, item: QTreeWidgetItem, col: int):
        editable_map = {
            4: ("status", "combo"),
            5: ("billing_date", "entry"),
            6: ("weld_amount", "entry"),
            7: ("material_amount", "entry"),
            11: ("remark", "entry"),
        }
        if col not in editable_map:
            return
        field, etype = editable_map[col]

        def cb(it, c, val):
            if c in [6, 7] and val:
                try:
                    val = f"${float(val):,.0f}"
                except Exception:
                    pass
            it.setText(c, val)
            rid = it.text(0)
            for r in self.records:
                if str(r["report_id"]) == rid:
                    raw = str(val).replace("$", "").replace(",", "") if c in [6, 7] else val
                    if field == "status":
                        raw = normalize_billing_status(raw)
                        it.setText(c, raw)
                    r[field] = raw
                    if field in ("weld_amount", "material_amount"):
                        r[f"{field}_source"] = "manual" if raw else "empty"
                        self._refresh_total_for_row(r, it)
                    break
            self.modified = True
            self.modified_label.setText("⚠️ 有未儲存的變更")
            self._update_statistics()

        if etype == "combo":
            self._editor.start_combo(self.tree, item, col, self.BILLING_STATUS, cb)
        else:
            self._editor.start_entry(self.tree, item, col, cb)

    # ────────────────── 請款批次 ──────────────────
    @reentry_guard("_billing_batch_create_in_progress", _show_reentry_notice)
    def _create_batch_from_selection(self):
        if self.modified:
            QMessageBox.warning(self, "請先儲存", "請先儲存目前請款變更，再建立請款批次。")
            return

        report_ids = self._selected_report_ids()
        if not report_ids:
            QMessageBox.information(self, "提示", "請先在請款表選取要加入批次的修改單。")
            return

        blocked_statuses = self._selected_batch_blocked_statuses()
        if blocked_statuses:
            QMessageBox.warning(
                self,
                "無法建立請款批次",
                self._format_batch_status_blocks(blocked_statuses),
            )
            return

        unresolved = self._selected_unresolved_materials()
        if unresolved:
            QMessageBox.warning(
                self,
                "材料尚未定價",
                self._format_unresolved_material_blocks(unresolved),
            )
            return

        active_batches = self._load_active_batch_index()
        conflicts = {
            report_id: active_batches[report_id]
            for report_id in report_ids
            if report_id in active_batches
        }
        if conflicts:
            QMessageBox.warning(
                self,
                "無法建立請款批次",
                self._format_batch_conflicts(conflicts),
            )
            return

        period_default = datetime.now().strftime("%Y-%m")
        period, ok = QInputDialog.getText(
            self,
            "請款批次",
            "估驗期數 / 期間：",
            text=period_default,
        )
        if not ok:
            return

        client, ok = QInputDialog.getText(
            self,
            "請款批次",
            "業主 / 請款對象（可空白）：",
            text="",
        )
        if not ok:
            return

        if QMessageBox.question(
            self,
            "確認建立請款批次",
            self._format_batch_create_confirmation(report_ids, period, client),
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        ) != QMessageBox.StandardButton.Yes:
            return

        try:
            batch = create_billing_batch(
                report_ids,
                period=period,
                client=client,
            )
            self.load_data()
            QMessageBox.information(
                self,
                "請款批次已建立",
                self._format_batch_created_message(batch),
            )
        except BillingBatchError as exc:
            QMessageBox.warning(
                self,
                "無法建立請款批次",
                self._format_batch_issues(exc.issues),
            )
        except Exception as exc:
            QMessageBox.critical(self, "錯誤", f"建立請款批次失敗：{exc}")

    def _refresh_batch_tree(self):
        if not hasattr(self, "batch_tree"):
            return
        self.batch_tree.clear()
        try:
            store = self._load_batch_store()
        except Exception as exc:
            self.batch_summary_label.setText(f"批次載入失敗：{exc}")
            return

        batches = list(store.get("batches", []) or [])
        batches.sort(key=lambda b: str(b.get("created_at", "")), reverse=True)
        active_count = 0
        active_report_count = 0

        for batch in batches:
            report_ids = batch_report_ids(batch)
            status = str(batch.get("status", "") or "草稿")
            if is_active_batch_status(status):
                active_count += 1
                active_report_count += len(report_ids)
            item = QTreeWidgetItem(self.batch_tree, [
                str(batch.get("batch_id", "")),
                status,
                str(batch.get("period", "")),
                str(batch.get("client", "")),
                str(len(report_ids)),
                str(batch.get("created_at", "")),
                str(batch.get("updated_at", "")),
            ])
            item.setData(0, Qt.ItemDataRole.UserRole, str(batch.get("batch_id", "")))
            if status in {"已結案", "作廢"}:
                for col in range(7):
                    item.setForeground(col, QBrush(QColor(Colors.TEXT_MUTED)))
            elif status == "草稿":
                item.setForeground(1, QBrush(QColor(Colors.INFO)))
            else:
                item.setForeground(1, QBrush(QColor(Colors.WARNING)))
            tooltip = self._format_batch_detail_tooltip(batch)
            for col in range(7):
                item.setToolTip(col, tooltip)

        self.batch_summary_label.setText(
            f"共 {len(batches)} 批，活躍 {active_count} 批，鎖定 {active_report_count} 張修改單"
        )

    def _on_batch_double_click(self, item: QTreeWidgetItem, col: int):
        if col == 1:
            self._update_selected_batch_status()

    @reentry_guard("_billing_batch_status_in_progress", _show_reentry_notice)
    def _update_selected_batch_status(self):
        batch_id = self._selected_batch_id()
        if not batch_id:
            QMessageBox.information(self, "提示", "請先選擇一個請款批次。")
            return
        batch = self._find_batch(batch_id)
        if not batch:
            QMessageBox.warning(self, "找不到批次", f"找不到請款批次：{batch_id}")
            return

        old_status = str(batch.get("status", "") or "草稿")
        new_status, ok = QInputDialog.getItem(
            self,
            "更新請款批次狀態",
            f"批次 {batch_id}\n目前狀態：{old_status}\n新狀態：",
            list(BATCH_STATUS_OPTIONS),
            list(BATCH_STATUS_OPTIONS).index(old_status) if old_status in BATCH_STATUS_OPTIONS else 0,
            False,
        )
        if not ok or new_status == old_status:
            return

        if QMessageBox.question(
            self,
            "確認批次狀態變更",
            self._format_batch_status_confirmation(batch, new_status),
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        ) != QMessageBox.StandardButton.Yes:
            return

        try:
            updated = update_billing_batch_status(self._batch_store, batch_id, new_status)
            save_billing_batches(updated)
            self._batch_store = updated
            self._apply_filter()
            self._refresh_batch_tree()
            QMessageBox.information(
                self,
                "批次狀態已更新",
                f"{batch_id}\n{old_status} → {new_status}",
            )
        except BillingBatchError as exc:
            QMessageBox.warning(self, "無法更新批次狀態", self._format_batch_issues(exc.issues))
        except Exception as exc:
            QMessageBox.critical(self, "錯誤", f"更新批次狀態失敗：{exc}")

    def _selected_batch_id(self) -> str:
        selected = self.batch_tree.selectedItems() if hasattr(self, "batch_tree") else []
        if not selected:
            return ""
        return str(selected[0].data(0, Qt.ItemDataRole.UserRole) or selected[0].text(0)).strip()

    def _find_batch(self, batch_id: str) -> dict | None:
        store = self._batch_store or {}
        for batch in store.get("batches", []) or []:
            if str(batch.get("batch_id", "")) == str(batch_id):
                return batch
        try:
            store = self._load_batch_store()
        except Exception:
            return None
        for batch in store.get("batches", []) or []:
            if str(batch.get("batch_id", "")) == str(batch_id):
                return batch
        return None

    def _selected_report_ids(self) -> list[str]:
        return self._unique_report_ids_from_values(
            item.text(0) for item in self.tree.selectedItems()
        )

    def _selected_batch_blocked_statuses(self) -> dict[str, str]:
        blocked: dict[str, str] = {}
        for item in self.tree.selectedItems():
            report_id = item.text(0).strip()
            status = normalize_billing_status(item.text(4))
            if report_id and status in {"已結案", "作廢"}:
                blocked[report_id] = status
        return blocked

    def _selected_unresolved_materials(self) -> dict[str, dict[str, str]]:
        selected_ids = set(self._selected_report_ids())
        unresolved: dict[str, dict[str, str]] = {}
        for row in self.records:
            report_id = str(row.get("report_id", "")).strip()
            if report_id not in selected_ids:
                continue
            total = int(str(row.get("unresolved_material_total", "0") or "0"))
            if total:
                unresolved[report_id] = {
                    "total": str(total),
                    "missing_price": str(row.get("missing_price_count", "0") or "0"),
                    "missing_pricebook": str(row.get("missing_pricebook_count", "0") or "0"),
                }
        return unresolved

    @staticmethod
    def _unique_report_ids_from_values(values) -> list[str]:
        seen: set[str] = set()
        report_ids: list[str] = []
        for value in values or []:
            report_id = str(value).strip()
            if not report_id or report_id in seen:
                continue
            seen.add(report_id)
            report_ids.append(report_id)
        return report_ids

    @staticmethod
    def _format_batch_conflicts(conflicts: dict[str, str]) -> str:
        lines = [
            "以下修改單已經在活躍請款批次中，不能重複加入：",
            "",
        ]
        for report_id, batch_id in list(conflicts.items())[:12]:
            lines.append(f"- {report_id}: {batch_id}")
        if len(conflicts) > 12:
            lines.append(f"...還有 {len(conflicts) - 12} 筆")
        lines.append("")
        lines.append("請先將原批次結案或作廢，再建立新的請款批次。")
        return "\n".join(lines)

    @staticmethod
    def _format_batch_status_blocks(blocked: dict[str, str]) -> str:
        lines = [
            "以下修改單狀態不可加入新的請款批次：",
            "",
        ]
        for report_id, status in list(blocked.items())[:12]:
            lines.append(f"- {report_id}: {status}")
        if len(blocked) > 12:
            lines.append(f"...還有 {len(blocked) - 12} 筆")
        return "\n".join(lines)

    @staticmethod
    def _format_unresolved_material_tooltip(row: dict) -> str:
        total = int(str(row.get("unresolved_material_total", "0") or "0"))
        missing_price = int(str(row.get("missing_price_count", "0") or "0"))
        missing_pricebook = int(str(row.get("missing_pricebook_count", "0") or "0"))
        lines = [f"此修改單有 {total} 筆材料尚未可請款。"]
        if missing_price:
            lines.append(f"未定價（待補價）：{missing_price} 筆")
        if missing_pricebook:
            lines.append(f"查無價目（待建料）：{missing_pricebook} 筆")
        lines.append("請先補價/建料並套用補價，再建立請款批次。")
        return "\n".join(lines)

    @staticmethod
    def _format_unresolved_material_blocks(unresolved: dict[str, dict[str, str]]) -> str:
        lines = [
            "以下修改單仍有材料未定價，暫不允許加入請款批次：",
            "",
        ]
        for report_id, counts in list(unresolved.items())[:12]:
            lines.append(
                f"- {report_id}: 共 {counts.get('total', '0')} 筆"
                f"（待補價 {counts.get('missing_price', '0')}、待建料 {counts.get('missing_pricebook', '0')}）"
            )
        if len(unresolved) > 12:
            lines.append(f"...還有 {len(unresolved) - 12} 張")
        lines.append("")
        lines.append("請先到材料價目表補價/建料並套用補價，再建立請款批次。")
        return "\n".join(lines)

    @staticmethod
    def _format_batch_issues(issues) -> str:
        lines = ["請款批次尚未建立：", ""]
        for issue in issues[:12]:
            lines.append(f"- {issue.message}")
        if len(issues) > 12:
            lines.append(f"...還有 {len(issues) - 12} 筆")
        return "\n".join(lines)

    @staticmethod
    def _format_batch_create_confirmation(report_ids: list[str], period: str, client: str) -> str:
        lines = [
            f"即將建立請款批次，共 {len(report_ids)} 張修改單。",
            f"估驗期數 / 期間：{period or '空白'}",
            f"業主 / 請款對象：{client or '空白'}",
            "",
        ]
        for report_id in report_ids[:12]:
            lines.append(f"- {report_id}")
        if len(report_ids) > 12:
            lines.append(f"...還有 {len(report_ids) - 12} 筆")
        lines.append("")
        lines.append("建立後，這些修改單會被活躍批次鎖住，避免重複請款。")
        return "\n".join(lines)

    @staticmethod
    def _format_batch_created_message(batch: dict) -> str:
        report_ids = [
            str(item.get("report_id", ""))
            for item in batch.get("items", [])
            if isinstance(item, dict) and item.get("report_id")
        ]
        lines = [
            f"批次編號：{batch.get('batch_id', '')}",
            f"狀態：{batch.get('status', '')}",
            f"修改單數量：{len(report_ids)}",
        ]
        if batch.get("period"):
            lines.append(f"估驗期數 / 期間：{batch.get('period')}")
        if batch.get("client"):
            lines.append(f"業主 / 請款對象：{batch.get('client')}")
        return "\n".join(lines)

    @staticmethod
    def _format_batch_status_confirmation(batch: dict, new_status: str) -> str:
        report_ids = batch_report_ids(batch)
        lines = [
            f"批次編號：{batch.get('batch_id', '')}",
            f"狀態：{batch.get('status', '')} → {new_status}",
            f"修改單數量：{len(report_ids)}",
        ]
        if batch.get("period"):
            lines.append(f"估驗期數 / 期間：{batch.get('period')}")
        if batch.get("client"):
            lines.append(f"業主 / 請款對象：{batch.get('client')}")
        lines.append("")
        lines.append("此動作只更新批次狀態，不會自動修改各修改單的請款狀態。")
        return "\n".join(lines)

    @staticmethod
    def _format_batch_detail_tooltip(batch: dict) -> str:
        report_ids = batch_report_ids(batch)
        lines = [
            f"批次：{batch.get('batch_id', '')}",
            f"狀態：{batch.get('status', '')}",
            f"修改單：{len(report_ids)} 張",
        ]
        if batch.get("period"):
            lines.append(f"期數：{batch.get('period')}")
        if batch.get("client"):
            lines.append(f"業主：{batch.get('client')}")
        if report_ids:
            lines.append("")
            for report_id in report_ids[:8]:
                lines.append(f"- {report_id}")
            if len(report_ids) > 8:
                lines.append(f"...還有 {len(report_ids) - 8} 張")
        return "\n".join(lines)

    # ────────────────── 儲存 ──────────────────
    @reentry_guard("_billing_save_in_progress", _show_reentry_notice)
    def _save_changes(self):
        if not self.modified:
            QMessageBox.information(self, "提示", "沒有需要儲存的變更")
            return
        journal = None
        try:
            from operation_journal import OperationJournal

            journal = OperationJournal(
                BASE_DIR,
                "billing_save_changes",
                {
                    "visible_rows": self.tree.topLevelItemCount(),
                    "records": len(self.records),
                    "legacy_manual_total_count": sum(
                        1 for r in self.records if r.get("manual_total")
                    ),
                },
            )
            old_billing = self._load_billing()
            billing = self._build_billing_payload()
            status_issues = validate_billing_status_changes(old_billing, billing)
            if status_issues:
                QMessageBox.warning(
                    self,
                    "請款狀態不可直接跳轉",
                    self._format_status_issues(status_issues),
                )
                return

            audit_events = build_billing_change_events(
                old_billing,
                billing,
                operation_id=journal.journal_id,
            )
            if not self._confirm_sensitive_billing_changes(audit_events):
                return

            journal.begin()
            journal.step("load_existing_billing", row_count=len(old_billing))
            journal.step(
                "prepare_billing_json",
                row_count=len(billing),
                audit_events=len(audit_events),
            )
            self._save_billing_json(billing)
            journal.step("save_billing_json", path=BILLING_JSON_PATH, row_count=len(billing))
            audit_path, audit_count = append_billing_audit(BASE_DIR, audit_events)
            if audit_count:
                journal.step(
                    "append_billing_audit",
                    path=str(audit_path),
                    event_count=audit_count,
                )
            journal.complete()
            journal = None
            self.modified = False
            self.modified_label.setText("")
            self.load_data()
            QMessageBox.information(self, "成功", "變更已儲存")
        except Exception as e:
            if journal is not None:
                journal.fail(str(e))
            QMessageBox.critical(self, "錯誤", f"儲存失敗：{e}")

    def _build_billing_payload(self) -> dict:
        billing = {}
        for r in self.records:
            rid = r["report_id"]
            if not rid:
                continue
            billing[rid] = {
                "status": r.get("status", ""),
                "billing_date": r.get("billing_date", ""),
                "weld_amount": self._billing_amount_for_save(r, "weld_amount"),
                "material_amount": self._billing_amount_for_save(r, "material_amount"),
                "total": self._billing_amount_for_save(r, "total"),
                "remark": r.get("remark", ""),
            }
        return billing

    @staticmethod
    def _format_status_issues(issues) -> str:
        lines = ["以下請款狀態轉換不符合狀態機規則，尚未儲存：", ""]
        for issue in issues[:10]:
            lines.append(f"- {issue.report_id}: {issue.message}")
        if len(issues) > 10:
            lines.append(f"...還有 {len(issues) - 10} 筆")
        lines.append("")
        lines.append("請依序走：未請款 → 請款中 → 已請款 → 已付款/已結案。")
        return "\n".join(lines)

    def _confirm_sensitive_billing_changes(self, audit_events: list[dict]) -> bool:
        sensitive_events = [
            event for event in audit_events
            if {"amount", "status"} & set(event.get("change_types", []))
        ]
        if not sensitive_events:
            return True

        message = self._format_billing_confirmation(sensitive_events)
        return QMessageBox.question(
            self,
            "確認請款變更",
            message,
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        ) == QMessageBox.StandardButton.Yes

    @staticmethod
    def _format_billing_confirmation(events: list[dict]) -> str:
        labels = {
            "status": "狀態",
            "billing_date": "請款日期",
            "weld_amount": "焊口金額",
            "material_amount": "材料金額",
            "total": "未稅小計",
            "remark": "備註",
        }
        lines = [
            "以下請款狀態或金額將被寫入，並留下 billing_audit 稽核紀錄：",
            "",
        ]
        for event in events[:8]:
            parts = []
            for field, change in event.get("changes", {}).items():
                if field not in ("status", "weld_amount", "material_amount", "total"):
                    continue
                old = change.get("old", "") or "空白"
                new = change.get("new", "") or "空白"
                parts.append(f"{labels.get(field, field)} {old} → {new}")
            if parts:
                lines.append(f"- {event.get('report_id', '')}: " + "；".join(parts))
        if len(events) > 8:
            lines.append(f"...還有 {len(events) - 8} 筆")
        lines.append("")
        lines.append("確定要儲存這些請款變更嗎？")
        return "\n".join(lines)

    @staticmethod
    def _billing_amount_for_save(row: dict, field: str) -> str:
        if field == "total":
            return ""
        if row.get(f"{field}_source") == "calculated":
            return ""
        return row.get(field, "")

    def _refresh_total_for_row(self, row: dict, item: QTreeWidgetItem):
        weld = parse_billing_amount(row.get("weld_amount")) or Decimal("0")
        material = parse_billing_amount(row.get("material_amount")) or Decimal("0")
        row["total"] = money_to_text(weld + material)
        row["total_source"] = "calculated" if row["total"] else "empty"
        subtotal = parse_billing_amount(row.get("total")) or Decimal("0")
        tax = calculate_tax_amount(subtotal)
        grand_total = subtotal + tax
        row["subtotal"] = row["total"]
        row["tax_rate"] = tax_rate_to_text()
        row["tax_amount"] = money_to_text(tax)
        row["grand_total"] = money_to_text(grand_total)
        row["tax_mode"] = BILLING_TAX_MODE
        row["currency"] = BILLING_CURRENCY
        row["rounding_rule"] = BILLING_ROUNDING_RULE
        row["manual_total"] = ""
        row["total_mismatch"] = ""
        row["total_mismatch_amount"] = ""
        item.setText(8, self._format_amount(row["total"]))
        item.setText(9, self._format_amount(row["tax_amount"]))
        item.setText(10, self._format_amount(row["grand_total"]))
        if row["total"]:
            item.setForeground(8, QBrush(QColor(Colors.INFO)))
            item.setToolTip(8, "未稅小計由焊口金額 + 材料金額自動計算，不可手動覆蓋")
        else:
            item.setToolTip(8, "")
        item.setForeground(9, QBrush(QColor(Colors.INFO)))
        item.setToolTip(9, "稅額依未稅小計 x 5% 外加計算，四捨五入到元")
        item.setForeground(10, QBrush(QColor(Colors.INFO)))
        item.setToolTip(10, "含稅總額 = 未稅小計 + 稅額")

    # ────────────────── 匯出 ──────────────────
    @reentry_guard("_billing_export_in_progress", _show_reentry_notice)
    def _export_report(self):
        fp, _ = QFileDialog.getSaveFileName(
            self, "匯出統計報表", f"請款統計報表_{datetime.now().strftime('%Y%m%d')}.xlsx",
            "Excel 檔案 (*.xlsx)")
        if not fp:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            from record_manager import _apply_excel_styles

            wb = Workbook()
            ws = wb.active
            ws.title = "請款統計"
            ws.sheet_properties.tabColor = "1F3864"

            hds = ["報告編號", "修改日期", "Series", "說明", "請款狀態",
                   "請款日期", "焊口金額", "材料金額", "未稅小計",
                   "稅額", "含稅總額", "備註"]

            # 表頭 (row 2), row 1 留給標題
            for i, h in enumerate(hds, 1):
                ws.cell(row=2, column=i, value=h)

            ri = 3
            for idx in range(self.tree.topLevelItemCount()):
                it = self.tree.topLevelItem(idx)
                for ci in range(12):
                    val = it.text(ci)
                    # 金額欄嘗試轉為數值
                    if ci in (6, 7, 8, 9, 10):
                        try:
                            val = float(val.replace('$', '').replace(',', '')) if val else ""
                        except (ValueError, TypeError):
                            pass
                    ws.cell(row=ri, column=ci + 1, value=val)
                ri += 1

            data_end = max(3, ri - 1)
            _apply_excel_styles(
                ws, header_row=2, data_start=3, data_end=data_end,
                col_count=12,
                col_widths=[14, 11, 10, 32, 10, 11, 14, 14, 14, 12, 14, 22],
                title=f"管線修改單請款統計報表 — {datetime.now().strftime('%Y/%m/%d')}",
                money_cols=[7, 8, 9, 10, 11],
            )

            # 合計列
            total_row = data_end + 1
            NAVY = "1F3864"
            TOTAL_BG = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            bold_f = Font(name="Microsoft JhengHei UI", bold=True, size=10, color=NAVY)
            side_thin = Side(style="thin", color="B4C6E7")
            bdr = Border(left=side_thin, right=side_thin, top=side_thin, bottom=side_thin)

            ws.cell(row=total_row, column=1, value="合計")
            for ci in range(1, 13):
                c = ws.cell(row=total_row, column=ci)
                c.fill = TOTAL_BG
                c.font = bold_f
                c.border = bdr
            # 金額合計公式
            for col_idx in (7, 8, 9, 10, 11):
                letter = get_column_letter(col_idx)
                ws.cell(row=total_row, column=col_idx,
                        value=f"=SUM({letter}3:{letter}{data_end})")
                ws.cell(row=total_row, column=col_idx).number_format = '#,##0'
                ws.cell(row=total_row, column=col_idx).alignment = Alignment(horizontal="right")

            atomic_save_wb(wb, fp)
            QMessageBox.information(self, "成功", f"報表已匯出：{fp}")
            os.startfile(fp)
        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"匯出失敗：{e}")

    @reentry_guard("_billing_export_in_progress", _show_reentry_notice)
    def _export_billing(self):
        pending = []
        for idx in range(self.tree.topLevelItemCount()):
            it = self.tree.topLevelItem(idx)
            if normalize_billing_status(it.text(4)) == "未請款":
                pending.append([it.text(c) for c in range(12)])
        if not pending:
            QMessageBox.information(self, "提示", "沒有未請款的項目")
            return

        fp, _ = QFileDialog.getSaveFileName(
            self, "匯出請款單", f"請款單_{datetime.now().strftime('%Y%m%d')}.xlsx",
            "Excel 檔案 (*.xlsx)")
        if not fp:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            from record_manager import _apply_excel_styles

            wb = Workbook()
            ws = wb.active
            ws.title = "請款單"
            ws.sheet_properties.tabColor = "C00000"

            hds = ["項次", "報告編號", "修改日期", "Series", "說明",
                   "焊口金額", "材料金額", "未稅小計", "稅額", "含稅小計"]
            for i, h in enumerate(hds, 1):
                ws.cell(row=2, column=i, value=h)

            total_amt = 0
            for idx_val, vals in enumerate(pending, 1):
                r = 2 + idx_val
                ws.cell(row=r, column=1, value=idx_val)
                ws.cell(row=r, column=2, value=vals[0])
                ws.cell(row=r, column=3, value=vals[1])
                ws.cell(row=r, column=4, value=vals[2])
                ws.cell(row=r, column=5, value=vals[3])
                w = self._parse_amount(vals[6])
                m = self._parse_amount(vals[7])
                sub = self._parse_amount(vals[8]) or (w + m)
                tax = self._parse_amount(vals[9])
                gross = self._parse_amount(vals[10]) or (sub + tax)
                ws.cell(row=r, column=6, value=w)
                ws.cell(row=r, column=7, value=m)
                ws.cell(row=r, column=8, value=sub)
                ws.cell(row=r, column=9, value=tax)
                ws.cell(row=r, column=10, value=gross)
                total_amt += gross

            data_end = max(3, 2 + len(pending))
            _apply_excel_styles(
                ws, header_row=2, data_start=3, data_end=data_end,
                col_count=10,
                col_widths=[7, 14, 11, 10, 34, 14, 14, 14, 12, 14],
                title=f"管線修改單請款單 — {datetime.now().strftime('%Y/%m/%d')}　共 {len(pending)} 筆",
                number_cols=[1],
                money_cols=[6, 7, 8, 9, 10],
            )

            # 合計列
            NAVY = "1F3864"
            total_row = data_end + 1
            TOTAL_BG = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            bold_f = Font(name="Microsoft JhengHei UI", bold=True, size=10, color=NAVY)
            red_f  = Font(name="Microsoft JhengHei UI", bold=True, size=11, color="C00000")
            side_thin = Side(style="thin", color="B4C6E7")
            bdr = Border(left=side_thin, right=side_thin, top=side_thin, bottom=side_thin)

            ws.cell(row=total_row, column=1, value="合計")
            for ci in range(1, 11):
                c = ws.cell(row=total_row, column=ci)
                c.fill = TOTAL_BG
                c.font = bold_f
                c.border = bdr

            for col_idx in (6, 7, 8, 9, 10):
                letter = get_column_letter(col_idx)
                ws.cell(row=total_row, column=col_idx,
                        value=f"=SUM({letter}3:{letter}{data_end})")
                ws.cell(row=total_row, column=col_idx).number_format = '#,##0'
                ws.cell(row=total_row, column=col_idx).alignment = Alignment(horizontal="right")
            # 含稅小計用紅色粗體
            ws.cell(row=total_row, column=10).font = red_f

            # 頁尾簽核區
            sign_row = total_row + 3
            sign_font = Font(name="Microsoft JhengHei UI", size=10, color="333333")
            for ci, label in [(1, "承辦人："), (5, "主管："), (8, "核准：")]:
                ws.cell(row=sign_row, column=ci, value=label).font = sign_font
                # 底線
                for offset in range(1, 3 if ci == 1 else 2):
                    uc = ws.cell(row=sign_row, column=ci + offset)
                    uc.border = Border(bottom=Side(style="thin", color="999999"))

            # ── 材料明細附表 ──
            pending_rids = set(v[0] for v in pending)
            store = _load_store()
            mat_rows = [m for m in store.get("materials", [])
                        if m.get("報告編號", "") in pending_rids]

            if mat_rows:
                ws_mat = wb.create_sheet("材料明細")
                ws_mat.sheet_properties.tabColor = "BF8F00"
                mat_hds = ["報告編號", "零件類型", "尺寸", "SCH", "材質",
                           "數量", "單位", "單價", "金額"]
                for i, h in enumerate(mat_hds, 1):
                    ws_mat.cell(row=2, column=i, value=h)
                for ri, m in enumerate(mat_rows, 3):
                    ws_mat.cell(row=ri, column=1, value=m.get("報告編號", ""))
                    ws_mat.cell(row=ri, column=2, value=m.get("零件類型", ""))
                    ws_mat.cell(row=ri, column=3, value=m.get("尺寸", ""))
                    ws_mat.cell(row=ri, column=4, value=m.get("SCH", ""))
                    ws_mat.cell(row=ri, column=5, value=m.get("材質", ""))
                    try:
                        ws_mat.cell(row=ri, column=6, value=float(m.get("數量", 0) or 0))
                    except (ValueError, TypeError):
                        ws_mat.cell(row=ri, column=6, value=m.get("數量", ""))
                    ws_mat.cell(row=ri, column=7, value=m.get("單位", ""))
                    try:
                        ws_mat.cell(row=ri, column=8, value=float(m.get("單價", 0) or 0) if m.get("單價") else "")
                    except (ValueError, TypeError):
                        ws_mat.cell(row=ri, column=8, value=m.get("單價", ""))
                    try:
                        ws_mat.cell(row=ri, column=9, value=float(m.get("金額", 0) or 0) if m.get("金額") else "")
                    except (ValueError, TypeError):
                        ws_mat.cell(row=ri, column=9, value=m.get("金額", ""))

                m_end = max(3, 2 + len(mat_rows))
                _apply_excel_styles(
                    ws_mat, header_row=2, data_start=3, data_end=m_end,
                    col_count=9,
                    col_widths=[14, 20, 8, 10, 10, 8, 6, 12, 14],
                    title=f"請款材料明細 — 共 {len(mat_rows)} 筆",
                    number_cols=[6],
                    money_cols=[8, 9],
                )

                # 材料合計
                m_total_row = m_end + 1
                ws_mat.cell(row=m_total_row, column=1, value="合計")
                for ci in range(1, 10):
                    c = ws_mat.cell(row=m_total_row, column=ci)
                    c.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                    c.font = bold_f
                    c.border = bdr
                amt_letter = get_column_letter(9)
                ws_mat.cell(row=m_total_row, column=9,
                            value=f"=SUM({amt_letter}3:{amt_letter}{m_end})")
                ws_mat.cell(row=m_total_row, column=9).number_format = '#,##0'
                ws_mat.cell(row=m_total_row, column=9).alignment = Alignment(horizontal="right")

            atomic_save_wb(wb, fp)
            QMessageBox.information(self, "成功", f"請款單已匯出：{fp}\n\n共 {len(pending)} 筆，合計 ${total_amt:,.0f}")
            os.startfile(fp)
        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"匯出失敗：{e}")
