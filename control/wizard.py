# -*- coding: utf-8 -*-
"""
wizard.py — 前置作業精靈（PyQt6 版）

功能：
- 互動式建立報告資料夾
- 步驟式引導：選擇日期 → 輸入 Series → 選擇模式 → 填寫焊口 → 選擇圖片 → 填寫說明 → 材料清單 → 確認
- 自動命名資料夾
- 檔案準備檢查清單
- 常用語句記憶功能
"""

import os
import re
import json
import shutil
from datetime import datetime
from typing import Optional, List, Dict

from PyQt6.QtWidgets import (
    QDialog, QWidget, QVBoxLayout, QHBoxLayout, QGridLayout, QFormLayout,
    QLabel, QPushButton, QLineEdit, QComboBox, QTextEdit, QGroupBox,
    QTreeWidget, QTreeWidgetItem, QHeaderView, QRadioButton, QButtonGroup,
    QMessageBox, QFileDialog, QScrollArea, QFrame, QSpinBox, QCheckBox,
    QStackedWidget, QSizePolicy, QApplication, QSplitter, QInputDialog,
)
from PyQt6.QtCore import Qt, QSize
from PyQt6.QtGui import QBrush, QColor, QFont, QIcon, QImage, QPixmap

from theme import Colors, Fonts, set_button_role, make_hint_label, make_separator
from utils import move_to_trash
from resources import resource_path

# 嘗試匯入專案模組
try:
    from config import BASE_DIR, ATTACHMENTS_ROOT
except ImportError:
    BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir))
    ATTACHMENTS_ROOT = os.path.join(BASE_DIR, "attachments")

try:
    from validator import validate_folder, FolderValidation
except ImportError:
    validate_folder = None
    FolderValidation = None

try:
    from image_processor import (
        analyze_folder_images as analyze_image_issues,
        preprocess_folder as prepare_folder_images,
        check_pillow,
    )
except ImportError:
    analyze_image_issues = None
    prepare_folder_images = None
    check_pillow = lambda: False


# PDF 縮圖渲染（可選）
try:
    import fitz as _fitz  # PyMuPDF
    _FITZ_OK = True
except ImportError:
    _fitz = None
    _FITZ_OK = False


# ---------------------------------------------------------------------------
#  Wizard 資料管理
# ---------------------------------------------------------------------------
WIZARD_DATA_PATH = resource_path("control", "wizard_data.json")

_FONT_TITLE = Fonts.heading()
_FONT_SUBTITLE = Fonts.subheading()
_FONT_CODE = Fonts.code(11)


def load_wizard_data() -> dict:
    if os.path.exists(WIZARD_DATA_PATH):
        try:
            with open(WIZARD_DATA_PATH, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"note_presets": {}, "note_history": [], "materials": {}, "materials_history": []}


def save_wizard_data(data: dict):
    try:
        with open(WIZARD_DATA_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"⚠️ 儲存精靈資料失敗: {e}")


def add_note_to_history(note: str, max_history: int = 20):
    if not note or len(note.strip()) < 3:
        return
    data = load_wizard_data()
    history = data.get("note_history", [])
    note = note.strip()
    if note in history:
        history.remove(note)
    history.insert(0, note)
    data["note_history"] = history[:max_history]
    save_wizard_data(data)


# =====================================================================
#  WizardStep 基底
# =====================================================================
class WizardStep:
    """精靈步驟基底類別"""

    def __init__(self, parent: QStackedWidget, wizard: "FolderWizard"):
        self.parent = parent
        self.wizard = wizard
        self.page = QWidget()
        self.layout = QVBoxLayout(self.page)
        parent.addWidget(self.page)

    def on_show(self):
        """步驟被顯示時呼叫（可 override）"""
        pass

    def validate(self) -> bool:
        return True

    def get_data(self) -> dict:
        return {}


# =====================================================================
#  Step1_DateSeries
# =====================================================================
class Step1_DateSeries(WizardStep):

    def __init__(self, parent, wizard):
        super().__init__(parent, wizard)
        self._build()

    def _build(self):
        lay = self.layout
        lbl = QLabel("步驟 1/7：基本資訊")
        lbl.setFont(_FONT_TITLE)
        lay.addWidget(lbl)
        lay.addSpacing(20)

        # 日期
        date_grp = QGroupBox("📅 日期")
        dlay = QHBoxLayout(date_grp)
        dlay.addWidget(QLabel("報告日期 (YYYYMMDD):"))
        self.date_edit = QLineEdit(datetime.now().strftime("%Y%m%d"))
        self.date_edit.setFixedWidth(120)
        dlay.addWidget(self.date_edit)
        btn_today = QPushButton("今天")
        btn_today.clicked.connect(lambda: self.date_edit.setText(datetime.now().strftime("%Y%m%d")))
        dlay.addWidget(btn_today)
        dlay.addStretch()
        lay.addWidget(date_grp)

        # 流水號
        series_grp = QGroupBox("🔢 流水號")
        slay = QHBoxLayout(series_grp)
        slay.addWidget(QLabel("流水號:"))
        self.series_edit = QLineEdit()
        self.series_edit.setFixedWidth(90)
        slay.addWidget(self.series_edit)
        slay.addWidget(QLabel("(如: 202, 0125, 1001)"))
        slay.addStretch()
        lay.addWidget(series_grp)

        # 提示
        hint_grp = QGroupBox()
        hint_grp.setFlat(True)
        hlay = QVBoxLayout(hint_grp)
        lb = QLabel("💡 提示:")
        lb.setFont(Fonts.body(10))
        lb.setStyleSheet(f"font-weight:bold; color: {Colors.TEXT}; border:none; background:transparent;")
        hlay.addWidget(lb)
        hlay.addWidget(QLabel("• 流水號會自動補零至 4 位數"))
        hlay.addWidget(QLabel("• 日期資料夾不存在會自動建立"))
        lay.addWidget(hint_grp)
        lay.addStretch()

    def validate(self) -> bool:
        d = self.date_edit.text().strip()
        s = self.series_edit.text().strip()
        if not re.fullmatch(r"\d{8}", d):
            QMessageBox.critical(self.page, "錯誤", "日期格式不正確，請輸入 YYYYMMDD")
            return False
        try:
            datetime.strptime(d, "%Y%m%d")
        except ValueError:
            QMessageBox.critical(self.page, "錯誤", "無效的日期")
            return False
        if not s:
            QMessageBox.critical(self.page, "錯誤", "請輸入流水號")
            return False
        if not re.fullmatch(r"\d+", s):
            QMessageBox.critical(self.page, "錯誤", "流水號必須是數字")
            return False
        return True

    def get_data(self) -> dict:
        s = self.series_edit.text().strip()
        return {"date": self.date_edit.text().strip(), "series": s.zfill(4)}


# =====================================================================
#  Step2_Mode
# =====================================================================
class Step2_Mode(WizardStep):

    def __init__(self, parent, wizard):
        super().__init__(parent, wizard)
        self._build()

    def _build(self):
        lay = self.layout
        lbl = QLabel("步驟 2/7：選擇模式")
        lbl.setFont(_FONT_TITLE)
        lay.addWidget(lbl)
        lay.addSpacing(20)

        self.btn_group = QButtonGroup(self.page)
        self.radio_single = QRadioButton("Single 模式（少量焊口，≤6 個）")
        self.radio_single.setChecked(True)
        self.btn_group.addButton(self.radio_single)
        lay.addWidget(self.radio_single)

        for txt in ["    📁 資料夾格式: {Series}_{焊口1}_{焊口2}_...",
                     "    📷 圖片: before.jpg, after.jpg",
                     "例: 202_15r1_12a0.5"]:
            lb = QLabel(txt)
            lb.setStyleSheet(f"color: {Colors.TEXT_MUTED}; border:none; background:transparent;" if "例" not in txt else f"color: {Colors.PRIMARY}; border:none; background:transparent;")
            lay.addWidget(lb)

        lay.addSpacing(15)

        self.radio_group = QRadioButton("Group 模式（大量焊口，>6 個）")
        self.btn_group.addButton(self.radio_group)
        lay.addWidget(self.radio_group)

        for txt in ["    📁 資料夾格式: {Series}_{X}G（X = A, B, C...）",
                     "    📷 圖片: before_1.jpg, before_2.jpg, after_1.jpg, after_2.jpg",
                     "    📄 必須有 GroupWeld.txt（每行一個焊口）",
                     "例: 125_AG, 125_BG"]:
            lb = QLabel(txt)
            lb.setStyleSheet(f"color: {Colors.TEXT_MUTED}; border:none; background:transparent;" if "例" not in txt else f"color: {Colors.PRIMARY}; border:none; background:transparent;")
            lay.addWidget(lb)

        lay.addSpacing(10)
        gl_row = QHBoxLayout()
        gl_row.addSpacing(20)
        gl_row.addWidget(QLabel("Group 字母:"))
        self.group_letter = QComboBox()
        self.group_letter.addItems(list("ABCDEFGHIJ"))
        self.group_letter.setFixedWidth(60)
        gl_row.addWidget(self.group_letter)
        gl_row.addStretch()
        lay.addLayout(gl_row)
        lay.addStretch()

    def get_data(self) -> dict:
        mode = "single" if self.radio_single.isChecked() else "group"
        return {
            "mode": mode,
            "group_letter": self.group_letter.currentText() if mode == "group" else None,
        }


# =====================================================================
#  Step3_Welds
# =====================================================================
class Step3_Welds(WizardStep):

    def __init__(self, parent, wizard):
        super().__init__(parent, wizard)
        self.welds_list: List[Dict] = []
        self._build()

    def _build(self):
        lay = self.layout
        lbl = QLabel("步驟 4/7：焊口資訊")
        lbl.setFont(_FONT_TITLE)
        lay.addWidget(lbl)

        # --- 輸入區 ---
        inp_grp = QGroupBox("➕ 新增焊口")
        ig = QVBoxLayout(inp_grp)

        r1 = QHBoxLayout()
        r1.addWidget(QLabel("焊口號:"))
        self.weld_no_edit = QLineEdit()
        self.weld_no_edit.setFixedWidth(60)
        self.weld_no_edit.returnPressed.connect(self._add_weld)
        r1.addWidget(self.weld_no_edit)

        r1.addSpacing(15)
        r1.addWidget(QLabel("標記:"))
        self.mark_group = QButtonGroup(self.page)
        self.mark_r = QRadioButton("r (裁切)")
        self.mark_r.setChecked(True)
        self.mark_a = QRadioButton("a (加長)")
        self.mark_b = QRadioButton("b (加長2)")
        for rb in (self.mark_r, self.mark_a, self.mark_b):
            self.mark_group.addButton(rb)
            r1.addWidget(rb)
        r1.addStretch()
        ig.addLayout(r1)

        r2 = QHBoxLayout()
        r2.addWidget(QLabel("尺寸:"))
        self.size_combo = QComboBox()
        self.size_combo.setEditable(True)
        self.size_combo.addItems(["0.5", "1", "1.5", "2", "2.5", "3", "4"])
        self.size_combo.setCurrentText("1")
        self.size_combo.setFixedWidth(80)
        r2.addWidget(self.size_combo)
        r2.addWidget(QLabel("吋"))
        r2.addStretch()
        ig.addLayout(r2)

        r3 = QHBoxLayout()
        r3.addWidget(QLabel("材質:"))
        self.material_combo = QComboBox()
        self.material_combo.setEditable(True)
        self.material_combo.addItems(["", "CS", "SS304", "SS316", "SS316L", "A106-B", "A312"])
        self.material_combo.setFixedWidth(100)
        r3.addWidget(self.material_combo)

        r3.addSpacing(15)
        r3.addWidget(QLabel("厚度:"))
        self.thickness_combo = QComboBox()
        self.thickness_combo.setEditable(True)
        self.thickness_combo.addItems(["", "SCH10", "SCH20", "SCH40", "SCH80", "SCH160", "3.0", "3.2", "4.0", "5.0", "6.0"])
        self.thickness_combo.setFixedWidth(90)
        r3.addWidget(self.thickness_combo)

        r3.addStretch()
        btn_add = QPushButton("➕ 加入")
        btn_add.clicked.connect(self._add_weld)
        r3.addWidget(btn_add)
        ig.addLayout(r3)

        hint = QLabel("💡 按 Enter 快速加入（材質/厚度可留空）")
        hint.setStyleSheet(f"color: {Colors.TEXT_MUTED}; border:none; background:transparent;")
        ig.addWidget(hint)
        lay.addWidget(inp_grp)

        # --- 焊口清單 ---
        list_grp = QGroupBox("📋 焊口清單")
        lg = QVBoxLayout(list_grp)
        self.tree = QTreeWidget()
        self.tree.setHeaderLabels(["#", "焊口號", "標記", "尺寸", "材質", "厚度", "焊口代碼"])
        self.tree.setAlternatingRowColors(True)
        hdr = self.tree.header()
        hdr.resizeSection(0, 30)
        hdr.resizeSection(1, 50)
        hdr.resizeSection(2, 70)
        hdr.resizeSection(3, 50)
        hdr.resizeSection(4, 60)
        hdr.resizeSection(5, 60)
        hdr.resizeSection(6, 80)
        lg.addWidget(self.tree)
        lay.addWidget(list_grp, stretch=1)

        # 按鈕列1
        br1 = QHBoxLayout()
        lb1 = QLabel("清單操作:")
        lb1.setStyleSheet(f"color: {Colors.TEXT_MUTED}; border:none; background:transparent;")
        br1.addWidget(lb1)
        for txt, fn in [("🗑️ 刪除", self._delete_selected), ("🧹 清空", self._clear_all),
                        ("⬆️", self._move_up), ("⬇️", self._move_down)]:
            b = QPushButton(txt)
            b.clicked.connect(fn)
            br1.addWidget(b)
        br1.addStretch()
        lay.addLayout(br1)

        # 按鈕列2
        br2 = QHBoxLayout()
        lb2 = QLabel("管制表:")
        lb2.setStyleSheet(f"color: {Colors.TEXT_MUTED}; border:none; background:transparent;")
        br2.addWidget(lb2)
        btn_show = QPushButton("📋 查看現有焊口")
        btn_show.clicked.connect(self._show_existing_welds)
        br2.addWidget(btn_show)
        btn_dup = QPushButton("🔍 檢查重複")
        btn_dup.clicked.connect(self._check_all_duplicates)
        br2.addWidget(btn_dup)
        br2.addStretch()
        lay.addLayout(br2)

        self.status_label = QLabel("已加入 0 個焊口")
        self.status_label.setStyleSheet(f"color: {Colors.PRIMARY}; font-weight:bold; border:none; background:transparent;")
        lay.addWidget(self.status_label)

    def on_show(self):
        self._update_status()
        self.weld_no_edit.setFocus()

    def _get_mark(self) -> str:
        if self.mark_r.isChecked():
            return "r"
        elif self.mark_a.isChecked():
            return "a"
        return "b"

    def _add_weld(self):
        weld_no = self.weld_no_edit.text().strip()
        mark = self._get_mark()
        size = self.size_combo.currentText().strip()
        material = self.material_combo.currentText().strip()
        thickness = self.thickness_combo.currentText().strip()

        if not weld_no:
            QMessageBox.warning(self.page, "提示", "請輸入焊口號")
            self.weld_no_edit.setFocus()
            return
        if not weld_no.isdigit():
            QMessageBox.warning(self.page, "提示", "焊口號必須是數字")
            self.weld_no_edit.setFocus()
            return
        if not size:
            QMessageBox.warning(self.page, "提示", "請輸入尺寸")
            return

        code = f"{weld_no}{mark}{size}"
        for w in self.welds_list:
            if w["code"] == code:
                QMessageBox.warning(self.page, "提示", f"焊口 {code} 已存在")
                return

        weld_id = f"{weld_no}{mark}"
        if not self._check_weld_duplicate(weld_id, code):
            return

        mark_names = {"r": "r (裁切)", "a": "a (加長)", "b": "b (加長2)"}
        item = {"weld_no": weld_no, "mark": mark, "size": size,
                "material": material, "thickness": thickness, "code": code}
        self.welds_list.append(item)

        idx = len(self.welds_list)
        self.tree.addTopLevelItem(QTreeWidgetItem([
            str(idx), weld_no, mark_names.get(mark, mark),
            f'{size}"', material or "-", thickness or "-", code,
        ]))
        self._update_status()
        self.weld_no_edit.clear()
        self.weld_no_edit.setFocus()

    def _delete_selected(self):
        items = self.tree.selectedItems()
        if not items:
            return
        indices = sorted([self.tree.indexOfTopLevelItem(it) for it in items], reverse=True)
        for idx in indices:
            if 0 <= idx < len(self.welds_list):
                self.welds_list.pop(idx)
            self.tree.takeTopLevelItem(idx)
        self._renumber()
        self._update_status()

    def _clear_all(self):
        if self.welds_list and QMessageBox.question(
            self.page, "確認", "確定要清空所有焊口？"
        ) == QMessageBox.StandardButton.Yes:
            self.welds_list.clear()
            self.tree.clear()
            self._update_status()

    def _move_up(self):
        items = self.tree.selectedItems()
        if not items or len(items) != 1:
            return
        idx = self.tree.indexOfTopLevelItem(items[0])
        if idx > 0:
            self.welds_list[idx], self.welds_list[idx - 1] = self.welds_list[idx - 1], self.welds_list[idx]
            self._rebuild_tree()
            self.tree.setCurrentItem(self.tree.topLevelItem(idx - 1))

    def _move_down(self):
        items = self.tree.selectedItems()
        if not items or len(items) != 1:
            return
        idx = self.tree.indexOfTopLevelItem(items[0])
        if idx < len(self.welds_list) - 1:
            self.welds_list[idx], self.welds_list[idx + 1] = self.welds_list[idx + 1], self.welds_list[idx]
            self._rebuild_tree()
            self.tree.setCurrentItem(self.tree.topLevelItem(idx + 1))

    def _renumber(self):
        for i in range(self.tree.topLevelItemCount()):
            self.tree.topLevelItem(i).setText(0, str(i + 1))

    def _rebuild_tree(self):
        self.tree.clear()
        mark_names = {"r": "r (裁切)", "a": "a (加長)", "b": "b (加長2)"}
        for i, w in enumerate(self.welds_list):
            self.tree.addTopLevelItem(QTreeWidgetItem([
                str(i + 1), w["weld_no"], mark_names.get(w["mark"], w["mark"]),
                f'{w["size"]}"', w.get("material", "") or "-",
                w.get("thickness", "") or "-", w["code"],
            ]))

    def _update_status(self):
        count = len(self.welds_list)
        mode = self.wizard.get_collected_data().get("mode", "single")
        if mode == "single" and count > 6:
            self.status_label.setText(f"⚠️ 已加入 {count} 個焊口（Single 模式建議 ≤6 個）")
        else:
            self.status_label.setText(f"已加入 {count} 個焊口")

    def _check_weld_duplicate(self, weld_id: str, full_code: str) -> bool:
        try:
            from settings_manager import get_weld_control_config
            config = get_weld_control_config()
            if not config.get("check_duplicate", True):
                return True
            serial_raw = self.wizard.get_collected_data().get("series", "")
            if not serial_raw:
                return True
            serial = self._format_serial(serial_raw, config)
            from weld_control import check_welds_exist
            exists = check_welds_exist(serial, [weld_id])
            if exists and exists.get(weld_id):
                return QMessageBox.question(
                    self.page, "焊口重複警告",
                    f"⚠️ 焊口 {weld_id} 已存在於焊口管制表！\n\n"
                    f"流水號: {serial}\n管制表焊口編號: {weld_id}\n本次輸入代碼: {full_code}\n\n是否仍要加入？"
                ) == QMessageBox.StandardButton.Yes
            return True
        except ImportError:
            return True
        except Exception as e:
            print(f"[警告] 焊口重複檢查失敗: {e}")
            return True

    def _format_serial(self, serial_raw: str, config: dict = None) -> str:
        if config is None:
            try:
                from settings_manager import get_weld_control_config
                config = get_weld_control_config()
            except Exception:
                config = {}
        fmt = config.get("serial_format", "raw")
        if fmt == "pad4":
            return serial_raw.zfill(4)
        return serial_raw.lstrip("0") or "0"

    def _check_all_duplicates(self):
        if not self.welds_list:
            QMessageBox.information(self.page, "提示", "尚未加入任何焊口")
            return
        try:
            from settings_manager import get_weld_control_config, get_weld_control_table_path
            table_path = get_weld_control_table_path()
            if not table_path:
                QMessageBox.warning(self.page, "提示", "尚未設定焊口管制表路徑\n請至「設定」頁籤進行設定")
                return
            if not os.path.exists(table_path):
                QMessageBox.warning(self.page, "提示", f"焊口管制表不存在:\n{table_path}")
                return

            config = get_weld_control_config()
            serial_raw = self.wizard.get_collected_data().get("series", "")
            if not serial_raw:
                QMessageBox.warning(self.page, "提示", "請先輸入流水號（步驟1）")
                return
            serial = self._format_serial(serial_raw, config)
            from weld_control import check_welds_exist

            weld_ids, weld_map = [], {}
            for w in self.welds_list:
                wid = f"{w['weld_no']}{w['mark']}"
                if wid not in weld_map:
                    weld_ids.append(wid)
                    weld_map[wid] = []
                weld_map[wid].append(w["code"])

            exists = check_welds_exist(serial, weld_ids)
            if not exists:
                QMessageBox.information(self.page, "檢查結果",
                    f"✅ 全部 {len(self.welds_list)} 個焊口皆為新焊口\n（流水號 {serial} 中不存在）")
            else:
                dup_list = []
                for wid in exists:
                    codes = weld_map.get(wid, [wid])
                    dup_list.append(f"  • {wid} → {', '.join(codes)}")
                msg = (f"⚠️ 發現 {len(exists)} 個焊口已存在於管制表：\n\n流水號: {serial}\n重複焊口:\n"
                       + "\n".join(dup_list[:10]))
                if len(dup_list) > 10:
                    msg += f"\n  ...還有 {len(dup_list) - 10} 個"
                msg += f"\n\n新焊口: {len(weld_ids) - len(exists)} 個"
                QMessageBox.warning(self.page, "檢查結果", msg)
        except ImportError as e:
            QMessageBox.critical(self.page, "錯誤", f"模組載入失敗: {e}")
        except Exception as e:
            QMessageBox.critical(self.page, "錯誤", f"檢查失敗: {e}")

    def _show_existing_welds(self):
        try:
            from settings_manager import get_weld_control_config, get_weld_control_table_path
            table_path = get_weld_control_table_path()
            if not table_path or not os.path.exists(table_path):
                QMessageBox.warning(self.page, "提示", "焊口管制表不存在或未設定")
                return
            config = get_weld_control_config()
            serial_raw = self.wizard.get_collected_data().get("series", "")
            if not serial_raw:
                QMessageBox.warning(self.page, "提示", "請先輸入流水號（步驟1）")
                return
            serial = self._format_serial(serial_raw, config)
            from weld_control import init_weld_manager_from_settings
            manager = init_weld_manager_from_settings()
            if not manager or not manager.is_configured():
                QMessageBox.warning(self.page, "提示", "焊口管制表未正確設定")
                return
            manager.load()
            existing = manager.get_all_welds_by_serial(serial)
            self._show_welds_dialog(serial, existing, config)
        except ImportError as e:
            QMessageBox.critical(self.page, "錯誤", f"模組載入失敗: {e}")
        except Exception as e:
            QMessageBox.critical(self.page, "錯誤", f"查詢失敗: {e}")

    def _show_welds_dialog(self, serial: str, welds: list, config: dict):
        dlg = QDialog(self.page)
        dlg.setWindowTitle(f"流水號 {serial} 的現有焊口")
        dlg.resize(500, 400)
        dlg.setModal(True)
        # 強制設定對話框底色與文字色，避免繼承衝突
        dlg.setStyleSheet(
            f"QDialog {{ background: {Colors.BG_WHITE}; color: {Colors.TEXT}; }}"
            f"QLabel {{ color: {Colors.TEXT}; background: transparent; border: none; }}"
            f"QTreeWidget {{ background: {Colors.BG_WHITE}; color: {Colors.TEXT}; "
            f"  alternate-background-color: {Colors.BG}; border: 1px solid {Colors.BORDER}; }}"
            f"QTreeWidget::item {{ color: {Colors.TEXT}; }}"
            f"QHeaderView::section {{ background: {Colors.BG}; color: {Colors.TEXT_SECONDARY}; "
            f"  border: none; border-bottom: 2px solid {Colors.BORDER}; padding: 6px 8px; font-weight: bold; }}"
        )
        dlay = QVBoxLayout(dlg)

        col_weld_no = config.get("col_weld_no", "焊口編號")

        if not welds:
            lb = QLabel(f"📋 流水號 {serial} 目前沒有任何焊口記錄")
            lb.setFont(_FONT_SUBTITLE)
            dlay.addWidget(lb)
            lb2 = QLabel("這是一個全新的流水號，可以放心新增焊口！")
            lb2.setStyleSheet(f"color: {Colors.SUCCESS}; background: transparent; border: none;")
            dlay.addWidget(lb2)
        else:
            lb = QLabel(f"📋 流水號 {serial} 已有 {len(welds)} 個焊口")
            lb.setFont(_FONT_SUBTITLE)
            dlay.addWidget(lb)
            lb2 = QLabel("以下焊口已存在於管制表中，請避免重複新增：")
            lb2.setStyleSheet(f"color: {Colors.WARNING}; background: transparent; border: none;")
            dlay.addWidget(lb2)

            # --- 解決設定欄位名與 Excel 實際欄名不一致的問題 ---
            actual_keys = set(welds[0].keys()) if welds else set()

            from utils import resolve_col
            col_weld_resolved = resolve_col(col_weld_no, actual_keys)

            # 決定要顯示的欄位
            display_cols = [col_weld_resolved]
            extra_candidates = [
                "LINE NO", "Line_num", "DWG NO",
                "SIZE", "尺寸",
                "SCH", "厚度",
                "登錄日期", "組銲完成日期",
                "修改單編號", "備註",
            ]
            for ec in extra_candidates:
                rc = resolve_col(ec, actual_keys)
                if rc in actual_keys and rc not in display_cols:
                    display_cols.append(rc)
                if len(display_cols) >= 5:
                    break

            tree = QTreeWidget()
            tree.setAlternatingRowColors(True)
            tree.setHeaderLabels(display_cols)
            _fg = QBrush(QColor(Colors.TEXT))
            for w in welds:
                vals = [str(v) if (v := w.get(c)) is not None else "" for c in display_cols]
                item = QTreeWidgetItem(vals)
                for col_i in range(len(display_cols)):
                    item.setForeground(col_i, _fg)
                tree.addTopLevelItem(item)
            for col_i in range(len(display_cols)):
                tree.resizeColumnToContents(col_i)
            dlay.addWidget(tree, stretch=1)

        btn_close = QPushButton("關閉")
        btn_close.clicked.connect(dlg.close)
        dlay.addWidget(btn_close, alignment=Qt.AlignmentFlag.AlignCenter)
        dlg.exec()

    def validate(self) -> bool:
        if not self.welds_list:
            QMessageBox.critical(self.page, "錯誤", "請至少加入一個焊口")
            return False
        mode = self.wizard.get_collected_data().get("mode", "single")
        if mode == "single" and len(self.welds_list) > 6:
            if QMessageBox.question(
                self.page, "警告",
                f"Single 模式建議焊口數 ≤6 個，目前有 {len(self.welds_list)} 個。\n是否繼續？"
            ) != QMessageBox.StandardButton.Yes:
                return False
        return True

    def get_data(self) -> dict:
        mode = self.wizard.get_collected_data().get("mode", "single")
        codes = [w["code"] for w in self.welds_list]
        result: dict = {"welds_list_full": self.welds_list.copy()}
        if mode == "single":
            result["welds_string"] = "_".join(codes)
        else:
            result["welds_list"] = codes
        return result


# =====================================================================
#  Staging 縮圖選取器（整合到 Step4）
# =====================================================================
_STAGING_THUMB_SIZE = 110

try:
    from staging_manager import scan_staging, make_thumbnail, ensure_staging_dir, StagingFile
    _STAGING_OK = True
except ImportError:
    _STAGING_OK = False


# 全域：hover 放大預覽圖片尺寸（px）
HOVER_ZOOM_SIZE = 480


class _ZoomPopup(QLabel):
    """滑鼠懸停時顯示的放大預覽浮窗"""

    def __init__(self):
        super().__init__(None, Qt.WindowType.ToolTip | Qt.WindowType.FramelessWindowHint)
        self.setAttribute(Qt.WidgetAttribute.WA_TranslucentBackground)
        self.setStyleSheet(
            "QLabel { background: white; border: 2px solid #2563eb; border-radius: 8px; padding: 4px; }"
        )
        self.setFixedSize(HOVER_ZOOM_SIZE + 8, HOVER_ZOOM_SIZE + 8)
        self.hide()

    def show_image(self, pixmap: QPixmap, global_pos):
        if pixmap and not pixmap.isNull():
            scaled = pixmap.scaled(
                HOVER_ZOOM_SIZE, HOVER_ZOOM_SIZE,
                Qt.AspectRatioMode.KeepAspectRatio,
                Qt.TransformationMode.SmoothTransformation,
            )
            self.setPixmap(scaled)
            # 偏移顯示，避免擋住滑鼠
            self.move(global_pos.x() + 16, global_pos.y() + 16)
            self.show()


# 全局單例，所有縮圖共用
_zoom_popup: Optional[_ZoomPopup] = None

def _get_zoom_popup() -> _ZoomPopup:
    global _zoom_popup
    if _zoom_popup is None:
        _zoom_popup = _ZoomPopup()
    return _zoom_popup


class _StagingThumb(QFrame):
    """Staging 照片縮圖卡片 — 附帶角色下拉"""

    def __init__(self, sf, role_items: list, parent=None):
        super().__init__(parent)
        self.sf = sf
        self._full_pixmap: Optional[QPixmap] = None  # 完整解析度圖片（延遲載入）
        self.setFrameShape(QFrame.Shape.StyledPanel)
        self.setStyleSheet(
            f"_StagingThumb {{ background: white; border: 1px solid #ddd; border-radius: 6px; }}"
        )
        self.setFixedSize(_STAGING_THUMB_SIZE + 20, _STAGING_THUMB_SIZE + 70)
        self.setMouseTracking(True)

        lay = QVBoxLayout(self)
        lay.setContentsMargins(4, 4, 4, 2)
        lay.setSpacing(2)

        # 縮圖
        thumb_lbl = QLabel()
        thumb_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        thumb_lbl.setFixedSize(_STAGING_THUMB_SIZE, _STAGING_THUMB_SIZE)

        if sf.file_type == "pdf":
            thumb_lbl.setText("📄 PDF")
            thumb_lbl.setFont(QFont("Segoe UI", 14))
            thumb_lbl.setStyleSheet("background: #e3f2fd; border-radius: 6px;")
        elif _STAGING_OK:
            data = make_thumbnail(sf.path, _STAGING_THUMB_SIZE)
            if data:
                img = QImage()
                img.loadFromData(data)
                pm = QPixmap.fromImage(img)
                thumb_lbl.setPixmap(pm.scaled(
                    _STAGING_THUMB_SIZE, _STAGING_THUMB_SIZE,
                    Qt.AspectRatioMode.KeepAspectRatio,
                    Qt.TransformationMode.SmoothTransformation,
                ))
            else:
                thumb_lbl.setText("🖼️")
                thumb_lbl.setFont(QFont("Segoe UI", 18))
        else:
            thumb_lbl.setText("🖼️")
            thumb_lbl.setFont(QFont("Segoe UI", 18))
        lay.addWidget(thumb_lbl)

        # 檔名
        name_lbl = QLabel(sf.filename)
        name_lbl.setFont(QFont("Segoe UI", 7))
        name_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        name_lbl.setWordWrap(True)
        name_lbl.setToolTip(sf.filename)
        name_lbl.setMaximumHeight(26)
        lay.addWidget(name_lbl)

        # 角色下拉
        self.role_combo = QComboBox()
        self.role_combo.addItems(["(跳過)"] + role_items)
        self.role_combo.setFixedWidth(_STAGING_THUMB_SIZE + 10)
        self.role_combo.setFont(QFont("Segoe UI", 8))
        lay.addWidget(self.role_combo)

    def get_assignment(self):
        """回傳 (role_key, path) 或 None"""
        role = self.role_combo.currentText()
        if role == "(跳過)":
            return None
        return (role, self.sf.path)

    # ---- hover 放大預覽 ----
    def _ensure_full_pixmap(self):
        if self._full_pixmap is None and self.sf.file_type != "pdf":
            pm = QPixmap(self.sf.path)
            if not pm.isNull():
                self._full_pixmap = pm

    def enterEvent(self, event):
        super().enterEvent(event)
        self.setStyleSheet(
            "_StagingThumb { background: white; border: 2px solid #2563eb; border-radius: 6px; }"
        )
        if self.sf.file_type != "pdf":
            self._ensure_full_pixmap()
            if self._full_pixmap:
                popup = _get_zoom_popup()
                popup.show_image(self._full_pixmap, self.mapToGlobal(self.rect().topRight()))

    def leaveEvent(self, event):
        super().leaveEvent(event)
        self.setStyleSheet(
            "_StagingThumb { background: white; border: 1px solid #ddd; border-radius: 6px; }"
        )
        _get_zoom_popup().hide()

    def mouseMoveEvent(self, event):
        super().mouseMoveEvent(event)
        popup = _get_zoom_popup()
        if popup.isVisible() and self._full_pixmap:
            popup.move(event.globalPosition().toPoint().x() + 16,
                       event.globalPosition().toPoint().y() + 16)


class StagingPickerDialog(QDialog):
    """從 staging/ 選取照片並指定角色的對話框"""

    def __init__(self, role_map: List[tuple], parent=None):
        """role_map: [(key, label), ...] 如 [("before", "修改前"), ...]"""
        super().__init__(parent)
        self.setWindowTitle("📷 從 staging/ 選取照片")
        self.resize(680, 520)
        self.setMinimumSize(500, 400)
        self.result_map: Dict[str, str] = {}     # key → file path
        self._role_map = role_map
        self._role_labels = [label for _, label in role_map]
        self._key_by_label = {label: key for key, label in role_map}
        self._thumbs: List[_StagingThumb] = []
        self._build_ui()

    def _build_ui(self):
        lay = QVBoxLayout(self)

        # 頂部
        header = QHBoxLayout()
        self.count_label = QLabel()
        self.count_label.setFont(_FONT_SUBTITLE)
        header.addWidget(self.count_label)
        header.addStretch()
        btn_refresh = QPushButton("🔄 重新掃描")
        btn_refresh.clicked.connect(self._load)
        header.addWidget(btn_refresh)
        btn_open = QPushButton("📂 開啟 staging/")
        btn_open.clicked.connect(self._open_staging)
        header.addWidget(btn_open)
        btn_cleanup = QPushButton("🗑️ 清空 staging")
        btn_cleanup.setToolTip("刪除 staging/ 中所有已不需要的暫存檔案")
        btn_cleanup.clicked.connect(self._cleanup_staging)
        set_button_role(btn_cleanup, "danger")
        header.addWidget(btn_cleanup)
        lay.addLayout(header)

        hint = QLabel("💡 用每張照片下方的下拉選單指定角色，不需要的保持「跳過」")
        hint.setStyleSheet(f"color: {Colors.TEXT_MUTED}; border:none; background:transparent;")
        lay.addWidget(hint)

        # 捲動區
        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll_content = QWidget()
        self.scroll_inner = QVBoxLayout(self.scroll_content)
        self.scroll_inner.setAlignment(Qt.AlignmentFlag.AlignTop)
        self.scroll.setWidget(self.scroll_content)
        lay.addWidget(self.scroll, stretch=1)

        # 底部按鈕
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_cancel = QPushButton("取消")
        btn_cancel.clicked.connect(self.reject)
        btn_row.addWidget(btn_cancel)
        btn_ok = QPushButton("✅ 確認匯入")
        set_button_role(btn_ok, "primary")
        btn_ok.clicked.connect(self._accept)
        btn_row.addWidget(btn_ok)
        lay.addLayout(btn_row)

        self._load()

    def _load(self):
        self._thumbs.clear()
        while self.scroll_inner.count():
            child = self.scroll_inner.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

        if not _STAGING_OK:
            self.count_label.setText("staging 模組載入失敗")
            return

        staging_dir = os.path.join(BASE_DIR, "staging")
        ensure_staging_dir(BASE_DIR)
        files = scan_staging(staging_dir)

        if not files:
            self.count_label.setText("staging/ 為空")
            empty = QLabel("📂 staging/ 資料夾中沒有檔案\n\n"
                           "請先將照片／PDF 丟入 staging/ 資料夾\n"
                           "再點「重新掃描」")
            empty.setAlignment(Qt.AlignmentFlag.AlignCenter)
            empty.setStyleSheet(f"color: {Colors.TEXT_MUTED}; padding: 40px;")
            self.scroll_inner.addWidget(empty)
            return

        self.count_label.setText(f"staging/ 共 {len(files)} 個檔案")

        # 網格排列
        grid = QWidget()
        grid_lay = QGridLayout(grid)
        grid_lay.setSpacing(6)
        cols = 4
        for i, sf in enumerate(files):
            thumb = _StagingThumb(sf, self._role_labels, self)
            self._thumbs.append(thumb)
            grid_lay.addWidget(thumb, i // cols, i % cols)
        self.scroll_inner.addWidget(grid)
        self.scroll_inner.addStretch()

    def _open_staging(self):
        staging_dir = os.path.join(BASE_DIR, "staging")
        os.makedirs(staging_dir, exist_ok=True)
        os.startfile(staging_dir)

    def _cleanup_staging(self):
        """清除 staging/ 中所有檔案（由使用者確認後執行）"""
        staging_dir = os.path.join(BASE_DIR, "staging")
        if not os.path.isdir(staging_dir):
            return
        files = [f for f in os.listdir(staging_dir)
                 if os.path.isfile(os.path.join(staging_dir, f))]
        if not files:
            QMessageBox.information(self, "提示", "staging/ 已經是空的")
            return
        reply = QMessageBox.question(
            self, "清除 staging",
            f"確定要將 staging/ 資料夾中的 {len(files)} 個檔案移到隔離區嗎？\n\n"
            "檔案會移到 .trash/，需要時仍可人工還原。",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        removed = 0
        for f in files:
            try:
                if move_to_trash(os.path.join(staging_dir, f), BASE_DIR, reason="cleanup_staging"):
                    removed += 1
            except OSError:
                pass
        QMessageBox.information(self, "完成", f"已將 {removed} 個檔案移到隔離區")
        self._load()

    def _accept(self):
        self.result_map = {}
        conflicts = []
        for thumb in self._thumbs:
            assignment = thumb.get_assignment()
            if not assignment:
                continue
            label, path = assignment
            key = self._key_by_label.get(label, label)
            if key in self.result_map:
                conflicts.append(label)
            self.result_map[key] = path

        if conflicts:
            QMessageBox.warning(
                self, "角色衝突",
                f"以下角色被指派了多次：\n{'、'.join(set(conflicts))}\n\n每個角色只能指派一個檔案，後者會覆蓋前者。"
            )
        self.accept()


# =====================================================================
#  Step4_Images
# =====================================================================
class Step4_Images(WizardStep):

    def __init__(self, parent, wizard):
        super().__init__(parent, wizard)
        self.image_paths: Dict[str, str] = {}
        self._edits: Dict[str, QLineEdit] = {}
        self._build()

    def _build(self):
        lay = self.layout
        lbl = QLabel("步驟 3/7：選擇圖片")
        lbl.setFont(_FONT_TITLE)
        lay.addWidget(lbl)
        lay.addSpacing(6)

        # === Staging 快速匯入 ===
        staging_row = QHBoxLayout()
        self.btn_staging = QPushButton("📷 從 staging/ 匯入照片")
        self.btn_staging.setToolTip("開啟縮圖選取器，從 staging 資料夾批量匯入")
        set_button_role(self.btn_staging, "primary")
        self.btn_staging.clicked.connect(self._open_staging_picker)
        staging_row.addWidget(self.btn_staging)
        staging_row.addStretch()
        lbl_st = QLabel("可先把照片丟入 staging/ 資料夾")
        lbl_st.setStyleSheet(f"color: {Colors.TEXT_MUTED}; border:none; background:transparent;")
        staging_row.addWidget(lbl_st)
        lay.addLayout(staging_row)

        sep1 = QFrame()
        sep1.setFrameShape(QFrame.Shape.HLine)
        sep1.setFrameShadow(QFrame.Shadow.Sunken)
        lay.addWidget(sep1)

        # === 檔案欄位（動態建立於 on_show） ===
        self.images_container = QVBoxLayout()
        lay.addLayout(self.images_container)
        lay.addStretch()

    def _get_role_map(self) -> List[tuple]:
        """依模式回傳 [(key, label), ...]"""
        mode = self.wizard.get_collected_data().get("mode", "single")
        if mode == "single":
            return [("before", "修改前 (before)"), ("after", "修改後 (after)"), ("pdf", "圖面 PDF")]
        else:
            return [
                ("before_1", "修改前 1"), ("before_2", "修改前 2"),
                ("after_1", "修改後 1"), ("after_2", "修改後 2"),
                ("pdf", "圖面 PDF"),
            ]

    def _open_staging_picker(self):
        """開啟 staging 縮圖選取器"""
        role_map = self._get_role_map()
        dlg = StagingPickerDialog(role_map, self.page)
        if dlg.exec() == QDialog.DialogCode.Accepted and dlg.result_map:
            for key, path in dlg.result_map.items():
                self.image_paths[key] = path
                if key in self._edits:
                    self._edits[key].setText(path)

    def _create_image_row(self, label: str, key: str, file_type: str = "image"):
        row = QHBoxLayout()
        lb = QLabel(label)
        lb.setFixedWidth(120)
        row.addWidget(lb)
        edit = QLineEdit()
        edit.setMinimumWidth(250)
        row.addWidget(edit, stretch=1)

        def browse():
            if file_type == "pdf":
                filt = "PDF 檔案 (*.pdf)"
            else:
                filt = "圖片檔 (*.jpg *.jpeg *.png);;所有檔案 (*.*)"
            path, _ = QFileDialog.getOpenFileName(self.page, f"選擇 {label}", "", filt)
            if path:
                edit.setText(path)
                self.image_paths[key] = path

        btn = QPushButton("瀏覽...")
        btn.clicked.connect(browse)
        row.addWidget(btn)

        btn_ann = QPushButton("✏️ 標註")
        btn_ann.setToolTip("開啟標註工具，在圖片上畫線/箭頭/文字")
        btn_ann.setEnabled(False)
        btn_ann.clicked.connect(lambda _, k=key, ft=file_type: self._open_annotator(k, ft == "pdf"))
        row.addWidget(btn_ann)

        self.image_paths[key] = ""
        self._edits[key] = edit

        def _on_text_changed(t, k=key, b=btn_ann):
            self.image_paths[k] = t
            b.setEnabled(bool(t.strip()) and os.path.exists(t.strip()))
            self._notify_image_change()

        edit.textChanged.connect(_on_text_changed)
        self.images_container.addLayout(row)

    def on_show(self):
        # 清空重建
        while self.images_container.count():
            child = self.images_container.takeAt(0)
            if child.widget():
                child.widget().deleteLater()
            elif child.layout():
                while child.layout().count():
                    sub = child.layout().takeAt(0)
                    if sub.widget():
                        sub.widget().deleteLater()

        self.image_paths.clear()
        self._edits.clear()
        mode = self.wizard.get_collected_data().get("mode", "single")

        lbl_manual = QLabel("📁 手動瀏覽選取（或使用上方 staging 匯入）：")
        lbl_manual.setStyleSheet(f"color: {Colors.TEXT}; font-weight: bold; border:none; background:transparent;")
        self.images_container.addWidget(lbl_manual)

        if mode == "single":
            self._create_image_row("修改前 (before)", "before")
            self._create_image_row("修改後 (after)", "after")
        else:
            self._create_image_row("修改前 1", "before_1")
            self._create_image_row("修改前 2", "before_2")
            self._create_image_row("修改後 1", "after_1")
            self._create_image_row("修改後 2", "after_2")

        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setFrameShadow(QFrame.Shadow.Sunken)
        self.images_container.addWidget(sep)
        lb = QLabel("📎 附件 PDF（必選）:")
        self.images_container.addWidget(lb)
        self._create_image_row("圖面 PDF *", "pdf", file_type="pdf")

        hint = QLabel("💡 圖片可稍後手動複製到資料夾")
        hint.setStyleSheet(f"color: {Colors.TEXT_MUTED}; border:none; background:transparent;")
        self.images_container.addWidget(hint)

    def validate(self) -> bool:
        pdf = self.image_paths.get("pdf", "")
        if not pdf:
            QMessageBox.warning(self.page, "提示", "請選擇圖面 PDF 檔案")
            return False
        if not pdf.lower().endswith(".pdf"):
            QMessageBox.warning(self.page, "提示", "圖面附件必須是 PDF 格式")
            return False
        return True

    def _notify_image_change(self):
        """通知精靈更新圖片預覽側邊欄"""
        if hasattr(self.wizard, '_img_preview'):
            self.wizard._img_preview.update_images(self.image_paths)

    def _open_annotator(self, key: str, is_pdf: bool):
        """開啟標註對話框"""
        path = self.image_paths.get(key, "").strip()
        if not path or not os.path.exists(path):
            QMessageBox.warning(self.page, "提示", "請先選擇檔案")
            return
        from gui_annotator import AnnotationDialog
        dlg = AnnotationDialog(path, is_pdf=is_pdf, parent=self.page)
        if not dlg._load_ok:
            QMessageBox.warning(self.page, "無法載入", f"無法開啟標註工具:\n{path}")
            return
        dlg.exec()
        if dlg.was_saved and dlg.saved_path:
            new_path = dlg.saved_path
            self.image_paths[key] = new_path
            if key in self._edits:
                self._edits[key].setText(new_path)

    def get_data(self) -> dict:
        return {"images": self.image_paths.copy()}


# =====================================================================
#  Step5_Note
# =====================================================================
class Step5_Note(WizardStep):

    def __init__(self, parent, wizard):
        super().__init__(parent, wizard)
        self._build()

    def _build(self):
        lay = self.layout
        lbl = QLabel("步驟 5/7：修改原因說明")
        lbl.setFont(_FONT_TITLE)
        lay.addWidget(lbl)

        wdata = load_wizard_data()

        # 預設語句
        preset_grp = QGroupBox("📝 預設語句（點擊插入）")
        pg = QVBoxLayout(preset_grp)
        presets = wdata.get("note_presets", {})
        cats = {"cut_shorten": "✂️ 裁切/縮短", "extend": "📏 延長",
                "rework": "🔧 返工/修正", "other": "📋 其他"}
        for ck, cn in cats.items():
            if ck in presets:
                row = QHBoxLayout()
                lb = QLabel(cn)
                lb.setFixedWidth(100)
                row.addWidget(lb)
                for phrase in presets[ck]:
                    b = QPushButton(phrase)
                    b.clicked.connect(lambda checked, p=phrase: self._insert_text(p))
                    row.addWidget(b)
                row.addStretch()
                pg.addLayout(row)
        lay.addWidget(preset_grp)

        # 歷史
        hist_row = QHBoxLayout()
        hist_row.addWidget(QLabel("📜 歷史記錄:"))
        self.history_combo = QComboBox()
        self.history_combo.addItems(wdata.get("note_history", []))
        self.history_combo.setMinimumWidth(300)
        self.history_combo.activated.connect(self._on_history_select)
        hist_row.addWidget(self.history_combo, stretch=1)
        lay.addLayout(hist_row)

        # 文字區
        inp_grp = QGroupBox("✏️ 修改原因（可自行編輯）")
        ig = QVBoxLayout(inp_grp)
        self.note_text = QTextEdit()
        ig.addWidget(self.note_text)
        lay.addWidget(inp_grp, stretch=1)

        hint = QLabel("💡 此內容將寫入 note.txt，可留空")
        hint.setStyleSheet(f"color: {Colors.TEXT_MUTED}; border:none; background:transparent;")
        lay.addWidget(hint)

    def _insert_text(self, text: str):
        cur = self.note_text.toPlainText().strip()
        if cur:
            self.note_text.append(text)
        else:
            self.note_text.setPlainText(text)

    def _on_history_select(self, index: int):
        txt = self.history_combo.itemText(index)
        if txt:
            self.note_text.setPlainText(txt)

    def get_data(self) -> dict:
        note = self.note_text.toPlainText().strip()
        if note:
            add_note_to_history(note)
        return {"note_content": note}


# =====================================================================
#  Step6_Materials
# =====================================================================

# 材質 → 預設 Schedule 對應
_DEFAULT_SCH = {"ss": "SCH 10", "cs": "SCH 40"}

class Step6_Materials(WizardStep):

    _BTN_SIZE = 96          # 圖片按鈕邊長 (px)
    _GRID_COLS = 8          # 每列按鈕數

    def __init__(self, parent, wizard):
        super().__init__(parent, wizard)
        self.materials_list: List[Dict] = []
        self._components_data: list = []
        self._build()

    def _build(self):
        lay = self.layout
        lbl = QLabel("步驟 6/7：使用材料清單")
        lbl.setFont(_FONT_TITLE)
        lay.addWidget(lbl)

        wdata = load_wizard_data()
        mdata = wdata.get("materials", {})
        self._components_data = mdata.get("components", [])
        pipe_types = mdata.get("pipe_types", [])
        sizes = mdata.get("sizes_inch", [])
        schedules = mdata.get("schedules", [])

        # ── 新增材料區 ──
        add_grp = QGroupBox("➕ 新增材料")
        ag = QVBoxLayout(add_grp)

        # ─── 零件圖片按鈕矩陣 ───
        img_dir = resource_path("control", "image")
        self._comp_btn_group = QButtonGroup(self.page)
        self._comp_btn_group.setExclusive(True)

        grid_widget = QWidget()
        grid_lay = QGridLayout(grid_widget)
        grid_lay.setSpacing(4)
        grid_lay.setContentsMargins(2, 2, 2, 2)

        icon_sz = QSize(self._BTN_SIZE - 12, self._BTN_SIZE - 12)
        for i, comp in enumerate(self._components_data):
            btn = QPushButton()
            btn.setCheckable(True)
            btn.setFixedSize(self._BTN_SIZE, self._BTN_SIZE)
            btn.setToolTip(comp["name"])

            img_path = os.path.join(img_dir, comp.get("image", ""))
            if os.path.isfile(img_path):
                btn.setIcon(QIcon(img_path))
                btn.setIconSize(icon_sz)

            btn.setStyleSheet(
                "QPushButton { font-size:11px; padding:4px; border:2px solid #555; border-radius:6px; }"
                "QPushButton:checked { border:3px solid #4FC3F7; background:#1A3A4A; }"
                "QPushButton:hover { background:#2A2A2A; }"
            )
            row, col = divmod(i, self._GRID_COLS)
            grid_lay.addWidget(btn, row, col)
            self._comp_btn_group.addButton(btn, i)

        ag.addWidget(grid_widget)

        self._comp_btn_group.idClicked.connect(self._on_component_changed)

        # ─── 詳細規格面板 ───

        # Row 1: 子類型 + Rating + 尺寸 (+ 出口尺寸)
        r1 = QHBoxLayout()

        self._subtype_lbl = QLabel("類型:")
        self._subtype_combo = QComboBox()
        self._subtype_combo.setFixedWidth(100)
        r1.addWidget(self._subtype_lbl)
        r1.addWidget(self._subtype_combo)
        self._subtype_combo.currentTextChanged.connect(self._on_subtype_changed)

        self._rating_lbl = QLabel("Rating:")
        self._rating_combo = QComboBox()
        self._rating_combo.setFixedWidth(80)
        r1.addWidget(self._rating_lbl)
        r1.addWidget(self._rating_combo)

        r1.addSpacing(8)
        r1.addWidget(QLabel("尺寸:"))
        self.size_combo = QComboBox()
        self.size_combo.addItems(sizes)
        self.size_combo.setFixedWidth(80)
        r1.addWidget(self.size_combo)

        self._size2_lbl = QLabel("→")
        self._size2_combo = QComboBox()
        self._size2_combo.addItems(sizes)
        self._size2_combo.setFixedWidth(80)
        r1.addWidget(self._size2_lbl)
        r1.addWidget(self._size2_combo)

        r1.addStretch()
        ag.addLayout(r1)

        # Row 1.5 (動態規格): 製法 + Radius + 端口 + 面型
        r15 = QHBoxLayout()

        self._mfg_lbl = QLabel("製法:")
        self._mfg_combo = QComboBox()
        self._mfg_combo.setFixedWidth(80)
        r15.addWidget(self._mfg_lbl)
        r15.addWidget(self._mfg_combo)

        self._radius_lbl = QLabel("Radius:")
        self._radius_combo = QComboBox()
        self._radius_combo.setFixedWidth(65)
        r15.addWidget(self._radius_lbl)
        r15.addWidget(self._radius_combo)

        self._ends_lbl = QLabel("端口:")
        self._ends_combo = QComboBox()
        self._ends_combo.setFixedWidth(90)
        r15.addWidget(self._ends_lbl)
        r15.addWidget(self._ends_combo)

        self._face_lbl = QLabel("面型:")
        self._face_combo = QComboBox()
        self._face_combo.setFixedWidth(70)
        r15.addWidget(self._face_lbl)
        r15.addWidget(self._face_combo)

        r15.addStretch()
        self._spec_row = r15
        ag.addLayout(r15)

        # Row 2: 材質 + SCH + 數量 + 單位 + 備註 + 加入
        r2 = QHBoxLayout()
        r2.addWidget(QLabel("材質:"))
        self._pipe_types = pipe_types
        self.mat_combo = QComboBox()
        for pt in pipe_types:
            self.mat_combo.addItem(pt["name"], pt["id"])
        self.mat_combo.setFixedWidth(170)
        self.mat_combo.currentIndexChanged.connect(self._on_material_changed)
        r2.addWidget(self.mat_combo)

        r2.addSpacing(6)
        self._sch_lbl = QLabel("SCH:")
        self._sch_combo = QComboBox()
        self._sch_combo.addItems(schedules)
        self._sch_combo.setFixedWidth(90)
        r2.addWidget(self._sch_lbl)
        r2.addWidget(self._sch_combo)

        r2.addSpacing(6)
        r2.addWidget(QLabel("數量:"))
        self.qty_spin = QSpinBox()
        self.qty_spin.setRange(1, 999)
        self.qty_spin.setValue(1)
        self.qty_spin.setFixedWidth(72)
        r2.addWidget(self.qty_spin)

        self._unit_lbl = QLabel("")
        self._unit_lbl.setFixedWidth(30)
        self._unit_lbl.setStyleSheet(f"color: {Colors.TEXT_MUTED};")
        r2.addWidget(self._unit_lbl)

        r2.addSpacing(6)
        r2.addWidget(QLabel("備註:"))
        self._remark_edit = QLineEdit()
        self._remark_edit.setPlaceholderText("選填")
        self._remark_edit.setFixedWidth(140)
        r2.addWidget(self._remark_edit)

        r2.addSpacing(8)
        btn_add = QPushButton("➕ 加入")
        btn_add.setFixedHeight(28)
        btn_add.clicked.connect(self._add_material)
        r2.addWidget(btn_add)
        r2.addStretch()
        ag.addLayout(r2)

        lay.addWidget(add_grp)

        # ── 清單 ──
        list_grp = QGroupBox("📋 材料清單（雙擊數量/備註可編輯）")
        lg = QVBoxLayout(list_grp)
        self.tree = QTreeWidget()
        self.tree.setHeaderLabels(["零件", "尺寸", "SCH", "材質", "數量", "備註"])
        self.tree.setAlternatingRowColors(True)
        hdr = self.tree.header()
        hdr.resizeSection(0, 200)
        hdr.resizeSection(1, 100)
        hdr.resizeSection(2, 70)
        hdr.resizeSection(3, 90)
        hdr.resizeSection(4, 70)
        hdr.resizeSection(5, 140)
        self.tree.setItemsExpandable(False)
        self.tree.setRootIsDecorated(False)
        self.tree.itemDoubleClicked.connect(self._on_item_double_click)
        lg.addWidget(self.tree)
        lay.addWidget(list_grp, stretch=1)

        br = QHBoxLayout()
        btn_del = QPushButton("🗑️ 刪除選中")
        btn_del.clicked.connect(self._delete_selected)
        br.addWidget(btn_del)
        btn_clr = QPushButton("🧹 清空")
        btn_clr.clicked.connect(self._clear_all)
        br.addWidget(btn_clr)
        br.addStretch()

        hint = QLabel("💡 此內容將寫入 materials.txt，可留空跳過")
        hint.setStyleSheet(f"color: {Colors.TEXT_MUTED}; border:none; background:transparent;")
        br.addWidget(hint)
        lay.addLayout(br)

        # 預設選第一個按鈕，初始化動態欄位
        first_btn = self._comp_btn_group.button(0)
        if first_btn:
            first_btn.setChecked(True)
        self._on_component_changed()
        self._on_material_changed()

    # ── 動態欄位 ──

    def _get_current_component(self) -> dict:
        idx = self._comp_btn_group.checkedId()
        if 0 <= idx < len(self._components_data):
            return self._components_data[idx]
        return {}

    def _on_component_changed(self):
        """根據零件類型顯示/隱藏動態子欄位"""
        comp = self._get_current_component()
        comp_id = comp.get("id", "")

        # 子類型（彎頭角度 / 法蘭類型 / 閥類型 / 大小頭類型 / Tee / Coupling / Gasket）
        angles = comp.get("angles", [])
        types = comp.get("types", [])
        subtypes = angles or types
        has_sub = bool(subtypes)
        self._subtype_lbl.setVisible(has_sub)
        self._subtype_combo.setVisible(has_sub)
        if has_sub:
            self._subtype_combo.blockSignals(True)
            self._subtype_combo.clear()
            self._subtype_combo.addItems(subtypes)
            self._subtype_combo.blockSignals(False)
            self._subtype_lbl.setText("角度:" if angles else "類型:")

        # Rating（法蘭 + 閥）
        ratings = comp.get("ratings", [])
        has_rat = bool(ratings)
        self._rating_lbl.setVisible(has_rat)
        self._rating_combo.setVisible(has_rat)
        if has_rat:
            self._rating_combo.clear()
            self._rating_combo.addItems(ratings)

        # 出口尺寸（Reducer + Reducing Tee）
        needs_size2 = comp_id == "reducer" or (
            comp_id == "tee" and self._subtype_combo.currentText() == "Reducing")
        self._size2_lbl.setVisible(needs_size2)
        self._size2_combo.setVisible(needs_size2)

        # === Row 1.5 動態規格 ===
        # 製法（Pipe only: SMLS/ERW）
        mfg = comp.get("mfg", [])
        has_mfg = bool(mfg)
        self._mfg_lbl.setVisible(has_mfg)
        self._mfg_combo.setVisible(has_mfg)
        if has_mfg:
            self._mfg_combo.clear()
            self._mfg_combo.addItems(mfg)

        # Radius（Elbow only: LR/SR）
        radius = comp.get("radius", [])
        has_rad = bool(radius)
        self._radius_lbl.setVisible(has_rad)
        self._radius_combo.setVisible(has_rad)
        if has_rad:
            self._radius_combo.clear()
            self._radius_combo.addItems(radius)

        # 端口（Pipe/Cap/Coupling/Union/Valve）
        ends = comp.get("ends", [])
        has_ends = bool(ends)
        self._ends_lbl.setVisible(has_ends)
        self._ends_combo.setVisible(has_ends)
        if has_ends:
            self._ends_combo.clear()
            self._ends_combo.addItems(ends)

        # 面型（Flange only: RF/FF/RTJ）
        faces = comp.get("faces", [])
        has_face = bool(faces)
        self._face_lbl.setVisible(has_face)
        self._face_combo.setVisible(has_face)
        if has_face:
            self._face_combo.clear()
            self._face_combo.addItems(faces)

        # SCH 顯示（管件類需要，消耗品/墊片/其他不需要）
        no_sch = comp_id in ("gasket", "other", "electrode", "filler_wire",
                             "seal_tape", "bolt_nut", "weld_pad")
        self._sch_lbl.setVisible(not no_sch)
        self._sch_combo.setVisible(not no_sch)

        # 單位提示
        unit_map = {"pipe": "M", "electrode": "kg", "filler_wire": "kg",
                    "seal_tape": "卷"}
        self._unit_lbl.setText(unit_map.get(comp_id, "個"))

    def _on_subtype_changed(self, text: str):
        """Reducing Tee 需要出口尺寸"""
        comp = self._get_current_component()
        if comp.get("id") == "tee":
            needs = (text == "Reducing")
            self._size2_lbl.setVisible(needs)
            self._size2_combo.setVisible(needs)

    def _on_material_changed(self):
        """材質變更 → 自動調整預設 SCH"""
        mat_id = self.mat_combo.currentData()
        default_sch = _DEFAULT_SCH.get(mat_id, "SCH 40")
        idx = self._sch_combo.findText(default_sch)
        if idx >= 0:
            self._sch_combo.setCurrentIndex(idx)

    # ── 加入 / 編輯 / 刪除 ──

    def _add_material(self):
        comp = self._get_current_component()
        comp_id = comp.get("id", "")
        comp_name = comp.get("name", "")
        size = self.size_combo.currentText()
        mat_data = None
        mat_idx = self.mat_combo.currentIndex()
        if 0 <= mat_idx < len(self._pipe_types):
            mat_data = self._pipe_types[mat_idx]
        mat_name = mat_data["name"] if mat_data else ""
        mat_abbr = mat_data.get("abbr", "") if mat_data else ""
        qty = self.qty_spin.value()
        _unit_map = {"pipe": "M", "electrode": "kg", "filler_wire": "kg",
                     "seal_tape": "卷"}
        unit = _unit_map.get(comp_id, "個")
        sch = self._sch_combo.currentText() if self._sch_combo.isVisible() else ""
        remark = self._remark_edit.text().strip()

        # 收集所有規格標籤
        specs = []

        # 子類型
        subtype = ""
        if self._subtype_combo.isVisible() and self._subtype_combo.currentText():
            subtype = self._subtype_combo.currentText()
            specs.append(subtype)

        # Radius（Elbow）
        radius = ""
        if self._radius_combo.isVisible() and self._radius_combo.currentText():
            radius = self._radius_combo.currentText()
            specs.append(radius)

        # 製法（Pipe）
        mfg = ""
        if self._mfg_combo.isVisible() and self._mfg_combo.currentText():
            mfg = self._mfg_combo.currentText()
            specs.append(mfg)

        # 端口
        ends = ""
        if self._ends_combo.isVisible() and self._ends_combo.currentText():
            ends = self._ends_combo.currentText()
            specs.append(ends)

        # Rating
        rating = ""
        if self._rating_combo.isVisible() and self._rating_combo.currentText():
            rating = self._rating_combo.currentText()
            specs.append(rating)

        # 面型（Flange）
        face = ""
        if self._face_combo.isVisible() and self._face_combo.currentText():
            face = self._face_combo.currentText()
            specs.append(face)

        # 組合顯示名稱: "specs comp_name" e.g. "90° LR Elbow (彎頭)"
        if specs:
            display_name = f"{' '.join(specs)} {comp_name}"
        else:
            display_name = comp_name

        # 尺寸字串
        size_str = size
        if self._size2_combo.isVisible():
            size_str = f"{size} × {self._size2_combo.currentText()}"

        item = {
            "component": display_name,
            "component_id": comp_id,
            "size": size_str,
            "schedule": sch,
            "material": mat_name,
            "material_abbr": mat_abbr,
            "qty": qty,
            "unit": unit,
            "subtype": subtype,
            "mfg": mfg,
            "radius": radius,
            "ends": ends,
            "rating": rating,
            "face": face,
            "remark": remark,
        }
        self.materials_list.append(item)

        # 材質簡寫
        mat_short = f"{mat_name.split(' ')[0]}({mat_abbr})" if mat_abbr else mat_name
        tree_item = QTreeWidgetItem([
            display_name, size_str, sch, mat_short,
            f"{qty} {unit}", remark,
        ])
        self.tree.addTopLevelItem(tree_item)

        # Sticky: 只清數量和備註
        self.qty_spin.setValue(1)
        self._remark_edit.clear()

        # 法蘭提示加墊片
        if comp_id == "flange":
            r = QMessageBox.question(
                self.page, "配件提示",
                f"已加入 {display_name}，是否一併加入 Gasket (墊片)？",
            )
            if r == QMessageBox.StandardButton.Yes:
                gasket = {
                    "component": "Gasket (墊片)", "component_id": "gasket",
                    "size": size, "schedule": "", "material": mat_name,
                    "material_abbr": mat_abbr, "qty": qty, "unit": "個",
                    "subtype": "", "mfg": "", "radius": "", "ends": "",
                    "rating": "", "face": "",
                    "remark": f"配 {display_name}",
                }
                self.materials_list.append(gasket)
                gi = QTreeWidgetItem([
                    "Gasket (墊片)", size, "", mat_short,
                    f"{qty} 個", f"配 {display_name}",
                ])
                self.tree.addTopLevelItem(gi)

    def _on_item_double_click(self, item: QTreeWidgetItem, column: int):
        """雙擊數量或備註欄位可編輯"""
        if column == 4:  # 數量欄
            idx = self.tree.indexOfTopLevelItem(item)
            if idx < 0 or idx >= len(self.materials_list):
                return
            old = self.materials_list[idx]
            new_qty, ok = QInputDialog.getInt(
                self.page, "修改數量", f"{old['component']} 數量：",
                old["qty"], 1, 999,
            )
            if ok:
                old["qty"] = new_qty
                item.setText(4, f"{new_qty} {old['unit']}")
        elif column == 5:  # 備註欄
            idx = self.tree.indexOfTopLevelItem(item)
            if idx < 0 or idx >= len(self.materials_list):
                return
            old = self.materials_list[idx]
            new_rem, ok = QInputDialog.getText(
                self.page, "修改備註", "備註：", text=old.get("remark", ""),
            )
            if ok:
                old["remark"] = new_rem.strip()
                item.setText(5, new_rem.strip())

    def _delete_selected(self):
        items = self.tree.selectedItems()
        if not items:
            return
        indices = sorted([self.tree.indexOfTopLevelItem(it) for it in items], reverse=True)
        for idx in indices:
            if 0 <= idx < len(self.materials_list):
                self.materials_list.pop(idx)
            self.tree.takeTopLevelItem(idx)

    def _clear_all(self):
        if self.materials_list and QMessageBox.question(
            self.page, "確認", "確定要清空材料清單？"
        ) == QMessageBox.StandardButton.Yes:
            self.materials_list.clear()
            self.tree.clear()

    def get_data(self) -> dict:
        return {"materials_list": self.materials_list.copy()}


# =====================================================================
#  Step7_Confirm
# =====================================================================
class Step7_Confirm(WizardStep):

    def __init__(self, parent, wizard):
        super().__init__(parent, wizard)
        self._folder_name = ""
        self._folder_path = ""
        self._build()

    def _build(self):
        lay = self.layout
        lbl = QLabel("步驟 7/7：確認並建立")
        lbl.setFont(_FONT_TITLE)
        lay.addWidget(lbl)

        # 預覽
        preview_grp = QGroupBox("📋 預覽")
        pg = QVBoxLayout(preview_grp)
        self.preview_text = QTextEdit()
        self.preview_text.setReadOnly(True)
        pg.addWidget(self.preview_text)
        lay.addWidget(preview_grp, stretch=1)

        # 選項
        opt_grp = QGroupBox("⚙️ 選項")
        og = QVBoxLayout(opt_grp)
        self.chk_preprocess = QCheckBox("預處理圖片（調整尺寸）")
        self.chk_preprocess.setChecked(True)
        og.addWidget(self.chk_preprocess)

        self.chk_write_weld = QCheckBox("📝 將新焊口寫入焊口管制表")
        self.chk_write_weld.toggled.connect(self._toggle_weld_options)
        og.addWidget(self.chk_write_weld)

        self.weld_options = QWidget()
        wo_lay = QVBoxLayout(self.weld_options)
        wo_lay.setContentsMargins(20, 0, 0, 0)

        wr = QHBoxLayout()
        btn_preview_w = QPushButton("🔍 預覽新增焊口")
        btn_preview_w.clicked.connect(self._preview_new_welds)
        wr.addWidget(btn_preview_w)
        self.new_weld_label = QLabel("")
        self.new_weld_label.setStyleSheet(f"color: {Colors.PRIMARY}; border:none; background:transparent;")
        wr.addWidget(self.new_weld_label)
        wr.addStretch()
        wo_lay.addLayout(wr)

        nr = QHBoxLayout()
        nr.addWidget(QLabel("另存檔名:"))
        self.output_name_edit = QLineEdit("焊口管制表_更新")
        self.output_name_edit.setFixedWidth(200)
        nr.addWidget(self.output_name_edit)
        nr.addWidget(QLabel(f"_{datetime.now().strftime('%Y.%m.%d')}.xlsx"))
        nr.addStretch()
        wo_lay.addLayout(nr)

        self.weld_options.hide()
        og.addWidget(self.weld_options)
        lay.addWidget(opt_grp)

    def _toggle_weld_options(self, checked: bool):
        self.weld_options.setVisible(checked)
        if checked:
            self._check_new_welds()

    def _check_new_welds(self):
        try:
            nw = self._get_new_welds()
            if nw is None:
                self.new_weld_label.setText("⚠️ 未設定管制表")
                self.new_weld_label.setStyleSheet(f"color: {Colors.WARNING}; border:none; background:transparent;")
            elif len(nw) == 0:
                self.new_weld_label.setText("✅ 無新焊口（全部已存在）")
                self.new_weld_label.setStyleSheet(f"color: {Colors.TEXT_MUTED}; border:none; background:transparent;")
            else:
                self.new_weld_label.setText(f"📋 {len(nw)} 個新焊口待寫入")
                self.new_weld_label.setStyleSheet(f"color: {Colors.PRIMARY}; border:none; background:transparent;")
        except Exception as e:
            self.new_weld_label.setText(f"⚠️ 檢查失敗: {e}")
            self.new_weld_label.setStyleSheet(f"color: {Colors.DANGER}; border:none; background:transparent;")

    def _get_new_welds(self):
        try:
            from settings_manager import get_weld_control_config, get_weld_control_table_path
            from weld_control import init_weld_manager_from_settings

            tp = get_weld_control_table_path()
            if not tp or not os.path.exists(tp):
                return None
            config = get_weld_control_config()
            data = self.wizard.get_collected_data()
            sr = data.get("series", "")
            if not sr:
                return []
            fmt = config.get("serial_format", "raw")
            serial = sr.zfill(4) if fmt == "pad4" else (sr.lstrip("0") or "0")
            welds_full = data.get("welds_list_full", [])
            if not welds_full:
                return []
            manager = init_weld_manager_from_settings()
            if not manager or not manager.load():
                return None
            new = []
            for w in welds_full:
                wid = f"{w['weld_no']}{w['mark']}"
                if not manager.check_exists(serial, wid):
                    new.append({"serial": serial, "weld_id": wid, "code": w["code"],
                                "size": w["size"], "mark": w["mark"], "weld_no": w["weld_no"]})
            return new
        except Exception:
            return None

    def _preview_new_welds(self):
        nw = self._get_new_welds()
        if nw is None:
            QMessageBox.warning(self.page, "提示", "尚未設定焊口管制表路徑")
            return
        if not nw:
            QMessageBox.information(self.page, "預覽", "✅ 所有焊口皆已存在於管制表中\n無需新增")
            return

        dlg = QDialog(self.page)
        dlg.setWindowTitle("📋 新增焊口預覽")
        dlg.resize(500, 400)
        dlg.setModal(True)
        dlay = QVBoxLayout(dlg)
        lb = QLabel(f"以下 {len(nw)} 個焊口將寫入管制表：")
        lb.setFont(_FONT_SUBTITLE)
        dlay.addWidget(lb)
        tree = QTreeWidget()
        tree.setHeaderLabels(["流水號", "焊口編號", "尺寸", "完整代碼"])
        for w in nw:
            tree.addTopLevelItem(QTreeWidgetItem([w["serial"], w["weld_id"], f'{w["size"]}"', w["code"]]))
        dlay.addWidget(tree, stretch=1)
        btn_ok = QPushButton("確定")
        btn_ok.clicked.connect(dlg.close)
        dlay.addWidget(btn_ok, alignment=Qt.AlignmentFlag.AlignCenter)
        dlg.exec()

    def on_show(self):
        self._update_preview()
        if self.chk_write_weld.isChecked():
            self._check_new_welds()

    def _update_preview(self):
        data = self.wizard.get_collected_data()
        if data.get("mode") == "single":
            fn = f"{data['series']}_{data.get('welds_string', '')}"
        else:
            fn = f"{data['series']}_{data.get('group_letter', 'A')}G"
        fp = os.path.join(ATTACHMENTS_ROOT, data["date"], fn)

        lines = [
            "📁 資料夾路徑:", f"   {fp}", "",
            f"📅 日期: {data['date']}", f"🔢 Series: {data['series']}",
            f"📋 模式: {data['mode']}", "",
        ]
        if data.get("mode") == "single":
            lines.append(f"🔧 焊口: {data.get('welds_string', '')}")
        else:
            wl = data.get("welds_list", [])
            lines.append(f"🔧 焊口 ({len(wl)} 個):")
            for w in wl[:10]:
                lines.append(f"   • {w}")
            if len(wl) > 10:
                lines.append(f"   ... 還有 {len(wl) - 10} 個")

        lines.append("")
        lines.append("📷 圖片:")
        images = data.get("images", {})
        any_img = False
        for k, v in images.items():
            if v:
                lines.append(f"   • {k}: {os.path.basename(v)}")
                any_img = True
        if not any_img:
            lines.append("   （稍後手動複製）")

        note = data.get("note_content", "")
        if note:
            lines += ["", "📝 修改原因:"]
            for ln in note.split("\n")[:2]:
                lines.append(f"   {ln}")
            if note.count("\n") > 1:
                lines.append("   ...")

        mats = data.get("materials_list", [])
        if mats:
            lines += ["", f"🔧 材料清單 ({len(mats)} 項)"]

        self.preview_text.setPlainText("\n".join(lines))
        self._folder_name = fn
        self._folder_path = fp

    def get_data(self) -> dict:
        return {
            "folder_name": self._folder_name,
            "folder_path": self._folder_path,
            "preprocess_images": self.chk_preprocess.isChecked(),
            "write_weld_control": self.chk_write_weld.isChecked(),
            "weld_output_name": self.output_name_edit.text() if self.chk_write_weld.isChecked() else "",
        }


# =====================================================================
#  _StagingSidebar — staging/ 待處理檔案預覽
# =====================================================================
_STAGING_THUMB = 64

_TOGGLE_STYLE = (
    "QPushButton { font-size: 9pt; padding: 1px 4px;"
    " border: 1px solid #aaa; border-radius: 3px; background: #f0f0f0; }"
    " QPushButton:hover { background: #ddd; }"
)


class _StagingSidebar(QFrame):
    """顯示 staging/ 資料夾中尚未處理的檔案清單"""

    SIDEBAR_W = 240
    SIDEBAR_MIN = 160

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setMinimumWidth(self.SIDEBAR_MIN)
        self.setMaximumWidth(600)
        self.resize(self.SIDEBAR_W, 0)
        self.setFrameShape(QFrame.Shape.StyledPanel)
        self.setStyleSheet(
            f"_StagingSidebar {{ background: {Colors.BG}; "
            f"border-left: 2px solid {Colors.BORDER}; }}"
        )
        self._collapsed = False
        self._build_ui()

    def _build_ui(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(6, 8, 6, 4)
        outer.setSpacing(4)

        # 標題行
        hdr = QHBoxLayout()
        self._title_lbl = QLabel("Staging 待處理")
        self._title_lbl.setFont(Fonts.subheading())
        self._title_lbl.setStyleSheet(
            f"color: {Colors.TEXT}; border: none; background: transparent;"
        )
        hdr.addWidget(self._title_lbl)
        hdr.addStretch()
        self._toggle_btn = QPushButton("收")
        self._toggle_btn.setFixedSize(32, 24)
        self._toggle_btn.setStyleSheet(_TOGGLE_STYLE)
        self._toggle_btn.setToolTip("收合/展開")
        self._toggle_btn.clicked.connect(self._toggle)
        hdr.addWidget(self._toggle_btn)
        outer.addLayout(hdr)

        # 工具列：匯入照片 + 開啟資料夾
        tool_row = QHBoxLayout()
        tool_row.setSpacing(4)

        self._btn_import = QPushButton("匯入照片")
        self._btn_import.setFont(QFont("Segoe UI", 8))
        self._btn_import.setFixedHeight(24)
        self._btn_import.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_import.clicked.connect(self._do_import)
        tool_row.addWidget(self._btn_import)

        self._btn_open_folder = QPushButton("開啟資料夾")
        self._btn_open_folder.setFont(QFont("Segoe UI", 8))
        self._btn_open_folder.setFixedHeight(24)
        self._btn_open_folder.setCursor(Qt.CursorShape.PointingHandCursor)
        self._btn_open_folder.clicked.connect(self._do_open_folder)
        tool_row.addWidget(self._btn_open_folder)

        outer.addLayout(tool_row)

        # 計數
        self._count_lbl = QLabel("")
        self._count_lbl.setFont(QFont("Segoe UI", 8))
        self._count_lbl.setStyleSheet(
            f"color: {Colors.TEXT_MUTED}; border: none; background: transparent;"
        )
        outer.addWidget(self._count_lbl)

        # 可捲動的卡片區
        self._scroll = QScrollArea()
        self._scroll.setWidgetResizable(True)
        self._scroll.setStyleSheet(
            "QScrollArea { border: none; background: transparent; }"
        )
        self._scroll_content = QWidget()
        self._scroll_layout = QVBoxLayout(self._scroll_content)
        self._scroll_layout.setContentsMargins(0, 0, 0, 0)
        self._scroll_layout.setSpacing(6)
        self._scroll_layout.addStretch()
        self._scroll.setWidget(self._scroll_content)
        outer.addWidget(self._scroll, stretch=1)

        # 空白提示
        self._empty_lbl = QLabel("staging/ 資料夾為空\n無待處理檔案")
        self._empty_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._empty_lbl.setStyleSheet(
            f"color: {Colors.TEXT_MUTED}; border: none; background: transparent;"
            " padding: 20px 0;"
        )
        self._empty_lbl.setWordWrap(True)
        outer.addWidget(self._empty_lbl)
        self._scroll.hide()

    # ── 收合 / 展開 ──

    def _toggle(self):
        self._collapsed = not self._collapsed
        if self._collapsed:
            self._scroll.hide()
            self._empty_lbl.hide()
            self._count_lbl.hide()
            self._btn_import.hide()
            self._btn_open_folder.hide()
            self._toggle_btn.setText("展")
            self.setMinimumWidth(36)
            self.setMaximumWidth(36)
            self._title_lbl.hide()
        else:
            self._toggle_btn.setText("收")
            self.setMinimumWidth(self.SIDEBAR_MIN)
            self.setMaximumWidth(600)
            self._title_lbl.show()
            self._count_lbl.show()
            self._btn_import.show()
            self._btn_open_folder.show()
            self._refresh_visibility()

    def _refresh_visibility(self):
        has = self._scroll_layout.count() > 1
        self._scroll.setVisible(has)
        self._empty_lbl.setVisible(not has)

    # ── 頂層工具 ──

    @staticmethod
    def _staging_dir() -> str:
        return os.path.join(BASE_DIR, "staging")

    def _do_import(self):
        """匯入複數照片到 staging/"""
        files, _ = QFileDialog.getOpenFileNames(
            self, "匯入照片到 staging",
            "", "圖片檔案 (*.jpg *.jpeg *.png *.bmp);;所有檔案 (*)"
        )
        if not files:
            return
        staging = self._staging_dir()
        os.makedirs(staging, exist_ok=True)
        for src in files:
            dst = os.path.join(staging, os.path.basename(src))
            if os.path.abspath(src) == os.path.abspath(dst):
                continue
            shutil.copy2(src, dst)
        self.refresh()

    def _do_open_folder(self):
        """用系統檔案總管打開 staging/ 資料夾"""
        staging = self._staging_dir()
        os.makedirs(staging, exist_ok=True)
        os.startfile(staging)

    # ── 卡片內操作 ──

    def _do_open_file(self, fpath: str):
        """開啟檔案"""
        if os.path.exists(fpath):
            os.startfile(fpath)

    def _do_delete_file(self, fpath: str):
        """刪除檔案"""
        fname = os.path.basename(fpath)
        if QMessageBox.question(
            self, "移到隔離區", f"確定要將檔案移到隔離區？\n{fname}",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        ) != QMessageBox.StandardButton.Yes:
            return
        try:
            move_to_trash(fpath, BASE_DIR, reason="delete_staging_file")
        except Exception as e:
            QMessageBox.warning(self, "移動失敗", str(e))
        self.refresh()

    def _do_rename_file(self, fpath: str):
        """改名檔案"""
        staging = self._staging_dir()
        fname = os.path.basename(fpath)
        new_name, ok = QInputDialog.getText(
            self, "檔案改名", "新檔名:", text=fname
        )
        if not ok or not new_name.strip() or new_name.strip() == fname:
            return
        new_path = os.path.join(staging, new_name.strip())
        if os.path.exists(new_path):
            QMessageBox.warning(self, "錯誤", f"檔案已存在：{new_name.strip()}")
            return
        try:
            os.rename(fpath, new_path)
        except Exception as e:
            QMessageBox.warning(self, "改名失敗", str(e))
        self.refresh()

    def _do_annotate(self, fpath: str):
        """開啟標註對話框"""
        if not os.path.exists(fpath):
            return
        is_pdf = fpath.lower().endswith(".pdf")
        from gui_annotator import AnnotationDialog
        dlg = AnnotationDialog(fpath, is_pdf=is_pdf, parent=self)
        if not dlg._load_ok:
            QMessageBox.warning(self, "無法載入", f"無法開啟標註工具:\n{fpath}")
            return
        dlg.exec()
        self.refresh()

    # ── 公開方法 ──

    def refresh(self):
        """掃描 staging/ 資料夾並重建卡片"""
        self._clear_cards()
        staging_dir = self._staging_dir()
        if not os.path.isdir(staging_dir):
            self._count_lbl.setText("staging/ 資料夾不存在")
            self._refresh_visibility()
            return

        files = sorted(
            f for f in os.listdir(staging_dir)
            if os.path.isfile(os.path.join(staging_dir, f))
        )
        if not files:
            self._count_lbl.setText("無待處理檔案")
            self._refresh_visibility()
            return

        # 分類
        images = [f for f in files if f.lower().endswith((".jpg", ".jpeg", ".png", ".bmp"))]
        pdfs = [f for f in files if f.lower().endswith(".pdf")]
        others = [f for f in files if f not in images and f not in pdfs]

        self._count_lbl.setText(
            f"共 {len(files)} 個檔案  "
            f"(圖{len(images)} PDF{len(pdfs)}"
            + (f" 其他{len(others)}" if others else "")
            + ")"
        )

        # 圖片
        for fname in images:
            fpath = os.path.join(staging_dir, fname)
            card = self._make_file_card(fname, fpath, "image")
            self._scroll_layout.insertWidget(
                self._scroll_layout.count() - 1, card
            )

        # PDF
        for fname in pdfs:
            fpath = os.path.join(staging_dir, fname)
            card = self._make_file_card(fname, fpath, "pdf")
            self._scroll_layout.insertWidget(
                self._scroll_layout.count() - 1, card
            )

        # 其他
        for fname in others:
            fpath = os.path.join(staging_dir, fname)
            card = self._make_file_card(fname, fpath, "other")
            self._scroll_layout.insertWidget(
                self._scroll_layout.count() - 1, card
            )

        self._refresh_visibility()

    # ── 內部 ──

    def _clear_cards(self):
        while self._scroll_layout.count() > 1:
            item = self._scroll_layout.takeAt(0)
            w = item.widget()
            if w:
                w.deleteLater()

    def _make_file_card(self, fname: str, fpath: str, ftype: str) -> QFrame:
        card = QFrame()
        card.setStyleSheet(
            f"QFrame {{ background: {Colors.BG_WHITE};"
            f" border: 1px solid {Colors.BORDER_LIGHT};"
            f" border-radius: 6px; }}"
        )
        clay = QVBoxLayout(card)
        clay.setContentsMargins(4, 4, 4, 4)
        clay.setSpacing(3)

        # 檔名標題
        name_lbl = QLabel(fname)
        name_lbl.setFont(QFont("Segoe UI", 7))
        name_lbl.setWordWrap(True)
        name_lbl.setToolTip(fpath)
        name_lbl.setStyleSheet(
            f"color: {Colors.TEXT}; border: none; background: transparent;"
        )
        clay.addWidget(name_lbl)

        # 縮圖（使用 _HistoryThumb 支援 hover 放大）
        if ftype == "image":
            pm = QPixmap(fpath)
            if not pm.isNull():
                thumb = _HistoryThumb(pm, fname)
                clay.addWidget(thumb)
        elif ftype == "pdf":
            pdf_pm = _render_pdf_page(fpath, 320)
            if pdf_pm and not pdf_pm.isNull():
                thumb = _HistoryThumb(pdf_pm, fname)
                clay.addWidget(thumb)
            else:
                icon_lbl = QLabel("PDF")
                icon_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
                icon_lbl.setStyleSheet(
                    f"color: {Colors.TEXT_MUTED}; border: none; background: transparent;"
                )
                clay.addWidget(icon_lbl)
        else:
            icon_lbl = QLabel("檔案")
            icon_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            icon_lbl.setStyleSheet(
                f"color: {Colors.TEXT_MUTED}; border: none; background: transparent;"
            )
            clay.addWidget(icon_lbl)

        # 操作按鈕列
        btn_row = QHBoxLayout()
        btn_row.setSpacing(2)
        _btn_font = QFont("Segoe UI", 7)

        btn_open = QPushButton("開啟")
        btn_open.setFont(_btn_font)
        btn_open.setFixedHeight(20)
        btn_open.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_open.clicked.connect(lambda _, p=fpath: self._do_open_file(p))
        btn_row.addWidget(btn_open)

        btn_rename = QPushButton("改名")
        btn_rename.setFont(_btn_font)
        btn_rename.setFixedHeight(20)
        btn_rename.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_rename.clicked.connect(lambda _, p=fpath: self._do_rename_file(p))
        btn_row.addWidget(btn_rename)

        btn_del = QPushButton("刪除")
        btn_del.setFont(_btn_font)
        btn_del.setFixedHeight(20)
        btn_del.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_del.setStyleSheet("QPushButton { color: #c0392b; }")
        btn_del.clicked.connect(lambda _, p=fpath: self._do_delete_file(p))
        btn_row.addWidget(btn_del)

        # 標註按鈕（僅圖片 / PDF）
        if ftype in ("image", "pdf"):
            btn_anno = QPushButton("標註")
            btn_anno.setFont(_btn_font)
            btn_anno.setFixedHeight(20)
            btn_anno.setCursor(Qt.CursorShape.PointingHandCursor)
            btn_anno.clicked.connect(lambda _, p=fpath: self._do_annotate(p))
            btn_row.addWidget(btn_anno)

        clay.addLayout(btn_row)

        return card


# =====================================================================
#  _HistoryThumb — 歷史圖片/PDF 縮圖（hover 放大）
# =====================================================================
_HIST_THUMB = 80


class _HistoryThumb(QLabel):
    """小縮圖 label，hover 時使用 _ZoomPopup 放大"""

    def __init__(self, pixmap: QPixmap, tag: str = "", parent=None):
        super().__init__(parent)
        self._full_pixmap = pixmap
        self.setFixedSize(_HIST_THUMB, _HIST_THUMB)
        self.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.setStyleSheet(
            f"border: 1px solid {Colors.BORDER}; border-radius: 4px; background: white;"
        )
        self.setToolTip(tag)
        if pixmap and not pixmap.isNull():
            self.setPixmap(pixmap.scaled(
                _HIST_THUMB, _HIST_THUMB,
                Qt.AspectRatioMode.KeepAspectRatio,
                Qt.TransformationMode.SmoothTransformation,
            ))
        else:
            self.setText(tag or "—")
            self.setFont(QFont("Segoe UI", 9))
        self.setMouseTracking(True)

    def enterEvent(self, event):
        super().enterEvent(event)
        self.setStyleSheet(
            f"border: 2px solid {Colors.PRIMARY}; border-radius: 4px; background: white;"
        )
        if self._full_pixmap and not self._full_pixmap.isNull():
            popup = _get_zoom_popup()
            popup.show_image(self._full_pixmap, self.mapToGlobal(self.rect().topRight()))

    def leaveEvent(self, event):
        super().leaveEvent(event)
        self.setStyleSheet(
            f"border: 1px solid {Colors.BORDER}; border-radius: 4px; background: white;"
        )
        _get_zoom_popup().hide()

    def mouseMoveEvent(self, event):
        super().mouseMoveEvent(event)
        popup = _get_zoom_popup()
        if popup.isVisible() and self._full_pixmap:
            popup.move(event.globalPosition().toPoint().x() + 16,
                       event.globalPosition().toPoint().y() + 16)


def _render_pdf_page(pdf_path: str, size: int = 320) -> Optional[QPixmap]:
    """用 PyMuPDF 將 PDF 第一頁渲染為 QPixmap"""
    if not _FITZ_OK or not os.path.exists(pdf_path):
        return None
    try:
        doc = _fitz.open(pdf_path)
        page = doc[0]
        # 計算適當的 zoom 使短邊 ≈ size
        rect = page.rect
        zoom = size / min(rect.width, rect.height)
        mat = _fitz.Matrix(zoom, zoom)
        pix = page.get_pixmap(matrix=mat, alpha=False)
        img = QImage(pix.samples, pix.width, pix.height, pix.stride, QImage.Format.Format_RGB888)
        pm = QPixmap.fromImage(img)
        doc.close()
        return pm
    except Exception:
        return None


# =====================================================================
#  _HistorySidebar — 流水號歷史紀錄側邊欄
# =====================================================================
class _HistorySidebar(QFrame):
    """顯示指定流水號的所有歷史 attachments 紀錄"""

    SIDEBAR_W = 290
    SIDEBAR_MIN = 180

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setMinimumWidth(self.SIDEBAR_MIN)
        self.setMaximumWidth(600)
        self.resize(self.SIDEBAR_W, 0)
        self.setFrameShape(QFrame.Shape.StyledPanel)
        self.setStyleSheet(
            f"_HistorySidebar {{ background: {Colors.BG}; "
            f"border-left: 2px solid {Colors.BORDER}; }}"
        )

        self._current_serial = ""
        self._build_ui()

    def _build_ui(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(6, 8, 6, 4)
        outer.setSpacing(4)

        # 標題行
        hdr = QHBoxLayout()
        self._title_lbl = QLabel("📜 歷史紀錄")
        self._title_lbl.setFont(Fonts.subheading())
        self._title_lbl.setStyleSheet(f"color: {Colors.TEXT}; border: none; background: transparent;")
        hdr.addWidget(self._title_lbl)
        hdr.addStretch()
        self._toggle_btn = QPushButton("收")
        self._toggle_btn.setFixedSize(32, 24)
        self._toggle_btn.setStyleSheet(_TOGGLE_STYLE)
        self._toggle_btn.setToolTip("收合/展開")
        self._toggle_btn.clicked.connect(self._toggle)
        hdr.addWidget(self._toggle_btn)
        outer.addLayout(hdr)

        # 計數
        self._count_lbl = QLabel("")
        self._count_lbl.setFont(QFont("Segoe UI", 8))
        self._count_lbl.setStyleSheet(f"color: {Colors.TEXT_MUTED}; border: none; background: transparent;")
        outer.addWidget(self._count_lbl)

        # 可捲動的卡片區
        self._scroll = QScrollArea()
        self._scroll.setWidgetResizable(True)
        self._scroll.setStyleSheet("QScrollArea { border: none; background: transparent; }")
        self._scroll_content = QWidget()
        self._scroll_layout = QVBoxLayout(self._scroll_content)
        self._scroll_layout.setContentsMargins(0, 0, 0, 0)
        self._scroll_layout.setSpacing(6)
        self._scroll_layout.addStretch()
        self._scroll.setWidget(self._scroll_content)
        outer.addWidget(self._scroll, stretch=1)

        # 空白提示
        self._empty_lbl = QLabel("輸入流水號後\n自動顯示歷史紀錄")
        self._empty_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._empty_lbl.setStyleSheet(
            f"color: {Colors.TEXT_MUTED}; border: none; background: transparent; padding: 30px 0;"
        )
        self._empty_lbl.setWordWrap(True)
        outer.addWidget(self._empty_lbl)
        self._scroll.hide()

        self._collapsed = False

    def _toggle(self):
        self._collapsed = not self._collapsed
        if self._collapsed:
            self._scroll.hide()
            self._empty_lbl.hide()
            self._count_lbl.hide()
            self._toggle_btn.setText("展")
            self.setMinimumWidth(36)
            self.setMaximumWidth(36)
            self._title_lbl.hide()
        else:
            self._toggle_btn.setText("收")
            self.setMinimumWidth(self.SIDEBAR_MIN)
            self.setMaximumWidth(600)
            self._title_lbl.show()
            self._count_lbl.show()
            self._refresh_visibility()

    def _refresh_visibility(self):
        has = self._scroll_layout.count() > 1  # >1 because of the stretch
        self._scroll.setVisible(has)
        self._empty_lbl.setVisible(not has)

    # ── 公開方法 ──

    def update_serial(self, serial_raw: str):
        """根據流水號掃描 attachments/ 並重建卡片"""
        serial = serial_raw.strip()
        if serial == self._current_serial:
            return
        self._current_serial = serial
        self._clear_cards()

        if not serial:
            self._title_lbl.setText("📜 歷史紀錄")
            self._count_lbl.setText("")
            self._refresh_visibility()
            return

        # 掃描 attachments/ 中所有日期子資料夾
        records = self._scan_attachments(serial)
        self._title_lbl.setText(f"📜 流水號 {serial}")

        if not records:
            self._count_lbl.setText("尚無歷史紀錄")
            self._refresh_visibility()
            return

        self._count_lbl.setText(f"共 {len(records)} 筆歷史")

        for rec in records:
            card = self._make_card(rec)
            # insert before the stretch
            self._scroll_layout.insertWidget(self._scroll_layout.count() - 1, card)

        self._refresh_visibility()

    def _clear_cards(self):
        while self._scroll_layout.count() > 1:
            item = self._scroll_layout.takeAt(0)
            w = item.widget()
            if w:
                w.deleteLater()

    def _scan_attachments(self, serial: str) -> list:
        """掃描 attachments/ 找出符合流水號的資料夾"""
        results = []
        serial_num = serial.lstrip("0") or "0"

        if not os.path.isdir(ATTACHMENTS_ROOT):
            return results

        for date_dir in sorted(os.listdir(ATTACHMENTS_ROOT), reverse=True):
            date_path = os.path.join(ATTACHMENTS_ROOT, date_dir)
            if not os.path.isdir(date_path) or not re.fullmatch(r"\d{8}", date_dir):
                continue
            for folder in sorted(os.listdir(date_path)):
                folder_path = os.path.join(date_path, folder)
                if not os.path.isdir(folder_path):
                    continue
                # 資料夾名稱格式: {series}_{焊口}  or {series}_{letter}G
                parts = folder.split("_", 1)
                folder_num = parts[0].lstrip("0") or "0"
                if folder_num == serial_num:
                    # Read weld info
                    welds_str = parts[1] if len(parts) > 1 else ""
                    note = ""
                    note_path = os.path.join(folder_path, "note.txt")
                    if os.path.exists(note_path):
                        try:
                            with open(note_path, "r", encoding="utf-8") as f:
                                note = f.read(200).strip()
                                if note.startswith("#"):
                                    note = ""
                        except Exception:
                            pass
                    results.append({
                        "date": date_dir,
                        "folder": folder,
                        "folder_path": folder_path,
                        "welds": welds_str,
                        "note": note,
                    })
        return results

    @staticmethod
    def _find_images(folder_path: str) -> list:
        """找出 before/after 圖片（支援 single + group 模式）

        Returns: [(path, label), ...]
        """
        result = []
        files = set(os.listdir(folder_path)) if os.path.isdir(folder_path) else set()
        # single: before.jpg / after.jpg
        if "before.jpg" in files:
            result.append((os.path.join(folder_path, "before.jpg"), "修改前"))
        if "after.jpg" in files:
            result.append((os.path.join(folder_path, "after.jpg"), "修改後"))
        # group: before_1.jpg, before_2.jpg, ... / after_1.jpg, after_2.jpg, ...
        idx = 1
        while True:
            bf = f"before_{idx}.jpg"
            af = f"after_{idx}.jpg"
            found = False
            if bf in files:
                result.append((os.path.join(folder_path, bf), f"修改前 {idx}"))
                found = True
            if af in files:
                result.append((os.path.join(folder_path, af), f"修改後 {idx}"))
                found = True
            if not found:
                break
            idx += 1
        return result

    def _make_card(self, rec: dict) -> QFrame:
        """建立一張歷史紀錄卡片"""
        card = QFrame()
        card.setStyleSheet(
            f"QFrame {{ background: {Colors.BG_WHITE}; border: 1px solid {Colors.BORDER_LIGHT};"
            f" border-radius: 6px; }}"
        )
        clay = QVBoxLayout(card)
        clay.setContentsMargins(6, 6, 6, 6)
        clay.setSpacing(4)

        # 日期 + 焊口
        date_lbl = QLabel(f"📅 {rec['date']}")
        date_lbl.setFont(QFont("Segoe UI", 9, QFont.Weight.Bold))
        date_lbl.setStyleSheet(f"color: {Colors.PRIMARY}; border:none; background:transparent;")
        clay.addWidget(date_lbl)

        welds_lbl = QLabel(f"🔧 {rec['welds']}")
        welds_lbl.setFont(QFont("Segoe UI", 8))
        welds_lbl.setWordWrap(True)
        welds_lbl.setStyleSheet(f"color: {Colors.TEXT}; border:none; background:transparent;")
        clay.addWidget(welds_lbl)

        if rec["note"]:
            note_lbl = QLabel(f"💬 {rec['note'][:60]}")
            note_lbl.setFont(QFont("Segoe UI", 7))
            note_lbl.setWordWrap(True)
            note_lbl.setStyleSheet(f"color: {Colors.TEXT_SECONDARY}; border:none; background:transparent;")
            note_lbl.setToolTip(rec["note"])
            clay.addWidget(note_lbl)

        # 縮圖列: before/after (支援 group 模式) + PDF
        thumb_row = QHBoxLayout()
        thumb_row.setSpacing(4)

        for img_path, label in self._find_images(rec["folder_path"]):
            if os.path.exists(img_path):
                pm = QPixmap(img_path)
                thumb = _HistoryThumb(pm, label)
            else:
                thumb = _HistoryThumb(QPixmap(), f"無{label}")
            thumb_row.addWidget(thumb)

        # PDF — 找資料夾中的 .pdf 檔
        pdf_files = [f for f in os.listdir(rec["folder_path"]) if f.lower().endswith(".pdf")]
        if pdf_files:
            pdf_path = os.path.join(rec["folder_path"], pdf_files[0])
            pdf_pm = _render_pdf_page(pdf_path, 320)
            if pdf_pm and not pdf_pm.isNull():
                thumb = _HistoryThumb(pdf_pm, f"📄 {pdf_files[0]}")
            else:
                thumb = _HistoryThumb(QPixmap(), "📄 PDF")
            thumb.setCursor(Qt.CursorShape.PointingHandCursor)
            thumb.mousePressEvent = lambda e, p=pdf_path: os.startfile(p)
            thumb_row.addWidget(thumb)

        thumb_row.addStretch()
        clay.addLayout(thumb_row)

        # 開啟資料夾按鈕
        btn_open = QPushButton("📂 開啟")
        btn_open.setFixedHeight(22)
        btn_open.setFont(QFont("Segoe UI", 7))
        btn_open.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_open.clicked.connect(lambda _, p=rec["folder_path"]: os.startfile(p))
        clay.addWidget(btn_open)

        return card


# =====================================================================
#  _ImagePreviewSidebar — 已選圖片預覽側邊欄
# =====================================================================
_PREVIEW_THUMB = 100


class _ImagePreviewSidebar(QFrame):
    """顯示精靈已選取的照片/PDF 縮圖，附標註與替換功能"""

    SIDEBAR_W = 200
    SIDEBAR_MIN = 140

    def __init__(self, parent=None):
        super().__init__(parent)
        self._wizard = None          # 稍後由 FolderWizard 設定
        self._image_paths: dict = {}  # key → path 參照
        self.setMinimumWidth(self.SIDEBAR_MIN)
        self.setMaximumWidth(500)
        self.resize(self.SIDEBAR_W, 0)
        self.setFrameShape(QFrame.Shape.StyledPanel)
        self.setStyleSheet(
            f"_ImagePreviewSidebar {{ background: {Colors.BG}; "
            f"border-left: 2px solid {Colors.BORDER}; }}"
        )
        self._build_ui()
        self.hide()

    def set_wizard(self, wizard):
        self._wizard = wizard

    def _build_ui(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(6, 8, 6, 4)
        outer.setSpacing(4)

        title = QLabel("🖼️ 已選圖片")
        title.setFont(Fonts.subheading())
        title.setStyleSheet(f"color: {Colors.TEXT}; border: none; background: transparent;")
        outer.addWidget(title)

        self._scroll = QScrollArea()
        self._scroll.setWidgetResizable(True)
        self._scroll.setStyleSheet("QScrollArea { border: none; background: transparent; }")
        self._scroll_content = QWidget()
        self._container = QVBoxLayout(self._scroll_content)
        self._container.setContentsMargins(0, 0, 0, 0)
        self._container.setSpacing(8)
        self._container.addStretch()
        self._scroll.setWidget(self._scroll_content)
        outer.addWidget(self._scroll, stretch=1)

    def _clear(self):
        while self._container.count() > 1:  # keep trailing stretch
            child = self._container.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

    # ---- 公開 ----

    def update_images(self, image_paths: dict):
        """根據 image_paths 重建縮圖卡片"""
        self._image_paths = image_paths
        self._clear()

        _labels = {
            "before": "修改前", "after": "修改後", "pdf": "📄 圖面 PDF",
            "before_1": "修改前 1", "before_2": "修改前 2",
            "after_1": "修改後 1", "after_2": "修改後 2",
        }
        _order = ["before", "before_1", "before_2", "after", "after_1", "after_2", "pdf"]

        any_shown = False
        for key in _order:
            path = image_paths.get(key, "").strip()
            if not path or not os.path.exists(path):
                continue
            card = self._make_card(key, path, _labels.get(key, key))
            self._container.insertWidget(self._container.count() - 1, card)
            any_shown = True

        self.setVisible(any_shown)

    def _make_card(self, key: str, path: str, display: str) -> QFrame:
        card = QFrame()
        card.setStyleSheet(
            f"QFrame {{ background: {Colors.BG_WHITE}; border: 1px solid {Colors.BORDER_LIGHT};"
            f" border-radius: 6px; }}"
        )
        clay = QVBoxLayout(card)
        clay.setContentsMargins(4, 4, 4, 4)
        clay.setSpacing(3)

        # 標題
        lbl = QLabel(display)
        lbl.setFont(QFont("Segoe UI", 8, QFont.Weight.Bold))
        lbl.setStyleSheet(f"color: {Colors.TEXT_SECONDARY}; border: none; background: transparent;")
        clay.addWidget(lbl)

        # 縮圖
        is_pdf = (key == "pdf")
        if is_pdf:
            pm = _render_pdf_page(path, 320)
            if not pm or pm.isNull():
                pm = QPixmap()
        else:
            pm = QPixmap(path)

        thumb = QLabel()
        thumb.setFixedSize(_PREVIEW_THUMB, _PREVIEW_THUMB)
        thumb.setAlignment(Qt.AlignmentFlag.AlignCenter)
        thumb.setStyleSheet(
            f"border: 1px solid {Colors.BORDER}; border-radius: 4px; background: white;"
        )
        thumb.setMouseTracking(True)
        if pm and not pm.isNull():
            thumb.setPixmap(pm.scaled(
                _PREVIEW_THUMB, _PREVIEW_THUMB,
                Qt.AspectRatioMode.KeepAspectRatio,
                Qt.TransformationMode.SmoothTransformation,
            ))
            thumb._full_pixmap = pm

            def _make_handlers(t):
                def _enter(e):
                    t.setStyleSheet(
                        f"border: 2px solid {Colors.PRIMARY}; border-radius: 4px; background: white;"
                    )
                    if hasattr(t, '_full_pixmap') and t._full_pixmap:
                        _get_zoom_popup().show_image(
                            t._full_pixmap, t.mapToGlobal(t.rect().topRight())
                        )
                def _leave(e):
                    t.setStyleSheet(
                        f"border: 1px solid {Colors.BORDER}; border-radius: 4px; background: white;"
                    )
                    _get_zoom_popup().hide()
                def _move(e):
                    if _get_zoom_popup().isVisible():
                        _get_zoom_popup().move(
                            e.globalPosition().toPoint().x() + 16,
                            e.globalPosition().toPoint().y() + 16,
                        )
                return _enter, _leave, _move

            _enter, _leave, _move = _make_handlers(thumb)
            thumb.enterEvent = _enter
            thumb.leaveEvent = _leave
            thumb.mouseMoveEvent = _move
        else:
            thumb.setText("🖼️" if not is_pdf else "📄")
            thumb.setFont(QFont("Segoe UI", 16))

        clay.addWidget(thumb)

        # 按鈕列：標註 + 替換
        btn_row = QHBoxLayout()
        btn_row.setSpacing(4)
        btn_ann = QPushButton("標註")
        btn_ann.setFixedHeight(22)
        btn_ann.setFont(QFont("Segoe UI", 7))
        btn_ann.setToolTip("開啟標註工具")
        btn_ann.clicked.connect(lambda _, k=key, p=is_pdf: self._do_annotate(k, p))
        btn_row.addWidget(btn_ann)

        btn_rep = QPushButton("替換")
        btn_rep.setFixedHeight(22)
        btn_rep.setFont(QFont("Segoe UI", 7))
        btn_rep.setToolTip("替換圖片")
        btn_rep.clicked.connect(lambda _, k=key, p=is_pdf: self._do_replace(k, p))
        btn_row.addWidget(btn_rep)
        btn_row.addStretch()
        clay.addLayout(btn_row)

        return card

    # ---- 操作 ----

    def _do_annotate(self, key: str, is_pdf: bool):
        path = self._image_paths.get(key, "").strip()
        if not path or not os.path.exists(path):
            return
        from gui_annotator import AnnotationDialog
        dlg = AnnotationDialog(path, is_pdf=is_pdf, parent=self)
        if not dlg._load_ok:
            QMessageBox.warning(self, "無法載入", f"無法開啟標註工具:\n{path}")
            return
        dlg.exec()
        if dlg.was_saved and dlg.saved_path:
            new_path = dlg.saved_path
            self._image_paths[key] = new_path
            # 同步回 Step 的 edit
            if self._wizard:
                step_img = self._wizard._get_images_step()
                if step_img and key in step_img._edits:
                    step_img._edits[key].setText(new_path)
            self.update_images(self._image_paths)

    def _do_replace(self, key: str, is_pdf: bool):
        if is_pdf:
            filt = "PDF 檔案 (*.pdf)"
        else:
            filt = "圖片檔 (*.jpg *.jpeg *.png);;所有檔案 (*.*)"
        path, _ = QFileDialog.getOpenFileName(self, f"替換 {key}", "", filt)
        if not path:
            return
        self._image_paths[key] = path
        # 同步回 Step4 的 edit
        if self._wizard:
            step_img = self._wizard._get_images_step()
            if step_img and key in step_img._edits:
                step_img._edits[key].setText(path)
        self.update_images(self._image_paths)


# =====================================================================
#  FolderWizard — 主精靈
# =====================================================================
class FolderWizard(QDialog):
    """前置作業精靈主類別"""

    def __init__(self, parent: Optional[QWidget] = None):
        super().__init__(parent)
        self.setWindowTitle("📁 報告資料夾建立精靈")
        self.resize(960, 640)
        self.setMinimumSize(780, 620)

        self.collected_data: dict = {}
        self.current_step = 0
        self.steps: List[WizardStep] = []

        self._build_ui()
        self._create_steps()
        self._show_step(0)

    def _build_ui(self):
        root = QVBoxLayout(self)

        # 橫向：左側精靈 + 右側側邊欄（可拉伸）
        self._splitter = QSplitter(Qt.Orientation.Horizontal)
        self._splitter.setChildrenCollapsible(False)
        self._splitter.setHandleWidth(4)
        self._splitter.setStyleSheet(
            "QSplitter::handle { background: #ccc; }"
            " QSplitter::handle:hover { background: #999; }"
        )

        self.stack = QStackedWidget()
        self._splitter.addWidget(self.stack)

        self._sidebar = _HistorySidebar()
        self._splitter.addWidget(self._sidebar)
        self._img_preview = _ImagePreviewSidebar()
        self._img_preview.set_wizard(self)
        self._splitter.addWidget(self._img_preview)
        self._staging_sidebar = _StagingSidebar()
        self._splitter.addWidget(self._staging_sidebar)

        # 設定初始寬度比例：stack 佔大部分
        self._splitter.setStretchFactor(0, 1)   # stack
        self._splitter.setStretchFactor(1, 0)   # history
        self._splitter.setStretchFactor(2, 0)   # img preview
        self._splitter.setStretchFactor(3, 0)   # staging

        root.addWidget(self._splitter, stretch=1)

        # 按鈕列
        btn_row = QHBoxLayout()
        self.prev_btn = QPushButton("← 上一步")
        self.prev_btn.clicked.connect(self._prev_step)
        btn_row.addWidget(self.prev_btn)

        self.progress_label = QLabel("步驟 1/7")
        btn_row.addSpacing(20)
        btn_row.addWidget(self.progress_label)
        btn_row.addStretch()

        self.cancel_btn = QPushButton("取消")
        self.cancel_btn.clicked.connect(self._cancel)
        btn_row.addWidget(self.cancel_btn)

        self.next_btn = QPushButton("下一步 →")
        set_button_role(self.next_btn, "primary")
        self.next_btn.clicked.connect(self._next_step)
        btn_row.addWidget(self.next_btn)
        root.addLayout(btn_row)

    def _create_steps(self):
        self.steps = [
            Step1_DateSeries(self.stack, self),
            Step2_Mode(self.stack, self),
            Step4_Images(self.stack, self),
            Step3_Welds(self.stack, self),
            Step5_Note(self.stack, self),
            Step6_Materials(self.stack, self),
            Step7_Confirm(self.stack, self),
        ]
        # 連接 Step1 流水號變更 → sidebar 更新
        step1 = self.steps[0]
        step1.series_edit.textChanged.connect(self._sidebar.update_serial)

    def _show_step(self, index: int):
        self.current_step = index
        self.stack.setCurrentIndex(index)
        self.steps[index].on_show()

        self.prev_btn.setEnabled(index > 0)
        self.next_btn.setText("✅ 建立" if index == len(self.steps) - 1 else "下一步 →")
        self.progress_label.setText(f"步驟 {index + 1}/{len(self.steps)}")

        # Staging sidebar 僅在 Step1 顯示
        if index == 0:
            self._staging_sidebar.show()
            self._staging_sidebar.refresh()
        else:
            self._staging_sidebar.hide()

    def _next_step(self):
        step = self.steps[self.current_step]
        if not step.validate():
            return
        self.collected_data.update(step.get_data())
        if self.current_step == len(self.steps) - 1:
            self._create_folder()
        else:
            self._show_step(self.current_step + 1)

    def _prev_step(self):
        if self.current_step > 0:
            self._show_step(self.current_step - 1)

    def _cancel(self):
        if QMessageBox.question(
            self, "確認", "確定要取消嗎？"
        ) == QMessageBox.StandardButton.Yes:
            self.reject()

    def get_collected_data(self) -> dict:
        return self.collected_data.copy()

    def _get_images_step(self):
        """回傳 Step4_Images 實例（供 sidebar 同步路徑）"""
        for s in self.steps:
            if isinstance(s, Step4_Images):
                return s
        return None

    def _create_folder(self):
        self.collected_data.update(self.steps[-1].get_data())
        # 從 Step4_Images 重新取得最新圖片路徑（標註後可能已更新）
        img_step = self._get_images_step()
        if img_step:
            self.collected_data.update(img_step.get_data())
        data = self.collected_data
        folder_path = data["folder_path"]

        try:
            date_folder = os.path.dirname(folder_path)
            os.makedirs(date_folder, exist_ok=True)

            if os.path.exists(folder_path):
                if QMessageBox.question(
                    self, "警告", f"資料夾已存在：\n{folder_path}\n\n是否覆蓋？"
                ) != QMessageBox.StandardButton.Yes:
                    return

            os.makedirs(folder_path, exist_ok=True)

            # GroupWeld.txt
            if data.get("mode") == "group" and data.get("welds_list"):
                with open(os.path.join(folder_path, "GroupWeld.txt"), "w", encoding="utf-8") as f:
                    for weld in data["welds_list"]:
                        f.write(weld + "\n")

            # weld_info.json
            welds_full = data.get("welds_list_full", [])
            if welds_full:
                weld_info = {"series": data.get("series", ""), "date": data.get("date", ""), "welds": welds_full}
                with open(os.path.join(folder_path, "weld_info.json"), "w", encoding="utf-8") as f:
                    json.dump(weld_info, f, ensure_ascii=False, indent=2)

            # note.txt
            note = data.get("note_content", "")
            with open(os.path.join(folder_path, "note.txt"), "w", encoding="utf-8") as f:
                f.write(note if note else "# 請填寫修改原因說明\n")

            # materials.txt
            mats = data.get("materials_list", [])
            if mats:
                with open(os.path.join(folder_path, "materials.txt"), "w", encoding="utf-8") as f:
                    f.write("# 使用材料清單\n# 零件, 尺寸, SCH, 材質, 數量, 備註\n" + "-" * 50 + "\n")
                    for m in mats:
                        comp = m.get("component", "")
                        # 移除 emoji icon（若有）
                        if comp and len(comp) > 2 and comp[1] == " ":
                            comp = comp[2:].strip() or comp
                        size = m.get("size", "")
                        sch = m.get("schedule", "")
                        mat_abbr = m.get("material_abbr", "")
                        mat_name = m.get("material", "").split(" ")[0] if m.get("material") else ""
                        mat_str = f"{mat_name}({mat_abbr})" if mat_abbr else mat_name
                        qty = m.get("qty", 1)
                        unit = m.get("unit", "個")
                        remark = m.get("remark", "")
                        f.write(f"{comp}, {size}, {sch}, {mat_str}, {qty} {unit}, {remark}\n")

            # 複製圖片（來自 staging/ 的自動搬移，其餘複製）
            staging_dir = os.path.join(BASE_DIR, "staging")
            moved_from_staging = []
            images = data.get("images", {})
            for key, src in images.items():
                if not src or not os.path.exists(src):
                    continue
                if key == "pdf":
                    dst_name = os.path.basename(src)
                else:
                    dst_name = f"{key}.jpg"
                dst_full = os.path.join(folder_path, dst_name)
                shutil.copy2(src, dst_full)
                # 若來自 staging/ 則記錄，稍後詢問是否刪除
                try:
                    if os.path.normcase(os.path.abspath(os.path.dirname(src))) == \
                       os.path.normcase(os.path.abspath(staging_dir)):
                        moved_from_staging.append(src)
                except Exception:
                    pass

            # 預處理圖片
            if data.get("preprocess_images") and prepare_folder_images:
                results = prepare_folder_images(folder_path)
                if results.get("processed"):
                    print(f"✅ 已預處理 {len(results['processed'])} 張圖片")

            # 寫入焊口管制表
            weld_result = None
            if data.get("write_weld_control"):
                weld_result = self._write_welds_to_control_table(data)

            msg = f"✅ 資料夾已建立：\n{folder_path}\n\n已建立以下檔案：\n"
            msg += "• before.jpg / after.jpg（圖片）\n"
            msg += f"• {data.get('series', '[Series]')}.*.pdf（圖面）\n"
            msg += "• note.txt（修改原因）\n"
            if mats:
                msg += "• materials.txt（材料清單）\n"
            if data.get("mode") == "group":
                msg += "• GroupWeld.txt（焊口清單）\n"
            if weld_result:
                msg += f"\n{weld_result}"

            QMessageBox.information(self, "成功", msg)

            # 自動清除已分派的 staging 照片
            if moved_from_staging:
                reply = QMessageBox.question(
                    self, "清除 staging 照片",
                    f"已從 staging/ 複製 {len(moved_from_staging)} 個檔案到目標資料夾。\n"
                    "是否刪除這些 staging 中的原始檔案？",
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                )
                if reply == QMessageBox.StandardButton.Yes:
                    for sf in moved_from_staging:
                        try:
                            move_to_trash(sf, BASE_DIR, reason="wizard_consumed_staging")
                        except OSError:
                            pass

            if QMessageBox.question(
                self, "開啟資料夾", "是否開啟資料夾？"
            ) == QMessageBox.StandardButton.Yes:
                os.startfile(folder_path)

            self.accept()

        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"建立資料夾失敗：\n{e}")

    def _write_welds_to_control_table(self, data: dict) -> str:
        try:
            from settings_manager import get_weld_control_config, get_weld_control_table_path
            from weld_control import init_weld_manager_from_settings
            from openpyxl import load_workbook

            table_path = get_weld_control_table_path()
            if not table_path or not os.path.exists(table_path):
                return "⚠️ 焊口管制表未設定"

            config = get_weld_control_config()
            sr = data.get("series", "")
            if not sr:
                return "⚠️ 缺少流水號"

            fmt = config.get("serial_format", "raw")
            serial = sr.zfill(4) if fmt == "pad4" else (sr.lstrip("0") or "0")

            welds_full = data.get("welds_list_full", [])
            if not welds_full:
                return "⚠️ 無焊口資料"

            manager = init_weld_manager_from_settings()
            if not manager or not manager.load():
                return "⚠️ 無法載入焊口管制表"

            col_serial = config.get("col_serial", "流水號")
            col_weld_no = config.get("col_weld_no", "焊口編號")

            new_welds = []
            for w in welds_full:
                wid = f"{w['weld_no']}{w['mark']}"
                if not manager.check_exists(serial, wid):
                    new_welds.append({
                        col_serial: serial, col_weld_no: wid,
                        "SIZE": w.get("size", ""),
                        "material": w.get("material", ""),
                        "thickness": w.get("thickness", ""),
                    })

            if not new_welds:
                return "ℹ️ 所有焊口已存在，無需寫入"

            wb = load_workbook(table_path)
            sheet_name = config.get("sheet_name", "焊口編號明細")
            if sheet_name not in wb.sheetnames:
                return f"⚠️ 找不到工作表: {sheet_name}"
            ws = wb[sheet_name]

            headers = [cell.value for cell in ws[1]]
            col_map = {h: i + 1 for i, h in enumerate(headers) if h}

            next_row = ws.max_row + 1
            today = datetime.now().strftime("%Y/%m/%d")

            for wd in new_welds:
                for field in [col_serial, col_weld_no]:
                    if field in col_map and field in wd:
                        ws.cell(row=next_row, column=col_map[field], value=wd[field])

                sv = wd.get("SIZE", "")
                if sv:
                    for sc in ["SIZE", "尺寸", "size", "Size"]:
                        if sc in col_map:
                            ws.cell(row=next_row, column=col_map[sc], value=sv)
                            break

                mv = wd.get("material", "")
                if mv:
                    for mc in ["材質", "MATERIAL", "Material", "material"]:
                        if mc in col_map:
                            ws.cell(row=next_row, column=col_map[mc], value=mv)
                            break

                tv = wd.get("thickness", "")
                if tv:
                    for tc in ["厚度", "SCH", "THICKNESS", "Thickness", "thickness"]:
                        if tc in col_map:
                            ws.cell(row=next_row, column=col_map[tc], value=tv)
                            break

                col_date = config.get("col_date", "登錄日期")
                if col_date in col_map:
                    ws.cell(row=next_row, column=col_map[col_date], value=today)

                next_row += 1

            output_name = data.get("weld_output_name", "焊口管制表_更新")
            date_str = datetime.now().strftime("%Y.%m.%d")
            base_dir = os.path.dirname(table_path)
            output_path = os.path.join(base_dir, f"{output_name}_{date_str}.xlsx")
            if os.path.exists(output_path):
                for i in range(1, 100):
                    output_path = os.path.join(base_dir, f"{output_name}_{date_str}_{i}.xlsx")
                    if not os.path.exists(output_path):
                        break

            wb.save(output_path)
            wb.close()
            manager.invalidate_cache()

            return f"📝 已寫入 {len(new_welds)} 個新焊口\n另存至: {os.path.basename(output_path)}"

        except Exception as e:
            return f"❌ 寫入焊口管制表失敗: {e}"


def launch_wizard(parent: Optional[QWidget] = None):
    """啟動精靈"""
    dlg = FolderWizard(parent)
    dlg.exec()
    return dlg


if __name__ == "__main__":
    from PyQt6.QtWidgets import QApplication
    from theme import apply_theme
    import sys
    app = QApplication(sys.argv)
    apply_theme(app)
    launch_wizard()
