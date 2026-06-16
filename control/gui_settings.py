# -*- coding: utf-8 -*-
"""
gui_settings.py — 設定面板 + 說明工具 (PyQt6)

從 gui.py 拆分而來，包含 HelpTooltip、HELP_TEXTS、SettingsPanel。
"""

import os

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGroupBox, QLabel, QPushButton,
    QLineEdit, QCheckBox, QRadioButton, QButtonGroup, QTabWidget,
    QTreeWidget, QTreeWidgetItem, QDialog, QScrollArea, QComboBox,
    QMessageBox, QFileDialog, QFrame, QTextEdit, QHeaderView,
)
from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtGui import QFont, QColor

from settings_manager import remember_browse_directory
from theme import Colors, Fonts, set_button_role, make_hint_label

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ========== HelpTooltip ==========
class HelpTooltip:
    """說明提示元件 — 顯示 ❓ 標籤，點擊後彈出詳細說明（PyQt6 版）"""

    def __init__(self, parent: QWidget, title: str, text: str, width: int = 400):
        self.parent_widget = parent
        self.title = title
        self.text = text
        self.width = width
        self.popup = None

        self.label = QLabel("❓")
        self.label.setStyleSheet(
            "color: #0066cc; font-size: 14px; padding: 0 4px; cursor: pointer;"
        )
        self.label.setCursor(Qt.CursorShape.PointingHandCursor)
        self.label.mousePressEvent = self._show_help

    def widget(self) -> QLabel:
        """回傳可加入 layout 的 QLabel"""
        return self.label

    def _show_help(self, event=None):
        if self.popup and self.popup.isVisible():
            self.popup.close()
            self.popup = None
            return

        self.popup = QDialog(self.parent_widget)
        self.popup.setWindowTitle(f"📖 說明: {self.title}")
        self.popup.resize(self.width, 340)

        layout = QVBoxLayout(self.popup)

        title_lbl = QLabel(f"📌 {self.title}")
        title_lbl.setFont(QFont("Microsoft JhengHei UI", 12, QFont.Weight.Bold))
        layout.addWidget(title_lbl)

        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        layout.addWidget(sep)

        text_edit = QTextEdit()
        text_edit.setReadOnly(True)
        text_edit.setFont(QFont("Microsoft JhengHei UI", 10))
        text_edit.setStyleSheet("background: #f8f9fa; border: none; padding: 10px;")
        text_edit.setPlainText(self.text)
        layout.addWidget(text_edit)

        btn_ok = QPushButton("✓ 了解")
        btn_ok.clicked.connect(self.popup.close)
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        btn_row.addWidget(btn_ok)
        layout.addLayout(btn_row)

        self.popup.show()


# ========== 設定說明文字 ==========
HELP_TEXTS = {
    "weld_table": """【焊口管制表是什麼？】
焊口管制表是一份 Excel 檔案，記錄了所有焊口編號的登記狀態。

【為什麼需要設定？】
• 當您處理修改單時，系統會自動將新增的焊口編號寫入此表
• 這樣可以追蹤哪些焊口已經登記、哪些還沒有
• 也可以記錄每個焊口的相關資訊（如 LINE NO、SIZE 等）

【如何設定？】
1. 點擊「瀏覽」選擇您的焊口管制表 Excel 檔案
2. 設定工作表名稱（預設為「焊口編號明細」）
3. 確認主鍵欄位名稱正確（流水號、焊口編號）

【注意事項】
• 檔案必須是 .xlsx 或 .xlsm 格式
• 處理時請確保 Excel 沒有開啟此檔案
• 建議定期備份此檔案""",
    "sheet_name": "【工作表名稱】\n指定焊口管制表中要使用的工作表名稱。\n預設為「焊口編號明細」，如果您的工作表名稱不同，請修改此設定。",
    "pk_fields": "【主鍵欄位設定】\n流水號欄位：用來識別不同案件（如 632、633…）\n焊口編號欄位：記錄焊口編號（如 1001a、1002r…）",
    "serial_format": "【流水號格式】\n• 純數字：632、633（最常見）\n• 補零4位：0632、0001",
    "dynamic_columns": "【動態欄位對應】\n設定要同步到管制表的額外欄位，如 LINE NO、SIZE、SCH、登錄日期等。",
    "auto_sync": "【自動同步新增焊口】\n開啟：處理修改單時自動將新焊口寫入管制表。\n關閉：只產出報告，不修改管制表。",
    "check_dup": "【檢查重複焊口】\n新增前先查詢管制表中是否已存在相同焊口。建議保持開啟。",
    "dwg_list": "【DWG LIST（圖號清單）】\n包含流水號與圖號對應的 Excel 檔案，用於自動填入報告欄位。",
    "dwg_list_sheet": "【DWG LIST 工作表名稱】\n預設為「DRAWING LIST」。",
    "dwg_list_pk": "【DWG LIST 主鍵欄位】\n指定用來查詢的主鍵欄位名稱（通常是「NO」或「流水號」）。",
    "dwg_dynamic_columns": "【DWG LIST 欄位對應】\n設定要從 DWG LIST 取得哪些欄位資訊（如 DWG NO、DWG名稱、REV）。",
    "export_pdf": "【產出 PDF】\n同時產出 PDF 和 Excel 兩種格式。",
    "skip_unchanged": "【略過未修改的資料夾】\n比對指紋，未變更則跳過。",
    "auto_preprocess": "【自動預處理圖片】\n自動壓縮過大圖片，原圖保留 .orig 備份。",
    "debug_mode": "【除錯模式】\n顯示更詳細的執行資訊，幫助診斷問題。",
    "image_settings": "【圖片壓縮設定】\n最大邊長（預設 1280）與 JPEG 品質（預設 85）。",
    "prefab_drawing": "【預製圖路徑】\n指定一個資料夾，內含預製圖 PDF 檔案。\n\n命名慣例：{流水號}.DW-xxxx-xx-xxxx-xx-x.pdf\n例如：243.DW-1302-25-AA1B-NA-3.pdf\n\n處理修改單時，系統會依流水號自動從此資料夾\n複製對應的 PDF 到附件目錄，省去手動搬檔。",
}


# ========== SettingsPanel ==========
class SettingsPanel(QWidget):
    """系統設定面板 — 管理路徑與焊口管制表設定 (PyQt6)"""

    def __init__(self, parent: QWidget = None):
        super().__init__(parent)
        self._build_ui()
        self._load_settings()

    # ────────────────── UI ──────────────────
    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)

        # 子頁籤
        self.notebook = QTabWidget()
        root.addWidget(self.notebook)

        # 頁籤1：焊口管制表
        weld_tab = QWidget()
        self._build_weld_tab(weld_tab)
        self.notebook.addTab(weld_tab, "🔩 焊口管制表")

        # 頁籤2：DWG LIST
        dwg_tab = QWidget()
        self._build_dwg_tab(dwg_tab)
        self.notebook.addTab(dwg_tab, "📐 DWG LIST")

        # 頁籤3：執行選項
        runtime_tab = QWidget()
        self._build_runtime_tab(runtime_tab)
        self.notebook.addTab(runtime_tab, "⚡ 執行選項")

        # 頁籤4：關於
        about_tab = QWidget()
        self._build_about_tab(about_tab)
        self.notebook.addTab(about_tab, "ℹ️ 關於")

        # 底部按鈕列
        btn_frame = QHBoxLayout()
        btn_frame.setSpacing(8)
        btn_save = QPushButton("💾 儲存所有設定")
        btn_save.setProperty("role", "primary")
        btn_save.setMinimumWidth(140)
        btn_save.clicked.connect(self._save_settings)
        btn_frame.addWidget(btn_save)
        btn_reload = QPushButton("🔄 重新載入")
        btn_reload.clicked.connect(self._load_settings)
        btn_frame.addWidget(btn_reload)
        btn_reset = QPushButton("↩️ 恢復預設")
        btn_reset.clicked.connect(self._reset_defaults)
        btn_frame.addWidget(btn_reset)
        self.save_status_label = QLabel("")
        self.save_status_label.setStyleSheet(f"color: {Colors.SUCCESS}; font-weight: bold; border:none; background:transparent;")
        btn_frame.addWidget(self.save_status_label)
        btn_frame.addStretch()
        root.addLayout(btn_frame)

    # ── helpers ──
    def _make_scroll_area(self, inner: QWidget) -> QScrollArea:
        sa = QScrollArea()
        sa.setWidgetResizable(True)
        sa.setWidget(inner)
        return sa

    def _h_with_help(self, parent_layout, label_text: str, help_key: str):
        """建立一列帶 ❓ 的標籤"""
        row = QHBoxLayout()
        row.addWidget(QLabel(label_text))
        ht = HelpTooltip(self, help_key, HELP_TEXTS.get(help_key, ""))
        row.addWidget(ht.widget())
        row.addStretch()
        parent_layout.addLayout(row)

    # ── 焊口管制表頁籤 ──
    def _build_weld_tab(self, parent: QWidget):
        scroll_content = QWidget()
        vbox = QVBoxLayout(scroll_content)

        # 標題
        header = QHBoxLayout()
        header.addWidget(QLabel("設定焊口管制表以同步新增的焊口編號"))
        ht = HelpTooltip(self, "焊口管制表", HELP_TEXTS["weld_table"])
        header.addWidget(ht.widget())
        self.weld_status_label = QLabel("")
        self.weld_status_label.setStyleSheet("color: red;")
        header.addStretch()
        header.addWidget(self.weld_status_label)
        vbox.addLayout(header)

        # 檔案路徑
        path_group = QGroupBox("📂 檔案路徑")
        path_row = QHBoxLayout(path_group)
        self.weld_table_path_edit = QLineEdit()
        path_row.addWidget(self.weld_table_path_edit, 1)
        btn_browse = QPushButton("瀏覽")
        btn_browse.setFixedWidth(70)
        btn_browse.clicked.connect(self._browse_weld_table)
        path_row.addWidget(btn_browse)
        btn_open = QPushButton("開啟")
        btn_open.setFixedWidth(70)
        btn_open.clicked.connect(self._open_weld_table)
        path_row.addWidget(btn_open)
        vbox.addWidget(path_group)

        # 工作表與主鍵
        config_group = QGroupBox("📋 工作表與主鍵設定")
        config_vbox = QVBoxLayout(config_group)

        sheet_row = QHBoxLayout()
        sheet_row.addWidget(QLabel("工作表名稱:"))
        self.sheet_name_edit = QLineEdit("焊口編號明細")
        self.sheet_name_edit.setFixedWidth(180)
        sheet_row.addWidget(self.sheet_name_edit)
        ht2 = HelpTooltip(self, "工作表名稱", HELP_TEXTS["sheet_name"])
        sheet_row.addWidget(ht2.widget())
        sheet_row.addStretch()
        config_vbox.addLayout(sheet_row)

        pk_row = QHBoxLayout()
        pk_row.addWidget(QLabel("流水號欄位:"))
        self.col_serial_edit = QLineEdit("流水號")
        self.col_serial_edit.setFixedWidth(120)
        pk_row.addWidget(self.col_serial_edit)
        pk_row.addSpacing(15)
        pk_row.addWidget(QLabel("焊口編號欄位:"))
        self.col_weld_no_edit = QLineEdit("焊口編號")
        self.col_weld_no_edit.setFixedWidth(120)
        pk_row.addWidget(self.col_weld_no_edit)
        ht3 = HelpTooltip(self, "主鍵欄位", HELP_TEXTS["pk_fields"])
        pk_row.addWidget(ht3.widget())
        pk_row.addStretch()
        config_vbox.addLayout(pk_row)

        fmt_row = QHBoxLayout()
        fmt_row.addWidget(QLabel("流水號格式:"))
        self.serial_fmt_raw = QRadioButton("原始格式")
        self.serial_fmt_pad4 = QRadioButton("補零4位")
        self.serial_fmt_group = QButtonGroup(self)
        self.serial_fmt_group.addButton(self.serial_fmt_raw, 0)
        self.serial_fmt_group.addButton(self.serial_fmt_pad4, 1)
        self.serial_fmt_raw.setChecked(True)
        fmt_row.addWidget(self.serial_fmt_raw)
        fmt_row.addWidget(self.serial_fmt_pad4)
        ht4 = HelpTooltip(self, "流水號格式", HELP_TEXTS["serial_format"])
        fmt_row.addWidget(ht4.widget())
        fmt_row.addStretch()
        config_vbox.addLayout(fmt_row)

        vbox.addWidget(config_group)

        # 動態欄位
        dyn_group = QGroupBox("📊 動態欄位對應")
        dyn_vbox = QVBoxLayout(dyn_group)
        dyn_header = QHBoxLayout()
        dyn_header.addWidget(QLabel("設定要同步到管制表的額外欄位"))
        ht5 = HelpTooltip(self, "動態欄位", HELP_TEXTS["dynamic_columns"])
        dyn_header.addWidget(ht5.widget())
        dyn_header.addStretch()
        dyn_vbox.addLayout(dyn_header)

        self.dyn_tree = QTreeWidget()
        self.dyn_tree.setHeaderLabels(["Excel 欄位", "資料來源", "必填"])
        self.dyn_tree.setColumnWidth(0, 140)
        self.dyn_tree.setColumnWidth(1, 140)
        self.dyn_tree.setColumnWidth(2, 50)
        self.dyn_tree.setMaximumHeight(160)
        self.dyn_tree.setAlternatingRowColors(True)
        dyn_vbox.addWidget(self.dyn_tree)

        dyn_btns = QHBoxLayout()
        btn_load_excel = QPushButton("📥 從Excel載入")
        btn_load_excel.clicked.connect(self._load_excel_columns)
        dyn_btns.addWidget(btn_load_excel)
        btn_add = QPushButton("➕ 新增")
        btn_add.clicked.connect(self._add_dynamic_column)
        dyn_btns.addWidget(btn_add)
        btn_edit = QPushButton("✏️ 編輯")
        btn_edit.clicked.connect(self._edit_dynamic_column)
        dyn_btns.addWidget(btn_edit)
        btn_del = QPushButton("➖ 刪除")
        btn_del.clicked.connect(self._delete_dynamic_column)
        dyn_btns.addWidget(btn_del)
        dyn_btns.addStretch()
        dyn_vbox.addLayout(dyn_btns)
        vbox.addWidget(dyn_group)

        # 同步選項
        sync_group = QGroupBox("⚙️ 同步選項")
        sync_row = QHBoxLayout(sync_group)
        self.auto_sync_chk = QCheckBox("自動同步新增焊口")
        self.auto_sync_chk.setChecked(True)
        sync_row.addWidget(self.auto_sync_chk)
        ht6 = HelpTooltip(self, "自動同步", HELP_TEXTS["auto_sync"])
        sync_row.addWidget(ht6.widget())
        sync_row.addSpacing(30)
        self.check_dup_chk = QCheckBox("檢查重複焊口")
        self.check_dup_chk.setChecked(True)
        sync_row.addWidget(self.check_dup_chk)
        ht7 = HelpTooltip(self, "檢查重複", HELP_TEXTS["check_dup"])
        sync_row.addWidget(ht7.widget())
        sync_row.addStretch()
        vbox.addWidget(sync_group)

        # 測試
        test_row = QHBoxLayout()
        btn_test = QPushButton("🔍 測試連線")
        btn_test.setFixedWidth(110)
        btn_test.clicked.connect(self._test_weld_table)
        test_row.addWidget(btn_test)
        self.test_result_label = QLabel("")
        test_row.addWidget(self.test_result_label)
        test_row.addStretch()
        vbox.addLayout(test_row)
        vbox.addStretch()

        sa = self._make_scroll_area(scroll_content)
        outer = QVBoxLayout(parent)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.addWidget(sa)

    # ── DWG LIST 頁籤 ──
    def _build_dwg_tab(self, parent: QWidget):
        scroll_content = QWidget()
        vbox = QVBoxLayout(scroll_content)

        header = QHBoxLayout()
        header.addWidget(QLabel("設定 DWG LIST 以查詢流水號對應的圖號資訊"))
        ht = HelpTooltip(self, "DWG LIST", HELP_TEXTS["dwg_list"])
        header.addWidget(ht.widget())
        self.dwg_status_label = QLabel("")
        self.dwg_status_label.setStyleSheet("color: red;")
        header.addStretch()
        header.addWidget(self.dwg_status_label)
        vbox.addLayout(header)

        # 啟用
        self.dwg_enabled_chk = QCheckBox("啟用 DWG LIST 查詢功能")
        self.dwg_enabled_chk.setChecked(True)
        vbox.addWidget(self.dwg_enabled_chk)

        # 路徑
        path_group = QGroupBox("📂 檔案路徑")
        path_row = QHBoxLayout(path_group)
        self.dwg_list_path_edit = QLineEdit()
        path_row.addWidget(self.dwg_list_path_edit, 1)
        btn_b = QPushButton("瀏覽")
        btn_b.setFixedWidth(70)
        btn_b.clicked.connect(self._browse_dwg_list)
        path_row.addWidget(btn_b)
        btn_o = QPushButton("開啟")
        btn_o.setFixedWidth(70)
        btn_o.clicked.connect(self._open_dwg_list)
        path_row.addWidget(btn_o)
        vbox.addWidget(path_group)

        # 工作表與主鍵
        cfg_group = QGroupBox("📋 工作表與主鍵設定")
        cfg_vbox = QVBoxLayout(cfg_group)

        r1 = QHBoxLayout()
        r1.addWidget(QLabel("工作表名稱:"))
        self.dwg_sheet_name_edit = QLineEdit("DRAWING LIST")
        self.dwg_sheet_name_edit.setFixedWidth(180)
        r1.addWidget(self.dwg_sheet_name_edit)
        r1.addStretch()
        cfg_vbox.addLayout(r1)

        r2 = QHBoxLayout()
        r2.addWidget(QLabel("流水號欄位:"))
        self.dwg_col_serial_edit = QLineEdit("NO")
        self.dwg_col_serial_edit.setFixedWidth(120)
        r2.addWidget(self.dwg_col_serial_edit)
        r2.addStretch()
        cfg_vbox.addLayout(r2)

        r3 = QHBoxLayout()
        r3.addWidget(QLabel("流水號格式:"))
        self.dwg_fmt_raw = QRadioButton("原始格式")
        self.dwg_fmt_pad4 = QRadioButton("補零4位")
        self.dwg_fmt_group = QButtonGroup(self)
        self.dwg_fmt_group.addButton(self.dwg_fmt_raw, 0)
        self.dwg_fmt_group.addButton(self.dwg_fmt_pad4, 1)
        self.dwg_fmt_raw.setChecked(True)
        r3.addWidget(self.dwg_fmt_raw)
        r3.addWidget(self.dwg_fmt_pad4)
        r3.addStretch()
        cfg_vbox.addLayout(r3)

        vbox.addWidget(cfg_group)

        # 動態欄位
        dyn_group = QGroupBox("📊 欄位對應")
        dyn_vbox = QVBoxLayout(dyn_group)
        dyn_header = QHBoxLayout()
        dyn_header.addWidget(QLabel("設定要從 DWG LIST 取得的欄位"))
        ht2 = HelpTooltip(self, "DWG 欄位對應", HELP_TEXTS["dwg_dynamic_columns"])
        dyn_header.addWidget(ht2.widget())
        dyn_header.addStretch()
        dyn_vbox.addLayout(dyn_header)

        self.dwg_dyn_tree = QTreeWidget()
        self.dwg_dyn_tree.setHeaderLabels(["Excel 欄位名稱", "對應變數"])
        self.dwg_dyn_tree.setColumnWidth(0, 180)
        self.dwg_dyn_tree.setColumnWidth(1, 150)
        self.dwg_dyn_tree.setMaximumHeight(160)
        self.dwg_dyn_tree.setAlternatingRowColors(True)
        dyn_vbox.addWidget(self.dwg_dyn_tree)

        dyn_btns = QHBoxLayout()
        for txt, slot in [("📥 從Excel載入", self._load_dwg_excel_columns),
                          ("➕ 新增", self._add_dwg_dynamic_column),
                          ("✏️ 編輯", self._edit_dwg_dynamic_column),
                          ("➖ 刪除", self._delete_dwg_dynamic_column)]:
            b = QPushButton(txt)
            b.clicked.connect(slot)
            dyn_btns.addWidget(b)
        dyn_btns.addStretch()
        dyn_vbox.addLayout(dyn_btns)
        vbox.addWidget(dyn_group)

        # 測試
        test_row = QHBoxLayout()
        btn_test = QPushButton("🔍 測試連線")
        btn_test.setFixedWidth(110)
        btn_test.clicked.connect(self._test_dwg_list)
        test_row.addWidget(btn_test)
        self.dwg_test_result_label = QLabel("")
        test_row.addWidget(self.dwg_test_result_label)
        test_row.addStretch()
        vbox.addLayout(test_row)
        vbox.addStretch()

        sa = self._make_scroll_area(scroll_content)
        outer = QVBoxLayout(parent)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.addWidget(sa)

    # ── 執行選項頁籤 ──
    def _build_runtime_tab(self, parent: QWidget):
        vbox = QVBoxLayout(parent)

        # 預製圖路徑
        prefab_group = QGroupBox("📐 預製圖路徑")
        pfg = QVBoxLayout(prefab_group)
        pf_header = QHBoxLayout()
        pf_header.addWidget(QLabel("指定預製圖 PDF 來源資料夾，處理時自動複製到附件目錄"))
        ht_pf = HelpTooltip(self, "預製圖路徑", HELP_TEXTS["prefab_drawing"])
        pf_header.addWidget(ht_pf.widget())
        pf_header.addStretch()
        pfg.addLayout(pf_header)
        pf_row = QHBoxLayout()
        self.prefab_dir_edit = QLineEdit()
        self.prefab_dir_edit.setPlaceholderText("尚未設定（選填）")
        pf_row.addWidget(self.prefab_dir_edit, 1)
        btn_prefab_browse = QPushButton("瀏覽")
        btn_prefab_browse.setFixedWidth(70)
        btn_prefab_browse.clicked.connect(self._browse_prefab_dir)
        pf_row.addWidget(btn_prefab_browse)
        btn_prefab_open = QPushButton("開啟")
        btn_prefab_open.setFixedWidth(70)
        btn_prefab_open.clicked.connect(self._open_prefab_dir)
        pf_row.addWidget(btn_prefab_open)
        pfg.addLayout(pf_row)
        vbox.addWidget(prefab_group)

        report_group = QGroupBox("📄 報告產出")
        rg = QVBoxLayout(report_group)
        r1 = QHBoxLayout()
        self.export_pdf_chk = QCheckBox("產出 PDF 檔案")
        self.export_pdf_chk.setChecked(True)
        r1.addWidget(self.export_pdf_chk)
        ht1 = HelpTooltip(self, "產出 PDF", HELP_TEXTS["export_pdf"])
        r1.addWidget(ht1.widget())
        r1.addStretch()
        rg.addLayout(r1)
        vbox.addWidget(report_group)

        proc_group = QGroupBox("⚙️ 處理選項")
        pg = QVBoxLayout(proc_group)
        r2 = QHBoxLayout()
        self.skip_unchanged_chk = QCheckBox("略過未變更的資料夾（依指紋比對）")
        self.skip_unchanged_chk.setChecked(True)
        r2.addWidget(self.skip_unchanged_chk)
        ht2 = HelpTooltip(self, "略過未修改", HELP_TEXTS["skip_unchanged"])
        r2.addWidget(ht2.widget())
        r2.addStretch()
        pg.addLayout(r2)
        r3 = QHBoxLayout()
        self.auto_preprocess_chk = QCheckBox("自動預處理過大的圖片")
        self.auto_preprocess_chk.setChecked(True)
        r3.addWidget(self.auto_preprocess_chk)
        ht3 = HelpTooltip(self, "自動預處理", HELP_TEXTS["auto_preprocess"])
        r3.addWidget(ht3.widget())
        r3.addStretch()
        pg.addLayout(r3)
        vbox.addWidget(proc_group)

        img_group = QGroupBox("🖼️ 圖片壓縮設定")
        ig = QVBoxLayout(img_group)
        img_header = QHBoxLayout()
        img_header.addWidget(QLabel("預處理時的圖片壓縮參數"))
        ht4 = HelpTooltip(self, "圖片設定", HELP_TEXTS["image_settings"])
        img_header.addWidget(ht4.widget())
        img_header.addStretch()
        ig.addLayout(img_header)
        img_row = QHBoxLayout()
        img_row.addWidget(QLabel("最大邊長:"))
        self.max_edge_edit = QLineEdit("1280")
        self.max_edge_edit.setFixedWidth(70)
        img_row.addWidget(self.max_edge_edit)
        img_row.addWidget(QLabel("像素"))
        img_row.addSpacing(20)
        img_row.addWidget(QLabel("JPEG 品質:"))
        self.quality_edit = QLineEdit("85")
        self.quality_edit.setFixedWidth(50)
        img_row.addWidget(self.quality_edit)
        img_row.addWidget(QLabel("% (1-100)"))
        img_row.addStretch()
        ig.addLayout(img_row)
        vbox.addWidget(img_group)

        debug_group = QGroupBox("🔧 除錯")
        dg = QVBoxLayout(debug_group)
        r4 = QHBoxLayout()
        self.debug_chk = QCheckBox("除錯模式（顯示詳細執行資訊）")
        r4.addWidget(self.debug_chk)
        ht5 = HelpTooltip(self, "除錯模式", HELP_TEXTS["debug_mode"])
        r4.addWidget(ht5.widget())
        r4.addStretch()
        dg.addLayout(r4)
        vbox.addWidget(debug_group)
        vbox.addStretch()

    # ── 關於頁籤 ──
    def _build_about_tab(self, parent: QWidget):
        vbox = QVBoxLayout(parent)
        vbox.setSpacing(16)

        # 標題卡片
        header_card = QFrame()
        header_card.setStyleSheet(
            f"QFrame {{ background: qlineargradient(x1:0,y1:0,x2:1,y2:1,"
            f" stop:0 {Colors.PRIMARY}, stop:1 {Colors.ACCENT});"
            f" border-radius: 12px; padding: 20px; }}"
        )
        hc_lay = QVBoxLayout(header_card)
        title_lbl = QLabel("📋 管線修改單產出系統")
        title_lbl.setFont(Fonts.heading(18))
        title_lbl.setStyleSheet("color: white; border: none; background: transparent;")
        title_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        hc_lay.addWidget(title_lbl)
        ver_lbl = QLabel("v2.3 • PyQt6 Edition")
        ver_lbl.setFont(Fonts.body(11))
        ver_lbl.setStyleSheet("color: rgba(255,255,255,0.85); border: none; background: transparent;")
        ver_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        hc_lay.addWidget(ver_lbl)
        vbox.addWidget(header_card)

        info_group = QGroupBox("📁 系統資訊")
        ig = QVBoxLayout(info_group)
        ig.setSpacing(6)
        for label, value in [("設定檔位置", "settings.json"), ("附件根目錄", "attachments/"),
                             ("輸出目錄", "output/"), ("PDF 目錄", "pdf/")]:
            row = QHBoxLayout()
            l = QLabel(f"{label}：")
            l.setStyleSheet(f"color: {Colors.TEXT_SECONDARY}; border:none; background:transparent;")
            l.setFixedWidth(90)
            row.addWidget(l)
            v = QLabel(value)
            v.setFont(Fonts.code())
            v.setStyleSheet(f"color: {Colors.TEXT}; border:none; background:transparent;")
            row.addWidget(v)
            row.addStretch()
            ig.addLayout(row)
        vbox.addWidget(info_group)

        tip_group = QGroupBox("💡 使用提示")
        tg = QVBoxLayout(tip_group)
        tg.setSpacing(6)
        for tip in ["• 點擊 ❓ 圖示可查看詳細說明",
                     "• 修改設定後記得點擊「儲存所有設定」",
                     "• 焊口補登工具已移至「記錄管理」頁面"]:
            lbl = QLabel(tip)
            lbl.setStyleSheet(f"color: {Colors.PRIMARY}; border:none; background:transparent;")
            tg.addWidget(lbl)
        vbox.addWidget(tip_group)
        vbox.addStretch()

    # ────────────────── 載入 / 儲存 ──────────────────
    def _load_settings(self):
        from settings_manager import (
            get_settings, get_weld_control_table_path, get_weld_control_config,
            get_drawing_list_path, get_weld_dynamic_columns,
            get_dwg_list_config, get_dwg_dynamic_columns,
            get_prefab_drawing_dir
        )
        sm = get_settings()

        self.weld_table_path_edit.setText(get_weld_control_table_path())
        config = get_weld_control_config()
        self.sheet_name_edit.setText(config.get("sheet_name", "焊口編號明細"))
        self.auto_sync_chk.setChecked(config.get("auto_sync", True))
        self.check_dup_chk.setChecked(config.get("check_duplicate", True))
        self.col_serial_edit.setText(config.get("col_serial", "流水號"))
        self.col_weld_no_edit.setText(config.get("col_weld_no", "焊口編號"))
        fmt = config.get("serial_format", "raw")
        self.serial_fmt_pad4.setChecked(fmt == "pad4")
        self.serial_fmt_raw.setChecked(fmt != "pad4")

        self._load_dynamic_columns(get_weld_dynamic_columns())

        self.dwg_list_path_edit.setText(get_drawing_list_path())
        dwg_config = get_dwg_list_config()
        self.dwg_enabled_chk.setChecked(dwg_config.get("enabled", True))
        self.dwg_sheet_name_edit.setText(dwg_config.get("sheet_name", "DRAWING LIST"))
        self.dwg_col_serial_edit.setText(dwg_config.get("col_serial", "NO"))
        dwg_fmt = dwg_config.get("serial_format", "raw")
        self.dwg_fmt_pad4.setChecked(dwg_fmt == "pad4")
        self.dwg_fmt_raw.setChecked(dwg_fmt != "pad4")
        self._load_dwg_dynamic_columns(get_dwg_dynamic_columns())

        self.prefab_dir_edit.setText(get_prefab_drawing_dir())

        self.export_pdf_chk.setChecked(sm.get_runtime("export_pdf", True))
        self.skip_unchanged_chk.setChecked(sm.get_runtime("skip_unchanged", True))
        self.auto_preprocess_chk.setChecked(sm.get_runtime("auto_preprocess_images", True))
        self.debug_chk.setChecked(sm.get_runtime("debug_mode", False))
        self.max_edge_edit.setText(str(sm.get_runtime("preprocess_max_edge", 1280)))
        self.quality_edit.setText(str(sm.get_runtime("preprocess_quality", 85)))

        self._update_weld_status()
        self._update_dwg_status()

    def _save_settings(self):
        from settings_manager import (
            get_settings, set_weld_control_table_path, set_weld_control_config,
            set_weld_dynamic_columns, set_dwg_list_config, set_dwg_dynamic_columns,
            set_drawing_list_path, set_prefab_drawing_dir
        )
        sm = get_settings()

        set_weld_control_table_path(self.weld_table_path_edit.text().strip())
        set_prefab_drawing_dir(self.prefab_dir_edit.text().strip())
        config = {
            "sheet_name": self.sheet_name_edit.text().strip(),
            "auto_sync": self.auto_sync_chk.isChecked(),
            "check_duplicate": self.check_dup_chk.isChecked(),
            "col_serial": self.col_serial_edit.text().strip(),
            "col_weld_no": self.col_weld_no_edit.text().strip(),
            "serial_format": "pad4" if self.serial_fmt_pad4.isChecked() else "raw",
        }
        set_weld_control_config(config)
        set_weld_dynamic_columns(self._get_dynamic_columns_data())

        set_drawing_list_path(self.dwg_list_path_edit.text().strip())
        dwg_config = {
            "enabled": self.dwg_enabled_chk.isChecked(),
            "sheet_name": self.dwg_sheet_name_edit.text().strip(),
            "col_serial": self.dwg_col_serial_edit.text().strip(),
            "serial_format": "pad4" if self.dwg_fmt_pad4.isChecked() else "raw",
        }
        set_dwg_list_config(dwg_config)
        set_dwg_dynamic_columns(self._get_dwg_dynamic_columns_data())

        sm.set_runtime("export_pdf", self.export_pdf_chk.isChecked(), auto_save=False)
        sm.set_runtime("skip_unchanged", self.skip_unchanged_chk.isChecked(), auto_save=False)
        sm.set_runtime("auto_preprocess_images", self.auto_preprocess_chk.isChecked(), auto_save=False)
        sm.set_runtime("debug_mode", self.debug_chk.isChecked(), auto_save=False)
        try:
            sm.set_runtime("preprocess_max_edge", int(self.max_edge_edit.text()), auto_save=False)
            sm.set_runtime("preprocess_quality", int(self.quality_edit.text()), auto_save=False)
        except ValueError:
            pass
        sm.save()

        self._update_weld_status()
        self._update_dwg_status()
        self.save_status_label.setText("✅ 設定已儲存")
        self.save_status_label.setStyleSheet("color: green;")
        QTimer.singleShot(3000, lambda: self.save_status_label.setText(""))

    def _reset_defaults(self):
        if QMessageBox.question(self, "確認", "確定要恢復所有設定為預設值嗎？") != QMessageBox.StandardButton.Yes:
            return
        self.sheet_name_edit.setText("焊口編號明細")
        self.col_serial_edit.setText("流水號")
        self.col_weld_no_edit.setText("焊口編號")
        self.auto_sync_chk.setChecked(True)
        self.check_dup_chk.setChecked(True)
        self.serial_fmt_raw.setChecked(True)
        default_columns = [
            {"name": "LINE NO", "source": "line_no", "required": False},
            {"name": "SIZE", "source": "size", "required": False},
            {"name": "SCH", "source": "sch", "required": False},
            {"name": "登錄日期", "source": "auto_date", "required": False},
            {"name": "修改單編號", "source": "report_id", "required": False},
            {"name": "備註", "source": "remark", "required": False},
        ]
        self._load_dynamic_columns(default_columns)
        self.dwg_enabled_chk.setChecked(False)
        self.dwg_sheet_name_edit.setText("Drawing List")
        self.dwg_col_serial_edit.setText("流水號")
        self.dwg_fmt_raw.setChecked(True)
        self._load_dwg_dynamic_columns([])
        self.prefab_dir_edit.clear()
        self.export_pdf_chk.setChecked(True)
        self.skip_unchanged_chk.setChecked(True)
        self.auto_preprocess_chk.setChecked(True)
        self.debug_chk.setChecked(False)
        self.max_edge_edit.setText("1280")
        self.quality_edit.setText("85")

    # ────────────────── 瀏覽 / 開啟 ──────────────────
    def _browse_weld_table(self):
        fp, _ = QFileDialog.getOpenFileName(self, "選擇焊口管制表", "", "Excel 檔案 (*.xlsx *.xlsm);;所有檔案 (*.*)")
        if fp:
            self.weld_table_path_edit.setText(fp)
            remember_browse_directory(fp)
            self._update_weld_status()

    def _open_weld_table(self):
        p = self.weld_table_path_edit.text().strip()
        if p and os.path.exists(p):
            os.startfile(p)
        else:
            QMessageBox.warning(self, "提示", "檔案不存在")

    def _browse_dwg_list(self):
        fp, _ = QFileDialog.getOpenFileName(self, "選擇 DWG LIST", "", "Excel 檔案 (*.xlsx *.xlsm);;所有檔案 (*.*)")
        if fp:
            self.dwg_list_path_edit.setText(fp)
            remember_browse_directory(fp)
            self._update_dwg_status()

    def _open_dwg_list(self):
        p = self.dwg_list_path_edit.text().strip()
        if p and os.path.exists(p):
            os.startfile(p)
        else:
            QMessageBox.warning(self, "提示", "檔案不存在")

    def _browse_prefab_dir(self):
        d = QFileDialog.getExistingDirectory(self, "選擇預製圖 PDF 來源資料夾", "")
        if d:
            self.prefab_dir_edit.setText(d)
            remember_browse_directory(d)

    def _open_prefab_dir(self):
        p = self.prefab_dir_edit.text().strip()
        if p and os.path.isdir(p):
            os.startfile(p)
        else:
            QMessageBox.warning(self, "提示", "資料夾不存在")

    # ────────────────── 狀態 ──────────────────
    def _update_weld_status(self):
        p = self.weld_table_path_edit.text().strip()
        if not p:
            self.weld_status_label.setText("⚠️ 尚未設定路徑")
            self.weld_status_label.setStyleSheet("color: orange;")
        elif not os.path.exists(p):
            self.weld_status_label.setText("❌ 檔案不存在")
            self.weld_status_label.setStyleSheet("color: red;")
        else:
            self.weld_status_label.setText("✅ 檔案存在")
            self.weld_status_label.setStyleSheet("color: green;")

    def _update_dwg_status(self):
        p = self.dwg_list_path_edit.text().strip()
        if not p:
            self.dwg_status_label.setText("⚠️ 尚未設定路徑")
            self.dwg_status_label.setStyleSheet("color: orange;")
        elif not os.path.exists(p):
            self.dwg_status_label.setText("❌ 檔案不存在")
            self.dwg_status_label.setStyleSheet("color: red;")
        else:
            self.dwg_status_label.setText("✅ 檔案存在")
            self.dwg_status_label.setStyleSheet("color: green;")

    # ────────────────── 測試連線 ──────────────────
    def _test_weld_table(self):
        path = self.weld_table_path_edit.text().strip()
        sheet_name = self.sheet_name_edit.text().strip()
        if not path:
            self.test_result_label.setText("❌ 請先設定檔案路徑")
            self.test_result_label.setStyleSheet("color: red;")
            return
        if not os.path.exists(path):
            self.test_result_label.setText("❌ 檔案不存在")
            self.test_result_label.setStyleSheet("color: red;")
            return
        try:
            wb = load_workbook(path, read_only=True)
            if sheet_name not in wb.sheetnames:
                self.test_result_label.setText(f"❌ 找不到工作表 '{sheet_name}'")
                self.test_result_label.setStyleSheet("color: red;")
                wb.close()
                return
            ws = wb[sheet_name]
            headers = [c.value for c in ws[1] if c.value]
            wb.close()
            required = [self.col_serial_edit.text(), self.col_weld_no_edit.text()]
            missing = [r for r in required if r and r not in headers]
            if missing:
                self.test_result_label.setText(f"⚠️ 缺少欄位: {', '.join(missing)}")
                self.test_result_label.setStyleSheet("color: orange;")
            else:
                self.test_result_label.setText(f"✅ 連線成功！找到 {len(headers)} 個欄位")
                self.test_result_label.setStyleSheet("color: green;")
        except PermissionError:
            self.test_result_label.setText("❌ 檔案被開啟中")
            self.test_result_label.setStyleSheet("color: red;")
        except Exception as e:
            self.test_result_label.setText(f"❌ 錯誤: {e}")
            self.test_result_label.setStyleSheet("color: red;")

    def _test_dwg_list(self):
        path = self.dwg_list_path_edit.text().strip()
        sheet_name = self.dwg_sheet_name_edit.text().strip()
        if not path:
            self.dwg_test_result_label.setText("❌ 請先設定檔案路徑")
            self.dwg_test_result_label.setStyleSheet("color: red;")
            return
        if not os.path.exists(path):
            self.dwg_test_result_label.setText("❌ 檔案不存在")
            self.dwg_test_result_label.setStyleSheet("color: red;")
            return
        try:
            wb = load_workbook(path, read_only=True)
            if sheet_name not in wb.sheetnames:
                self.dwg_test_result_label.setText(f"❌ 找不到工作表 '{sheet_name}'")
                self.dwg_test_result_label.setStyleSheet("color: red;")
                wb.close()
                return
            ws = wb[sheet_name]
            headers = [c.value for c in ws[1] if c.value]
            wb.close()
            col_serial = self.dwg_col_serial_edit.text()
            if col_serial not in headers:
                self.dwg_test_result_label.setText(f"⚠️ 缺少欄位: {col_serial}")
                self.dwg_test_result_label.setStyleSheet("color: orange;")
            else:
                self.dwg_test_result_label.setText(f"✅ 連線成功！找到 {len(headers)} 個欄位")
                self.dwg_test_result_label.setStyleSheet("color: green;")
        except PermissionError:
            self.dwg_test_result_label.setText("❌ 檔案被開啟中")
            self.dwg_test_result_label.setStyleSheet("color: red;")
        except Exception as e:
            self.dwg_test_result_label.setText(f"❌ 錯誤: {e}")
            self.dwg_test_result_label.setStyleSheet("color: red;")

    # ────────────────── 動態欄位管理 ──────────────────
    def _load_dynamic_columns(self, columns: list):
        self.dyn_tree.clear()
        for col in columns:
            req = "✓" if col.get("required", False) else ""
            QTreeWidgetItem(self.dyn_tree, [col.get("name", ""), col.get("source", ""), req])

    def _get_dynamic_columns_data(self) -> list:
        cols = []
        for i in range(self.dyn_tree.topLevelItemCount()):
            it = self.dyn_tree.topLevelItem(i)
            cols.append({"name": it.text(0), "source": it.text(1), "required": it.text(2) == "✓"})
        return cols

    def _load_dwg_dynamic_columns(self, columns: list):
        self.dwg_dyn_tree.clear()
        for col in columns:
            QTreeWidgetItem(self.dwg_dyn_tree, [col.get("name", ""), col.get("target", "")])

    def _get_dwg_dynamic_columns_data(self) -> list:
        cols = []
        for i in range(self.dwg_dyn_tree.topLevelItemCount()):
            it = self.dwg_dyn_tree.topLevelItem(i)
            cols.append({"name": it.text(0), "target": it.text(1)})
        return cols

    # ── 焊口動態欄位 CRUD ──
    def _load_excel_columns(self):
        path = self.weld_table_path_edit.text().strip()
        sheet = self.sheet_name_edit.text().strip()
        if not path or not os.path.exists(path):
            QMessageBox.warning(self, "提示", "請先設定有效的焊口管制表路徑")
            return
        try:
            wb = load_workbook(path, read_only=True)
            if sheet not in wb.sheetnames:
                wb.close()
                QMessageBox.warning(self, "提示", f"找不到工作表 '{sheet}'")
                return
            ws = wb[sheet]
            headers = [c.value for c in ws[1] if c.value]
            wb.close()
            pk = [self.col_serial_edit.text(), self.col_weld_no_edit.text()]
            other = [h for h in headers if h not in pk]
            self._show_column_select_dialog(other)
        except PermissionError:
            QMessageBox.critical(self, "錯誤", "檔案被開啟中，請先關閉 Excel")
        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"讀取失敗: {e}")

    def _show_column_select_dialog(self, headers: list):
        dlg = QDialog(self)
        dlg.setWindowTitle("選擇要同步的欄位")
        dlg.resize(500, 400)
        layout = QVBoxLayout(dlg)
        layout.addWidget(QLabel("請勾選要同步到焊口管制表的欄位："))

        existing = {c.get("name", ""): c.get("source", "") for c in self._get_dynamic_columns_data()}
        source_map = {"LINE NO": "line_no", "SIZE": "size", "SCH": "sch",
                      "登錄日期": "auto_date", "修改單編號": "report_id", "備註": "remark",
                      "圖號": "dwg_no", "DWG NO": "dwg_no", "圖名": "dwg_name"}

        sa = QScrollArea()
        sa.setWidgetResizable(True)
        inner = QWidget()
        inner_layout = QVBoxLayout(inner)

        chk_map = {}
        combo_map = {}
        for h in headers:
            row = QHBoxLayout()
            chk = QCheckBox(h)
            chk.setChecked(h in existing)
            chk_map[h] = chk
            row.addWidget(chk)
            row.addWidget(QLabel("→ 來源:"))
            combo = QComboBox()
            combo.setEditable(True)
            combo.addItems(["line_no", "size", "sch", "dwg_no", "dwg_name", "auto_date", "report_id", "remark", ""])
            combo.setCurrentText(existing.get(h, source_map.get(h, "")))
            combo_map[h] = combo
            row.addWidget(combo)
            row.addStretch()
            inner_layout.addLayout(row)
        inner_layout.addStretch()
        sa.setWidget(inner)
        layout.addWidget(sa)

        btns = QHBoxLayout()
        btn_ok = QPushButton("✅ 確定")
        btn_cancel = QPushButton("❌ 取消")
        btns.addStretch()
        btns.addWidget(btn_ok)
        btns.addWidget(btn_cancel)
        layout.addLayout(btns)

        def apply_sel():
            new_cols = [{"name": h, "source": combo_map[h].currentText(), "required": False}
                        for h in headers if chk_map[h].isChecked()]
            self._load_dynamic_columns(new_cols)
            dlg.accept()

        btn_ok.clicked.connect(apply_sel)
        btn_cancel.clicked.connect(dlg.reject)
        dlg.exec()

    def _add_dynamic_column(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("新增動態欄位")
        dlg.resize(350, 180)
        layout = QVBoxLayout(dlg)
        r1 = QHBoxLayout(); r1.addWidget(QLabel("Excel 欄位名稱:")); name_e = QLineEdit(); r1.addWidget(name_e); layout.addLayout(r1)
        r2 = QHBoxLayout(); r2.addWidget(QLabel("資料來源:"))
        src_combo = QComboBox(); src_combo.setEditable(True)
        src_combo.addItems(["line_no","size","sch","dwg_no","dwg_name","auto_date","report_id","remark",""])
        r2.addWidget(src_combo); layout.addLayout(r2)
        req_chk = QCheckBox("必填欄位"); layout.addWidget(req_chk)
        btns = QHBoxLayout(); btn_ok = QPushButton("➕ 新增"); btn_cancel = QPushButton("取消")
        btns.addStretch(); btns.addWidget(btn_ok); btns.addWidget(btn_cancel); layout.addLayout(btns)
        def add_col():
            n = name_e.text().strip()
            if not n: QMessageBox.warning(dlg, "提示", "請輸入欄位名稱"); return
            QTreeWidgetItem(self.dyn_tree, [n, src_combo.currentText(), "✓" if req_chk.isChecked() else ""])
            dlg.accept()
        btn_ok.clicked.connect(add_col); btn_cancel.clicked.connect(dlg.reject); dlg.exec()

    def _edit_dynamic_column(self):
        items = self.dyn_tree.selectedItems()
        if not items: QMessageBox.information(self, "提示", "請先選擇要編輯的欄位"); return
        it = items[0]
        dlg = QDialog(self); dlg.setWindowTitle("編輯動態欄位"); dlg.resize(350, 180)
        layout = QVBoxLayout(dlg)
        r1 = QHBoxLayout(); r1.addWidget(QLabel("Excel 欄位名稱:")); name_e = QLineEdit(it.text(0)); r1.addWidget(name_e); layout.addLayout(r1)
        r2 = QHBoxLayout(); r2.addWidget(QLabel("資料來源:"))
        src_combo = QComboBox(); src_combo.setEditable(True)
        src_combo.addItems(["line_no","size","sch","dwg_no","dwg_name","auto_date","report_id","remark",""])
        src_combo.setCurrentText(it.text(1)); r2.addWidget(src_combo); layout.addLayout(r2)
        req_chk = QCheckBox("必填欄位"); req_chk.setChecked(it.text(2) == "✓"); layout.addWidget(req_chk)
        btns = QHBoxLayout(); btn_ok = QPushButton("💾 儲存"); btn_cancel = QPushButton("取消")
        btns.addStretch(); btns.addWidget(btn_ok); btns.addWidget(btn_cancel); layout.addLayout(btns)
        def save_col():
            n = name_e.text().strip()
            if not n: QMessageBox.warning(dlg, "提示", "請輸入欄位名稱"); return
            it.setText(0, n); it.setText(1, src_combo.currentText()); it.setText(2, "✓" if req_chk.isChecked() else "")
            dlg.accept()
        btn_ok.clicked.connect(save_col); btn_cancel.clicked.connect(dlg.reject); dlg.exec()

    def _delete_dynamic_column(self):
        items = self.dyn_tree.selectedItems()
        if not items: QMessageBox.information(self, "提示", "請先選擇要刪除的欄位"); return
        if QMessageBox.question(self, "確認", "確定要刪除選中的欄位嗎？") == QMessageBox.StandardButton.Yes:
            for it in items:
                idx = self.dyn_tree.indexOfTopLevelItem(it)
                self.dyn_tree.takeTopLevelItem(idx)

    # ── DWG 動態欄位 CRUD ──
    def _load_dwg_excel_columns(self):
        path = self.dwg_list_path_edit.text().strip()
        sheet = self.dwg_sheet_name_edit.text().strip()
        if not path or not os.path.exists(path):
            QMessageBox.warning(self, "提示", "請先設定有效的 DWG LIST 路徑"); return
        try:
            wb = load_workbook(path, read_only=True)
            if sheet not in wb.sheetnames: wb.close(); QMessageBox.warning(self, "提示", f"找不到工作表 '{sheet}'"); return
            ws = wb[sheet]; headers = [c.value for c in ws[1] if c.value]; wb.close()
            pk = self.dwg_col_serial_edit.text(); other = [h for h in headers if h != pk]
            self._show_dwg_column_select_dialog(other)
        except PermissionError:
            QMessageBox.critical(self, "錯誤", "檔案被開啟中，請先關閉 Excel")
        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"讀取失敗: {e}")

    def _show_dwg_column_select_dialog(self, headers: list):
        dlg = QDialog(self); dlg.setWindowTitle("選擇要對應的欄位"); dlg.resize(450, 350)
        layout = QVBoxLayout(dlg); layout.addWidget(QLabel("請勾選要從 DWG LIST 取得的欄位："))
        existing = {c.get("name",""): c.get("target","") for c in self._get_dwg_dynamic_columns_data()}
        target_map = {"DWG NO":"dwg_no","DWG名稱":"dwg_name","REV":"rev","圖號":"dwg_no","圖名":"dwg_name","版次":"rev"}
        sa = QScrollArea(); sa.setWidgetResizable(True); inner = QWidget(); inner_layout = QVBoxLayout(inner)
        chk_map = {}; combo_map = {}
        for h in headers:
            row = QHBoxLayout()
            chk = QCheckBox(h); chk.setChecked(h in existing); chk_map[h] = chk; row.addWidget(chk)
            row.addWidget(QLabel("→"))
            combo = QComboBox(); combo.setEditable(True); combo.addItems(["dwg_no","dwg_name","rev",""])
            combo.setCurrentText(existing.get(h, target_map.get(h, ""))); combo_map[h] = combo
            row.addWidget(combo); row.addStretch(); inner_layout.addLayout(row)
        inner_layout.addStretch(); sa.setWidget(inner); layout.addWidget(sa)
        btns = QHBoxLayout(); btn_ok = QPushButton("✅ 確定"); btn_cancel = QPushButton("❌ 取消")
        btns.addStretch(); btns.addWidget(btn_ok); btns.addWidget(btn_cancel); layout.addLayout(btns)
        def apply_sel():
            new_cols = [{"name": h, "target": combo_map[h].currentText()} for h in headers if chk_map[h].isChecked()]
            self._load_dwg_dynamic_columns(new_cols); dlg.accept()
        btn_ok.clicked.connect(apply_sel); btn_cancel.clicked.connect(dlg.reject); dlg.exec()

    def _add_dwg_dynamic_column(self):
        dlg = QDialog(self); dlg.setWindowTitle("新增欄位對應"); dlg.resize(320, 150)
        layout = QVBoxLayout(dlg)
        r1 = QHBoxLayout(); r1.addWidget(QLabel("Excel 欄位名稱:")); name_e = QLineEdit(); r1.addWidget(name_e); layout.addLayout(r1)
        r2 = QHBoxLayout(); r2.addWidget(QLabel("對應變數:"))
        combo = QComboBox(); combo.setEditable(True); combo.addItems(["dwg_no","dwg_name","rev",""])
        r2.addWidget(combo); layout.addLayout(r2)
        btns = QHBoxLayout(); btn_ok = QPushButton("➕ 新增"); btn_cancel = QPushButton("取消")
        btns.addStretch(); btns.addWidget(btn_ok); btns.addWidget(btn_cancel); layout.addLayout(btns)
        def add_col():
            n = name_e.text().strip()
            if not n: QMessageBox.warning(dlg, "提示", "請輸入欄位名稱"); return
            QTreeWidgetItem(self.dwg_dyn_tree, [n, combo.currentText()]); dlg.accept()
        btn_ok.clicked.connect(add_col); btn_cancel.clicked.connect(dlg.reject); dlg.exec()

    def _edit_dwg_dynamic_column(self):
        items = self.dwg_dyn_tree.selectedItems()
        if not items: QMessageBox.information(self, "提示", "請先選擇要編輯的欄位"); return
        it = items[0]
        dlg = QDialog(self); dlg.setWindowTitle("編輯欄位對應"); dlg.resize(320, 150)
        layout = QVBoxLayout(dlg)
        r1 = QHBoxLayout(); r1.addWidget(QLabel("Excel 欄位名稱:")); name_e = QLineEdit(it.text(0)); r1.addWidget(name_e); layout.addLayout(r1)
        r2 = QHBoxLayout(); r2.addWidget(QLabel("對應變數:"))
        combo = QComboBox(); combo.setEditable(True); combo.addItems(["dwg_no","dwg_name","rev",""])
        combo.setCurrentText(it.text(1)); r2.addWidget(combo); layout.addLayout(r2)
        btns = QHBoxLayout(); btn_ok = QPushButton("💾 儲存"); btn_cancel = QPushButton("取消")
        btns.addStretch(); btns.addWidget(btn_ok); btns.addWidget(btn_cancel); layout.addLayout(btns)
        def save_col():
            n = name_e.text().strip()
            if not n: QMessageBox.warning(dlg, "提示", "請輸入欄位名稱"); return
            it.setText(0, n); it.setText(1, combo.currentText()); dlg.accept()
        btn_ok.clicked.connect(save_col); btn_cancel.clicked.connect(dlg.reject); dlg.exec()

    def _delete_dwg_dynamic_column(self):
        items = self.dwg_dyn_tree.selectedItems()
        if not items: QMessageBox.information(self, "提示", "請先選擇要刪除的欄位"); return
        if QMessageBox.question(self, "確認", "確定要刪除選中的欄位嗎？") == QMessageBox.StandardButton.Yes:
            for it in items:
                idx = self.dwg_dyn_tree.indexOfTopLevelItem(it)
                self.dwg_dyn_tree.takeTopLevelItem(idx)

    # ── 焊口補登工具（公共介面保留）──
    def _open_weld_backfill_tool(self):
        table_path = self.weld_table_path_edit.text().strip()
        if not table_path or not os.path.exists(table_path):
            QMessageBox.warning(self, "提示", "請先設定有效的焊口管制表路徑"); return
        from gui_dialogs import WeldBackfillDialog
        dlg = WeldBackfillDialog(self, self)
        dlg.exec()
