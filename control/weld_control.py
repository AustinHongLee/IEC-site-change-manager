# -*- coding: utf-8 -*-
"""
weld_control.py — 焊口管制表管理模組（優化版）

功能：
- 讀取焊口管制表現有焊口
- 檢查焊口是否存在（PK = 流水號_焊口編號）
- 新增焊口到管制表
- 動態欄位支援

優化功能：
- JSON 快照快取（比 Excel 快 10-50 倍）
- 檔案修改時間檢查（只有 Excel 變更時才重新載入）
- 流水號索引（O(1) 查詢特定流水號的焊口）
- 記憶體快取（同一 session 內不重複載入）
"""

import os
import json
import hashlib
import time
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime

from resources import project_path

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# 快取目錄
CACHE_DIR = ".weld_cache"


def _norm_label(value: Any) -> str:
    """Normalize sheet/header labels for fuzzy matching."""
    return str(value or "").replace(" ", "").replace("\n", "").strip().lower()


def _header_text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


class WeldControlManager:
    """焊口管制表管理器（優化版）"""
    
    def __init__(self, config: dict = None):
        """
        初始化管理器
        
        Args:
            config: 設定字典，包含：
                - file_path: 焊口管制表路徑
                - sheet_name: 工作表名稱
                - pk_fields: 主鍵欄位列表 [流水號欄位, 焊口編號欄位]
                - columns: 動態欄位配置列表
        """
        self.config = config or {}
        self._cache: Dict[str, dict] = {}  # {pk: row_data}
        self._serial_index: Dict[str, List[str]] = {}  # {serial: [pk1, pk2, ...]}
        self._headers: List[str] = []
        self._col_map: Dict[str, int] = {}  # {欄位名: 欄位索引}
        self._loaded = False
        self._file_path = ""
        self._sheet_name = ""
        self._file_mtime: float = 0  # 檔案修改時間
        self._load_time: float = 0   # 載入時間（用於效能統計）
        self._header_row: int = 1
        self._cache_dirty: bool = False  # 快取是否需要儲存
        self._auto_save_cache: bool = True  # 是否自動儲存快取
    
    def configure(self, config: dict):
        """更新設定"""
        self.config = config
        self._loaded = False
        self._cache.clear()
        self._serial_index.clear()
        self._cache_dirty = False
    
    @property
    def file_path(self) -> str:
        return self.config.get("file_path", "")
    
    @property
    def sheet_name(self) -> str:
        return self.config.get("sheet_name", "焊口編號明細")
    
    @property
    def pk_fields(self) -> Tuple[str, str]:
        """取得主鍵欄位名稱 (流水號, 焊口編號)"""
        return (
            self.config.get("col_serial", "流水號"),
            self.config.get("col_weld_no", "焊口編號")
        )
    
    @property
    def columns(self) -> List[dict]:
        """取得動態欄位配置"""
        return self.config.get("columns", [])
    
    def is_configured(self) -> bool:
        """檢查是否已設定"""
        return bool(self.file_path) and os.path.exists(self.file_path)
    
    def make_pk(self, serial: str, weld_no: str) -> str:
        """建立主鍵"""
        return f"{serial}_{weld_no}"
    
    def parse_pk(self, pk: str) -> Tuple[str, str]:
        """解析主鍵"""
        parts = pk.split("_", 1)
        if len(parts) == 2:
            return parts[0], parts[1]
        return pk, ""
    
    # ========= 快取管理 =========
    
    def _get_cache_dir(self) -> str:
        """取得快取目錄路徑"""
        if self.file_path:
            base_dir = os.path.dirname(self.file_path)
            return os.path.join(base_dir, CACHE_DIR)
        return ""
    
    def _get_cache_path(self) -> str:
        """取得快取檔案路徑"""
        if not self.file_path:
            return ""
        
        # 使用檔案路徑 hash 作為快取檔名
        file_hash = hashlib.md5(self.file_path.encode('utf-8')).hexdigest()[:12]
        sheet_safe = self.sheet_name.replace("/", "_").replace("\\", "_")
        cache_name = f"weld_cache_{sheet_safe}_{file_hash}.json"
        
        return os.path.join(self._get_cache_dir(), cache_name)
    
    def _get_file_mtime(self) -> float:
        """取得 Excel 檔案的修改時間"""
        if self.file_path and os.path.exists(self.file_path):
            return os.path.getmtime(self.file_path)
        return 0
    
    def _is_cache_valid(self) -> bool:
        """檢查快取是否有效"""
        cache_path = self._get_cache_path()
        if not cache_path or not os.path.exists(cache_path):
            return False
        
        try:
            with open(cache_path, 'r', encoding='utf-8') as f:
                meta = json.load(f)
            
            # 檢查 Excel 檔案修改時間
            cached_mtime = meta.get("excel_mtime", 0)
            current_mtime = self._get_file_mtime()
            
            # 如果 Excel 沒有變更，快取有效
            if abs(cached_mtime - current_mtime) < 1:  # 1秒誤差容許
                return True
        except Exception:
            pass
        
        return False
    
    def _load_from_cache(self) -> bool:
        """從 JSON 快取載入"""
        cache_path = self._get_cache_path()
        if not cache_path:
            return False
        
        try:
            start_time = time.time()
            
            with open(cache_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            self._headers = data.get("headers", [])
            self._col_map = {h: i for i, h in enumerate(self._headers) if h}
            self._cache = data.get("cache", {})
            self._serial_index = data.get("serial_index", {})
            self._file_mtime = data.get("excel_mtime", 0)
            self._sheet_name = data.get("sheet_name") or self.sheet_name
            self._header_row = int(data.get("header_row") or 1)
            
            self._load_time = time.time() - start_time
            self._loaded = True
            
            print(f"✅ 從快取載入焊口: {len(self._cache)} 筆 ({self._load_time*1000:.1f}ms)")
            return True
            
        except Exception as e:
            print(f"⚠️ 快取載入失敗: {e}")
            return False
    
    def _save_to_cache(self):
        """儲存到 JSON 快取"""
        cache_path = self._get_cache_path()
        if not cache_path:
            return
        
        try:
            cache_dir = self._get_cache_dir()
            if not os.path.exists(cache_dir):
                os.makedirs(cache_dir)
            
            data = {
                "excel_mtime": self._file_mtime,
                "sheet_name": self._sheet_name or self.sheet_name,
                "header_row": self._header_row,
                "headers": self._headers,
                "cache": self._cache,
                "serial_index": self._serial_index,
                "created_at": datetime.now().isoformat(),
            }
            
            with open(cache_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            
            print(f"💾 已儲存快取: {cache_path}")
            
        except Exception as e:
            print(f"⚠️ 快取儲存失敗: {e}")
    
    def _build_serial_index(self):
        """建立流水號索引"""
        self._serial_index.clear()
        for pk in self._cache.keys():
            serial, _ = self.parse_pk(pk)
            if serial not in self._serial_index:
                self._serial_index[serial] = []
            self._serial_index[serial].append(pk)

    def _sheet_sort_key(self, sheet_name: str, order: int) -> tuple[int, int]:
        desired = self.sheet_name
        name_norm = _norm_label(sheet_name)
        desired_norm = _norm_label(desired)
        score = 0
        if sheet_name == desired:
            score += 1000
        if desired_norm and name_norm == desired_norm:
            score += 900
        if desired_norm and name_norm.startswith(desired_norm):
            score += 500
        if desired_norm and desired_norm in name_norm:
            score += 250

        lower_name = sheet_name.lower()
        if "new" in lower_name or "新版" in sheet_name or sheet_name.endswith("-NEW"):
            score += 100
        if "old" in lower_name or "舊" in sheet_name or sheet_name.endswith("-OLD"):
            score -= 100
        return (-score, order)

    def _candidate_sheet_names(self, wb) -> List[str]:
        ordered = sorted(
            enumerate(wb.sheetnames),
            key=lambda item: self._sheet_sort_key(item[1], item[0]),
        )
        return [name for _, name in ordered]

    def _find_header_in_sheet(self, ws, scan_rows: int = 12) -> Optional[dict]:
        from utils import resolve_col

        pk_serial, pk_weld = self.pk_fields
        max_row = ws.max_row or scan_rows
        for row_no, row in enumerate(
            ws.iter_rows(min_row=1, max_row=min(max_row, scan_rows), values_only=True),
            start=1,
        ):
            headers = [_header_text(value) for value in (row or [])]
            col_map = {}
            for index, header in enumerate(headers):
                if header and header not in col_map:
                    col_map[header] = index
            if not col_map:
                continue

            pk_serial_resolved = resolve_col(pk_serial, col_map.keys())
            pk_weld_resolved = resolve_col(pk_weld, col_map.keys())
            serial_idx = col_map.get(pk_serial_resolved)
            weld_idx = col_map.get(pk_weld_resolved)
            if serial_idx is None or weld_idx is None:
                continue

            return {
                "row_no": row_no,
                "headers": headers,
                "col_map": col_map,
                "serial_idx": serial_idx,
                "weld_idx": weld_idx,
            }
        return None

    def _resolve_sheet_and_header(self, wb) -> Optional[tuple[str, Any, dict]]:
        for sheet_name in self._candidate_sheet_names(wb):
            ws = wb[sheet_name]
            header_info = self._find_header_in_sheet(ws)
            if header_info:
                return sheet_name, ws, header_info
        return None
    
    def _load_from_excel(self) -> bool:
        """從 Excel 檔案載入"""
        if not OPENPYXL_AVAILABLE:
            print("⚠️ 需要 openpyxl 模組")
            return False
        
        try:
            start_time = time.time()
            
            wb = load_workbook(self.file_path, read_only=True, data_only=True)

            resolved = self._resolve_sheet_and_header(wb)
            if resolved is None:
                print(f"⚠️ 找不到可用的焊口工作表: {self.sheet_name}")
                print(f"   可用工作表: {', '.join(wb.sheetnames)}")
                wb.close()
                return False

            sheet_name, ws, header_info = resolved
            requested_sheet = self.sheet_name
            if sheet_name != requested_sheet:
                print(f"ℹ️ 自動改用焊口工作表: {sheet_name}（設定值: {requested_sheet}）")
                self.config["sheet_name"] = sheet_name

            self._sheet_name = sheet_name
            self._header_row = int(header_info["row_no"])
            self._headers = header_info["headers"]
            self._col_map = header_info["col_map"]
            serial_idx = header_info["serial_idx"]
            weld_idx = header_info["weld_idx"]
            
            # 載入所有資料
            self._cache.clear()
            self._serial_index.clear()
            
            for row in ws.iter_rows(min_row=self._header_row + 1, values_only=True):
                if not row or not row[serial_idx]:
                    continue
                
                serial = str(row[serial_idx] or "").strip()
                weld_no = str(row[weld_idx] or "").strip()
                
                if not serial or not weld_no:
                    continue
                
                pk = self.make_pk(serial, weld_no)
                
                # 儲存整列資料
                row_data = {}
                for i, h in enumerate(self._headers):
                    if not h:
                        continue
                    if i < len(row):
                        value = row[i]
                        # 確保可以 JSON 序列化
                        if value is not None:
                            if hasattr(value, 'isoformat'):  # datetime
                                value = value.isoformat()
                            elif not isinstance(value, (str, int, float, bool)):
                                value = str(value)
                        row_data[h] = value
                
                self._cache[pk] = row_data
                
                # 更新流水號索引
                if serial not in self._serial_index:
                    self._serial_index[serial] = []
                self._serial_index[serial].append(pk)
            
            wb.close()
            
            self._file_mtime = self._get_file_mtime()
            self._load_time = time.time() - start_time
            
            print(
                f"✅ 從 Excel 載入焊口: {len(self._cache)} 筆 "
                f"({self._sheet_name}, 第 {self._header_row} 列標題, {self._load_time*1000:.1f}ms)"
            )
            
            # 儲存快取
            self._save_to_cache()
            
            return True
            
        except PermissionError:
            print("❌ 焊口管制表被開啟中，請先關閉 Excel")
            return False
        except Exception as e:
            print(f"❌ 載入焊口管制表失敗: {e}")
            return False
    
    # ========= 主要載入方法 =========
    
    def load(self, force_reload: bool = False) -> bool:
        """
        載入焊口管制表（智慧快取）
        
        載入順序：
        1. 如果已在記憶體中且不需強制重載 → 直接使用
        2. 檢查 JSON 快取是否有效 → 從快取載入
        3. 否則從 Excel 載入並建立快取
        
        Returns:
            是否成功載入
        """
        if not self.is_configured():
            print("⚠️ 焊口管制表未設定或檔案不存在")
            return False
        
        # 情況1: 已載入且不需強制重載
        if self._loaded and not force_reload:
            # 檢查檔案是否被修改
            current_mtime = self._get_file_mtime()
            if abs(current_mtime - self._file_mtime) < 1:
                return True
            print("📄 偵測到 Excel 檔案變更，重新載入...")
        
        # 情況2: 嘗試從快取載入
        if not force_reload and self._is_cache_valid():
            if self._load_from_cache():
                return True
        
        # 情況3: 從 Excel 載入
        if self._load_from_excel():
            self._loaded = True
            self._file_path = self.file_path
            self._sheet_name = self.sheet_name
            return True
        
        return False
    
    def reload(self) -> bool:
        """強制重新載入"""
        return self.load(force_reload=True)
    
    def invalidate_cache(self):
        """清除快取（包括記憶體和檔案）"""
        self._cache.clear()
        self._serial_index.clear()
        self._loaded = False
        
        # 刪除快取檔案
        cache_path = self._get_cache_path()
        if cache_path and os.path.exists(cache_path):
            try:
                os.remove(cache_path)
                print(f"🗑️ 已刪除快取: {cache_path}")
            except Exception as e:
                print(f"⚠️ 刪除快取失敗: {e}")
    
    # ========= 查詢方法 =========
    
    def check_exists(self, serial: str, weld_no: str) -> bool:
        """
        檢查焊口是否存在
        
        Args:
            serial: 流水號
            weld_no: 焊口編號
            
        Returns:
            是否存在
        """
        if not self._loaded:
            self.load()
        
        pk = self.make_pk(str(serial).strip(), str(weld_no).strip())
        return pk in self._cache
    
    def check_exists_batch(self, welds: List[Tuple[str, str]]) -> Dict[str, bool]:
        """
        批次檢查焊口是否存在
        
        Args:
            welds: [(serial, weld_no), ...]
            
        Returns:
            {pk: exists}
        """
        if not self._loaded:
            self.load()
        
        result = {}
        for serial, weld_no in welds:
            pk = self.make_pk(str(serial).strip(), str(weld_no).strip())
            result[pk] = pk in self._cache
        return result
    
    def get_weld_info(self, serial: str, weld_no: str) -> Optional[dict]:
        """
        取得焊口資訊
        
        Returns:
            焊口資料字典，不存在則回傳 None
        """
        if not self._loaded:
            self.load()
        
        pk = self.make_pk(str(serial).strip(), str(weld_no).strip())
        return self._cache.get(pk)
    
    def get_all_welds_by_serial(self, serial: str) -> List[dict]:
        """
        取得指定流水號的所有焊口（使用索引，O(1) 查詢）
        
        Args:
            serial: 流水號
            
        Returns:
            焊口資料列表
        """
        if not self._loaded:
            self.load()
        
        serial = str(serial).strip()
        
        # 使用索引快速查詢
        pks = self._serial_index.get(serial, [])
        return [self._cache[pk] for pk in pks if pk in self._cache]
    
    def get_welds_count_by_serial(self, serial: str) -> int:
        """取得指定流水號的焊口數量（使用索引）"""
        if not self._loaded:
            self.load()
        
        serial = str(serial).strip()
        return len(self._serial_index.get(serial, []))
    
    def serial_exists(self, serial: str) -> bool:
        """檢查流水號是否有任何焊口（使用索引）"""
        if not self._loaded:
            self.load()
        
        serial = str(serial).strip()
        return serial in self._serial_index and len(self._serial_index[serial]) > 0
    
    # ========= 新增/修改方法 =========
    
    def add_weld(self, weld_data: dict) -> Tuple[bool, str]:
        """
        新增焊口到管制表
        
        Args:
            weld_data: 焊口資料字典，key 對應欄位名稱
            
        Returns:
            (成功與否, 訊息)
        """
        if not self.is_configured():
            return False, "焊口管制表未設定"
        
        if not self._loaded:
            if not self.load():
                return False, "無法載入焊口管制表"
        
        pk_serial, pk_weld = self.pk_fields
        serial = str(weld_data.get(pk_serial, "")).strip()
        weld_no = str(weld_data.get(pk_weld, "")).strip()
        
        if not serial or not weld_no:
            return False, f"缺少必要欄位: {pk_serial}, {pk_weld}"
        
        # 檢查是否已存在
        pk = self.make_pk(serial, weld_no)
        if pk in self._cache:
            return False, f"焊口已存在: {pk}"
        
        try:
            wb = load_workbook(self.file_path)
            try:
                ws = wb[self.sheet_name]
                
                # 找到下一個空白列
                next_row = ws.max_row + 1
                
                # 重新讀取欄位映射（確保與實際檔案同步）
                headers = [cell.value for cell in ws[1]]
                col_map = {h: i + 1 for i, h in enumerate(headers) if h}
                
                # 寫入資料
                for field, value in weld_data.items():
                    if field in col_map:
                        ws.cell(row=next_row, column=col_map[field], value=value)
                
                wb.save(self.file_path)
            finally:
                wb.close()
            
            # 更新快取
            self._cache[pk] = weld_data.copy()
            
            # 更新流水號索引
            if serial not in self._serial_index:
                self._serial_index[serial] = []
            self._serial_index[serial].append(pk)
            
            # 更新檔案修改時間
            self._file_mtime = self._get_file_mtime()
            
            # 標記快取需要更新（延遲寫入）
            self._cache_dirty = True
            if self._auto_save_cache:
                self._save_to_cache()
            
            return True, f"已新增焊口: {pk}"
            
        except PermissionError:
            return False, "焊口管制表被開啟中，請先關閉 Excel"
        except Exception as e:
            return False, f"新增失敗: {e}"
    
    def add_welds_batch(self, welds: List[dict]) -> Tuple[int, int, List[str]]:
        """
        批次新增焊口
        
        Args:
            welds: 焊口資料列表
            
        Returns:
            (成功數, 跳過數, 錯誤訊息列表)
        """
        if not self.is_configured():
            return 0, 0, ["焊口管制表未設定"]
        
        if not self._loaded:
            if not self.load():
                return 0, 0, ["無法載入焊口管制表"]
        
        pk_serial, pk_weld = self.pk_fields
        
        # 先篩選出要新增的（排除已存在的）
        to_add = []
        skipped = 0
        
        for weld_data in welds:
            serial = str(weld_data.get(pk_serial, "")).strip()
            weld_no = str(weld_data.get(pk_weld, "")).strip()
            
            if not serial or not weld_no:
                continue
            
            pk = self.make_pk(serial, weld_no)
            if pk in self._cache:
                skipped += 1
                continue
            
            to_add.append((pk, weld_data, serial))
        
        if not to_add:
            return 0, skipped, []
        
        errors = []
        added = 0
        
        try:
            wb = load_workbook(self.file_path)
            try:
                ws = wb[self.sheet_name]
                
                # 讀取欄位映射
                headers = [cell.value for cell in ws[1]]
                col_map = {h: i + 1 for i, h in enumerate(headers) if h}
                
                # 建立 field → resolved_header 對照（同義字模糊匹配）
                from utils import resolve_col
                _field_cache: dict[str, str] = {}

                def _resolve(field: str) -> str | None:
                    if field not in _field_cache:
                        r = resolve_col(field, col_map)
                        _field_cache[field] = r if r in col_map else None  # type: ignore
                    return _field_cache[field]

                next_row = ws.max_row + 1
                
                for pk, weld_data, serial in to_add:
                    # 寫入資料（含 fuzzy 欄名匹配）
                    for field, value in weld_data.items():
                        resolved = _resolve(field)
                        if resolved and resolved in col_map:
                            ws.cell(row=next_row, column=col_map[resolved], value=value)
                    
                    # 更新快取
                    self._cache[pk] = weld_data.copy()
                    
                    # 更新流水號索引
                    if serial not in self._serial_index:
                        self._serial_index[serial] = []
                    self._serial_index[serial].append(pk)
                    
                    next_row += 1
                    added += 1
                
                wb.save(self.file_path)
            finally:
                wb.close()
            
            # 更新檔案修改時間
            self._file_mtime = self._get_file_mtime()
            
            # 標記快取需要更新（延遲寫入）
            self._cache_dirty = True
            if self._auto_save_cache:
                self._save_to_cache()
            
        except PermissionError:
            errors.append("焊口管制表被開啟中，請先關閉 Excel")
        except Exception as e:
            errors.append(f"批次新增失敗: {e}")
        
        return added, skipped, errors
    
    # ========= 快取控制方法 =========
    
    def begin_batch(self):
        """
        開始批次操作（暫停自動儲存快取）
        
        使用方式：
            manager.begin_batch()
            for weld in welds:
                manager.add_weld(weld)
            manager.end_batch()  # 一次性儲存快取
        """
        self._auto_save_cache = False
    
    def end_batch(self):
        """結束批次操作（儲存快取並恢復自動儲存）"""
        self._auto_save_cache = True
        if self._cache_dirty:
            self._save_to_cache()
            self._cache_dirty = False
    
    def flush_cache(self):
        """強制儲存快取到磁碟"""
        if self._cache_dirty:
            self._save_to_cache()
            self._cache_dirty = False
    
    def get_headers(self) -> List[str]:
        """取得管制表的所有欄位名稱"""
        if not self._loaded:
            self.load()
        return self._headers.copy()
    
    def get_statistics(self) -> dict:
        """取得統計資訊"""
        if not self._loaded:
            self.load()
        
        return {
            "total_welds": len(self._cache),
            "total_serials": len(self._serial_index),
            "headers": self._headers,
            "load_time_ms": self._load_time * 1000,
            "file_mtime": self._file_mtime,
        }
    
    def get_performance_info(self) -> dict:
        """取得效能資訊"""
        cache_path = self._get_cache_path()
        cache_size = 0
        if cache_path and os.path.exists(cache_path):
            cache_size = os.path.getsize(cache_path)
        
        return {
            "loaded": self._loaded,
            "total_welds": len(self._cache),
            "total_serials": len(self._serial_index),
            "load_time_ms": self._load_time * 1000,
            "cache_valid": self._is_cache_valid(),
            "cache_dirty": self._cache_dirty,
            "cache_size_kb": cache_size / 1024,
            "cache_path": cache_path,
        }


# ========= 全域實例 =========
_weld_manager: Optional[WeldControlManager] = None


def get_weld_manager() -> WeldControlManager:
    """取得焊口管制管理器（單例）"""
    global _weld_manager
    if _weld_manager is None:
        _weld_manager = WeldControlManager()
    return _weld_manager


def init_weld_manager_from_settings():
    """從設定初始化焊口管制管理器"""
    try:
        from settings_manager import get_weld_control_table_path, get_weld_control_config
        
        config = get_weld_control_config()
        config["file_path"] = get_weld_control_table_path()
        
        manager = get_weld_manager()
        manager.configure(config)
        return manager
    except Exception as e:
        print(f"⚠️ 初始化焊口管制管理器失敗: {e}")
        return None


def check_welds_exist(serial: str, weld_nos: List[str]) -> Dict[str, Optional[dict]]:
    """
    便捷函數：檢查焊口是否存在
    
    Args:
        serial: 流水號
        weld_nos: 焊口編號列表
        
    Returns:
        {weld_no: weld_info or None}（只回傳存在的）
    """
    manager = init_weld_manager_from_settings()
    if not manager or not manager.is_configured():
        return {}
    
    manager.load()
    
    result = {}
    for weld_no in weld_nos:
        info = manager.get_weld_info(serial, weld_no)
        if info:
            result[weld_no] = info
    return result


def add_new_welds(serial: str, welds_data: List[dict], 
                  report_id: str = "", line_no: str = "") -> Tuple[int, int, List[str]]:
    """
    便捷函數：新增焊口到管制表
    
    Args:
        serial: 流水號
        welds_data: 焊口資料列表，每個包含 weld_no, size 等
        report_id: 修改單編號
        line_no: LINE NO
        
    Returns:
        (成功數, 跳過數, 錯誤訊息)
    """
    manager = init_weld_manager_from_settings()
    if not manager or not manager.is_configured():
        return 0, 0, ["焊口管制表未設定"]
    
    config = manager.config
    pk_serial = config.get("col_serial", "流水號")
    pk_weld = config.get("col_weld_no", "焊口編號")
    col_line = config.get("col_line_no", "LINE NO")
    col_date = config.get("col_date", "登錄日期")
    col_report = config.get("col_report_id", "修改單編號")
    
    # 準備資料
    to_add = []
    today = datetime.now().strftime("%Y/%m/%d")
    
    for wd in welds_data:
        weld_no = wd.get("weld_no", wd.get("焊口編號", ""))
        if not weld_no:
            continue
        
        row_data = {
            pk_serial: serial,
            pk_weld: weld_no,
        }
        
        # 填入其他欄位
        if line_no and col_line:
            row_data[col_line] = line_no
        if col_date:
            row_data[col_date] = today
        if report_id and col_report:
            row_data[col_report] = report_id
        
        # 從 welds_data 中取得其他欄位
        for key, value in wd.items():
            if key not in row_data and value:
                row_data[key] = value
        
        to_add.append(row_data)
    
    return manager.add_welds_batch(to_add)


# ========= 資料群掃描功能 =========

import re

def parse_folder_name(folder_name: str) -> Optional[Dict]:
    """
    解析資料夾名稱，取得流水號和焊口列表
    
    格式範例:
        - 72_11r2          -> serial=72, welds=[11r]
        - 243_12a1_12b1    -> serial=243, welds=[12a, 12b]
        - 632_AG           -> serial=632, welds=[] (Group 模式)
    
    Returns:
        {
            "serial": "72",
            "welds": [{"weld_no": "11", "mark": "r", "size": "2", "weld_id": "11r"}],
            "is_group": False,
            "folder_name": "72_11r2"
        }
    """
    if not folder_name or "_" not in folder_name:
        return None
    
    parts = folder_name.split("_")
    
    # 第一部分是流水號
    serial = parts[0]
    if not serial.isdigit():
        return None
    
    result = {
        "serial": serial,
        "welds": [],
        "is_group": False,
        "folder_name": folder_name
    }
    
    # 檢查是否是 Group 模式（AG, BG 等）
    if len(parts) == 2 and len(parts[1]) == 2 and parts[1].endswith("G"):
        result["is_group"] = True
        return result
    
    # 解析焊口代碼
    # 格式: 數字 + 標記(r/a/b) + 尺寸
    weld_pattern = re.compile(r'^(\d+)([rab])(.+)$')
    
    for part in parts[1:]:
        match = weld_pattern.match(part)
        if match:
            weld_no = match.group(1)
            mark = match.group(2)
            size = match.group(3)
            weld_id = f"{weld_no}{mark}"  # 焊口編號 = 焊口號 + 標記
            
            result["welds"].append({
                "weld_no": weld_no,
                "mark": mark,
                "size": size,
                "weld_id": weld_id,
                "code": part  # 完整代碼
            })
    
    return result


def scan_attachments_folder(attachments_root: str = None) -> List[Dict]:
    """
    掃描 attachments 資料夾，取得所有資料群的焊口資訊
    
    Returns:
        [
            {
                "date": "20250808",
                "folder_name": "72_11r2",
                "folder_path": "...",
                "serial": "72",
                "welds": [...],
                "is_group": False
            },
            ...
        ]
    """
    if attachments_root is None:
        try:
            from config import ATTACHMENTS_ROOT
            attachments_root = ATTACHMENTS_ROOT
        except ImportError:
            attachments_root = os.path.join(os.path.dirname(__file__), "..", "attachments")
    
    results = []
    
    if not os.path.exists(attachments_root):
        return results
    
    # 遍歷日期資料夾
    for date_dir in sorted(os.listdir(attachments_root)):
        date_path = os.path.join(attachments_root, date_dir)
        if not os.path.isdir(date_path):
            continue
        
        # 遍歷資料群資料夾
        for folder_name in sorted(os.listdir(date_path)):
            folder_path = os.path.join(date_path, folder_name)
            if not os.path.isdir(folder_path):
                continue
            
            # 解析資料夾名稱
            parsed = parse_folder_name(folder_name)
            if parsed:
                parsed["date"] = date_dir
                parsed["folder_path"] = folder_path
                results.append(parsed)
    
    return results


def find_missing_welds(attachments_root: str = None, 
                       serial_format: str = "raw") -> List[Dict]:
    """
    找出所有資料群中未登錄於焊口管制表的焊口
    
    Args:
        attachments_root: attachments 資料夾路徑
        serial_format: 流水號格式 ("raw" 或 "pad4")
    
    Returns:
        [
            {
                "date": "20250808",
                "folder_name": "72_11r2",
                "folder_path": "...",
                "serial": "72",              # 原始流水號
                "serial_formatted": "72",    # 格式化後的流水號
                "weld_id": "11r",
                "weld_no": "11",
                "mark": "r",
                "size": "2",
                "code": "11r2"
            },
            ...
        ]
    """
    # 載入焊口管制表
    manager = init_weld_manager_from_settings()
    if not manager or not manager.is_configured():
        print("⚠️ 焊口管制表未設定")
        return []
    
    if not manager.load():
        print("⚠️ 無法載入焊口管制表")
        return []
    
    # 取得流水號格式設定
    if serial_format is None:
        try:
            from settings_manager import get_weld_control_config
            config = get_weld_control_config()
            serial_format = config.get("serial_format", "raw")
        except Exception:
            serial_format = "raw"
    
    # 掃描所有資料群
    all_folders = scan_attachments_folder(attachments_root)
    
    missing = []
    
    for folder in all_folders:
        # 嘗試讀取 weld_info.json（包含完整焊口資訊：材質、厚度）
        weld_info_path = os.path.join(folder["folder_path"], "weld_info.json")
        weld_info_data = None
        
        if os.path.exists(weld_info_path):
            try:
                with open(weld_info_path, 'r', encoding='utf-8') as f:
                    weld_info_data = json.load(f)
            except Exception as e:
                print(f"⚠️ 讀取 {weld_info_path} 失敗: {e}")
        
        if folder["is_group"]:
            # Group 模式需要讀取 GroupWeld.txt 或 weld_info.json
            if weld_info_data and weld_info_data.get("welds"):
                # 優先使用 weld_info.json（包含材質、厚度）
                for w in weld_info_data["welds"]:
                    wn = str(w.get("weld_no", ""))
                    mk = str(w.get("mark", ""))
                    m_sp = re.match(r'^(\d+)([rab])', wn)
                    if m_sp:
                        wn = m_sp.group(1)
                        mk = mk or m_sp.group(2)
                    weld_id = f"{wn}{mk}"
                    folder["welds"].append({
                        "weld_no": wn,
                        "mark": mk,
                        "size": w.get("size", ""),
                        "material": w.get("material", ""),
                        "thickness": w.get("thickness", ""),
                        "weld_id": weld_id,
                        "code": w.get("code", weld_id)
                    })
            else:
                # 備用：讀取 GroupWeld.txt
                gw_path = os.path.join(folder["folder_path"], "GroupWeld.txt")
                if os.path.exists(gw_path):
                    try:
                        with open(gw_path, 'r', encoding='utf-8') as f:
                            for line in f:
                                line = line.strip()
                                if not line or line.startswith("#"):
                                    continue
                                # 解析 GroupWeld.txt 中的焊口代碼
                                match = re.match(r'^(\d+)([rab])(.+)$', line)
                                if match:
                                    weld_no = match.group(1)
                                    mark = match.group(2)
                                    size = match.group(3)
                                    weld_id = f"{weld_no}{mark}"
                                    
                                    folder["welds"].append({
                                        "weld_no": weld_no,
                                        "mark": mark,
                                        "size": size,
                                        "material": "",  # 無材質資訊
                                        "thickness": "",  # 無厚度資訊
                                        "weld_id": weld_id,
                                        "code": line
                                    })
                    except Exception as e:
                        print(f"⚠️ 讀取 {gw_path} 失敗: {e}")
        else:
            # Single 模式：從資料夾名稱解析或讀取 weld_info.json
            if weld_info_data and weld_info_data.get("welds"):
                # 優先使用 weld_info.json
                folder["welds"] = []
                for w in weld_info_data["welds"]:
                    wn = str(w.get("weld_no", ""))
                    mk = str(w.get("mark", ""))
                    m_sp = re.match(r'^(\d+)([rab])', wn)
                    if m_sp:
                        wn = m_sp.group(1)
                        mk = mk or m_sp.group(2)
                    weld_id = f"{wn}{mk}"
                    folder["welds"].append({
                        "weld_no": wn,
                        "mark": mk,
                        "size": w.get("size", ""),
                        "material": w.get("material", ""),
                        "thickness": w.get("thickness", ""),
                        "weld_id": weld_id,
                        "code": w.get("code", weld_id)
                    })
            else:
                # 沿用原本從資料夾名稱解析的結果（加入空的 material/thickness）
                for w in folder["welds"]:
                    w["material"] = ""
                    w["thickness"] = ""
        
        # ★ 按 weld_id 去重（歷史 weld_info.json 可能有髒重複）★
        seen_weld_ids: set = set()
        unique_welds = []
        for w in folder["welds"]:
            if w["weld_id"] not in seen_weld_ids:
                seen_weld_ids.add(w["weld_id"])
                unique_welds.append(w)
        folder["welds"] = unique_welds

        # 格式化流水號
        serial_raw = folder["serial"]
        if serial_format == "pad4":
            serial_fmt = serial_raw.zfill(4)
        else:
            serial_fmt = serial_raw.lstrip('0') or '0'
        
        # 檢查每個焊口
        for weld in folder["welds"]:
            weld_id = weld["weld_id"]  # 含 mark，如 "1001a"
            weld_no = weld["weld_no"]  # 不含 mark，如 "1001"
            mark = weld.get("mark", "")
            
            # Excel 管制表的焊口編號一定帶後綴 (r/a/b)，直接用 weld_id 查詢
            exists = manager.check_exists(serial_fmt, weld_id)
            
            if not exists:
                missing.append({
                    "date": folder["date"],
                    "folder_name": folder["folder_name"],
                    "folder_path": folder["folder_path"],
                    "serial": serial_raw,
                    "serial_formatted": serial_fmt,
                    "weld_id": weld_id,
                    "weld_no": weld["weld_no"],
                    "mark": weld["mark"],
                    "size": weld.get("size", ""),
                    "material": weld.get("material", ""),
                    "thickness": weld.get("thickness", ""),
                    "code": weld.get("code", weld_id)
                })
    
    return missing



def get_missing_welds_summary(missing_welds: List[Dict]) -> Dict:
    """
    取得缺少焊口的統計摘要
    
    Returns:
        {
            "total_missing": 10,
            "by_serial": {"72": 2, "243": 3, ...},
            "by_date": {"20250808": 5, ...}
        }
    """
    summary = {
        "total_missing": len(missing_welds),
        "by_serial": {},
        "by_date": {}
    }
    
    for w in missing_welds:
        serial = w["serial"]
        date = w["date"]
        
        summary["by_serial"][serial] = summary["by_serial"].get(serial, 0) + 1
        summary["by_date"][date] = summary["by_date"].get(date, 0) + 1
    
    return summary


# ========= 孤兒焊口稽查 =========

_MODIFICATION_RE = re.compile(r'^(\d+)([rab])', re.IGNORECASE)

def _is_modification_or_addition(weld_no: str) -> bool:
    """判斷焊口是否為修改口(r)/新增口(a,b)/1000+號"""
    weld_no = str(weld_no).strip()
    if _MODIFICATION_RE.match(weld_no):
        return True
    # 純數字 >= 1000 也是新增口
    try:
        return int(weld_no) >= 1000
    except ValueError:
        return False


def _build_record_index() -> Dict[str, List[str]]:
    """
    從 records.json 建立 (serial, weld_id) → [report_id, ...] 索引

    serial 存為去前導零的字串，weld_id 為小寫。
    """
    import json as _json
    from record_manager import RECORDS_JSON_PATH
    index: Dict[str, List[str]] = {}  # key = "serial_weldid" → [report_ids]

    if not os.path.isfile(RECORDS_JSON_PATH):
        return index

    with open(RECORDS_JSON_PATH, "r", encoding="utf-8") as f:
        store = _json.load(f)

    for rec in store.get("records", []):
        serial_raw = str(rec.get("Series NO", "")).strip()
        report_id = str(rec.get("報告編號", "")).strip()
        weld_list_str = str(rec.get("焊口清單", "")).strip()
        if not serial_raw or not weld_list_str:
            continue

        # 去除前導零以統一比較
        serial_norm = serial_raw.lstrip("0") or "0"

        # 焊口清單用 "、" 或 "," 分隔
        for part in re.split(r"[、,，]", weld_list_str):
            weld_id = part.strip().lower()
            if not weld_id:
                continue
            key = f"{serial_norm}_{weld_id}"
            index.setdefault(key, []).append(report_id)

    return index


def audit_orphan_welds(serial_format: str = "raw") -> List[Dict]:
    """
    孤兒焊口稽查：掃描焊口管制表中的修改/新增焊口，比對 records.json。

    對每條修改/新增焊口：
      - matched:   1:1 對應到一份報告 → 回傳 report_id
      - orphan:    管制表有但找不到報告 → 標記孤兒
      - duplicate: 管制表裡出現多次或有多份報告對應 → 標記重複

    Returns:
        [{
            "serial": "200",
            "weld_no": "1001a",
            "status": "matched" / "orphan" / "duplicate",
            "report_ids": ["20251230-01"],
            "remark": "..."
        }, ...]
    """
    manager = init_weld_manager_from_settings()
    if not manager or not manager.load():
        return []

    # 1) 建立 records 索引
    rec_index = _build_record_index()

    # 2) 遍歷管制表所有焊口
    pk_serial_col, pk_weld_col = manager.pk_fields
    results: List[Dict] = []
    # 追蹤管制表中是否有重複 PK
    seen_pks: Dict[str, int] = {}  # pk → count

    for pk, row_data in manager._cache.items():
        serial, weld_no = manager.parse_pk(pk)
        if not _is_modification_or_addition(weld_no):
            continue  # 排除基本焊口

        seen_pks[pk] = seen_pks.get(pk, 0) + 1

    # 3) 再次遍歷，做比對
    for pk, row_data in manager._cache.items():
        serial, weld_no = manager.parse_pk(pk)
        if not _is_modification_or_addition(weld_no):
            continue

        serial_norm = serial.lstrip("0") or "0"
        lookup_key = f"{serial_norm}_{weld_no.lower()}"

        matched_reports = rec_index.get(lookup_key, [])

        if len(matched_reports) == 0:
            status = "orphan"
            remark = "⚠️ 孤兒：管制表有此焊口，但找不到對應報告"
        elif len(matched_reports) == 1:
            status = "matched"
            remark = f"✅ 對應報告: {matched_reports[0]}"
        else:
            status = "duplicate"
            ids = ", ".join(matched_reports)
            remark = f"⚠️ 重複：{len(matched_reports)} 份報告對應 → {ids}"

        results.append({
            "serial": serial,
            "weld_no": weld_no,
            "status": status,
            "report_ids": matched_reports,
            "remark": remark,
        })

    # 排序：孤兒優先，然後重複，最後 matched；同狀態按流水號+焊口排序
    status_order = {"orphan": 0, "duplicate": 1, "matched": 2}
    results.sort(key=lambda r: (status_order.get(r["status"], 9),
                                r["serial"].zfill(6), r["weld_no"]))
    return results


# ========= 測試 =========
if __name__ == "__main__":
    # time already imported at module level
    
    print("=== 焊口管制模組測試（優化版）===\n")
    
    # 測試初始化
    manager = init_weld_manager_from_settings()
    
    if manager and manager.is_configured():
        print(f"檔案: {manager.file_path}")
        print(f"工作表: {manager.sheet_name}")
        
        # 測試第一次載入（從 Excel 或快取）
        print("\n--- 第一次載入 ---")
        start = time.time()
        if manager.load():
            elapsed = (time.time() - start) * 1000
            stats = manager.get_statistics()
            print(f"焊口總數: {stats['total_welds']}")
            print(f"流水號數: {stats['total_serials']}")
            print(f"載入時間: {elapsed:.1f}ms")
        
        # 測試第二次載入（從記憶體）
        print("\n--- 第二次載入（記憶體快取）---")
        start = time.time()
        manager.load()
        elapsed = (time.time() - start) * 1000
        print(f"載入時間: {elapsed:.1f}ms")
        
        # 測試強制重載（從 JSON 快取）
        print("\n--- 強制重載（JSON 快取）---")
        manager._loaded = False  # 模擬重新啟動
        start = time.time()
        manager.load()
        elapsed = (time.time() - start) * 1000
        print(f"載入時間: {elapsed:.1f}ms")
        
        # 測試查詢效能
        print("\n--- 查詢效能測試 ---")
        test_serial = "0001"
        
        # 查詢所有焊口
        start = time.time()
        welds = manager.get_all_welds_by_serial(test_serial)
        elapsed = (time.time() - start) * 1000
        print(f"流水號 {test_serial} 有 {len(welds)} 個焊口 ({elapsed:.2f}ms)")
        
        # 批次檢查 100 次
        start = time.time()
        for i in range(100):
            manager.check_exists(test_serial, f"{i}r1")
        elapsed = (time.time() - start) * 1000
        print(f"100 次檢查: {elapsed:.1f}ms (平均 {elapsed/100:.3f}ms/次)")
        
        # 效能資訊
        print("\n--- 效能資訊 ---")
        perf = manager.get_performance_info()
        print(f"快取有效: {perf['cache_valid']}")
        print(f"快取大小: {perf['cache_size_kb']:.1f}KB")
        
    else:
        print("焊口管制表未設定，請先在 GUI 設定頁籤進行設定")

# ========== 報告編號查詢函數 ==========

def find_record_xlsx_path() -> Optional[str]:
    """
    自動搜尋記錄清單檔案（舊版 Excel）。
    ⚠ 已棄用 — 僅供向後相容。新程式碼請使用 build_report_id_lookup()
    （直接讀取 records.json）。
    """
    standard_path = project_path("管線修改紀錄清單.xlsx")
    if os.path.exists(standard_path):
        return standard_path
    search_dirs = [os.path.dirname(standard_path), project_path("records")]
    for search_dir in search_dirs:
        if not os.path.exists(search_dir):
            continue
        for fname in os.listdir(search_dir):
            if fname.startswith("管線修改紀錄清單") and fname.endswith(".xlsx") and not fname.startswith("~$"):
                return os.path.join(search_dir, fname)
    return None


def build_report_id_lookup(record_xlsx_path: str = None) -> Dict[Tuple[str, str], str]:
    """
    建立 (流水號, 焊口編號) → 報告編號 對照表。

    優先從 records/records.json（新格式）讀取；
    若 JSON 不存在或 details 為空，則 fallback 到舊版 Excel。

    Returns:
        dict: {(series_no_padded, weld_no): report_id, ...}
    """
    # ---------- 1. 優先：讀 records.json ----------
    records_json_path = project_path("records", "records.json")

    if os.path.exists(records_json_path):
        try:
            with open(records_json_path, "r", encoding="utf-8") as f:
                store = json.load(f)
            details = store.get("details", [])
            if details:
                lookup: Dict[Tuple[str, str], str] = {}
                for det in details:
                    report_id = det.get("紀錄編號", "")
                    series_no = det.get("Series NO", "")
                    weld_no   = det.get("焊口編號", "")
                    if not (report_id and series_no and weld_no):
                        continue
                    try:
                        series_str = str(int(series_no))
                    except Exception:
                        series_str = str(series_no).strip().lstrip('0') or '0'
                    weld_str = str(weld_no).strip()
                    # 後面覆蓋前面（越後面越新）
                    lookup[(series_str, weld_str)] = str(report_id)
                print(f"[報告編號查詢] 從 records.json 載入 {len(lookup)} 筆對照")
                return lookup
        except Exception as e:
            print(f"[警告] 讀取 records.json 失敗: {e}")

    # ---------- 2. Fallback：舊版 Excel ----------
    if not OPENPYXL_AVAILABLE:
        print("[警告] openpyxl 未安裝且無 records.json，無法查詢報告編號")
        return {}

    if not record_xlsx_path:
        record_xlsx_path = find_record_xlsx_path()

    if not record_xlsx_path or not os.path.exists(record_xlsx_path):
        print("[警告] 找不到記錄清單（JSON / Excel 皆無）")
        return {}

    print(f"[報告編號查詢] fallback 使用舊版 Excel: {os.path.basename(record_xlsx_path)}")

    lookup = {}
    try:
        wb = load_workbook(record_xlsx_path, read_only=True, data_only=True)

        if '明細' not in wb.sheetnames:
            print("[警告] 記錄清單沒有「明細」工作表")
            wb.close()
            return {}

        ws = wb['明細']
        headers = {}
        for cell in ws[1]:
            if cell.value:
                headers[cell.value] = cell.column

        from utils import resolve_col
        col_report_id = headers.get(resolve_col('紀錄編號', headers))
        col_series    = headers.get(resolve_col('Series NO', headers))
        col_weld_no   = headers.get(resolve_col('焊口編號', headers))

        if not all([col_report_id, col_series, col_weld_no]):
            print(f"[警告] 明細工作表缺少必要欄位")
            wb.close()
            return {}

        for row in ws.iter_rows(min_row=2, values_only=True):
            report_id = row[col_report_id - 1] if col_report_id else None
            series_no = row[col_series - 1] if col_series else None
            weld_no   = row[col_weld_no - 1] if col_weld_no else None

            if report_id and series_no and weld_no:
                try:
                    series_str = f"{int(series_no):04d}"
                except Exception:
                    series_str = str(series_no)
                weld_str = str(weld_no).strip()
                lookup[(series_str, weld_str)] = str(report_id)

        wb.close()
        print(f"[報告編號查詢] 已建立 {len(lookup)} 筆對照 (Excel fallback)")
        return lookup

    except Exception as e:
        print(f"[錯誤] 讀取記錄清單失敗: {e}")
        return {}


def query_report_id(series_no: str, weld_no: str, lookup: Dict[Tuple[str, str], str] = None, 
                    record_xlsx_path: str = None) -> Optional[str]:
    """
    查詢單一焊口的報告編號
    
    Args:
        series_no: 流水號
        weld_no: 焊口編號  
        lookup: 已建立的對照表（可選，若無則重新建立）
        record_xlsx_path: 記錄清單路徑（可選）
        
    Returns:
        str: 報告編號，如 "20251020-01"，若查無則回傳 None
    """
    if lookup is None:
        lookup = build_report_id_lookup(record_xlsx_path)
    
    # 標準化流水號
    try:
        series_str = f"{int(series_no):04d}"
    except Exception:
        series_str = str(series_no)
    
    weld_str = str(weld_no).strip()
    key = (series_str, weld_str)
    
    return lookup.get(key)
