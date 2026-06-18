# -*- coding: utf-8 -*-
"""
record_manager.py — 紀錄清單管理模組

包含：
- 紀錄清單 CRUD
- 明細表 upsert
- DWG LIST 讀取
- 自動備份
"""

import os
import json
import shutil
from datetime import datetime
from typing import List, Dict, Tuple, Optional

from openpyxl import Workbook

from config import (
    RECORD_XLSX_PATH,
    RECORD_HEADER, DETAIL_HEADER, MATERIALS_HEADER
)
from utils import safe_load_workbook, atomic_save_wb, atomic_write_json, parse_seq_from_report_id
from billing_calculator import amount_to_text, parse_amount
from billing_status import is_billing_locked
from log_config import get_logger

logger = get_logger(__name__)

# ========= JSON 資料路徑 =========
_RECORDS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "records")
RECORDS_JSON_PATH = os.path.join(_RECORDS_DIR, "records.json")
BILLING_JSON_PATH = os.path.join(_RECORDS_DIR, "billing.json")
DWG_MAP_JSON_PATH = os.path.join(_RECORDS_DIR, "dwg_map.json")

# ========= 自動備份 =========
def auto_backup(path: str = None, max_backups: int = 5) -> str:
    """
    自動備份 JSON 或 Excel 檔案

    每次 upsert 前呼叫，備份到 backups/ 子目錄。
    自動清理超過 max_backups 個的舊備份。

    Returns:
        備份檔案路徑（若略過則回傳空字串）
    """
    if path is None:
        path = RECORDS_JSON_PATH

    if not os.path.exists(path):
        return ""

    backup_dir = os.path.join(os.path.dirname(path), "backups")
    os.makedirs(backup_dir, exist_ok=True)

    base_name = os.path.splitext(os.path.basename(path))[0]
    ext = os.path.splitext(path)[1]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"{base_name}_{timestamp}{ext}"
    backup_path = os.path.join(backup_dir, backup_name)

    try:
        shutil.copy2(path, backup_path)
        logger.info(f"📦 已備份: {backup_name}")
    except Exception as e:
        logger.warning(f"備份失敗: {e}")
        return ""

    # 清理舊備份（只保留最近 max_backups 個）
    try:
        pattern = f"{base_name}_"
        backups = sorted(
            [f for f in os.listdir(backup_dir)
             if f.startswith(pattern) and f.endswith(ext)],
            reverse=True
        )
        for old in backups[max_backups:]:
            os.remove(os.path.join(backup_dir, old))
            logger.debug(f"清理舊備份: {old}")
    except Exception:
        pass

    return backup_path


# ========= JSON 資料存取核心 =========
def _ensure_records_dir():
    """確保 records/ 目錄存在"""
    os.makedirs(_RECORDS_DIR, exist_ok=True)


def _load_store() -> dict:
    """載入 records.json（主資料庫），若不存在則自動從舊 Excel 遷移"""
    if not os.path.exists(RECORDS_JSON_PATH):
        # 若舊 Excel 存在，自動遷移
        if os.path.exists(RECORD_XLSX_PATH):
            logger.info("📦 偵測到舊 Excel 紀錄，自動遷移至 JSON...")
            migrate_excel_to_json()
            # 遷移完成後重新讀取
            if os.path.exists(RECORDS_JSON_PATH):
                with open(RECORDS_JSON_PATH, "r", encoding="utf-8") as f:
                    return json.load(f)
        return {"records": [], "details": [], "materials": [], "meta": {"version": "2.0"}}
    with open(RECORDS_JSON_PATH, "r", encoding="utf-8") as f:
        store = json.load(f)
    # 確保舊 JSON 也有 materials key
    if "materials" not in store:
        store["materials"] = []
    return store


def _save_store(store: dict):
    """儲存 records.json（原子寫入）"""
    _ensure_records_dir()
    store["meta"]["last_modified"] = datetime.now().isoformat()
    atomic_write_json(RECORDS_JSON_PATH, store)
    logger.debug(f"💾 records.json 已儲存 ({len(store['records'])} 筆)")


# ========= 紀錄索引載入 =========
def preload_record_index(path: str = None) -> Tuple[
    set,
    Dict[Tuple[str, str], int],
    Dict[Tuple[str, str], Dict[str, Optional[str]]],
    Dict[str, int]
]:
    """
    預載紀錄索引（從 JSON）

    Returns:
        (existing_key_set, key_to_row, key_to_meta, max_seq_by_date)
    """
    store = _load_store()  # 自動遷移已在 _load_store 內處理

    existing_key_set = set()
    key_to_row: Dict[Tuple[str, str], int] = {}
    key_to_meta: Dict[Tuple[str, str], Dict[str, Optional[str]]] = {}
    max_seq_by_date: Dict[str, int] = {}

    for i, rec in enumerate(store["records"]):
        date_str = str(rec.get("日期", ""))
        folder = str(rec.get("資料夾名", ""))
        report_id = rec.get("報告編號")
        fp = rec.get("內容指紋")

        if not date_str or not folder:
            continue

        key = (date_str, folder)
        existing_key_set.add(key)
        key_to_row[key] = i  # 用 list index 代替 row number
        key_to_meta[key] = {
            'report_id': str(report_id) if report_id else None,
            'fingerprint': str(fp) if fp else None,
        }

        seq = parse_seq_from_report_id(str(report_id)) if report_id else None
        if seq is not None:
            cur = max_seq_by_date.get(date_str, 0)
            if seq > cur:
                max_seq_by_date[date_str] = seq

    return existing_key_set, key_to_row, key_to_meta, max_seq_by_date


# ========= Upsert 操作（JSON） =========
REBUILD_FLAG_FIELDS = ("需重產", "需重產原因", "需重產時間")


def upsert_record(rows: List[Dict[str, str]], path: str = None):
    """插入或更新 record 紀錄（寫入 JSON）"""
    auto_backup(RECORDS_JSON_PATH)

    store = _load_store()
    records = store["records"]

    # 建立 key → index 映射
    key_idx: Dict[Tuple[str, str], int] = {}
    for i, rec in enumerate(records):
        d = str(rec.get("日期", ""))
        f = str(rec.get("資料夾名", ""))
        if d and f:
            key_idx[(d, f)] = i

    for row in rows:
        key = (str(row["日期"]), str(row["資料夾名"]))
        if key in key_idx:
            # 更新
            idx = key_idx[key]
            records[idx].update(row)
            _clear_rebuild_flags_after_successful_upsert(records[idx], row)
        else:
            # 新增
            records.append(dict(row))
            key_idx[key] = len(records) - 1

    _save_store(store)
    logger.info(f"📝 records.json 已更新 {len(rows)} 筆 record")


def _clear_rebuild_flags_after_successful_upsert(record: Dict[str, str], row: Dict[str, str]) -> None:
    """重新產出成功寫回 record 後，清除舊的需重產旗標。"""
    if str(row.get("需重產", "")).strip() == "1":
        return
    for field in REBUILD_FLAG_FIELDS:
        record.pop(field, None)


def upsert_detail_rows(detail_rows: List[Dict[str, str]], path: str = None):
    """插入或更新明細紀錄（寫入 JSON）"""
    if not detail_rows:
        return

    store = _load_store()
    details = store["details"]

    # 建立 key → index 映射
    key_idx: Dict[Tuple[str, str], int] = {}
    max_item = 0
    for i, det in enumerate(details):
        a = str(det.get("紀錄編號", ""))
        b = str(det.get("焊口編號", ""))
        if a and b:
            key_idx[(a, b)] = i
        try:
            v = det.get("項目")
            if v is not None:
                max_item = max(max_item, int(v))
        except Exception:
            pass

    for row in detail_rows:
        key = (str(row["紀錄編號"]), str(row["焊口編號"]))
        if key in key_idx:
            idx = key_idx[key]
            for k, v in row.items():
                if k != "項目":
                    details[idx][k] = v
        else:
            max_item += 1
            row["項目"] = max_item
            details.append(dict(row))
            key_idx[key] = len(details) - 1

    _save_store(store)
    logger.info(f"📝 records.json 已更新 {len(detail_rows)} 筆明細")


def upsert_materials_rows(materials_rows: List[Dict[str, str]], path: str = None):
    """插入或更新材料明細紀錄（寫入 JSON）"""
    if not materials_rows:
        return

    store = _load_store()
    materials = store["materials"]

    locked_report_ids = _load_locked_billing_report_ids()

    # 主鍵：報告編號 + 零件類型 + 尺寸 + SCH + 材質
    key_idx: Dict[Tuple[str, str, str, str, str], int] = {}
    max_item = 0
    for i, mat in enumerate(materials):
        rid = str(mat.get("報告編號", ""))
        comp = str(mat.get("零件類型", ""))
        sz = str(mat.get("尺寸", ""))
        sch = str(mat.get("SCH", ""))
        mt = str(mat.get("材質", ""))
        if rid and comp:
            key_idx[(rid, comp, sz, sch, mt)] = i
        try:
            v = mat.get("項目")
            if v is not None:
                max_item = max(max_item, int(v))
        except Exception:
            pass

    for row in materials_rows:
        rid = str(row.get("報告編號", ""))
        key = (
            rid,
            str(row.get("零件類型", "")),
            str(row.get("尺寸", "")),
            str(row.get("SCH", "")),
            str(row.get("材質", "")),
        )
        if key in key_idx:
            if rid in locked_report_ids:
                logger.warning(f"已請款修改單材料列已鎖定，略過更新: {rid} {key[1:]}")
                continue
            idx = key_idx[key]
            existing = materials[idx]
            for k, v in row.items():
                if k != "項目":
                    if k in (
                        "單價", "金額", "單價來源", "金額來源",
                        "價目表ID", "價目來源", "價目生效日", "配價狀態",
                    ) and _preserve_material_price(existing, row):
                        continue
                    existing[k] = v
            _recalculate_preserved_material_amount(existing, row)
        else:
            if rid in locked_report_ids:
                logger.warning(f"已請款修改單材料列已鎖定，略過新增: {rid} {key[1:]}")
                continue
            max_item += 1
            row["項目"] = max_item
            materials.append(dict(row))
            key_idx[key] = len(materials) - 1

    _save_store(store)
    logger.info(f"📝 records.json 已更新 {len(materials_rows)} 筆材料明細")


def _load_locked_billing_report_ids() -> set[str]:
    if not os.path.exists(BILLING_JSON_PATH):
        return set()
    try:
        with open(BILLING_JSON_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return set()
    billing = data.get("billing", {})
    if not isinstance(billing, dict):
        return set()

    locked: set[str] = set()
    for report_id, item in billing.items():
        if not isinstance(item, dict):
            continue
        status = item.get("status") or item.get("請款狀態") or ""
        if is_billing_locked(status):
            locked.add(str(report_id))
    return locked


def _preserve_material_price(existing: dict, incoming: dict) -> bool:
    """已有單價時，不讓重跑產出的空白/價目表/未配價/未定價值回沖歷史快照。"""
    existing_price = parse_amount(existing.get("單價"))
    if existing_price is None:
        return False
    incoming_source = str(incoming.get("單價來源", "")).strip()
    return incoming_source in ("", "pricebook", "missing_pricebook", "missing_price")


def _recalculate_preserved_material_amount(existing: dict, incoming: dict) -> None:
    if not _preserve_material_price(existing, incoming):
        return
    if str(existing.get("金額來源", "")).strip() == "manual":
        return
    qty = parse_amount(existing.get("數量"))
    price = parse_amount(existing.get("單價"))
    if qty is None or price is None:
        return
    existing["金額"] = amount_to_text(qty * price)
    existing["金額來源"] = "calculated"
    if not str(existing.get("單價來源", "")).strip():
        existing["單價來源"] = "manual"


# ========= Excel 遷移工具 =========
def migrate_excel_to_json(excel_path: str = None):
    """將舊 Excel 紀錄清單遷移到 records.json"""
    if excel_path is None:
        excel_path = RECORD_XLSX_PATH

    if not os.path.exists(excel_path):
        logger.warning(f"遷移來源不存在: {excel_path}")
        return

    wb = safe_load_workbook(excel_path, data_only=True)
    records = []
    details = []

    try:
        # record 表
        if 'record' in wb.sheetnames:
            ws = wb['record']
            headers = [c.value for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                rec = {}
                for i, h in enumerate(headers):
                    if h and i < len(row):
                        val = row[i]
                        rec[str(h)] = str(val) if val is not None else ""
                records.append(rec)

        # 明細表
        if '明細' in wb.sheetnames:
            ws = wb['明細']
            headers = [c.value for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                det = {}
                for i, h in enumerate(headers):
                    if h and i < len(row):
                        val = row[i]
                        det[str(h)] = str(val) if val is not None else ""
                details.append(det)

        # 材料明細表
        materials = []
        if '材料明細' in wb.sheetnames:
            ws = wb['材料明細']
            headers = [c.value for c in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                mat = {}
                for i, h in enumerate(headers):
                    if h and i < len(row):
                        val = row[i]
                        mat[str(h)] = str(val) if val is not None else ""
                materials.append(mat)
    finally:
        wb.close()

    store = {
        "records": records,
        "details": details,
        "materials": materials,
        "meta": {
            "version": "2.0",
            "migrated_from": os.path.basename(excel_path),
            "migrated_at": datetime.now().isoformat(),
        },
    }
    _save_store(store)
    logger.info(
        f"✅ 遷移完成: {len(records)} 筆 record, {len(details)} 筆明細, "
        f"{len(materials)} 筆材料 → records.json"
    )


# ========= 匯出 Excel（給長官的報告） =========

def _get_col_letter(col_idx: int) -> str:
    """1-based 欄位索引轉字母 (1→A, 27→AA)"""
    result = ""
    while col_idx > 0:
        col_idx, rem = divmod(col_idx - 1, 26)
        result = chr(65 + rem) + result
    return result


def _apply_excel_styles(ws, *, header_row: int, data_start: int, data_end: int,
                        col_count: int, col_widths: list, title: str = None,
                        number_cols: list = None, money_cols: list = None):
    """
    統一套用專業 Excel 樣式

    Args:
        ws: 工作表
        header_row: 表頭所在列
        data_start: 資料起始列
        data_end: 資料結束列 (含)
        col_count: 欄位總數
        col_widths: 各欄寬度 list
        title: 若有值則在 header_row-1 列合併顯示標題
        number_cols: 數值欄位索引 (1-based) — 靠右對齊
        money_cols: 金額欄位索引 (1-based) — 千分位格式
    """
    from openpyxl.styles import (Font, Alignment, PatternFill,
                                  Border, Side, numbers)
    from openpyxl.utils import get_column_letter

    # ── 色彩定義 ──
    NAVY      = "1F3864"   # 深藍 — 表頭
    ACCENT    = "D6E4F0"   # 淺藍 — 斑馬紋偶數列
    WHITE     = "FFFFFF"
    BORDER_C  = "B4C6E7"   # 邊框淺灰藍
    TOTAL_BG  = "FFF2CC"   # 合計列底色（淡黃）

    hdr_font  = Font(name="Microsoft JhengHei UI", bold=True, size=10, color=WHITE)
    hdr_fill  = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    data_font = Font(name="Microsoft JhengHei UI", size=9)
    side_thin = Side(style="thin", color=BORDER_C)
    border    = Border(left=side_thin, right=side_thin,
                       top=side_thin, bottom=side_thin)
    center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_al   = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_al  = Alignment(horizontal="right", vertical="center")
    money_fmt = '#,##0'

    number_cols = set(number_cols or [])
    money_cols  = set(money_cols or [])

    # ── 標題列 ──
    if title:
        title_row = header_row - 1
        ws.merge_cells(start_row=title_row, start_column=1,
                       end_row=title_row, end_column=col_count)
        tc = ws.cell(row=title_row, column=1, value=title)
        tc.font = Font(name="Microsoft JhengHei UI", bold=True, size=13, color=NAVY)
        tc.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[title_row].height = 30

    # ── 表頭 ──
    ws.row_dimensions[header_row].height = 24
    for ci in range(1, col_count + 1):
        c = ws.cell(row=header_row, column=ci)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center
        c.border = border

    # ── 資料列: 斑馬紋 + 框線 ──
    stripe = PatternFill(start_color=ACCENT, end_color=ACCENT, fill_type="solid")
    for ri in range(data_start, data_end + 1):
        is_even = (ri - data_start) % 2 == 0
        ws.row_dimensions[ri].height = 20
        for ci in range(1, col_count + 1):
            c = ws.cell(row=ri, column=ci)
            c.font = data_font
            c.border = border
            if is_even:
                c.fill = stripe
            # 對齊
            if ci in money_cols:
                c.alignment = right_al
                c.number_format = money_fmt
            elif ci in number_cols:
                c.alignment = right_al
            else:
                c.alignment = left_al

    # ── 欄寬 ──
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── 凍結窗格：凍結表頭 ──
    ws.freeze_panes = ws.cell(row=data_start, column=1)

    # ── 自動篩選 ──
    last_letter = get_column_letter(col_count)
    ws.auto_filter.ref = f"A{header_row}:{last_letter}{data_end}"

    # ── 列印設定 ──
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = f"{header_row}:{header_row}"


def export_records_to_excel(output_path: str = None) -> str:
    """從 JSON 匯出專業格式的 Excel 紀錄清單（含材料明細與彙總）"""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from collections import defaultdict

    if output_path is None:
        output_path = os.path.join(
            os.path.dirname(RECORDS_JSON_PATH),
            f"管線修改紀錄清單_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )

    store = _load_store()
    wb = Workbook()

    NAVY = "1F3864"
    TOTAL_BG = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    bold = Font(name="Microsoft JhengHei UI", bold=True, size=10, color=NAVY)
    red_bold = Font(name="Microsoft JhengHei UI", bold=True, size=11, color="C00000")
    side_thin = Side(style="thin", color="B4C6E7")
    border = Border(left=side_thin, right=side_thin, top=side_thin, bottom=side_thin)

    def _write_total_row(ws, total_row, col_count, money_col_indices):
        """寫合計列（黃底＋SUM 公式）"""
        ws.cell(row=total_row, column=1, value="合計")
        for ci in range(1, col_count + 1):
            c = ws.cell(row=total_row, column=ci)
            c.fill = TOTAL_BG
            c.font = bold
            c.border = border
        for col_idx in money_col_indices:
            letter = get_column_letter(col_idx)
            data_start = 3  # row 2 = header, row 3 = first data
            ws.cell(row=total_row, column=col_idx,
                    value=f"=SUM({letter}{data_start}:{letter}{total_row - 1})")
            ws.cell(row=total_row, column=col_idx).number_format = '#,##0'
            ws.cell(row=total_row, column=col_idx).alignment = Alignment(
                horizontal="right", vertical="center")

    # ━━━━ Sheet 1: record 紀錄 ━━━━
    ws = wb.active
    ws.title = "record"
    ws.sheet_properties.tabColor = NAVY

    for i, h in enumerate(RECORD_HEADER, 1):
        ws.cell(row=2, column=i, value=h)

    for ri, rec in enumerate(store["records"], 3):
        for ci, h in enumerate(RECORD_HEADER, 1):
            ws.cell(row=ri, column=ci, value=rec.get(h, ""))

    data_end = max(3, 2 + len(store["records"]))
    col_widths = [11, 15, 12, 18, 12, 10, 22, 22, 34, 16, 14, 20, 9, 9, 18]
    while len(col_widths) < len(RECORD_HEADER):
        col_widths.append(12)

    _apply_excel_styles(
        ws, header_row=2, data_start=3, data_end=data_end,
        col_count=len(RECORD_HEADER), col_widths=col_widths,
        title=f"管線修改紀錄清單 — 匯出日期 {datetime.now().strftime('%Y/%m/%d')}",
    )

    # ━━━━ Sheet 2: 焊口明細 ━━━━
    ws2 = wb.create_sheet("焊口明細")
    ws2.sheet_properties.tabColor = "2E75B6"

    for i, h in enumerate(DETAIL_HEADER, 1):
        ws2.cell(row=2, column=i, value=h)

    for ri, det in enumerate(store["details"], 3):
        for ci, h in enumerate(DETAIL_HEADER, 1):
            val = det.get(h, "")
            if h in ("係數", "單價/DB", "金額", "焊口尺寸"):
                try:
                    val = float(val) if val else ""
                except (ValueError, TypeError):
                    pass
            ws2.cell(row=ri, column=ci, value=val)

    detail_end = max(3, 2 + len(store["details"]))
    detail_widths = [7, 15, 11, 28, 12, 12, 14, 10, 8, 10, 12, 16]
    while len(detail_widths) < len(DETAIL_HEADER):
        detail_widths.append(12)

    _apply_excel_styles(
        ws2, header_row=2, data_start=3, data_end=detail_end,
        col_count=len(DETAIL_HEADER), col_widths=detail_widths,
        title=f"焊口明細表 — 共 {len(store['details'])} 筆",
        number_cols=[8, 9],
        money_cols=[10, 11],
    )

    if store["details"]:
        _write_total_row(ws2, detail_end + 1, len(DETAIL_HEADER), [11])

    # ━━━━ Sheet 3: 材料明細 ━━━━
    materials = store.get("materials", [])
    ws3 = wb.create_sheet("材料明細")
    ws3.sheet_properties.tabColor = "BF8F00"

    for i, h in enumerate(MATERIALS_HEADER, 1):
        ws3.cell(row=2, column=i, value=h)

    for ri, mat in enumerate(materials, 3):
        for ci, h in enumerate(MATERIALS_HEADER, 1):
            val = mat.get(h, "")
            if h in ("數量", "單價", "金額"):
                try:
                    val = float(val) if val else ""
                except (ValueError, TypeError):
                    pass
            ws3.cell(row=ri, column=ci, value=val)

    mat_end = max(3, 2 + len(materials))
    mat_widths = [7, 15, 11, 10, 20, 8, 10, 10, 8, 6, 10, 12, 16]
    while len(mat_widths) < len(MATERIALS_HEADER):
        mat_widths.append(12)

    _apply_excel_styles(
        ws3, header_row=2, data_start=3, data_end=mat_end,
        col_count=len(MATERIALS_HEADER), col_widths=mat_widths,
        title=f"材料明細表 — 共 {len(materials)} 筆",
        number_cols=[9],
        money_cols=[11, 12],
    )

    if materials:
        _write_total_row(ws3, mat_end + 1, len(MATERIALS_HEADER), [12])

    # ━━━━ Sheet 4: 材料彙總 ━━━━
    ws4 = wb.create_sheet("材料彙總")
    ws4.sheet_properties.tabColor = "C00000"

    # 按 (零件類型, 尺寸, SCH, 材質) 彙總
    agg = defaultdict(lambda: {"數量": 0.0, "金額": 0.0, "reports": set()})
    for mat in materials:
        key = (
            mat.get("零件類型", ""),
            mat.get("尺寸", ""),
            mat.get("SCH", ""),
            mat.get("材質", ""),
        )
        try:
            qty = float(mat.get("數量", 0) or 0)
        except (ValueError, TypeError):
            qty = 0
        try:
            amt = float(mat.get("金額", 0) or 0)
        except (ValueError, TypeError):
            amt = 0
        unit = mat.get("單位", "個")
        agg[key]["數量"] += qty
        agg[key]["金額"] += amt
        agg[key]["單位"] = unit
        agg[key]["reports"].add(mat.get("報告編號", ""))

    agg_headers = ["項次", "零件類型", "尺寸", "SCH", "材質", "數量", "單位", "使用次數", "單價", "金額"]
    for i, h in enumerate(agg_headers, 1):
        ws4.cell(row=2, column=i, value=h)

    sorted_keys = sorted(agg.keys(), key=lambda k: (k[0], k[1], k[2], k[3]))
    for ri, key in enumerate(sorted_keys, 3):
        item = agg[key]
        ws4.cell(row=ri, column=1, value=ri - 2)        # 項次
        ws4.cell(row=ri, column=2, value=key[0])         # 零件類型
        ws4.cell(row=ri, column=3, value=key[1])         # 尺寸
        ws4.cell(row=ri, column=4, value=key[2])         # SCH
        ws4.cell(row=ri, column=5, value=key[3])         # 材質
        ws4.cell(row=ri, column=6, value=item["數量"])    # 數量
        ws4.cell(row=ri, column=7, value=item["單位"])    # 單位
        ws4.cell(row=ri, column=8, value=len(item["reports"]))  # 使用次數
        ws4.cell(row=ri, column=9, value="")              # 單價（待填）
        ws4.cell(row=ri, column=10, value=item["金額"] if item["金額"] else "")

    agg_end = max(3, 2 + len(sorted_keys))
    _apply_excel_styles(
        ws4, header_row=2, data_start=3, data_end=agg_end,
        col_count=len(agg_headers),
        col_widths=[7, 22, 10, 10, 12, 10, 6, 10, 12, 14],
        title=f"材料使用彙總表 — 共 {len(sorted_keys)} 類",
        number_cols=[1, 6, 8],
        money_cols=[9, 10],
    )

    if sorted_keys:
        _write_total_row(ws4, agg_end + 1, len(agg_headers), [10])

    # ━━━━ Sheet 5: 統計摘要 ━━━━
    ws5 = wb.create_sheet("摘要")
    ws5.sheet_properties.tabColor = "548235"
    summary_data = _build_summary(store)

    ws5.cell(row=1, column=1, value="管線修改單統計摘要")
    ws5.merge_cells("A1:D1")
    ws5.cell(row=1, column=1).font = Font(
        name="Microsoft JhengHei UI", bold=True, size=13, color=NAVY)
    ws5.row_dimensions[1].height = 30

    summary_headers = ["項目", "數量", "百分比", "備註"]
    for i, h in enumerate(summary_headers, 1):
        ws5.cell(row=3, column=i, value=h)

    for ri, row_data in enumerate(summary_data, 4):
        for ci, val in enumerate(row_data, 1):
            ws5.cell(row=ri, column=ci, value=val)

    _apply_excel_styles(
        ws5, header_row=3, data_start=4, data_end=3 + len(summary_data),
        col_count=4, col_widths=[24, 14, 10, 34],
        number_cols=[2],
    )

    atomic_save_wb(wb, output_path)
    wb.close()
    logger.info(f"📊 已匯出 Excel: {output_path}")
    return output_path


def _build_summary(store: dict) -> list:
    """從 store 建構統計摘要表資料（含材料統計）"""
    from collections import defaultdict
    records = store.get("records", [])
    details = store.get("details", [])
    materials = store.get("materials", [])
    total = len(records)

    # 按日期統計
    dates = set()
    change_types = {}
    for rec in records:
        d = rec.get("日期", "")
        if d:
            dates.add(d)
        ct = rec.get("變更類型", "未分類")
        change_types[ct] = change_types.get(ct, 0) + 1

    # 焊口金額統計
    weld_amount = 0
    for det in details:
        try:
            weld_amount += float(det.get("金額", 0) or 0)
        except (ValueError, TypeError):
            pass

    # 材料統計
    mat_amount = 0
    mat_types = defaultdict(float)  # 零件類型 → 總數量
    mat_report_count = len(set(m.get("報告編號", "") for m in materials))
    for mat in materials:
        try:
            mat_amount += float(mat.get("金額", 0) or 0)
        except (ValueError, TypeError):
            pass
        try:
            qty = float(mat.get("數量", 0) or 0)
        except (ValueError, TypeError):
            qty = 0
        mat_types[mat.get("零件類型", "其他")] += qty

    grand_total = weld_amount + mat_amount

    rows = [
        ["紀錄總筆數", total, "100%", ""],
        ["涵蓋日期數", len(dates), "",
         f"{min(dates, default='')} ~ {max(dates, default='')}"],
        ["焊口明細筆數", len(details), "", ""],
        ["焊口金額合計", f"${weld_amount:,.0f}", "", ""],
        ["", "", "", ""],
        ["材料明細筆數", len(materials), "",
         f"涵蓋 {mat_report_count} 份報告"],
        ["材料金額合計", f"${mat_amount:,.0f}", "", ""],
        ["材料種類數", len(mat_types), "", ""],
        ["", "", "", ""],
        ["總金額合計", f"${grand_total:,.0f}", "",
         f"焊口 ${weld_amount:,.0f} + 材料 ${mat_amount:,.0f}"],
        ["", "", "", ""],
        ["── 依變更類型分類 ──", "", "", ""],
    ]
    for ct, cnt in sorted(change_types.items(), key=lambda x: -x[1]):
        pct = f"{cnt / total * 100:.1f}%" if total else "0%"
        rows.append([f"  {ct}", cnt, pct, ""])

    if mat_types:
        rows.append(["", "", "", ""])
        rows.append(["── 材料使用分類 ──", "", "", ""])
        for comp, qty in sorted(mat_types.items(), key=lambda x: -x[1]):
            rows.append([f"  {comp}", f"{qty:g}", "", ""])

    return rows


# ========= 表頭管理（保留給 legacy / 匯入用） =========
def ensure_record_header(ws) -> Dict[str, int]:
    """確保 record 工作表有正確表頭"""
    if ws.max_row == 1 and all(ws.cell(row=1, column=i+1).value is None for i in range(len(RECORD_HEADER))):
        for i, name in enumerate(RECORD_HEADER, start=1):
            ws.cell(row=1, column=i, value=name)

    name_to_idx = {}
    for i in range(1, ws.max_column + 1):
        name = ws.cell(row=1, column=i).value
        if name:
            name_to_idx[str(name)] = i

    col = ws.max_column
    for name in RECORD_HEADER:
        if name not in name_to_idx:
            col += 1
            ws.cell(row=1, column=col, value=name)
            name_to_idx[name] = col

    return name_to_idx


def ensure_detail_header(ws) -> Dict[str, int]:
    """確保明細工作表有正確表頭"""
    if ws.max_row == 1 and all(ws.cell(row=1, column=i+1).value is None for i in range(len(DETAIL_HEADER))):
        for i, name in enumerate(DETAIL_HEADER, start=1):
            ws.cell(row=1, column=i, value=name)

    name_to_idx = {}
    for col in range(1, ws.max_column + 1):
        name = ws.cell(row=1, column=col).value
        if name:
            name_to_idx[str(name)] = col

    col = ws.max_column
    for name in DETAIL_HEADER:
        if name not in name_to_idx:
            col += 1
            ws.cell(row=1, column=col, value=name)
            name_to_idx[name] = col

    return name_to_idx


def ensure_workbook_scaffold(path: str = None):
    """確保工作簿有 record 和明細工作表（legacy，供匯入用）"""
    if path is None:
        path = RECORD_XLSX_PATH

    if os.path.exists(path):
        wb = safe_load_workbook(path)
    else:
        wb = Workbook()
        wb.active.title = "record"

    ws_record = wb['record'] if 'record' in wb.sheetnames else wb.active
    ensure_record_header(ws_record)

    if '明細' in wb.sheetnames:
        ws_detail = wb['明細']
    else:
        ws_detail = wb.create_sheet('明細')
    ensure_detail_header(ws_detail)

    atomic_save_wb(wb, path)
    wb.close()


# ========= DWG LIST 載入（含 JSON 快取） =========
def _dwg_cache_valid(excel_path: str) -> bool:
    """檢查 dwg_map.json 快取是否仍有效（比對 Excel mtime）"""
    if not os.path.exists(DWG_MAP_JSON_PATH):
        return False
    try:
        with open(DWG_MAP_JSON_PATH, "r", encoding="utf-8") as f:
            cache = json.load(f)
        cached_mtime = cache.get("source_mtime", 0)
        actual_mtime = os.path.getmtime(excel_path)
        return abs(cached_mtime - actual_mtime) < 1.0
    except Exception:
        return False


def _save_dwg_cache(mapping: Dict[str, Tuple[str, str]], excel_path: str):
    """將 DWG mapping 存為 JSON 快取"""
    data = {
        "source": os.path.basename(excel_path),
        "source_mtime": os.path.getmtime(excel_path),
        "updated": datetime.now().isoformat(),
        "count": len(mapping),
        "mapping": {k: list(v) for k, v in mapping.items()},
    }
    atomic_write_json(DWG_MAP_JSON_PATH, data)
    logger.info(f"💾 DWG MAP 快取已更新 ({len(mapping)} 筆)")


def _load_dwg_cache() -> Dict[str, Tuple[str, str]]:
    """從 JSON 快取載入 DWG mapping"""
    with open(DWG_MAP_JSON_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)
    return {k: tuple(v) for k, v in data["mapping"].items()}


def load_drawing_map(path: str = None) -> Dict[str, Tuple[str, str]]:
    """
    載入 DWG LIST（優先從 JSON 快取，Excel 變更時自動更新）

    Returns:
        {series_no: (line_number, dwg_no)}
    """
    if path is None:
        try:
            from settings_manager import get_drawing_list_path
            path = get_drawing_list_path()
        except ImportError:
            from config import DRAWING_LIST_PATH
            path = DRAWING_LIST_PATH

    if not os.path.exists(path):
        logger.warning(f"⚠️ 找不到 DWG LIST：{path}，LINE/DWG 將留空")
        # 仍嘗試讀取快取
        if os.path.exists(DWG_MAP_JSON_PATH):
            logger.info("📂 使用既有 DWG MAP 快取")
            return _load_dwg_cache()
        return {}

    # 快取有效 → 直接用
    if _dwg_cache_valid(path):
        logger.info("⚡ DWG MAP 從快取載入")
        return _load_dwg_cache()

    # 從 Excel 讀取
    logger.info("📖 從 Excel 載入 DWG LIST...")
    mapping = _load_drawing_map_from_excel(path)

    # 寫入快取
    try:
        _save_dwg_cache(mapping, path)
    except Exception as e:
        logger.warning(f"DWG MAP 快取寫入失敗: {e}")

    return mapping


def _load_drawing_map_from_excel(path: str) -> Dict[str, Tuple[str, str]]:
    """從 Excel 讀取原始 DWG mapping"""
    mapping: Dict[str, Tuple[str, str]] = {}

    wb = safe_load_workbook(path, data_only=True)
    try:
        # 尋找正確的工作表
        preferred_names = ["DRAWING LIST", "DWG LIST"]
        ws = None

        for pref in preferred_names:
            for name in wb.sheetnames:
                if str(name).strip().lower() == pref.lower():
                    ws = wb[name]
                    break
            if ws is not None:
                break

        if ws is None:
            for name in wb.sheetnames:
                sh = wb[name]
                try:
                    header_probe = [cell.value for cell in sh[1]]
                except Exception:
                    header_probe = []
                if header_probe and (
                    "Series NO" in header_probe and
                    (("LINE   NUMBER" in header_probe) or ("LINE NUMBER" in header_probe)) and
                    ("DWG NO" in header_probe)
                ):
                    ws = sh
                    break

        if ws is None:
            logger.warning("⚠️ 找不到名為 'DRAWING LIST' 的工作表，改用 active。")
            ws = wb.active

        header = [cell.value for cell in ws[1]]
        idx_series = header.index("Series NO")
        idx_line = header.index("LINE   NUMBER") if "LINE   NUMBER" in header else header.index("LINE NUMBER")
        idx_dwg = header.index("DWG NO")

        for row in ws.iter_rows(min_row=2, values_only=True):
            series = str(row[idx_series]).zfill(4) if row[idx_series] is not None else ""
            if not series:
                continue
            mapping[series] = (row[idx_line], row[idx_dwg])
    finally:
        wb.close()

    return mapping
