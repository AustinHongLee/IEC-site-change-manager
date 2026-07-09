import os
import subprocess
import sys

from openpyxl import Workbook

sys.path.insert(0, os.path.join(os.path.dirname(__file__), os.pardir, "control"))

from change_order import Spec  # noqa: E402
from weld_control import WeldControlManager  # noqa: E402
from weld_lookup import WeldLookup, normalize_series_raw  # noqa: E402


def _write_weld_control_fixture(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "焊口編號明細"
    ws.append(["流水號", "銲口編號", "尺寸", "厚度", "材質", "銲接型式", "屬性.1", "圖號", "DB數", "I.D", "預算編號"])
    ws.append([202, "1", '2"', "SCH40", "SUS304", "BW", "焊口", "DWG-202", 2, "52.5", "B-2"])
    ws.append([202, "2", '1"', "SCH80", "CS", "SW", "管牙製作安裝", "DWG-202", 1, "26.6", "B-1"])
    ws.append([202, "9", '3"', "SCH10", "SUS316", "RF", "VALVE安裝", "DWG-202", 3, "77.9", "B-3"])
    ws.append([202, "10", '4"', "SCH20", "CS", "RF", "法蘭安裝", "DWG-202", 4, "102.3", "B-4"])
    ws.append([203, "1", '6"', "SCH40", "SUS304", "BW", "焊口", "DWG-203", 6, "154.1", "B-6"])
    wb.save(path)
    wb.close()


def _lookup_for_fixture(tmp_path):
    workbook = tmp_path / "weld_control.xlsx"
    _write_weld_control_fixture(workbook)
    manager = WeldControlManager({
        "file_path": str(workbook),
        "sheet_name": "焊口編號明細",
        "col_serial": "流水號",
        "col_weld_no": "焊口編號",
    })
    return WeldLookup(manager=manager)


def _write_backfill_fixture(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "焊口編號明細"
    ws.append(["流水號", "焊口編號", "尺寸", "厚度", "材質", "銲接型式", "屬性", "焊口屬性"])
    ws.append([107, 1, 3, "S-40", "C.S", "FSW", None, "原圖焊口"])
    ws.append([107, 2, 3, "S-40", "C.S", "BW", None, "原圖焊口"])
    ws.append([108, "17r", 2, "S-80", "304", "SW", "修改", "修改"])
    ws.append([108, "1001a", 2, "S-80", "304", "BW", "新增", "新增"])
    wb.save(path)
    wb.close()


def _lookup_for_backfill_fixture(tmp_path):
    workbook = tmp_path / "weld_control_backfill.xlsx"
    _write_backfill_fixture(workbook)
    manager = WeldControlManager({
        "file_path": str(workbook),
        "sheet_name": "焊口編號明細",
        "col_serial": "流水號",
        "col_weld_no": "焊口編號",
    })
    return WeldLookup(manager=manager)


def test_manager_auto_resolves_suffixed_sheet_and_prefers_new(tmp_path):
    workbook = tmp_path / "weld_control_suffixed.xlsx"
    wb = Workbook()
    ws_old = wb.active
    ws_old.title = "焊口編號明細-OLD"
    ws_old.append(["流水號", "銲口編號", "尺寸", "厚度", "材質", "銲接型式", "屬性.1"])
    ws_old.append([149, "1", "1", "40S", "304L", "BW", "焊口"])
    ws_new = wb.create_sheet("焊口編號明細-NEW")
    ws_new.append(["流水號", "銲口編號", "尺寸", "厚度", "材質", "銲接型式", "屬性.1"])
    ws_new.append([149, "2", "2", "40S", "304L", "BW", "焊口"])
    wb.save(workbook)
    wb.close()

    manager = WeldControlManager({
        "file_path": str(workbook),
        "sheet_name": "焊口編號明細",
        "col_serial": "流水號",
        "col_weld_no": "焊口編號",
    })
    assert manager.load(force_reload=True) is True
    assert manager.sheet_name == "焊口編號明細-NEW"
    assert manager._sheet_name == "焊口編號明細-NEW"
    assert WeldLookup(manager=manager).existing_weld_ids("149") == ["2"]


def test_manager_finds_shifted_header_row(tmp_path):
    workbook = tmp_path / "weld_control_shifted_header.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "焊口編號明細"
    ws.append(["焊口管制表"])
    ws.append(["流水號", "銲口編號", "尺寸", "厚度", "材質", "銲接型式", "屬性.1"])
    ws.append([149, "3", "3", "40S", "304L", "BW", "焊口"])
    wb.save(workbook)
    wb.close()

    manager = WeldControlManager({
        "file_path": str(workbook),
        "sheet_name": "焊口編號明細",
        "col_serial": "流水號",
        "col_weld_no": "焊口編號",
    })
    assert manager.load(force_reload=True) is True
    assert manager._header_row == 2
    assert WeldLookup(manager=manager).existing_weld_ids("149") == ["3"]


def test_normalize_series_raw_strips_leading_zeroes():
    assert normalize_series_raw("0202") == "202"
    assert normalize_series_raw("202") == "202"
    assert normalize_series_raw(202) == "202"
    assert normalize_series_raw("") == "0"
    assert normalize_series_raw("0000") == "0"


def test_lookup_spec_maps_real_joint_row_to_spec(tmp_path):
    lookup = _lookup_for_fixture(tmp_path)

    assert lookup.lookup_spec("0202", "1") == Spec(
        size='2"',
        sch="SCH40",
        material="SUS304",
        weld_type="BW",
    )
    assert lookup.lookup_spec(202, "2") == Spec(
        size='1"',
        sch="SCH80",
        material="CS",
        weld_type="SW",
    )


def test_lookup_info_includes_budget_db_metadata(tmp_path):
    lookup = _lookup_for_fixture(tmp_path)

    assert lookup.lookup_info("0202", "1") == {
        "size": '2"',
        "sch": "SCH40",
        "material": "SUS304",
        "weld_type": "BW",
        "db": "2",
        "budget_no": "B-2",
        "inside_diameter": "52.5",
    }


def test_lookup_spec_returns_none_for_missing_or_non_real_rows(tmp_path):
    lookup = _lookup_for_fixture(tmp_path)

    assert lookup.lookup_spec("0202", "404") is None
    assert lookup.lookup_spec("0202", "9") is None
    assert lookup.lookup_spec("0202", "10") is None


def test_existing_weld_ids_filters_installation_accounting_rows(tmp_path):
    lookup = _lookup_for_fixture(tmp_path)

    assert lookup.existing_weld_ids("0202") == ["1", "2"]
    assert lookup.existing_weld_ids("0203") == ["1"]


def test_existing_weld_ids_accepts_backfill_sheet_without_attribute_1(tmp_path):
    lookup = _lookup_for_backfill_fixture(tmp_path)

    assert lookup.existing_weld_ids("107") == ["1", "2"]
    assert lookup.lookup_spec("107", "1") == Spec(
        size="3",
        sch="S-40",
        material="C.S",
        weld_type="FSW",
    )
    assert lookup.existing_weld_ids("108") == ["17r", "1001a"]


def test_exists_checks_manager_primary_key_with_normalized_series(tmp_path):
    lookup = _lookup_for_fixture(tmp_path)

    assert lookup.exists("0202", "1") is True
    assert lookup.exists(202, "2") is True
    assert lookup.exists("0202", "404") is False


def test_weld_lookup_import_is_headless():
    control_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), os.pardir, "control"))
    result = subprocess.run(
        [
            sys.executable,
            "-c",
            (
                "import sys; "
                f"sys.path.insert(0, {control_dir!r}); "
                "import weld_lookup; "
                "print('PyQt6' in sys.modules)"
            ),
        ],
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=False,
    )

    assert result.returncode == 0, result.stderr
    assert result.stdout.strip() == "False"
