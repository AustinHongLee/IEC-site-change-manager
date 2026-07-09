# -*- coding: utf-8 -*-

import os
import shutil
import sys
from pathlib import Path
from zipfile import ZipFile

from openpyxl import load_workbook
from PIL import Image
from pypdf import PdfWriter


sys.path.insert(0, os.path.join(os.path.dirname(__file__), os.pardir, "control"))

from owner_data_report import build_owner_data_report_package


def _write_pdf(path: Path) -> None:
    writer = PdfWriter()
    writer.add_blank_page(width=595, height=842)
    with open(path, "wb") as f:
        writer.write(f)


def _report_set(tmp_path: Path) -> dict:
    source = tmp_path / "source" / "20260702" / "720_3r2_3a2"
    source.mkdir(parents=True)
    before = source / "before_1.jpg"
    after = source / "after_1.jpg"
    pdf = source / "720.CA-2007-100-AA1B-NA-1.pdf"
    Image.new("RGB", (80, 160), (180, 40, 40)).save(before)
    Image.new("RGB", (180, 80), (40, 90, 180)).save(after)
    _write_pdf(pdf)
    return {
        "schema_version": "report_set.v1",
        "project": {
            "root": str(tmp_path),
            "attachments_root": str(tmp_path / "source"),
            "collected_at": "2026-07-02T10:00:00",
        },
        "reports": [{
            "report": {
                "report_id": "R-720",
                "date": "2026-07-02",
                "date_raw": "20260702",
                "series": "0720",
                "dwg_no": "CA-2007-100-AA1B-NA",
                "line_number": "L-100",
                "folder": "720_3r2_3a2",
                "folder_path": str(source),
                "status": "produced",
                "change_type": "現場修改",
                "description": "因現場管線干涉，切除原焊口並新增焊口。",
            },
            "welds": {
                "summary": "3r、3a（共2口）",
                "count": 2,
                "rows": [
                    {"code": "3r", "size": 3, "material": "C.S", "thickness": "S-40", "mark": "r"},
                    {"code": "3a", "size": 3, "material": "C.S", "thickness": "S-40", "mark": "a"},
                ],
            },
            "materials": {
                "summary": "管材×1",
                "count": 1,
                "rows": [{
                    "component": "管材",
                    "size": "DN80",
                    "sch": "SCH40",
                    "material": "C.S",
                    "qty": 1,
                    "unit": "米",
                    "remark": "",
                }],
            },
            "photos": {
                "before": [{"name": before.name, "path": str(before), "w": 80, "h": 160}],
                "after": [{"name": after.name, "path": str(after), "w": 180, "h": 80}],
            },
            "attachment_pdf": {"path": str(pdf), "name": pdf.name, "exists": True, "pages": 1},
            "completeness": {"level": "complete", "missing": []},
        }],
        "aggregates": {
            "report_count": 1,
            "weld_count": 2,
            "material_row_count": 1,
            "photo_count": 2,
            "before_photo_count": 1,
            "after_photo_count": 1,
        },
        "issues": [],
    }


def test_owner_data_report_package_copies_assets_and_builds_index(tmp_path):
    output = tmp_path / "out"

    result = build_owner_data_report_package(output, _report_set(tmp_path), weld_lookup=None)

    package = Path(result["package_root"])
    index = Path(result["index_xlsx"])
    report_dir = package / "R-720"
    assert package.exists()
    assert index.exists()
    assert (report_dir / "before").is_dir()
    assert (report_dir / "after").is_dir()
    assert (report_dir / "pdf").is_dir()
    assert list((report_dir / "before").glob("*.jpg"))
    assert list((report_dir / "after").glob("*.jpg"))
    assert list((report_dir / "pdf").glob("*.pdf"))

    wb = load_workbook(index, data_only=False)
    try:
        assert wb.sheetnames == ["資料索引", "焊口統計", "用料統計"]
        assert [wb["資料索引"].cell(2, col).value for col in range(1, 16)] == [
            "項次",
            "工務修改確認單編號",
            "日期",
            "ISO流編",
            "圖號",
            "Line No.",
            "新增或修改說明",
            "新增修改焊口詳細",
            "材料新增或修改摘要",
            "修改前相片",
            "修改後相片",
            "相關圖說",
            "Before檔",
            "After檔",
            "圖面PDF",
        ]
        assert wb["資料索引"]["A3"].value == 1
        assert wb["資料索引"]["B3"].value == "R-720"
        assert wb["資料索引"]["D3"].value == "0720"
        assert wb["資料索引"]["E3"].value == "CA-2007-100-AA1B-NA"
        assert wb["資料索引"]["G3"].value == "因現場管線干涉，切除原焊口並新增焊口。"
        assert not wb["資料索引"]["J3"].value
        assert wb["資料索引"]["J3"].hyperlink is None
        assert not wb["資料索引"]["K3"].value
        assert wb["資料索引"]["K3"].hyperlink is None
        assert not wb["資料索引"]["L3"].value
        assert wb["資料索引"]["L3"].hyperlink is None
        assert wb["資料索引"]["B3"].hyperlink.target == "R-720"
        assert wb["資料索引"]["M3"].hyperlink.target == "R-720/before"
        assert wb["資料索引"]["N3"].hyperlink.target == "R-720/after"
        assert wb["資料索引"]["O3"].hyperlink.target.endswith(".pdf")
        assert wb["資料索引"]["H3"].value == '3r（3" / C.S / S-40 / DB 3）\n3a（3" / C.S / S-40 / DB 3）\n（共2口）'
        assert wb["資料索引"]["I3"].value == "管材×1"
        assert wb["資料索引"]["H3"].alignment.vertical == "top"
        assert wb["資料索引"]["H3"].alignment.wrap_text is True
        assert wb["資料索引"]["A3"].alignment.horizontal == "center"
        assert wb["資料索引"]["B3"].alignment.horizontal == "center"
        assert wb["資料索引"]["M3"].alignment.horizontal == "center"
        assert wb["資料索引"]["N3"].alignment.horizontal == "center"
        assert wb["資料索引"]["O3"].alignment.horizontal == "center"
        assert wb["焊口統計"]["E3"].value == '3"'
        assert wb["焊口統計"]["A1"].value == "HP6精濾區配管工事-工務修改確認單 - 焊口統計"
        assert [wb["焊口統計"].cell(2, col).value for col in range(1, 12)] == [
            "工務修改確認單編號",
            "日期",
            "ISO流編",
            "焊口編號",
            "尺寸",
            "材質",
            "厚度",
            "新增或修改",
            "新增或修改係數",
            "DB",
            "預算編號",
        ]
        assert wb["焊口統計"]["A3"].value == "R-720"
        assert wb["焊口統計"]["C3"].value == "0720"
        assert wb["焊口統計"]["D3"].value == "3r"
        assert wb["焊口統計"]["H3"].value == "修改"
        assert wb["焊口統計"]["I3"].value == 1.5
        assert wb["焊口統計"]["J3"].value == "3"
        assert wb["焊口統計"]["H4"].value == "新增"
        assert wb["焊口統計"]["I4"].value == 1
        assert wb["資料索引"].row_dimensions[3].height >= 118
    finally:
        wb.close()

    moved_package = tmp_path / "moved_owner_data_report"
    shutil.copytree(package, moved_package)
    moved_index = moved_package / "owner_data_index.xlsx"
    moved_wb = load_workbook(moved_index, data_only=False)
    try:
        ws = moved_wb["資料索引"]
        for coordinate in ("B3", "M3", "N3", "O3"):
            target = ws[coordinate].hyperlink.target
            assert not os.path.isabs(target)
            assert (moved_package / target).exists()
    finally:
        moved_wb.close()

    with ZipFile(index) as archive:
        media = [name for name in archive.namelist() if name.startswith("xl/media/")]
    assert len(media) >= 3


def test_owner_data_report_displays_clean_report_labels_for_folder_ids(tmp_path):
    output = tmp_path / "out"
    report_set = _report_set(tmp_path)
    report_set["reports"][0]["report"]["report_id"] = ""
    report_set["reports"][0]["report"]["folder"] = "107_20260701_01"
    report_set["reports"][0]["report"]["series"] = "0107"
    report_set["reports"][0]["report"]["date_raw"] = "20260701"

    result = build_owner_data_report_package(output, report_set, weld_lookup=None)

    package = Path(result["package_root"])
    index = Path(result["index_xlsx"])
    report_dir = package / "CO-107-20260701-01"
    assert report_dir.is_dir()

    wb = load_workbook(index, data_only=False)
    try:
        ws = wb["資料索引"]
        assert ws["B3"].value == "CO-107-20260701-01"
        assert ws["B3"].hyperlink.target == "CO-107-20260701-01"
        assert "資料夾" not in [ws.cell(2, col).value for col in range(1, ws.max_column + 1)]
    finally:
        wb.close()


def test_owner_data_report_backfills_dwg_and_line_from_drawing_lookup(tmp_path):
    report_set = _report_set(tmp_path)
    report_set["reports"][0]["report"]["series"] = "150"
    report_set["reports"][0]["report"]["dwg_no"] = ""
    report_set["reports"][0]["report"]["line_number"] = ""

    result = build_owner_data_report_package(
        tmp_path / "out",
        report_set,
        weld_lookup=None,
        drawing_lookup={"0150": ("AI-00001", "1-S11U-AI-00001-150")},
    )

    wb = load_workbook(result["index_xlsx"], data_only=False)
    try:
        ws = wb["資料索引"]
        assert ws["D3"].value == "150"
        assert ws["E3"].value == "1-S11U-AI-00001-150"
        assert ws["F3"].value == "AI-00001"
    finally:
        wb.close()


def test_owner_data_report_weld_summary_prefers_lookup_db_value(tmp_path):
    class FakeLookup:
        def lookup_info(self, series, base):
            assert series == "0720"
            return {"db": f"B-{base}", "budget_no": f"PB-{base}"}

    output = tmp_path / "out"

    result = build_owner_data_report_package(output, _report_set(tmp_path), weld_lookup=FakeLookup())

    wb = load_workbook(Path(result["index_xlsx"]), data_only=False)
    try:
        assert wb["資料索引"]["H3"].value == '3r（3" / C.S / S-40 / DB B-3 / 預算 PB-3）\n3a（3" / C.S / S-40 / DB B-3 / 預算 PB-3）\n（共2口）'
        assert wb["焊口統計"]["J3"].value == "B-3"
        assert wb["焊口統計"]["K3"].value == "PB-3"
    finally:
        wb.close()


def test_owner_data_report_weld_sheet_prefers_lookup_db_over_size_inference(tmp_path):
    class FakeLookup:
        def lookup_info(self, series, base):
            assert series == "0720"
            return {"size": "0.5", "sch": "S-40", "material": "SUS304", "db": "1", "budget_no": "1.5"}

    output = tmp_path / "out"
    report_set = _report_set(tmp_path)
    report_set["reports"][0]["welds"]["rows"] = [
        {"code": "3r", "size": "0.5", "material": "C.S", "thickness": "S-40", "mark": "r"},
    ]
    report_set["reports"][0]["welds"]["count"] = 1

    result = build_owner_data_report_package(output, report_set, weld_lookup=FakeLookup())

    wb = load_workbook(Path(result["index_xlsx"]), data_only=False)
    try:
        assert wb["資料索引"]["H3"].value == '3r（0.5" / SUS304 / S-40 / DB 1 / 預算 1.5）\n（共1口）'
        assert wb["焊口統計"]["E3"].value == '0.5"'
        assert wb["焊口統計"]["F3"].value == "SUS304"
        assert wb["焊口統計"]["J3"].value == "1"
        assert wb["焊口統計"]["K3"].value == "1.5"
    finally:
        wb.close()


def test_owner_data_report_lookup_uses_renamed_weld_code_before_stale_base(tmp_path):
    class FakeLookup:
        def lookup_info(self, series, base):
            assert series == "0720"
            assert base == "20"
            return {"size": "1", "sch": "40S", "material": "304L", "db": "1"}

    output = tmp_path / "out"
    report_set = _report_set(tmp_path)
    report_set["reports"][0]["welds"]["rows"] = [{
        "code": "20A",
        "weld_no": "1001",
        "size": "1",
        "material": "304L",
        "thickness": "40S",
        "mark": "a",
    }]
    report_set["reports"][0]["welds"]["count"] = 1

    result = build_owner_data_report_package(output, report_set, weld_lookup=FakeLookup())

    wb = load_workbook(Path(result["index_xlsx"]), data_only=False)
    try:
        assert wb["資料索引"]["H3"].value == '20A（1" / 304L / 40S / DB 1）\n（共1口）'
        assert wb["焊口統計"]["D3"].value == "20A"
        assert wb["焊口統計"]["J3"].value == "1"
    finally:
        wb.close()
