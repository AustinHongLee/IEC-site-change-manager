# -*- coding: utf-8 -*-

import os
import sys

from openpyxl import load_workbook


sys.path.insert(0, os.path.join(os.path.dirname(__file__), os.pardir, "control"))

from demo_smoke import run_demo_edge_matrix, run_demo_output_smoke


def test_demo_output_smoke_creates_repeatable_outputs(tmp_path):
    output_dir = tmp_path / "demo"

    result = run_demo_output_smoke(output_dir, overwrite=False)

    assert result["ok"] is True
    assert result["report_count"] == 1
    assert os.path.exists(result["files"]["report_set"])
    assert os.path.exists(result["files"]["template"])
    assert os.path.exists(result["files"]["pdf_overlay_template"])
    assert os.path.exists(result["files"]["pdf_overlay_base"])
    assert os.path.exists(result["files"]["rendered_xlsx"])
    assert os.path.exists(result["files"]["site_statistics_xlsx"])
    assert result["xlsx_template"]["post_validation"]["ok"] is True
    assert result["pdf_overlay_template"]["ok"] is True
    assert result["xlsx_template"]["summary"]["image"] == 2
    assert result["xlsx_template"]["summary"]["table"] == 2

    wb = load_workbook(result["files"]["rendered_xlsx"], data_only=False)
    try:
        ws = wb["現場修改單"]
        assert ws["A1"].value == "0547_AG"
        assert ws["C1"].value == "0547"
        assert ws["A19"].value
        assert ws["E19"].value
    finally:
        wb.close()


def test_demo_output_smoke_refuses_to_overwrite_unmarked_folder(tmp_path):
    output_dir = tmp_path / "not_demo"
    output_dir.mkdir()
    (output_dir / "important.txt").write_text("keep", encoding="utf-8")

    try:
        run_demo_output_smoke(output_dir, overwrite=True)
    except RuntimeError as exc:
        assert "拒絕覆寫非 demo 資料夾" in str(exc)
    else:
        raise AssertionError("expected RuntimeError")


def test_demo_edge_matrix_creates_expected_failure_cases(tmp_path):
    output_dir = tmp_path / "edge"

    result = run_demo_edge_matrix(output_dir, overwrite=False)

    assert result["ok"] is True
    assert result["report_count"] == 5
    assert os.path.exists(result["files"]["pdf_overlay_rotated_template"])
    assert os.path.exists(result["files"]["pdf_overlay_rotated_base"])
    by_folder = {case["folder"]: case for case in result["cases"]}
    assert by_folder["0601_NO_AFTER"]["expectation_ok"] is True
    assert "missing_image_value" in by_folder["0601_NO_AFTER"]["issue_codes"]
    assert "table_overflow" in by_folder["0602_MATERIAL_OVERFLOW"]["issue_codes"]
    assert "table_overflow" in by_folder["0603_MANY_PHOTOS"]["issue_codes"]
    assert by_folder["0604_MULTI_PAGE_PDF"]["attachment_pdf_pages"] == 2
