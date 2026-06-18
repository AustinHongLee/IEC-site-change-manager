# -*- coding: utf-8 -*-

import json
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


def test_cli_export_site_statistics_creates_workbook(tmp_path):
    repo = Path(__file__).resolve().parents[1]
    output = tmp_path / "site_statistics.xlsx"

    result = subprocess.run(
        [
            sys.executable,
            str(repo / "tools" / "export_site_statistics.py"),
            "--output",
            str(output),
        ],
        cwd=repo,
        text=True,
        encoding="utf-8",
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=False,
    )

    assert result.returncode == 0
    assert "已匯出現場修改統計單" in result.stdout
    wb = load_workbook(output, read_only=True)
    try:
        assert "總覽" in wb.sheetnames
        assert "照片索引" in wb.sheetnames
        assert "照片表" in wb.sheetnames
    finally:
        wb.close()


def test_cli_export_site_statistics_pdf_output_reports_missing_libreoffice(tmp_path):
    repo = Path(__file__).resolve().parents[1]
    output = tmp_path / "site_statistics.xlsx"
    pdf_output = tmp_path / "site_statistics.pdf"
    missing_soffice = tmp_path / "missing_soffice.exe"

    result = subprocess.run(
        [
            sys.executable,
            str(repo / "tools" / "export_site_statistics.py"),
            "--output",
            str(output),
            "--pdf-output",
            str(pdf_output),
            "--soffice",
            str(missing_soffice),
            "--json",
        ],
        cwd=repo,
        text=True,
        encoding="utf-8",
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=False,
    )

    assert result.returncode == 1
    assert output.exists()
    assert not pdf_output.exists()
    data = json.loads(result.stdout)
    assert data["ok"] is False
    assert data["result_schema_version"] == "output_result.v1"
    assert data["xlsx"] == str(output)
    assert data["outputs"][0]["kind"] == "site_statistics_xlsx"
    assert data["pdf_conversion"]["ok"] is False
    assert data["pdf_conversion"]["issues"][0]["code"] == "libreoffice_unavailable"
