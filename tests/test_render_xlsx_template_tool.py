# -*- coding: utf-8 -*-

import json
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


def test_render_xlsx_template_cli_creates_workbook_from_report_set_json(tmp_path):
    repo = Path(__file__).resolve().parents[1]
    template_path = tmp_path / "template.json"
    report_set_path = tmp_path / "report_set.json"
    output = tmp_path / "rendered.xlsx"
    template_path.write_text(
        json.dumps({
            "sheet": "修改單",
            "fields": [
                {"type": "text", "source": "report.report_id", "cell": "A1"},
                {"type": "table", "source": "materials.rows", "start_cell": "A5", "max_rows": 5, "columns": ["component", "qty"]},
            ],
        }, ensure_ascii=False),
        encoding="utf-8",
    )
    report_set_path.write_text(
        json.dumps({
            "reports": [{
                "report": {"report_id": "R-CLI", "folder": "001_1r2"},
                "materials": {"rows": [{"component": "Pipe (管)", "qty": 3}]},
            }]
        }, ensure_ascii=False),
        encoding="utf-8",
    )

    result = subprocess.run(
        [
            sys.executable,
            str(repo / "tools" / "render_xlsx_template.py"),
            str(template_path),
            str(output),
            "--report-set",
            str(report_set_path),
        ],
        cwd=repo,
        text=True,
        encoding="utf-8",
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=False,
    )

    assert result.returncode == 0
    assert "已渲染 xlsx_template" in result.stdout
    assert output.exists()
    wb = load_workbook(output, data_only=False)
    try:
        ws = wb["修改單"]
        assert ws["A1"].value == "R-CLI"
        assert ws["A5"].value == "Pipe (管)"
        assert ws["B5"].value == 3
    finally:
        wb.close()


def test_render_xlsx_template_cli_can_select_report_by_folder(tmp_path):
    repo = Path(__file__).resolve().parents[1]
    template_path = tmp_path / "template.json"
    report_set_path = tmp_path / "report_set.json"
    output = tmp_path / "rendered.xlsx"
    template_path.write_text(
        json.dumps({"fields": [{"type": "text", "source": "report.folder", "cell": "A1"}]}, ensure_ascii=False),
        encoding="utf-8",
    )
    report_set_path.write_text(
        json.dumps({
            "reports": [
                {"report": {"report_id": "R-1", "folder": "first"}},
                {"report": {"report_id": "R-2", "folder": "target_folder"}},
            ]
        }, ensure_ascii=False),
        encoding="utf-8",
    )

    result = subprocess.run(
        [
            sys.executable,
            str(repo / "tools" / "render_xlsx_template.py"),
            str(template_path),
            str(output),
            "--report-set",
            str(report_set_path),
            "--report",
            "target_folder",
        ],
        cwd=repo,
        text=True,
        encoding="utf-8",
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=False,
    )

    assert result.returncode == 0
    wb = load_workbook(output, data_only=False)
    try:
        assert wb.active["A1"].value == "target_folder"
    finally:
        wb.close()


def test_render_xlsx_template_cli_json_includes_post_validation(tmp_path):
    repo = Path(__file__).resolve().parents[1]
    template_path = tmp_path / "template.json"
    report_set_path = tmp_path / "report_set.json"
    output = tmp_path / "rendered.xlsx"
    template_path.write_text(
        json.dumps({
            "fields": [{"type": "text", "source": "report.report_id", "cell": "A1"}],
        }, ensure_ascii=False),
        encoding="utf-8",
    )
    report_set_path.write_text(
        json.dumps({"reports": [{"report": {"report_id": "R-JSON"}}]}, ensure_ascii=False),
        encoding="utf-8",
    )

    result = subprocess.run(
        [
            sys.executable,
            str(repo / "tools" / "render_xlsx_template.py"),
            str(template_path),
            str(output),
            "--report-set",
            str(report_set_path),
            "--json",
        ],
        cwd=repo,
        text=True,
        encoding="utf-8",
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=False,
    )

    assert result.returncode == 0
    data = json.loads(result.stdout)
    assert data["result_schema_version"] == "output_result.v1"
    assert data["outputs"][0]["kind"] == "xlsx_template"
    assert data["layout_validation"]["ok"] is True
    assert data["post_validation"]["ok"] is True
    assert data["post_validation"]["checked"]["text"] == 1


def test_render_xlsx_template_cli_pdf_output_reports_missing_libreoffice(tmp_path):
    repo = Path(__file__).resolve().parents[1]
    template_path = tmp_path / "template.json"
    report_set_path = tmp_path / "report_set.json"
    output = tmp_path / "rendered.xlsx"
    pdf_output = tmp_path / "rendered.pdf"
    missing_soffice = tmp_path / "missing_soffice.exe"
    template_path.write_text(
        json.dumps({
            "fields": [{"type": "text", "source": "report.report_id", "cell": "A1"}],
        }, ensure_ascii=False),
        encoding="utf-8",
    )
    report_set_path.write_text(
        json.dumps({"reports": [{"report": {"report_id": "R-PDF"}}]}, ensure_ascii=False),
        encoding="utf-8",
    )

    result = subprocess.run(
        [
            sys.executable,
            str(repo / "tools" / "render_xlsx_template.py"),
            str(template_path),
            str(output),
            "--report-set",
            str(report_set_path),
            "--pdf-output",
            str(pdf_output),
            "--soffice",
            str(missing_soffice),
            "--json",
        ],
        cwd=repo,
        text=True,
        encoding="utf-8",
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=False,
    )

    assert result.returncode == 1
    assert output.exists()
    assert not pdf_output.exists()
    data = json.loads(result.stdout)
    assert data["ok"] is False
    assert data["result_schema_version"] == "output_result.v1"
    assert data["outputs"][0]["kind"] == "xlsx_template"
    assert data["post_validation"]["ok"] is True
    assert data["pdf_conversion"]["ok"] is False
    assert data["pdf_conversion"]["issues"][0]["code"] == "libreoffice_unavailable"
    assert data["issues"][-1]["code"] == "libreoffice_unavailable"
    assert any(step["key"] == "workbook_pdf_conversion" for step in data["steps"])
