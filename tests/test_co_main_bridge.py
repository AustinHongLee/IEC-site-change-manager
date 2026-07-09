import os
import sys
import json
import base64

from PIL import Image
from openpyxl import Workbook, load_workbook
from pypdf import PdfWriter

sys.path.insert(0, os.path.join(os.path.dirname(__file__), os.pardir, "control"))

from co_main_bridge import MainBridge  # noqa: E402
import co_main_bridge  # noqa: E402
from material_taxonomy import material_family, normalize_material, normalize_schedule, normalize_size  # noqa: E402


def _write(root, name, obj):
    rec = root / "records"
    rec.mkdir(exist_ok=True)
    (rec / name).write_text(json.dumps(obj, ensure_ascii=False), encoding="utf-8")


def _write_root_json(root, name, obj):
    (root / name).write_text(json.dumps(obj, ensure_ascii=False), encoding="utf-8")


def _write_pdf(path):
    writer = PdfWriter()
    writer.add_blank_page(width=595, height=842)
    with open(path, "wb") as f:
        writer.write(f)


def _write_workbook(path, sheet_name, headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(path)


def test_info_envelope(tmp_path):
    res = MainBridge(tmp_path).info()
    assert set(res) >= {"ok", "data", "error"}
    assert res["ok"] is True and res["data"]["api_version"]


def test_pricebook_maps_real_columns(tmp_path):
    _write(tmp_path, "material_pricebook.json", {"items": [
        {"id": 'Pipe|2"', "零件類型": "Pipe (管)", "尺寸": '2"', "SCH": "SCH 40",
         "材質": "白鐵 (Stainless Steel)", "類別": "材料", "單位": "M",
         "單價": "100", "來源": "合約", "生效日": "", "備註": "ok"},
        {"id": "valve|3", "零件類型": "Valve", "尺寸": '3"', "SCH": "150#",
         "材質": "CS", "類別": "閥件", "單位": "EA", "單價": "", "來源": "", "生效日": "", "備註": ""},
    ]})
    res = MainBridge(tmp_path).pricebook()
    assert res["ok"] is True
    rows = res["data"]
    assert len(rows) == 2
    r0 = rows[0]
    assert r0["part"] == "Pipe (管)" and r0["size"] == "DN50" and r0["sch"] == "SCH40"
    assert r0["mat"] == "白鐵 (STAINLESS STEEL)" and r0["cat"] == "管材" and r0["unit"] == "M"
    assert r0["icon"] == "Pipe" and r0["material_family"] == "白鐵系"
    assert r0["src"] == "合約" and r0["remark"] == "ok"
    assert "price" not in r0              # 已去價：橋不再回傳單價


def test_pricebook_missing_file_returns_empty(tmp_path):
    res = MainBridge(tmp_path).pricebook()
    assert res["ok"] is True and res["data"] == []


def test_material_taxonomy_api_exposes_axes_and_fasteners(tmp_path):
    res = MainBridge(tmp_path).material_taxonomy()

    assert res["ok"] is True
    data = res["data"]
    sizes = data["axes"]["nominal_diameter"]["values"]
    assert sizes[0] == "DN15" and sizes[-1] == "DN500"
    assert {"SUS304", "SUS304L", "SUS316", "SUS321"} <= set(data["axes"]["material_grade"]["values"])
    assert {"螺栓", "螺帽", "華司", "U型螺栓"} <= set(data["options"]["cat"])
    assert any(p["code"] == "bolt" and "螺絲" in p["aliases"] for p in data["part_types"])


def test_pricebook_enriches_taxonomy_fields(tmp_path):
    _write(tmp_path, "material_pricebook.json", {"items": [
        {"id": "b1", "零件類型": "STUD BOLT", "尺寸": "", "SCH": "", "材質": "HDG",
         "類別": "", "單位": "", "來源": "", "備註": ""},
        {"id": "p1", "零件類型": "PIPE", "尺寸": '2"', "SCH": "S-40", "材質": "A106-B",
         "類別": "配管", "單位": "M", "來源": "", "備註": ""},
        {"id": "TYPE07-SLIDE-150X150X3T-AS", "零件類型": "Type07 滑板", "尺寸": "150x150x3t",
         "SCH": "", "材質": "A36/SS400", "類別": "滑板", "單位": "片", "來源": "",
         "備註": "", "Type": "Type07", "支撐級別": "零件", "規格": "Type07,Type07 滑板,150x150x3t,A36/SS400"},
    ]})

    rows = MainBridge(tmp_path).pricebook()["data"]

    assert rows[0]["cat"] == "螺栓"
    assert rows[0]["icon"] == "BoltNut"
    assert rows[0]["material_family"] == "鍍鋅系"
    assert rows[1]["part"] == "PIPE"
    assert rows[1]["size"] == "DN50"
    assert rows[1]["sch"] == "SCH40"
    assert rows[1]["mat"] == "A106 GR.B"
    assert rows[1]["cat"] == "管材"
    assert rows[1]["material_family"] == "黑鐵系"
    assert rows[2]["type"] == "Type07"
    assert rows[2]["level"] == "零件"
    assert rows[2]["spec"] == "Type07,Type07 滑板,150x150x3t,A36/SS400"


def test_material_catalog_rules_query_and_register_without_full_pricebook(tmp_path):
    bridge = MainBridge(tmp_path)

    summary = bridge.material_catalog_summary()
    assert summary["ok"] is True
    assert summary["data"]["total"] > 0
    assert summary["data"]["counts"]["Pipe"] > 0

    queried = bridge.material_catalog_query({"icon": "Pipe", "mat": "SUS304L"}, 0, 5)
    assert queried["ok"] is True
    assert queried["data"]["items"]
    first = queried["data"]["items"][0]
    assert first["part"] == "鋼管"
    assert first["mat"] == "SUS304L"
    assert "price" not in first

    reg = bridge.register_parts([first["id"]])
    assert reg["ok"] is True
    assert first["id"] in reg["data"]["registered"]

    parts = bridge.project_parts()
    assert parts["ok"] is True
    assert first["id"] in parts["data"]["registered"]
    assert any(row["id"] == first["id"] and row["mat"] == "SUS304L" for row in parts["data"]["items"])


def test_material_catalog_query_and_build_dual_dimension_part(tmp_path):
    bridge = MainBridge(tmp_path)

    queried = bridge.material_catalog_query({
        "part": "同心大小頭",
        "size1": "DN50",
        "size2": "DN25",
        "sch": "SCH40",
        "mat": "SUS304",
    }, 0, 5)

    assert queried["ok"] is True
    assert queried["data"]["total"] == 1
    row = queried["data"]["items"][0]
    assert row["part"] == "同心大小頭"
    assert row["size1"] == "DN50"
    assert row["size2"] == "DN25"
    assert row["size"] == "DN50xDN25"

    built = bridge.build_project_part({
        "part": "同心大小頭",
        "size1": "DN50",
        "size2": "DN25",
        "sch": "SCH40",
        "mat": "SUS304",
    })

    assert built["ok"] is True
    item = built["data"]["item"]
    assert item["id"] == row["id"]
    assert item["size1"] == "DN50" and item["size2"] == "DN25"
    assert item["id"] in built["data"]["registered"]
    parts = bridge.project_parts()["data"]["items"]
    assert any(p["id"] == item["id"] and p["size1"] == "DN50" and p["size2"] == "DN25" for p in parts)


def test_material_catalog_build_rejects_invalid_reducer_same_size(tmp_path):
    res = MainBridge(tmp_path).build_project_part({
        "part": "同心大小頭",
        "size1": "DN50",
        "size2": "DN50",
        "sch": "SCH40",
        "mat": "SUS304",
    })

    assert res["ok"] is False
    assert "尺寸1不可等於尺寸2" in res["error"]


def test_material_normalizers_cover_common_site_aliases():
    assert normalize_size('2"') == "DN50"
    assert normalize_size("DN80x15") == "DN80xDN15"
    assert normalize_size("L50*50*6") == "L50x50x6"
    assert normalize_size("M16x50") == "M16x50"
    assert normalize_size("1219x2438x12t") == "1219x2438x12t"
    assert normalize_schedule("S-40") == "SCH40"
    assert normalize_schedule("150LB") == "150#"
    assert normalize_schedule("CLASS150") == "150#"
    assert normalize_material("A182-F304.") == "A182 F304"
    assert normalize_material("304") == "SUS304"
    assert normalize_material("304L") == "SUS304L"
    assert normalize_material("316") == "SUS316"
    assert normalize_material("321") == "SUS321"
    assert material_family("304L") == "白鐵系"
    assert material_family("A234 GR.WPB") == "黑鐵系"
    assert material_family("A36/SS400") == "黑鐵系"


def test_records_missing_attachments(tmp_path):
    assert MainBridge(tmp_path).records()["data"] == []


def test_weld_table_setting_round_trips_to_settings_json(tmp_path):
    _write_root_json(tmp_path, "settings.json", {
        "paths": {
            "weld_control_table": "old.xlsx",
            "drawing_list": "dwg.xlsx",
            "prefab_drawing_dir": "old-pdfs",
        },
        "meta": {"version": "test"},
    })
    b = MainBridge(tmp_path)

    settings = b.app_settings()
    assert settings["ok"] is True
    assert settings["data"]["weld_control_table"] == "old.xlsx"
    assert settings["data"]["prefab_drawing_dir"] == "old-pdfs"

    res = b.save_setting("weld_control_table", "new-control.xlsx")
    assert res["ok"] is True
    saved = json.loads((tmp_path / "settings.json").read_text(encoding="utf-8"))
    assert saved["paths"]["weld_control_table"] == "new-control.xlsx"
    assert json.loads((tmp_path / "records" / "app_settings.json").read_text(encoding="utf-8"))[
        "weld_control_table"
    ] == "new-control.xlsx"

    res = b.save_setting("prefab_drawing_dir", "drawing-pdfs")
    assert res["ok"] is True
    saved = json.loads((tmp_path / "settings.json").read_text(encoding="utf-8"))
    assert saved["paths"]["prefab_drawing_dir"] == "drawing-pdfs"
    assert json.loads((tmp_path / "records" / "app_settings.json").read_text(encoding="utf-8"))[
        "prefab_drawing_dir"
    ] == "drawing-pdfs"

    res = b.save_setting("project_name", "測試專案")
    assert res["ok"] is True
    saved = json.loads((tmp_path / "settings.json").read_text(encoding="utf-8"))
    assert saved["project"]["name"] == "測試專案"
    assert json.loads((tmp_path / "records" / "app_settings.json").read_text(encoding="utf-8"))[
        "project_name"
    ] == "測試專案"
    assert b.app_settings()["data"]["project_name"] == "測試專案"


def test_app_settings_reports_source_health_counts(tmp_path):
    dwg = tmp_path / "dwg.xlsx"
    weld = tmp_path / "weld.xlsx"
    pdf_dir = tmp_path / "pdfs"
    pdf_dir.mkdir()
    _write_workbook(dwg, "DRAWING LIST", ["NO", "DWG NO"], [["100", "A-100"], ["101", "A-101"], ["", "SKIP"]])
    _write_workbook(weld, "焊口編號明細", ["流水號", "焊口編號", "DB數"], [["100", "1a", 0.5], ["100", "2r", 0.75], ["", "x", 1]])
    _write_pdf(pdf_dir / "100-ISO.pdf")
    _write_pdf(pdf_dir / "999-ISO.pdf")
    folder = tmp_path / "attachments" / "100_2026-07-01_01"
    folder.mkdir(parents=True)
    (folder / "change_order.json").write_text(json.dumps({
        "id": "100_2026-07-01_01",
        "date": "2026-07-01",
        "series": "100",
        "status": "完整",
        "welds": [],
    }, ensure_ascii=False), encoding="utf-8")
    _write_root_json(tmp_path, "settings.json", {
        "paths": {
            "drawing_list": str(dwg),
            "weld_control_table": str(weld),
            "prefab_drawing_dir": str(pdf_dir),
        },
        "dwg_list": {"sheet_name": "DRAWING LIST", "col_serial": "NO"},
        "weld_control": {"sheet_name": "焊口編號明細", "col_serial": "流水號", "col_weld_no": "焊口編號"},
    })

    data = MainBridge(tmp_path).app_settings()["data"]
    health = data["source_health"]
    assert health["dwg"]["count"] == 2
    assert health["weld"]["count"] == 2
    assert health["drawingpdf"]["total"] == 2
    assert health["drawingpdf"]["matched"] == 1
    assert health["drawingpdf"]["partial"] is False


def test_pdf_source_health_scans_nested_folder_with_limit(tmp_path, monkeypatch):
    pdf_dir = tmp_path / "pdfs"
    nested = pdf_dir / "nested"
    nested.mkdir(parents=True)
    _write_pdf(nested / "100-ISO.pdf")
    _write_pdf(pdf_dir / "101-ISO.pdf")
    folder = tmp_path / "attachments" / "100_2026-07-01_01"
    folder.mkdir(parents=True)
    (folder / "change_order.json").write_text(json.dumps({
        "id": "100_2026-07-01_01",
        "date": "2026-07-01",
        "series": "100",
        "status": "完整",
        "welds": [],
    }, ensure_ascii=False), encoding="utf-8")
    _write_root_json(tmp_path, "settings.json", {"paths": {"prefab_drawing_dir": str(pdf_dir)}})

    health = MainBridge(tmp_path).app_settings()["data"]["source_health"]["drawingpdf"]

    assert health["matched"] == 1
    assert health["total"] == 2
    assert health["scanned_dirs"] == 2

    monkeypatch.setattr(co_main_bridge, "PDF_HEALTH_SCAN_FILE_LIMIT", 1)
    limited = MainBridge(tmp_path).app_settings()["data"]["source_health"]["drawingpdf"]
    assert limited["partial"] is True
    assert limited["label"] == "部分讀取"
    assert limited["scan_limit"] == 1


def test_app_settings_auto_detects_shifted_source_headers_and_aliases(tmp_path):
    dwg = tmp_path / "dwg_new_format.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "DWG NO.ALL"
    ws.append(["新版圖號清單"])
    ws.append([])
    ws.append(["圖號", "管線序號"])
    ws.append(["A-100", "100"])
    ws.append(["A-101", "101"])
    wb.save(dwg)
    _write_root_json(tmp_path, "settings.json", {
        "paths": {"drawing_list": str(dwg)},
        "dwg_list": {"sheet_name": "DRAWING LIST", "col_serial": "NO"},
    })

    health = MainBridge(tmp_path).app_settings()["data"]["source_health"]["dwg"]
    assert health["state"] == "ok"
    assert health["count"] == 2
    assert health["sheet"] == "DWG NO.ALL"
    assert health["header_row"] == 3
    assert health["fields"][0]["actual"] == "管線序號"
    assert "自動辨識" in health["message"]


def test_save_source_schema_updates_settings_json(tmp_path):
    dwg = tmp_path / "dwg.xlsx"
    _write_workbook(dwg, "NEW SHEET", ["圖號", "管線序號"], [["A-100", "100"]])
    _write_root_json(tmp_path, "settings.json", {
        "paths": {"drawing_list": str(dwg)},
        "dwg_list": {"sheet_name": "DRAWING LIST", "col_serial": "NO"},
    })
    bridge = MainBridge(tmp_path)

    res = bridge.save_source_schema("dwg", {"sheet_name": "NEW SHEET", "serial_column": "管線序號"})
    assert res["ok"] is True
    saved = json.loads((tmp_path / "settings.json").read_text(encoding="utf-8"))
    assert saved["dwg_list"]["sheet_name"] == "NEW SHEET"
    assert saved["dwg_list"]["col_serial"] == "管線序號"
    health = bridge.app_settings()["data"]["source_health"]["dwg"]
    assert health["state"] == "ok"
    assert health["count"] == 1
    assert "設定值" not in health["message"]

    res = bridge.save_source_schema("weld", {
        "sheet_name": "焊口",
        "serial_column": "ISO流編",
        "weld_no_column": "焊口碼",
        "size_column": "口徑",
        "material_column": "鋼種",
        "thickness_column": "管厚",
        "db_column": "DB值",
        "budget_no_column": "預算碼",
    })
    assert res["ok"] is True
    saved = json.loads((tmp_path / "settings.json").read_text(encoding="utf-8"))
    assert saved["weld_control"]["sheet_name"] == "焊口"
    assert saved["weld_control"]["col_serial"] == "ISO流編"
    assert saved["weld_control"]["col_weld_no"] == "焊口碼"
    assert saved["weld_control"]["col_size"] == "口徑"
    assert saved["weld_control"]["col_material"] == "鋼種"
    assert saved["weld_control"]["col_thickness"] == "管厚"
    assert saved["weld_control"]["col_db"] == "DB值"
    assert saved["weld_control"]["col_budget_no"] == "預算碼"


def test_source_excel_preview_maps_workbook_grid_and_detected_fields(tmp_path):
    weld = tmp_path / "weld_preview.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "新版焊口表"
    ws.append(["說明列"])
    ws.append(["流水號", "銲口編號", "尺寸", "材質", "厚度", "DB數", "預算編號"])
    ws.append(["100", "1a", "0.5", "304L", "40S", 0.5, "PB-1"])
    ws.append(["100", "2r", "0.75", "316", "80S", 0.75, "PB-2"])
    wb.save(weld)
    _write_root_json(tmp_path, "settings.json", {
        "paths": {"weld_control_table": str(weld)},
        "weld_control": {"sheet_name": "焊口編號明細", "col_serial": "流水號", "col_weld_no": "焊口編號"},
    })

    res = MainBridge(tmp_path).source_excel_preview("weld")
    assert res["ok"] is True
    data = res["data"]
    assert data["sheet"] == "新版焊口表"
    assert data["header_row"] == 2
    assert data["columns"][0]["letter"] == "A"
    assert data["rows"][1]["is_header"] is True
    assert data["fields"][0]["actual"] == "流水號"
    assert data["fields"][1]["actual"] == "銲口編號"
    fields = {field["key"]: field for field in data["fields"]}
    assert fields["size_column"]["actual"] == "尺寸"
    assert fields["material_column"]["actual"] == "材質"
    assert fields["thickness_column"]["actual"] == "厚度"
    assert fields["db_column"]["actual"] == "DB數"
    assert fields["budget_no_column"]["actual"] == "預算編號"
    assert any(role["key"] == "budget_no_column" for role in data["roles"])


def test_records_reads_change_orders(tmp_path):
    # 精靈出單：attachments/<id>/change_order.json → records() 映成前端記錄
    folder = tmp_path / "attachments" / "55_2026-07-01_01"
    folder.mkdir(parents=True)
    (folder / "before_1.jpg").write_bytes(b"jpg")
    (folder / "change_order.json").write_text(json.dumps({
        "id": "55_2026-07-01_01", "date": "2026-07-01", "series": "55",
        "status": "待補", "reason": "現場改管",
        "welds": [{"code": "1001", "op": "裁切",
                   "spec": {"size": "DN50", "sch": "SCH40", "material": "A106 GR.B"}}],
        "materials": [{"component_id": "0002", "component": "90°彎頭", "size": "DN50",
                       "schedule": "SCH40", "material": "A234 GR.WPB", "qty": 2,
                       "unit": "個", "remark": ""}],
        "photos": [{"role": "before", "file": "before_1.jpg"}],
    }, ensure_ascii=False), encoding="utf-8")
    res = MainBridge(tmp_path).records()
    assert res["ok"] is True and len(res["data"]) == 1
    r = res["data"][0]
    assert r["series"] == "55" and r["status"] == "pending" and r["reason"] == "現場改管"
    assert r["welds"][0] == {"code": "1001", "mark": "重焊", "size": "DN50",
                             "mat": "A106 GR.B", "sch": "SCH40", "coef": ""}
    assert r["mats"][0]["id"] == "0002"
    assert r["mats"][0]["part"] == "90°彎頭" and r["mats"][0]["qty"] == 2
    assert r["mats"][0]["unit"] == "個"
    assert r["photos"][0]["label"] == "修改前"
    assert r["photos"][0]["src"] == str((folder / "before_1.jpg").resolve())
    bill = MainBridge(tmp_path).billing()
    assert bill["ok"] and len(bill["data"]["rows"]) == 1
    assert bill["data"]["rows"][0]["status"] == "未請款"


def test_dates_include_change_order_drafts(tmp_path):
    folder = tmp_path / "attachments" / "55_2026-07-01_01"
    folder.mkdir(parents=True)
    (folder / "change_order.json").write_text(json.dumps({
        "id": "55_2026-07-01_01", "date": "2026-07-01", "series": "55",
        "status": "完整",
        "welds": [{"code": "1001", "spec": {}}, {"code": "1002", "spec": {}}],
    }, ensure_ascii=False), encoding="utf-8")

    res = MainBridge(tmp_path).dates()

    assert res["ok"] is True
    assert res["data"][0]["date"] == "20260701"
    item = res["data"][0]["items"][0]
    assert item["series"] == "55"
    assert item["welds"] == 2
    assert item["weld_codes"] == ["1001", "1002"]
    assert item["status"] == "done"
    assert item["record_id"] == "55_2026-07-01_01"
    assert item["source"] == "change_order"


def test_save_record_updates_change_order_json(tmp_path):
    folder = tmp_path / "attachments" / "55_20260701_01"
    folder.mkdir(parents=True)
    (folder / "change_order.json").write_text(json.dumps({
        "id": "55_20260701_01", "date": "20260701", "series": "55",
        "status": "完整",
        "welds": [{"code": "1001", "op": "加長", "spec": {"size": "DN50", "sch": "SCH40", "material": "CS"}}],
        "materials": [],
    }, ensure_ascii=False), encoding="utf-8")
    rec = MainBridge(tmp_path).records()["data"][0]
    rec["welds"][0]["size"] = "DN80"
    rec["welds"][0]["mat"] = "A106 GR.B"
    rec["mats"].append({
        "id": "0002", "part": "90°彎頭", "size": "DN80", "sch": "SCH40",
        "mat": "A234 WPB", "qty": "2", "unit": "個", "remark": "現場追加",
    })

    res = MainBridge(tmp_path).save_record(rec)

    assert res["ok"] is True
    saved = json.loads((folder / "change_order.json").read_text(encoding="utf-8"))
    assert saved["welds"][0]["spec"]["size"] == "DN80"
    assert saved["welds"][0]["op"] == "重焊"
    assert saved["welds"][0]["base"] == "1001"
    assert saved["welds"][0]["origin"] == "existing"
    assert saved["welds"][0]["spec"]["material"] == "A106 GR.B"
    assert saved["materials"][0]["component_id"] == "0002"
    assert saved["materials"][0]["qty"] == 2
    assert saved["audit"]["history"][-1]["action"] == "saved_from_main"
    assert res["data"]["mats"][0]["unit"] == "個"


def test_save_record_recomputes_weld_identity_when_code_is_renamed(tmp_path):
    folder = tmp_path / "attachments" / "150_20260709_01"
    folder.mkdir(parents=True)
    (folder / "change_order.json").write_text(json.dumps({
        "id": "150_20260709_01",
        "date": "20260709",
        "series": "150",
        "status": "完整",
        "welds": [{
            "code": "1001",
            "base": None,
            "origin": "new",
            "op": "新焊",
            "spec": {"size": "1", "sch": "40S", "material": "304L"},
        }],
        "materials": [],
    }, ensure_ascii=False), encoding="utf-8")
    rec = MainBridge(tmp_path).records()["data"][0]
    rec["welds"][0]["code"] = "20A"
    rec["welds"][0]["mark"] = "新焊"

    res = MainBridge(tmp_path).save_record(rec)

    assert res["ok"] is True
    saved = json.loads((folder / "change_order.json").read_text(encoding="utf-8"))
    assert saved["welds"][0]["code"] == "20A"
    assert saved["welds"][0]["base"] is None
    assert saved["welds"][0]["origin"] == "new"
    assert saved["welds"][0]["op"] == "新焊"
    assert res["data"]["welds"][0]["code"] == "20A"
    assert res["data"]["welds"][0]["mark"] == "新焊"


def test_open_record_folder_and_pdf_use_injected_opener(tmp_path):
    folder = tmp_path / "attachments" / "55_20260701_01"
    folder.mkdir(parents=True)
    (folder / "drawing.pdf").write_bytes(b"%PDF-1.4")
    (folder / "change_order.json").write_text(json.dumps({
        "id": "55_20260701_01",
        "drawing_pdf": {"file": "drawing.pdf"},
    }, ensure_ascii=False), encoding="utf-8")
    opened = []
    b = MainBridge(tmp_path)
    b._open_path_fn = opened.append

    folder_res = b.open_record_folder("55_20260701_01")
    pdf_res = b.open_record_pdf("55_20260701_01")

    assert folder_res["ok"] and folder_res["data"]["path"] == str(folder)
    assert pdf_res["ok"] and pdf_res["data"]["path"] == str(folder / "drawing.pdf")
    assert opened == [str(folder), str(folder / "drawing.pdf")]


def test_open_wizard_uses_script_launcher_in_source_mode(tmp_path, monkeypatch):
    app = tmp_path / "control" / "co_wizard_app.py"
    app.parent.mkdir()
    app.write_text("# smoke\n", encoding="utf-8")
    calls = []

    monkeypatch.setattr(co_main_bridge.sys, "frozen", False, raising=False)
    monkeypatch.setattr(co_main_bridge.sys, "executable", "python.exe")
    monkeypatch.setattr(co_main_bridge.subprocess, "Popen", lambda command, cwd=None: calls.append((command, cwd)))

    res = MainBridge(tmp_path).open_wizard()

    assert res["ok"] is True
    assert res["data"]["mode"] == "source"
    assert calls == [(["python.exe", str(app)], str(tmp_path))]


def test_open_wizard_uses_wizard_arg_when_frozen(tmp_path, monkeypatch):
    calls = []

    monkeypatch.setattr(co_main_bridge.sys, "frozen", True, raising=False)
    monkeypatch.setattr(co_main_bridge.sys, "executable", r"C:\dist\IEC-site-change-manager.exe")
    monkeypatch.setattr(co_main_bridge.subprocess, "Popen", lambda command, cwd=None: calls.append((command, cwd)))

    res = MainBridge(tmp_path).open_wizard()

    assert res["ok"] is True
    assert res["data"]["mode"] == "frozen"
    assert calls == [([r"C:\dist\IEC-site-change-manager.exe", "--wizard"], str(tmp_path))]


def test_export_records_and_material_summary(tmp_path):
    folder = tmp_path / "attachments" / "55_20260701_01"
    folder.mkdir(parents=True)
    (folder / "change_order.json").write_text(json.dumps({
        "id": "55_20260701_01", "date": "20260701", "series": "55",
        "welds": [{"code": "1001", "spec": {}}],
        "materials": [{"component_id": "0002", "component": "90°彎頭", "qty": 2, "unit": "個"}],
    }, ensure_ascii=False), encoding="utf-8")
    b = MainBridge(tmp_path)
    records_path = tmp_path / "records.xlsx"
    mats_path = tmp_path / "materials.xlsx"

    rec_res = b.export_records(str(records_path))
    mat_res = b.export_record_materials(str(mats_path))

    assert rec_res["ok"] and rec_res["data"]["count"] == 1 and records_path.is_file()
    assert mat_res["ok"] and mat_res["data"]["count"] == 1 and mats_path.is_file()


def test_export_output_center_owner_data_package_from_selected_items(tmp_path):
    folder = tmp_path / "attachments" / "20260702" / "720_3r2_3a2"
    folder.mkdir(parents=True)
    (folder / "GroupWeld.txt").write_text("3r3\n3a3\n", encoding="utf-8")
    (folder / "note.txt").write_text("因現場管線干涉，切除原焊口並新增焊口。", encoding="utf-8")
    Image.new("RGB", (80, 160), (220, 80, 60)).save(folder / "before_1.jpg")
    Image.new("RGB", (180, 80), (80, 160, 90)).save(folder / "after_1.jpg")
    _write_pdf(folder / "720.CA-2007-100-AA1B-NA-1.pdf")

    res = MainBridge(tmp_path).export_output_center(
        "owner-data",
        [{"date": "20260702", "folder": "720_3r2_3a2", "series": "720"}],
    )

    assert res["ok"] is True
    files = res["data"]["files"]
    assert files["statistics_xlsx"] == ""
    assert files["owner_data_package"].endswith("owner_data_report")
    assert files["owner_data_index_xlsx"].endswith("owner_data_index.xlsx")
    assert (tmp_path / "staging" / "site_output_center_web" / "owner_data_report" / "CO-720-3r2-3a2").is_dir()


def test_export_output_center_owner_data_package_from_root_change_order_folder(tmp_path):
    _write_root_json(tmp_path, "settings.json", {"project": {"name": "測試工程"}})
    folder = tmp_path / "attachments" / "107_20260701_01"
    folder.mkdir(parents=True)
    Image.new("RGB", (80, 160), (220, 80, 60)).save(folder / "before_1.jpg")
    Image.new("RGB", (180, 80), (80, 160, 90)).save(folder / "after_1.jpg")
    _write_pdf(folder / "drawing.pdf")
    (folder / "change_order.json").write_text(json.dumps({
        "schema_version": "0.2",
        "id": "107_20260701_01",
        "status": "完整",
        "date": "20260701",
        "series": "107",
        "welds": [{
            "origin": "existing",
            "op": "重焊",
            "base": "3",
            "code": "3a",
            "spec": {"size": "3", "sch": "S-40", "material": "C.S", "weld_type": "BW"},
        }],
        "materials": [{
            "component": "支撐管",
            "size": "DN40",
            "schedule": "SCH80",
            "material": "SUS304",
            "qty": 1,
            "unit": "米",
        }],
        "photos": [{"role": "before", "file": "before_1.jpg"}, {"role": "after", "file": "after_1.jpg"}],
        "drawing_pdf": {"file": "drawing.pdf"},
        "reason": "因管線干涉新增支撐。",
    }, ensure_ascii=False), encoding="utf-8")

    res = MainBridge(tmp_path).export_output_center(
        "owner-data",
        [{"date": "20260701", "folder": str(folder), "record_id": "107_20260701_01", "series": "107"}],
    )

    assert res["ok"] is True
    files = res["data"]["files"]
    report_set = json.loads(open(files["report_set"], encoding="utf-8").read())
    report = report_set["reports"][0]
    assert report_set["project"]["name"] == "測試工程"
    assert report["report"]["folder"] == "107_20260701_01"
    assert report["report"]["status"] != "missing_folder"
    assert report["welds"]["count"] == 1
    assert report["welds"]["summary"] == "3a（3 / C.S / S-40）（共1口）"
    assert report["materials"]["count"] == 1
    assert report["photos"]["has_before"] is True
    assert report["photos"]["has_after"] is True
    assert report["attachment_pdf"]["exists"] is True
    assert (tmp_path / "staging" / "site_output_center_web" / "owner_data_report" / "CO-107-20260701-01").is_dir()
    wb = load_workbook(files["owner_data_index_xlsx"], read_only=True)
    try:
        assert wb["資料索引"]["A1"].value == "測試工程-工務修改確認單"
        assert wb["焊口統計"]["A1"].value == "測試工程-工務修改確認單 - 焊口統計"
    finally:
        wb.close()


def test_save_billing_status_round_trips(tmp_path):
    folder = tmp_path / "attachments" / "55_20260701_01"
    folder.mkdir(parents=True)
    (folder / "change_order.json").write_text(json.dumps({
        "id": "55_20260701_01", "date": "20260701", "series": "55",
    }, ensure_ascii=False), encoding="utf-8")
    b = MainBridge(tmp_path)

    res = b.save_billing([{"rec": 0, "status": "已請款", "billDate": "2026-07-01", "remark": "ok"}])
    bill = MainBridge(tmp_path).billing()

    assert res["ok"] and res["data"]["count"] == 1
    assert bill["data"]["rows"][0]["status"] == "已請款"
    assert bill["data"]["rows"][0]["billDate"] == "2026-07-01"
    assert bill["data"]["rows"][0]["remark"] == "ok"


def test_main_bridge_json_writes_leave_no_temp_files(tmp_path):
    folder = tmp_path / "attachments" / "55_20260701_01"
    folder.mkdir(parents=True)
    (folder / "change_order.json").write_text(json.dumps({
        "id": "55_20260701_01",
        "date": "20260701",
        "series": "55",
    }, ensure_ascii=False), encoding="utf-8")
    _write(tmp_path, "material_pricebook.json", {"items": [{"id": "PIPE-DN15-S40-CS"}]})

    bridge = MainBridge(tmp_path)
    assert bridge.save_setting("project_name", "測試工程")["ok"] is True
    assert bridge.save_billing([{"rec": 0, "status": "未請款"}])["ok"] is True
    assert bridge.register_parts(["PIPE-DN15-S40-CS"])["ok"] is True

    assert not list(tmp_path.rglob("*.tmp"))
    assert json.loads((tmp_path / "settings.json").read_text(encoding="utf-8"))["project"]["name"] == "測試工程"
    assert json.loads((tmp_path / "records" / "billing.json").read_text(encoding="utf-8"))["byId"]["55_20260701_01"]["status"] == "未請款"
    assert json.loads((tmp_path / "records" / "project_parts.json").read_text(encoding="utf-8"))["registered"] == ["PIPE-DN15-S40-CS"]


def test_replace_photo_copies_file_and_updates_change_order(tmp_path):
    folder = tmp_path / "attachments" / "55_20260701_01"
    folder.mkdir(parents=True)
    (folder / "before_1.jpg").write_bytes(b"old")
    (folder / "change_order.json").write_text(json.dumps({
        "id": "55_20260701_01",
        "photos": [{"role": "before", "file": "before_1.jpg"}],
    }, ensure_ascii=False), encoding="utf-8")
    replacement = tmp_path / "replacement.png"
    replacement.write_bytes(b"new")

    res = MainBridge(tmp_path).replace_photo("55_20260701_01", 0, str(replacement))

    assert res["ok"] is True
    data = json.loads((folder / "change_order.json").read_text(encoding="utf-8"))
    assert data["photos"][0]["file"] == "before_1.png"
    assert (folder / "before_1.png").read_bytes() == b"new"
    assert res["data"]["photo"]["label"] == "修改前"
    assert res["data"]["photo"]["src"] == str((folder / "before_1.png").resolve())


def test_save_photo_annotation_writes_png_and_updates_change_order(tmp_path):
    folder = tmp_path / "attachments" / "55_20260701_01"
    folder.mkdir(parents=True)
    (folder / "after_1.jpg").write_bytes(b"old")
    (folder / "change_order.json").write_text(json.dumps({
        "id": "55_20260701_01",
        "photos": [{"role": "after", "file": "after_1.jpg"}],
    }, ensure_ascii=False), encoding="utf-8")
    raw = b"\x89PNG\r\n\x1a\nannotated"
    data_url = "data:image/png;base64," + base64.b64encode(raw).decode("ascii")

    res = MainBridge(tmp_path).save_photo_annotation("55_20260701_01", 0, data_url)

    assert res["ok"] is True
    data = json.loads((folder / "change_order.json").read_text(encoding="utf-8"))
    assert data["photos"][0]["file"] == "after_1_annotated.png"
    assert (folder / "after_1_annotated.png").read_bytes() == raw
    assert res["data"]["photo"]["label"] == "修改後"


def test_image_data_url_accepts_record_photo_file_url(tmp_path):
    photo = tmp_path / "photo.jpg"
    photo.write_bytes(b"jpg-bytes")

    res = MainBridge(tmp_path).image_data_url(photo.as_uri())

    assert res["ok"] is True
    assert res["data"]["url"].startswith("data:image/jpeg;base64,")
    assert base64.b64decode(res["data"]["url"].split(",", 1)[1]) == b"jpg-bytes"


def test_project_parts_register_unregister(tmp_path):
    _write(tmp_path, "material_pricebook.json", {"items": [{"id": "0001"}, {"id": "0005"}]})
    b = MainBridge(tmp_path)
    assert b.project_parts()["data"]["registered"] == []          # 起始空
    r = b.register_parts(["0001", "0005", "0001"])                # 去重
    assert r["ok"] and set(r["data"]["registered"]) == {"0001", "0005"}
    r2 = b.unregister_parts(["0001"])
    assert r2["data"]["registered"] == ["0005"]
    # 換一個 bridge 實例仍讀得到 → 已持久化到 project_parts.json
    assert MainBridge(tmp_path).project_parts()["data"]["registered"] == ["0005"]


def test_project_parts_prunes_stale_ids_after_catalog_regeneration(tmp_path):
    _write(tmp_path, "material_pricebook.json", {"items": [{"id": "PIPE-DN15-S40-CS"}]})
    _write(tmp_path, "project_parts.json", {"registered": ["0001", "PIPE-DN15-S40-CS"]})

    res = MainBridge(tmp_path).project_parts()

    assert res["ok"] is True
    assert res["data"]["registered"] == ["PIPE-DN15-S40-CS"]
    assert res["data"]["dropped"] == ["0001"]
    saved = json.loads((tmp_path / "records" / "project_parts.json").read_text(encoding="utf-8"))
    assert saved["registered"] == ["PIPE-DN15-S40-CS"]


def test_upsert_project_parts_registers_custom_materials_and_pricebook(tmp_path):
    _write(tmp_path, "material_pricebook.json", {"items": [{"id": "PIPE-DN15-S40-CS", "零件類型": "PIPE"}]})
    b = MainBridge(tmp_path)
    support_pipe = {
        "id": "SUPPORT-01-2B-03C-01",
        "part": "支撐管",
        "size": "DN40",
        "sch": "SCH80",
        "mat": "SUS304",
        "qty": 0.171,
        "unit": "米",
        "cat": "管材",
        "type": "Type01",
        "level": "管架展開",
        "spec": '1-1/2"*SCH.80',
        "source_designation": "01-2B-03C",
        "remark": "from support calculator",
    }

    res = b.upsert_project_parts([support_pipe])

    assert res["ok"] is True
    assert res["data"]["added"] == ["SUPPORT-01-2B-03C-01"]
    project = b.project_parts()["data"]
    assert project["registered"] == ["SUPPORT-01-2B-03C-01"]
    assert project["custom"][0]["id"] == "SUPPORT-01-2B-03C-01"
    assert project["custom"][0]["project_only"] is True
    rows = b.pricebook()["data"]
    custom = next(r for r in rows if r["id"] == "SUPPORT-01-2B-03C-01")
    assert custom["part"] == "支撐管"
    assert custom["size"] == "DN40"
    assert custom["sch"] == "SCH80"
    assert custom["mat"] == "SUS304"
    assert custom["project_only"] is True
    assert custom["source_designation"] == "01-2B-03C"

    out = tmp_path / "project_parts.xlsx"
    export = b.export_project_parts(str(out))
    assert export["ok"] is True
    assert export["data"]["count"] == 1
    assert out.is_file()


def test_support_bom_expands_external_type_calculator_to_material_rows(tmp_path, monkeypatch):
    app = tmp_path / "for_iec_support" / "python_app"
    core = app / "core"
    core.mkdir(parents=True)
    (core / "__init__.py").write_text("", encoding="utf-8")
    (core / "calculator.py").write_text(
        """
class Entry:
    def __init__(self, item_no, name, spec, length, width, material, quantity, unit, category, role='', remark=''):
        self.item_no = item_no
        self.name = name
        self.spec = spec
        self.length = length
        self.width = width
        self.material = material
        self.quantity = quantity
        self.unit = unit
        self.category = category
        self.role = role
        self.remark = remark
        self.weight_output = 1.25 * item_no
        self.unit_weight = 1.25
        self.length_subtotal = length / 1000 if role == 'pipe' else 0
        self.qty_subtotal = quantity if role != 'pipe' else 0
        self.part_key = ''
        self.stock_id = ''
        self.item_class = ''
        self.manufacturing_type = 'raw_cut' if role == 'pipe' else 'plate_cut'

    @property
    def display_spec(self):
        return self.spec

    @property
    def display_remark(self):
        return self.remark


class Result:
    def __init__(self, fullstring):
        self.fullstring = fullstring
        self.error = ''
        self.warnings = []
        self.meta = {}
        self.entries = [
            Entry(1, '管路', '1-1/2"*SCH.80', 171, 0, 'SUS304', 1, 'M', '管路類', 'pipe'),
            Entry(2, 'Plate_a_無鑽孔', '9', 150, 150, 'A36/SS400', 1, 'PC', '鋼板類', 'base_plate'),
        ]

    @property
    def total_weight(self):
        return sum(e.weight_output for e in self.entries)


def analyze_single(fullstring, overrides=None):
    return Result(fullstring)
""",
        encoding="utf-8",
    )
    monkeypatch.setenv("CO_SUPPORT_APP_DIR", str(app))

    res = MainBridge(tmp_path).support_bom("01-2b-03c")

    assert res["ok"] is True
    data = res["data"]
    assert data["designation"] == "01-2B-03C"
    assert data["type"] == "01"
    assert data["total_weight"] == 3.75
    assert len(data["entries"]) == 2
    pipe, plate = data["materials"]
    assert pipe["part"] == "支撐管"
    assert pipe["size"] == "DN40"
    assert pipe["sch"] == "SCH80"
    assert pipe["qty"] == 0.171
    assert pipe["unit"] == "米"
    assert plate["cat"] == "底板"
    assert plate["size"] == "150x150x9t"
    assert plate["unit"] == "片"
    assert plate["type"] == "Type01"
