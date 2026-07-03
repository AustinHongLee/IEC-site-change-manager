# -*- coding: utf-8 -*-

import os
import sys
from zipfile import ZipFile

from openpyxl import load_workbook
import pytest


sys.path.insert(0, os.path.join(os.path.dirname(__file__), os.pardir, "control"))

from site_statistics_exporter import (
    build_overview_rows,
    export_site_statistics_workbook,
    flatten_material_summary_rows,
    flatten_photo_index_rows,
    flatten_report_rows,
    flatten_weld_summary_rows,
)


def _report_set():
    return {
        "schema_version": "report_set.v1",
        "project": {
            "root": "C:/project",
            "attachments_root": "C:/project/attachments",
            "collected_at": "2026-06-16T16:00:00",
        },
        "reports": [{
            "report": {
                "report_id": "R-1",
                "date": "2026-06-16",
                "series": "001",
                "folder": "001_1r2",
                "status": "unproduced",
                "change_type": "裁切重焊",
                "description": "現場修改",
                "folder_path": "C:/project/attachments/20260616/001_1r2",
            },
            "welds": {
                "summary": "1r（共1口）",
                "count": 1,
                "rows": [{"code": "1r", "size": 2, "weld_no": "1", "mark": "r"}],
            },
            "materials": {
                "count": 1,
                "rows": [{
                    "component": "Pipe (管)",
                    "size": '2"',
                    "sch": "SCH 40",
                    "material": "白鐵 (Stainless Steel)",
                    "category": "材料",
                    "qty": 3,
                    "unit": "M",
                }],
            },
            "photos": {
                "before": [{"name": "before.jpg", "path": "C:/p/before.jpg", "w": 100, "h": 80}],
                "after": [{"name": "after.jpg", "path": "C:/p/after.jpg", "w": 100, "h": 80}],
            },
            "attachment_pdf": {"exists": True},
            "completeness": {"level": "complete", "missing": []},
        }],
        "aggregates": {
            "report_count": 1,
            "weld_count": 1,
            "material_row_count": 1,
            "before_photo_count": 1,
            "after_photo_count": 1,
            "photo_count": 2,
            "status_counts": {"unproduced": 1},
            "completeness_counts": {"complete": 1},
        },
        "issues": [],
    }


def test_flatten_helpers_build_site_statistics_rows():
    report_set = _report_set()

    assert ["修改單總數", 1, ""] in build_overview_rows(report_set)
    assert flatten_report_rows(report_set)[0][0] == "R-1"
    assert flatten_weld_summary_rows(report_set)[0][0] == "2"
    assert len(flatten_photo_index_rows(report_set)) == 2
    assert flatten_material_summary_rows(report_set)[0][0] == "Pipe (管)"


def test_export_site_statistics_workbook_creates_expected_sheets(tmp_path):
    output = tmp_path / "site_statistics.xlsx"

    export_site_statistics_workbook(str(output), report_set=_report_set())

    wb = load_workbook(output, data_only=False)
    try:
        assert wb.sheetnames == [
            "目錄",
            "開發_資料總覽",
            "開發_修改單原始清單",
            "開發_照片索引",
            "開發_問題清單",
            "報告_總覽",
            "報告_修改單清單",
            "報告_焊口統計",
            "報告_用料統計",
            "報告_照片表",
        ]
        assert wb["目錄"]["A1"].value == "現場修改統計單 - 工作簿目錄"
        assert wb["目錄"]["A8"].hyperlink.target == "#'開發_資料總覽'!A1"
        assert wb["開發_資料總覽"]["A1"].value == "開發者檢查 - 資料總覽"
        assert wb["開發_修改單原始清單"]["A3"].value == "R-1"
        assert wb["開發_修改單原始清單"]["P3"].value == "C:/project/attachments/20260616/001_1r2"
        assert wb["開發_照片索引"]["E3"].value == "before"
        assert wb["報告_總覽"]["A1"].value == "報告 - 總覽"
        assert wb["報告_修改單清單"]["A3"].value == "R-1"
        assert wb["報告_照片表"]["A3"].value == "R-1"
        assert wb["報告_照片表"]["F3"].value == "找不到圖片"
        assert wb["報告_照片表"].max_column == 8
        assert wb["報告_用料統計"]["A3"].value == "Pipe (管)"
    finally:
        wb.close()


def test_export_site_statistics_workbook_embeds_before_after_images(tmp_path):
    pil_image = pytest.importorskip("PIL.Image")
    before = tmp_path / "before.png"
    after = tmp_path / "after.png"
    pil_image.new("RGB", (120, 80), (180, 40, 40)).save(before)
    pil_image.new("RGB", (80, 120), (40, 90, 180)).save(after)
    report_set = _report_set()
    report_set["reports"][0]["photos"] = {
        "before": [{"name": before.name, "path": str(before), "w": 120, "h": 80}],
        "after": [{"name": after.name, "path": str(after), "w": 80, "h": 120}],
    }
    output = tmp_path / "site_statistics.xlsx"

    export_site_statistics_workbook(str(output), report_set=report_set)

    wb = load_workbook(output, data_only=False)
    try:
        assert wb["報告_照片表"]["A3"].value == "R-1"
        assert wb["報告_照片表"].max_column == 8
        assert wb["開發_照片索引"]["J3"].value == str(before)
        assert wb["開發_照片索引"]["J4"].value == str(after)
    finally:
        wb.close()
    with ZipFile(output) as archive:
        embedded_images = [name for name in archive.namelist() if name.startswith("xl/media/")]
    assert len(embedded_images) == 2
