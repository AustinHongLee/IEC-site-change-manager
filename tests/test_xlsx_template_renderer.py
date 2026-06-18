# -*- coding: utf-8 -*-

import os
import sys
from zipfile import ZipFile

from openpyxl import load_workbook
import pytest


sys.path.insert(0, os.path.join(os.path.dirname(__file__), os.pardir, "control"))

from xlsx_template_renderer import (
    render_xlsx_template_for_report,
    validate_rendered_xlsx_workbook,
    validate_xlsx_template_layout,
)


def _report(before_path: str = ""):
    return {
        "report": {
            "report_id": "R-1",
            "date": "2026-06-17",
            "description": "現場修改",
        },
        "welds": {
            "rows": [
                {"code": "1r2", "size": 2},
                {"code": "2r2", "size": 2},
            ],
        },
        "photos": {
            "before": [{"path": before_path}],
        },
    }


def test_render_xlsx_template_writes_text_image_and_table(tmp_path):
    pil_image = pytest.importorskip("PIL.Image")
    before = tmp_path / "before.png"
    pil_image.new("RGB", (120, 80), (180, 40, 40)).save(before)
    output = tmp_path / "rendered.xlsx"
    template = {
        "sheet": "修改單",
        "fields": [
            {"type": "text", "source": "report.report_id", "cell": "A1"},
            {"type": "text", "source": "report.description", "cell": "A2"},
            {"type": "image", "source": "photos.before[0].path", "anchor": "D2", "max_width_px": 80},
            {"type": "table", "source": "welds.rows", "start_cell": "A10", "max_rows": 5, "columns": ["code", "size"]},
        ],
    }

    result = render_xlsx_template_for_report(_report(str(before)), template, str(output))

    assert result["ok"] is True
    assert result["summary"] == {"text": 2, "image": 1, "table": 1, "rows": 2}
    assert result["post_validation"]["ok"] is True
    assert result["post_validation"]["checked"] == {"text": 2, "image": 1, "table": 1, "rows": 2}
    wb = load_workbook(output, data_only=False)
    try:
        ws = wb["修改單"]
        assert ws["A1"].value == "R-1"
        assert ws["A2"].value == "現場修改"
        assert ws["A10"].value == "1r2"
        assert ws["B11"].value == 2
    finally:
        wb.close()
    with ZipFile(output) as archive:
        assert len([name for name in archive.namelist() if name.startswith("xl/media/")]) == 1


def test_render_xlsx_template_uses_existing_workbook_and_missing_image_placeholder(tmp_path):
    base = tmp_path / "base.xlsx"
    output = tmp_path / "rendered.xlsx"
    template_dir = tmp_path
    from openpyxl import Workbook
    wb = Workbook()
    wb.active.title = "既有格式"
    wb.active["A5"] = "保留"
    wb.save(base)
    wb.close()
    template = {
        "workbook": "base.xlsx",
        "sheet": "既有格式",
        "fields": [
            {"type": "text", "source": "report.report_id", "cell": "B1"},
            {"type": "image", "source": "photos.before[0].path", "anchor": "C3"},
        ],
    }

    result = render_xlsx_template_for_report(_report("C:/missing/before.jpg"), template, str(output), template_dir=template_dir)

    assert result["ok"] is True
    assert result["post_validation"]["ok"] is True
    assert any(issue["code"] == "missing_image_file" for issue in result["issues"])
    wb2 = load_workbook(output, data_only=False)
    try:
        ws = wb2["既有格式"]
        assert ws["A5"].value == "保留"
        assert ws["B1"].value == "R-1"
        assert ws["C3"].value == "找不到圖片"
    finally:
        wb2.close()


def test_render_xlsx_template_rejects_invalid_mapping_without_file(tmp_path):
    output = tmp_path / "bad.xlsx"

    result = render_xlsx_template_for_report(
        _report(),
        {"fields": [{"type": "text", "source": "report.bad_field", "cell": "A1"}]},
        str(output),
    )

    assert result["ok"] is False
    assert not output.exists()


def test_render_xlsx_template_rejects_table_overflow_without_overwriting_footer(tmp_path):
    output = tmp_path / "overflow.xlsx"
    template = {
        "sheet": "修改單",
        "fields": [
            {"type": "text", "source": "report.report_id", "cell": "A1"},
            {"type": "table", "source": "welds.rows", "start_cell": "A3", "max_rows": 1, "columns": ["code"]},
        ],
    }

    result = render_xlsx_template_for_report(_report(), template, str(output))

    assert result["ok"] is False
    assert any(issue["code"] == "table_overflow" for issue in result["issues"])
    assert not output.exists()


def test_validate_rendered_xlsx_workbook_detects_text_and_table_mismatch(tmp_path):
    output = tmp_path / "rendered.xlsx"
    template = {
        "sheet": "修改單",
        "fields": [
            {"type": "text", "source": "report.report_id", "cell": "A1"},
            {"type": "table", "source": "welds.rows", "start_cell": "A10", "max_rows": 5, "columns": ["code", "size"]},
        ],
    }
    render_xlsx_template_for_report(_report(), template, str(output))
    wb = load_workbook(output, data_only=False)
    try:
        ws = wb["修改單"]
        ws["A1"] = "BROKEN"
        ws["B10"] = 999
        wb.save(output)
    finally:
        wb.close()

    validation = validate_rendered_xlsx_workbook(str(output), _report(), template)

    assert validation["ok"] is False
    codes = {issue["code"] for issue in validation["issues"]}
    assert "post_validation_text_mismatch" in codes
    assert "post_validation_table_mismatch" in codes


def test_validate_xlsx_template_layout_detects_overlap():
    template = {
        "fields": [
            {"type": "image", "source": "photos.before[0].path", "anchor": "B2", "size_cells": [2, 4]},
            {"type": "table", "source": "welds.rows", "start_cell": "A3", "max_rows": 5, "columns": ["code", "size"]},
        ],
    }

    result = validate_xlsx_template_layout(_report(), template)

    assert result["ok"] is False
    assert any(issue["code"] == "layout_overlap" for issue in result["issues"])


def test_validate_xlsx_template_layout_detects_out_of_bounds():
    template = {
        "fields": [
            {"type": "table", "source": "welds.rows", "start_cell": "XFD1048576", "max_rows": 2, "columns": ["code"]},
        ],
    }

    result = validate_xlsx_template_layout(_report(), template)

    assert result["ok"] is False
    assert any(issue["code"] == "layout_out_of_bounds" for issue in result["issues"])


def test_render_xlsx_template_rejects_layout_overlap_without_file(tmp_path):
    output = tmp_path / "overlap.xlsx"
    template = {
        "fields": [
            {"type": "text", "source": "report.report_id", "cell": "A1"},
            {"type": "image", "source": "photos.before[0].path", "anchor": "A1", "size_cells": [1, 1]},
        ],
    }

    result = render_xlsx_template_for_report(_report(), template, str(output))

    assert result["ok"] is False
    assert result["layout_validation"]["ok"] is False
    assert any(issue["code"] == "layout_overlap" for issue in result["issues"])
    assert not output.exists()
